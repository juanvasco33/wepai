"""
test_excel_country.py — v15.5

Tests parametrizados por país × template para el refactor de soporte LATAM
de excel_controller. Cubre los 7 builders críticos refactorizados con cfg:

    payroll · quotation · sales · purchase_order · freelance_tracker
    income_statement · income_expense_book

y verifica que:
  • el archivo se genera sin error
  • el contenido del archivo refleja el país (etiquetas, tasas, moneda)
  • los builders no-críticos también funcionan cuando reciben cfg (no se rompen)

Espejo del enfoque de tests/test_word_builders.py (v15.4): smoke parametrizado.
"""
import os
import sys
import tempfile
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
os.environ.setdefault("ANTHROPIC_API_KEY", "dummy")

from openpyxl import load_workbook
from office import excel_controller as ec
from office.legal_config import EXCEL_FISCAL, LEGAL_CONFIG


# Subconjunto representativo (1 LATAM grande + 1 Andino + 1 Cono Sur + 1 Centroamérica + ES/US/GENERIC)
COUNTRIES = ["México", "Colombia", "Perú", "Argentina", "Chile",
             "España", "United States", "Panamá", "República Dominicana"]

# Los 7 builders refactorizados con cfg
COUNTRY_AWARE_TEMPLATES = [
    "payroll", "quotation", "sales", "purchase_order",
    "freelance_tracker", "income_statement", "income_expense_book",
]

# Builders no-críticos: deben seguir generando sin romperse aunque ahora
# reciban cfg implícito vía _dispatch
LEGACY_TEMPLATES = [
    "inventory", "accounts_receivable", "cashflow", "multi_currency",
    "dashboard", "budget", "break_even", "amortization", "balance_sheet",
    "client_directory", "supplier_directory", "employee_directory",
]


# ─── Smoke por país × template crítico ────────────────────────────────────────
@pytest.mark.parametrize("template_id", COUNTRY_AWARE_TEMPLATES)
@pytest.mark.parametrize("pais", COUNTRIES)
def test_country_aware_template_generates(tmp_path, template_id, pais):
    """Cada builder refactorizado genera sin error para cada país."""
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": template_id,
        "titulo": f"Test_{template_id}_{pais}",
        "datos": {"pais": pais},
    })
    assert os.path.exists(fp)
    assert os.path.getsize(fp) > 5000, "archivo demasiado pequeño — probable problema"
    wb = load_workbook(fp)
    assert len(wb.sheetnames) >= 1


# ─── Tasas IVA correctas por país ─────────────────────────────────────────────
@pytest.mark.parametrize("pais,iva_esperado", [
    ("México", 0.16),
    ("Colombia", 0.19),
    ("Perú", 0.18),
    ("Argentina", 0.21),
    ("Chile", 0.19),
    ("España", 0.21),
    ("Panamá", 0.07),
    ("República Dominicana", 0.18),
    ("United States", 0.00),
])
def test_quotation_uses_country_iva_rate(tmp_path, pais, iva_esperado):
    """quotation: la tasa default en Config debe ser la del país."""
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": "quotation", "titulo": "T",
        "datos": {"pais": pais},
    })
    wb = load_workbook(fp)
    ws = wb["Config"]
    iva_value = ws["B4"].value  # primera celda de configuración
    assert abs(iva_value - iva_esperado) < 1e-9, (
        f"{pais}: esperado IVA {iva_esperado}, encontrado {iva_value}"
    )


# ─── Etiqueta de impuesto correcta (IVA / IGV / ITBMS / ITBIS / Sales Tax) ────
@pytest.mark.parametrize("pais,label_esperado", [
    ("México", "IVA"),
    ("Colombia", "IVA"),
    ("Perú", "IGV"),
    ("Panamá", "ITBMS"),
    ("República Dominicana", "ITBIS"),
    ("Honduras", "ISV"),
    ("United States", "Sales Tax"),
])
def test_sales_uses_country_tax_label(tmp_path, pais, label_esperado):
    """sales: el header de la columna G debe usar la etiqueta del país."""
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": "sales", "titulo": "T",
        "datos": {"pais": pais},
    })
    wb = load_workbook(fp)
    ws = wb["Ventas"]
    header_g = ws["G4"].value
    assert label_esperado in header_g, (
        f"{pais}: esperado '{label_esperado}' en '{header_g}'"
    )


# ─── Deducciones de nómina específicas por país ───────────────────────────────
@pytest.mark.parametrize("pais,deduccion_esperada", [
    ("México",    "IMSS"),
    ("México",    "INFONAVIT"),
    ("Colombia",  "Salud EPS"),
    ("Colombia",  "Pensión AFP"),
    ("Perú",      "AFP / ONP"),
    ("Argentina", "Jubilación SIPA"),
    ("Argentina", "PAMI / INSSJP"),
    ("Chile",     "AFP"),
    ("Chile",     "Salud Isapre/Fonasa"),
    ("España",    "Seguridad Social"),
    ("Panamá",    "CSS aporte personal"),
])
def test_payroll_includes_country_deduction(tmp_path, pais, deduccion_esperada):
    """payroll: cada país debe incluir sus instituciones de deducción específicas."""
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": "payroll", "titulo": "T",
        "datos": {"pais": pais},
    })
    wb = load_workbook(fp)
    # En la hoja Parámetros, columna A son las labels de deducciones
    ws = wb["Parámetros"]
    labels = [ws.cell(r, 1).value for r in range(4, 15) if ws.cell(r, 1).value]
    assert deduccion_esperada in labels, (
        f"{pais}: esperaba '{deduccion_esperada}' en deducciones, encontré {labels}"
    )


# ─── Tabla LISR solo aparece para México ──────────────────────────────────────
def test_payroll_has_lisr_table_only_for_mexico(tmp_path):
    ec.OUTPUT_DIR = str(tmp_path)
    fp_mx, _ = ec.generate_excel({
        "template_id": "payroll", "titulo": "PayrollMX", "datos": {"pais": "México"}
    })
    fp_co, _ = ec.generate_excel({
        "template_id": "payroll", "titulo": "PayrollCO", "datos": {"pais": "Colombia"}
    })
    assert "Tabla ISR" in load_workbook(fp_mx).sheetnames
    assert "Tabla ISR" not in load_workbook(fp_co).sheetnames


# ─── income_expense_book: estructura MX vs general ────────────────────────────
def test_income_expense_book_mexico_specific_layout(tmp_path):
    ec.OUTPUT_DIR = str(tmp_path)
    fp_mx, _ = ec.generate_excel({
        "template_id": "income_expense_book", "titulo": "IEB_MX",
        "datos": {"pais": "México"},
    })
    fp_co, _ = ec.generate_excel({
        "template_id": "income_expense_book", "titulo": "IEB_CO",
        "datos": {"pais": "Colombia"},
    })
    wb_mx = load_workbook(fp_mx)
    wb_co = load_workbook(fp_co)

    # México tiene RESICO Calc; Colombia no
    assert "RESICO Calc" in wb_mx.sheetnames
    assert "RESICO Calc" not in wb_co.sheetnames

    # Headers diferentes
    mx_headers = [wb_mx["Registro"].cell(4, c).value for c in range(1, 11)]
    co_headers = [wb_co["Registro"].cell(4, c).value for c in range(1, 11)]
    assert "UUID CFDI" in mx_headers
    assert "RFC Tercero" in mx_headers
    assert "UUID CFDI" not in co_headers
    assert "Folio / Factura" in co_headers


# ─── income_statement: ISR label local ────────────────────────────────────────
@pytest.mark.parametrize("pais,fragmento_esperado", [
    ("México",    "ISR"),
    ("Argentina", "GANANCIAS"),
    ("España",    "SOCIEDADES"),
    ("United States", "FEDERAL INCOME TAX"),
])
def test_income_statement_uses_local_tax_label(tmp_path, pais, fragmento_esperado):
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": "income_statement", "titulo": "T",
        "datos": {"pais": pais},
    })
    wb = load_workbook(fp)
    ws = wb["Estado de Resultados"]
    # Buscar la fila "UTILIDAD ANTES DE …" en columna A
    found = False
    for r in range(20, 32):
        v = ws.cell(r, 1).value
        if v and "UTILIDAD ANTES DE" in str(v) and fragmento_esperado in str(v):
            found = True
            break
    assert found, f"{pais}: no se encontró '{fragmento_esperado}' en filas 20-32"


# ─── Builders no-críticos siguen funcionando con cfg ──────────────────────────
@pytest.mark.parametrize("template_id", LEGACY_TEMPLATES)
def test_legacy_template_still_works_with_cfg(tmp_path, template_id):
    """Los builders no-críticos no aceptan cfg pero el dispatcher debe seguir
    pasándoles solo lo que su firma soporta. No deben romperse."""
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": template_id, "titulo": "T",
        "datos": {"pais": "Colombia"},  # país no afecta a legacy
    })
    assert os.path.exists(fp)
    assert os.path.getsize(fp) > 4000


# ─── Detección de país desde keywords sin campo "pais" explícito ──────────────
@pytest.mark.parametrize("keyword,iva_esperado", [
    ("Vendí productos en CDMX, México",  0.16),
    ("Cliente en Bogotá, Colombia",       0.19),
    ("Cotización para Lima, Perú",         0.18),
    ("Empresa en Buenos Aires con CUIT",  0.21),
])
def test_country_detection_from_text(tmp_path, keyword, iva_esperado):
    """Cuando no se especifica país explícito, debe detectarse del texto."""
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": "quotation", "titulo": "T",
        "instrucciones": keyword,
        "datos": {},
    })
    wb = load_workbook(fp)
    iva_value = wb["Config"]["B4"].value
    assert abs(iva_value - iva_esperado) < 1e-9, (
        f"keyword='{keyword}': esperado {iva_esperado}, obtuvo {iva_value}"
    )


# ─── Consistencia entre Word y Excel ──────────────────────────────────────────
def test_word_and_excel_share_legal_config():
    """Word y Excel deben leer del mismo LEGAL_CONFIG."""
    from office import word_controller as wc
    assert wc.LEGAL_CONFIG is LEGAL_CONFIG, (
        "Word y Excel deben compartir la misma instancia de LEGAL_CONFIG"
    )
    # Verificar que tienen el mismo set de países
    assert set(wc.LEGAL_CONFIG.keys()) == set(EXCEL_FISCAL.keys())


# ─── Fallback a GENERIC si país no se reconoce ────────────────────────────────
def test_unknown_country_falls_back_to_generic(tmp_path):
    ec.OUTPUT_DIR = str(tmp_path)
    fp, _ = ec.generate_excel({
        "template_id": "quotation", "titulo": "T",
        "datos": {"pais": "Atlantida-Imaginaria"},  # país inexistente
    })
    # No debe crashear
    wb = load_workbook(fp)
    assert "Config" in wb.sheetnames
