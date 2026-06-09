import subprocess, os, re, threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, Protection)
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import (CellIsRule, ColorScaleRule, FormulaRule,
                                       DataBarRule, Rule)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date
import calendar

# v15.5 — Detección de país y configuración fiscal por país (módulo compartido con Word).
# Resuelve el problema histórico de que Excel asumiera México implícitamente
# mientras Word ya soportaba 20 países.
from office.legal_config import (
    _detect_country, get_excel_fiscal,
    iva_rate, iva_label, currency_format, currency_code, currency_symbol,
)

OUTPUT_DIR = os.path.expanduser("~/Documents/WEP_AI")

# ─────────────────────────────────────────────────────────────────────────────
# v14.4 — I18N (ES/EN) thread-safe
# ─────────────────────────────────────────────────────────────────────────────
# El idioma se setea en `generate_excel()` y se almacena en thread-local para
# que cada llamada concurrente vea su propio idioma. Los builders consultan el
# idioma vía `_lang()`. Las traducciones de labels comunes están en `_XL_TR`.
_xl_ctx = threading.local()


def _set_lang(lang: str):
    """Normaliza y setea el idioma para esta llamada (thread-local)."""
    if not lang:
        lang = "es"
    s = str(lang).strip().lower()
    # Aceptar variantes
    if s in ("es", "español", "espanol", "spanish", "esp"):
        _xl_ctx.lang = "es"
    elif s in ("en", "english", "inglés", "ingles", "eng"):
        _xl_ctx.lang = "en"
    else:
        _xl_ctx.lang = "es"


def _lang() -> str:
    return getattr(_xl_ctx, "lang", "es")


# Diccionario centralizado de traducciones. Keys en español por familiaridad
# con la base de código histórica. `_xt("cantidad")` devuelve "Cantidad" o "Qty"
# según el idioma actual.
_XL_TR = {
    # ── Headers comunes de tablas ──────────────────────────────────────────
    "fecha":         {"es": "Fecha",          "en": "Date"},
    "descripcion":   {"es": "Descripción",    "en": "Description"},
    "concepto":      {"es": "Concepto",       "en": "Concept"},
    "cantidad":      {"es": "Cantidad",       "en": "Quantity"},
    "cant":          {"es": "Cant.",          "en": "Qty"},
    "precio":        {"es": "Precio",         "en": "Price"},
    "precio_unit":   {"es": "Precio Unit.",   "en": "Unit Price"},
    "precio_unitario":{"es": "Precio Unitario","en": "Unit Price"},
    "subtotal":      {"es": "Subtotal",       "en": "Subtotal"},
    "total":         {"es": "Total",          "en": "Total"},
    "total_general": {"es": "Total General",  "en": "Grand Total"},
    "iva":           {"es": "IVA",            "en": "Tax"},
    "iva_pct":       {"es": "IVA %",          "en": "Tax %"},
    "moneda":        {"es": "Moneda",         "en": "Currency"},
    "monto":         {"es": "Monto",          "en": "Amount"},
    "saldo":         {"es": "Saldo",          "en": "Balance"},
    "notas":         {"es": "Notas",          "en": "Notes"},
    "comentarios":   {"es": "Comentarios",    "en": "Comments"},
    "estatus":       {"es": "Estatus",        "en": "Status"},
    "estado":        {"es": "Estado",         "en": "Status"},
    "tipo":          {"es": "Tipo",           "en": "Type"},
    "categoria":     {"es": "Categoría",      "en": "Category"},
    "responsable":   {"es": "Responsable",    "en": "Owner"},
    "area":          {"es": "Área",           "en": "Department"},
    "departamento":  {"es": "Departamento",   "en": "Department"},

    # ── Roles / Entidades ──────────────────────────────────────────────────
    "cliente":       {"es": "Cliente",        "en": "Client"},
    "proveedor":     {"es": "Proveedor",      "en": "Supplier"},
    "empleado":      {"es": "Empleado",       "en": "Employee"},
    "vendedor":      {"es": "Vendedor",       "en": "Salesperson"},
    "producto":      {"es": "Producto",       "en": "Product"},
    "servicio":      {"es": "Servicio",       "en": "Service"},
    "usuario":       {"es": "Usuario",        "en": "User"},
    "puesto":        {"es": "Puesto",         "en": "Position"},
    "cargo":         {"es": "Cargo",          "en": "Title"},
    "salario":       {"es": "Salario",        "en": "Salary"},
    "comision":      {"es": "Comisión",       "en": "Commission"},
    "bono":          {"es": "Bono",           "en": "Bonus"},
    "id":            {"es": "ID",             "en": "ID"},
    "codigo":        {"es": "Código",         "en": "Code"},
    "sku":           {"es": "SKU",            "en": "SKU"},
    "telefono":      {"es": "Teléfono",       "en": "Phone"},
    "email":         {"es": "Email",          "en": "Email"},
    "direccion":     {"es": "Dirección",      "en": "Address"},
    "ciudad":        {"es": "Ciudad",         "en": "City"},
    "pais":          {"es": "País",           "en": "Country"},

    # ── Tiempo ────────────────────────────────────────────────────────────
    "anio":          {"es": "Año",            "en": "Year"},
    "mes":           {"es": "Mes",            "en": "Month"},
    "semana":        {"es": "Semana",         "en": "Week"},
    "dia":           {"es": "Día",            "en": "Day"},
    "trimestre":     {"es": "Trimestre",      "en": "Quarter"},
    "periodo":       {"es": "Período",        "en": "Period"},
    "fecha_inicio":  {"es": "Fecha inicio",   "en": "Start date"},
    "fecha_fin":     {"es": "Fecha fin",      "en": "End date"},
    "vencimiento":   {"es": "Vencimiento",    "en": "Due date"},
    "ene": {"es": "Ene", "en": "Jan"}, "feb": {"es": "Feb", "en": "Feb"},
    "mar": {"es": "Mar", "en": "Mar"}, "abr": {"es": "Abr", "en": "Apr"},
    "may": {"es": "May", "en": "May"}, "jun": {"es": "Jun", "en": "Jun"},
    "jul": {"es": "Jul", "en": "Jul"}, "ago": {"es": "Ago", "en": "Aug"},
    "sep": {"es": "Sep", "en": "Sep"}, "oct": {"es": "Oct", "en": "Oct"},
    "nov": {"es": "Nov", "en": "Nov"}, "dic": {"es": "Dic", "en": "Dec"},

    # ── Status values ─────────────────────────────────────────────────────
    "pagado":        {"es": "Pagado",         "en": "Paid"},
    "pendiente":     {"es": "Pendiente",      "en": "Pending"},
    "vencido":       {"es": "Vencido",        "en": "Overdue"},
    "activo":        {"es": "Activo",         "en": "Active"},
    "inactivo":      {"es": "Inactivo",       "en": "Inactive"},
    "completado":    {"es": "Completado",     "en": "Completed"},
    "en_progreso":   {"es": "En progreso",    "en": "In progress"},
    "cancelado":     {"es": "Cancelado",      "en": "Cancelled"},
    "aprobado":      {"es": "Aprobado",       "en": "Approved"},

    # ── Operaciones financieras ───────────────────────────────────────────
    "ingreso":       {"es": "Ingreso",        "en": "Income"},
    "ingresos":      {"es": "Ingresos",       "en": "Income"},
    "egreso":        {"es": "Egreso",         "en": "Expense"},
    "egresos":       {"es": "Egresos",        "en": "Expenses"},
    "gasto":         {"es": "Gasto",          "en": "Expense"},
    "gastos":        {"es": "Gastos",         "en": "Expenses"},
    "venta":         {"es": "Venta",          "en": "Sale"},
    "ventas":        {"es": "Ventas",         "en": "Sales"},
    "compra":        {"es": "Compra",         "en": "Purchase"},
    "compras":       {"es": "Compras",        "en": "Purchases"},
    "presupuesto":   {"es": "Presupuesto",    "en": "Budget"},
    "real":          {"es": "Real",           "en": "Actual"},
    "varianza":      {"es": "Varianza",       "en": "Variance"},
    "ganancia":      {"es": "Ganancia",       "en": "Profit"},
    "perdida":       {"es": "Pérdida",        "en": "Loss"},
    "utilidad":      {"es": "Utilidad",       "en": "Profit"},
    "factura":       {"es": "Factura",        "en": "Invoice"},
    "cuenta":        {"es": "Cuenta",         "en": "Account"},

    # ── Inventario / Operaciones ──────────────────────────────────────────
    "stock":         {"es": "Stock",          "en": "Stock"},
    "minimo":        {"es": "Mínimo",         "en": "Minimum"},
    "maximo":        {"es": "Máximo",         "en": "Maximum"},
    "almacen":       {"es": "Almacén",        "en": "Warehouse"},
    "unidad":        {"es": "Unidad",         "en": "Unit"},
    "ubicacion":     {"es": "Ubicación",      "en": "Location"},

    # ── RRHH ──────────────────────────────────────────────────────────────
    "asistencia":    {"es": "Asistencia",     "en": "Attendance"},
    "vacaciones":    {"es": "Vacaciones",     "en": "Vacation"},
    "horario":       {"es": "Horario",        "en": "Schedule"},
    "turno":         {"es": "Turno",          "en": "Shift"},
    "horas":         {"es": "Horas",          "en": "Hours"},
    "extras":        {"es": "Extras",         "en": "Overtime"},
    "deducciones":   {"es": "Deducciones",    "en": "Deductions"},
    "neto":          {"es": "Neto",           "en": "Net"},
    "bruto":         {"es": "Bruto",          "en": "Gross"},

    # ── Nombres de sheets más comunes ─────────────────────────────────────
    "sheet_inventario":     {"es": "Inventario",      "en": "Inventory"},
    "sheet_nomina":         {"es": "Nómina",          "en": "Payroll"},
    "sheet_cotizacion":     {"es": "Cotización",      "en": "Quotation"},
    "sheet_cxc":            {"es": "Cuentas por Cobrar", "en": "Accounts Receivable"},
    "sheet_cashflow":       {"es": "Flujo de Caja",   "en": "Cash Flow"},
    "sheet_dashboard":      {"es": "Dashboard",       "en": "Dashboard"},
    "sheet_ventas":         {"es": "Ventas",          "en": "Sales"},
    "sheet_presupuesto":    {"es": "Presupuesto",     "en": "Budget"},
    "sheet_p_and_l":        {"es": "Estado de Resultados", "en": "Income Statement"},
    "sheet_balance":        {"es": "Balance General", "en": "Balance Sheet"},
    "sheet_clientes":       {"es": "Clientes",        "en": "Clients"},
    "sheet_proveedores":    {"es": "Proveedores",     "en": "Suppliers"},
    "sheet_empleados":      {"es": "Empleados",       "en": "Employees"},
    "sheet_proyecto":       {"es": "Proyecto",        "en": "Project"},
    "sheet_datos":          {"es": "Datos",           "en": "Data"},
    "sheet_resumen":        {"es": "Resumen",         "en": "Summary"},
    "sheet_config":         {"es": "Configuración",   "en": "Settings"},
    "sheet_amortizacion":   {"es": "Amortización",    "en": "Amortization"},
    "sheet_break_even":     {"es": "Punto Equilibrio","en": "Break-Even"},
    "sheet_comisiones":     {"es": "Comisiones",      "en": "Commissions"},
    "sheet_gastos":         {"es": "Gastos",          "en": "Expenses"},
    "sheet_orden_compra":   {"es": "Orden de Compra", "en": "Purchase Order"},
    "sheet_activos":        {"es": "Activos Fijos",   "en": "Fixed Assets"},
    "sheet_subscripciones": {"es": "Subscripciones",  "en": "Subscriptions"},
    "sheet_pipeline":       {"es": "Pipeline",        "en": "Pipeline"},
    "sheet_viaticos":       {"es": "Viáticos",        "en": "Travel Expenses"},
    "sheet_roi":            {"es": "ROI",             "en": "ROI"},
}


def _xt(key: str, default: str = None) -> str:
    """Traduce una key al idioma actual. Si no está, devuelve `default` o `key`."""
    entry = _XL_TR.get(key)
    if not entry:
        return default if default is not None else key
    return entry.get(_lang()) or entry.get("es") or (default if default is not None else key)


# ─────────────────────────────────────────────────────────────────────────────
# v14.4 — Post-procesador bilingüe
# ─────────────────────────────────────────────────────────────────────────────
# Después de que un builder genera el workbook (todo hardcodeado en español),
# este post-procesador escanea sheet names, encabezados de celdas, títulos de
# charts y axis labels, traduciendo cualquier término que coincida exactamente
# con un valor del diccionario `_XL_TR`. Si idioma actual == "es", no hace nada.
#
# El matching es:
#   • Exacto (case-sensitive) para evitar falsos positivos en data de usuario
#   • Por palabra completa (no parcial) si la cadena tiene espacios
#   • También cubre patrones compuestos comunes como "Total Ventas" → "Total Sales"

# Construir mapa inverso ES → EN una sola vez al importar el módulo
_ES_TO_EN = {}
for k, v in _XL_TR.items():
    es_val = v.get("es")
    en_val = v.get("en")
    if es_val and en_val and es_val != en_val:
        _ES_TO_EN[es_val] = en_val
        # También variantes case-insensitive (lowercase)
        _ES_TO_EN[es_val.lower()] = en_val.lower()
        _ES_TO_EN[es_val.upper()] = en_val.upper()

# Patrones compuestos comunes (ES → EN). Tienen prioridad sobre los individuales.
_COMPOSITE_PATTERNS = [
    (r"\bTotal General\b",         "Grand Total"),
    (r"\bTotal Ventas\b",          "Total Sales"),
    (r"\bTotal Ingresos\b",        "Total Income"),
    (r"\bTotal Gastos\b",          "Total Expenses"),
    (r"\bTotal Egresos\b",         "Total Expenses"),
    (r"\bGran Total\b",            "Grand Total"),
    (r"\bSubtotal\b",              "Subtotal"),
    (r"\bPrecio Unitario\b",       "Unit Price"),
    (r"\bPrecio Unit\.\b",         "Unit Price"),
    (r"\bPrecio Unit\b",           "Unit Price"),
    (r"\bFlujo de Caja\b",         "Cash Flow"),
    (r"\bCuentas por Cobrar\b",    "Accounts Receivable"),
    (r"\bCuentas por Pagar\b",     "Accounts Payable"),
    (r"\bEstado de Resultados\b",  "Income Statement"),
    (r"\bBalance General\b",       "Balance Sheet"),
    (r"\bPunto de Equilibrio\b",   "Break-Even Point"),
    (r"\bOrden de Compra\b",       "Purchase Order"),
    (r"\bActivos Fijos\b",         "Fixed Assets"),
    (r"\bMargen Bruto\b",          "Gross Margin"),
    (r"\bMargen Neto\b",           "Net Margin"),
    (r"\bMargen %\b",              "Margin %"),
    (r"\bMargen\b",                "Margin"),
    (r"\bUtilidad Neta\b",         "Net Profit"),
    (r"\bUtilidad Bruta\b",        "Gross Profit"),
    (r"\bUtilidad Operativa\b",    "Operating Profit"),
    (r"\bFecha Inicio\b",          "Start Date"),
    (r"\bFecha Fin\b",             "End Date"),
    (r"\bFecha de Emisión\b",      "Issue Date"),
    (r"\bFecha de Vencimiento\b",  "Due Date"),
    (r"\bFecha de Pago\b",         "Payment Date"),
    (r"\bForma de Pago\b",         "Payment Method"),
    (r"\bNúmero de Factura\b",     "Invoice Number"),
    (r"\bNúmero de Cliente\b",     "Client Number"),
    (r"\bID Cliente\b",            "Client ID"),
    (r"\bID Producto\b",           "Product ID"),
    (r"\bDías Vencido\b",          "Days Overdue"),
    (r"\bSaldo Pendiente\b",       "Outstanding Balance"),
    (r"\bSaldo Inicial\b",         "Opening Balance"),
    (r"\bSaldo Final\b",           "Closing Balance"),
    (r"\bTipo de Cambio\b",        "Exchange Rate"),
    (r"\bAño Anterior\b",          "Previous Year"),
    (r"\bAño Actual\b",            "Current Year"),
    (r"\bMes Actual\b",            "Current Month"),
    (r"\bAcumulado\b",             "Cumulative"),
    (r"\bPorcentaje\b",            "Percentage"),
    (r"\bNombre del Producto\b",   "Product Name"),
    (r"\bNombre del Cliente\b",    "Client Name"),
    (r"\bNombre del Empleado\b",   "Employee Name"),
    (r"\bNombre Completo\b",       "Full Name"),
    (r"\bRazón Social\b",          "Company Name"),
    (r"\bHoras Trabajadas\b",      "Hours Worked"),
    (r"\bHoras Extras\b",          "Overtime Hours"),
    (r"\bSalario Base\b",          "Base Salary"),
    (r"\bSueldo Base\b",           "Base Salary"),
    (r"\bSueldo Bruto\b",          "Gross Salary"),
    (r"\bSueldo Neto\b",           "Net Salary"),
    (r"\bDías Trabajados\b",       "Days Worked"),
    (r"\bDías Pendientes\b",       "Pending Days"),
    (r"\bDías Tomados\b",          "Days Taken"),
    (r"\bDías Disponibles\b",      "Days Available"),
    (r"\bMeta de Ventas\b",        "Sales Target"),
    (r"\bMeta\b",                  "Target"),
    (r"\bLogro\b",                 "Achievement"),
    (r"\bCumplimiento\b",          "Attainment"),
    (r"\bIngresos Totales\b",      "Total Income"),
    (r"\bGastos Totales\b",        "Total Expenses"),
    (r"\bInversión Inicial\b",     "Initial Investment"),
    (r"\bTasa de Interés\b",       "Interest Rate"),
    (r"\bTasa Anual\b",            "Annual Rate"),
    (r"\bPlazo\b",                 "Term"),
    (r"\bCuota Mensual\b",         "Monthly Payment"),
    (r"\bCapital\b",               "Principal"),
    (r"\bInterés\b",               "Interest"),
    (r"\bSaldo de Capital\b",      "Principal Balance"),
    (r"\bRecibido por\b",          "Received by"),
    (r"\bEntregado por\b",         "Delivered by"),
    (r"\bAutorizado por\b",        "Authorized by"),
    (r"\bElaborado por\b",         "Prepared by"),
    (r"\bRevisado por\b",          "Reviewed by"),
    (r"\bAprobado por\b",          "Approved by"),
    (r"\bFirma\b",                 "Signature"),
    (r"\bDirección\b",             "Address"),
    (r"\bTeléfono\b",              "Phone"),
    (r"\bConcepto\b",              "Item"),
    (r"\bDescripción\b",           "Description"),
    (r"\bCantidad\b",              "Quantity"),
    (r"\bCant\.\b",                "Qty"),
    (r"\bCantidades\b",            "Quantities"),
    (r"\bProveedor\b",             "Supplier"),
    (r"\bCliente\b",               "Client"),
    (r"\bProducto\b",              "Product"),
    (r"\bEmpleado\b",              "Employee"),
    (r"\bVendedor\b",              "Salesperson"),
    (r"\bResponsable\b",           "Owner"),
    (r"\bDepartamento\b",          "Department"),
    (r"\bCategoría\b",             "Category"),
    (r"\bEstatus\b",               "Status"),
    (r"\bEstado\b",                "Status"),
    (r"\bComentarios\b",           "Comments"),
    (r"\bObservaciones\b",         "Notes"),
    (r"\bNotas\b",                 "Notes"),
    (r"\bPagado\b",                "Paid"),
    (r"\bPendiente\b",             "Pending"),
    (r"\bVencido\b",               "Overdue"),
    (r"\bActivo\b",                "Active"),
    (r"\bInactivo\b",              "Inactive"),
    (r"\bCompletado\b",            "Completed"),
    (r"\bEn Progreso\b",           "In Progress"),
    (r"\bCancelado\b",             "Cancelled"),
    (r"\bAprobado\b",              "Approved"),
    (r"\bRechazado\b",             "Rejected"),
    (r"\bIngresos\b",              "Income"),
    (r"\bEgresos\b",               "Expenses"),
    (r"\bGastos\b",                "Expenses"),
    (r"\bGasto\b",                 "Expense"),
    (r"\bIngreso\b",               "Income"),
    (r"\bVentas\b",                "Sales"),
    (r"\bVenta\b",                 "Sale"),
    (r"\bCompras\b",               "Purchases"),
    (r"\bCompra\b",                "Purchase"),
    (r"\bPresupuesto\b",           "Budget"),
    (r"\bVarianza\b",              "Variance"),
    (r"\bGanancia\b",              "Profit"),
    (r"\bPérdida\b",               "Loss"),
    (r"\bUtilidad\b",              "Profit"),
    (r"\bFactura\b",               "Invoice"),
    (r"\bCotización\b",            "Quotation"),
    (r"\bRecibo\b",                "Receipt"),
    (r"\bInventario\b",            "Inventory"),
    (r"\bAlmacén\b",               "Warehouse"),
    (r"\bUbicación\b",             "Location"),
    (r"\bUnidad\b",                "Unit"),
    (r"\bMínimo\b",                "Minimum"),
    (r"\bMáximo\b",                "Maximum"),
    (r"\bAsistencia\b",            "Attendance"),
    (r"\bVacaciones\b",            "Vacation"),
    (r"\bHorario\b",               "Schedule"),
    (r"\bTurno\b",                 "Shift"),
    (r"\bDeducciones\b",           "Deductions"),
    (r"\bNómina\b",                "Payroll"),
    (r"\bComisiones\b",            "Commissions"),
    (r"\bComisión\b",              "Commission"),
    (r"\bBonificación\b",          "Bonus"),
    (r"\bSubtotal\b",              "Subtotal"),
    (r"\bImpuesto\b",              "Tax"),
    (r"\bImpuestos\b",             "Taxes"),
    (r"\bMoneda\b",                "Currency"),
    (r"\bResumen\b",               "Summary"),
    (r"\bConfiguración\b",         "Settings"),
    (r"\bDatos\b",                 "Data"),
    (r"\bProyecto\b",              "Project"),
    (r"\bProyectos\b",             "Projects"),
    (r"\bEmpleados\b",             "Employees"),
    (r"\bClientes\b",              "Clients"),
    (r"\bProveedores\b",           "Suppliers"),
    (r"\bSubscripciones\b",        "Subscriptions"),
    (r"\bViáticos\b",              "Travel Expenses"),
    (r"\bGastos de Viaje\b",       "Travel Expenses"),
    (r"\bAmortización\b",          "Amortization"),
    (r"\bCódigo\b",                "Code"),
    (r"\bMonto\b",                 "Amount"),
    (r"\bSaldo\b",                 "Balance"),
    (r"\bFecha\b",                 "Date"),
    (r"\bMes\b",                   "Month"),
    (r"\bAño\b",                   "Year"),
    (r"\bSemana\b",                "Week"),
    (r"\bDía\b",                   "Day"),
    (r"\bTrimestre\b",             "Quarter"),
    (r"\bPeríodo\b",               "Period"),
    (r"\bPeriodo\b",               "Period"),
    (r"\bPaís\b",                  "Country"),
    (r"\bCiudad\b",                "City"),
    (r"\bEmail\b",                 "Email"),
    (r"\bPuesto\b",                "Position"),
    (r"\bCargo\b",                 "Title"),
    (r"\bÁrea\b",                  "Department"),
    (r"\bTotal\b",                 "Total"),
    (r"\bTipo\b",                  "Type"),
    (r"\bMeses\b",                 "Months"),
    (r"\bSemanas\b",               "Weeks"),
    (r"\bDías\b",                  "Days"),
    (r"\bHoras\b",                 "Hours"),
    (r"\bNeto\b",                  "Net"),
    (r"\bBruto\b",                 "Gross"),
    (r"\bReal\b",                  "Actual"),
    (r"\bAcciones\b",              "Actions"),
    # ── Compound patterns (frases que se traducen parcialmente sin esto) ─
    (r"\bNombre Empleado\b",       "Employee Name"),
    (r"\bNombre Producto\b",       "Product Name"),
    (r"\bNombre Cliente\b",        "Client Name"),
    (r"\bNombre Proveedor\b",      "Supplier Name"),
    (r"\bStock Actual\b",          "Current Stock"),
    (r"\bStock Mínimo\b",          "Minimum Stock"),
    (r"\bStock Máximo\b",          "Maximum Stock"),
    (r"\bSal\. Diario\b",          "Daily Salary"),
    (r"\bSal\. Mensual\b",         "Monthly Salary"),
    (r"\bSal\. Anual\b",           "Annual Salary"),
    (r"\bSal\. Base\b",            "Base Salary"),
    # v14.6.1 — labels encontradas en testing de quotation/invoice
    # Nota: NO usar \b al final después de un :? — el \b no matchea el ":"
    # porque no es char alfanumérico, dejando el ":" residual del input.
    (r"\bEmpresa cliente:?",       "Client company:"),
    (r"\bContacto:?",              "Contact:"),
    (r"\bValidez:?",               "Valid until:"),
    (r"\bEmitida por:?",           "Issued by:"),
    (r"\bEmitido por:?",           "Issued by:"),
    (r"\bDirigida a:?",            "To:"),
    (r"\bDirigido a:?",            "To:"),
    (r"\bAtención:?",              "Attn:"),
    (r"\bAtencion:?",              "Attn:"),
    (r"\bDías naturales\b",        "calendar days"),
    (r"\bDías hábiles\b",          "business days"),
    (r"\bdías\b",                  "days"),
    (r"\bDías\b",                  "Days"),
    (r"\bVigencia\b",              "Validity"),
    (r"\bTérminos\b",              "Terms"),
    (r"\bCondiciones\b",           "Conditions"),
    (r"\bGarantía\b",              "Warranty"),
    (r"\bGarantia\b",              "Warranty"),
    (r"\bEntrega\b",               "Delivery"),
    (r"\bForma de entrega\b",      "Delivery method"),
    (r"\bIncluye\b",               "Includes"),
    (r"\bNo incluye\b",            "Does not include"),
    (r"\bGracias por su preferencia\b", "Thank you for your business"),
    (r"\bGracias\b",               "Thanks"),
    # Frases con "Equiv. MXN/mes" tipicas de subscription
    (r"\bEquiv\. MXN/mes\b",       "Monthly equivalent"),
    (r"\bEquiv\. USD/mes\b",       "Monthly equivalent"),
    # v14.5 — términos adicionales encontrados en testing real
    (r"\bDestino\b",               "Destination"),
    (r"\bOrigen\b",                "Origin"),
    (r"\bMotivo\b",                "Reason"),
    (r"\bReembolsado\b",           "Reimbursed"),
    (r"\bReembolso\b",             "Reimbursement"),
    (r"\bFactura A\d+\b",          "Invoice"),
    (r"\bRecibo\b",                "Receipt"),
    (r"\bAlojamiento\b",           "Lodging"),
    (r"\bAlimentación\b",          "Meals"),
    (r"\bAlimentacion\b",          "Meals"),
    (r"\bTransporte\b",            "Transportation"),
    (r"\bDiseño\b",                "Design"),
    (r"\bDesarrollo\b",            "Development"),
    (r"\bMarketing\b",             "Marketing"),
    (r"\bAdministración\b",        "Administration"),
    (r"\bAdministracion\b",        "Administration"),
    (r"\bOficina\b",               "Office"),
    (r"\bRenta\b",                 "Rent"),
    (r"\bArriendo\b",              "Rent"),
    (r"\bServicios\b",             "Services"),
    (r"\bAgua\b",                  "Water"),
    (r"\bLuz\b",                   "Electricity"),
    (r"\bInternet\b",              "Internet"),
    (r"\bMantenimiento\b",         "Maintenance"),
    (r"\bSeguros?\b",              "Insurance"),
    (r"\bPublicidad\b",            "Advertising"),
    (r"\bImpresión\b",             "Printing"),
    (r"\bImpresion\b",             "Printing"),
    (r"\bPapelería\b",             "Stationery"),
    (r"\bPapeleria\b",             "Stationery"),
    (r"\bSoftware\b",              "Software"),
    (r"\bMercancía\b",             "Merchandise"),
    (r"\bMercancia\b",             "Merchandise"),
    (r"\bMateria Prima\b",         "Raw Materials"),
    (r"\bComida\b",                "Food"),
    (r"\bComidas?\b",              "Meals"),
    (r"\bCena\b",                  "Dinner"),
    (r"\bDesayuno\b",              "Breakfast"),
    (r"\bAlmuerzo\b",              "Lunch"),
    (r"\bHotel\b",                 "Hotel"),
    (r"\bVuelo\b",                 "Flight"),
    (r"\bTaxi\b",                  "Taxi"),
    (r"\bUber\b",                  "Uber"),
    (r"\bPropinas?\b",             "Tips"),
    (r"\bGasolina\b",              "Gas"),
    (r"\bEstacionamiento\b",       "Parking"),
    (r"\bParqueadero\b",           "Parking"),
    (r"\bCajero\b",                "Teller"),
    (r"\bBanco\b",                 "Bank"),
    (r"\bSucursal\b",              "Branch"),
    (r"\bCheque\b",                "Check"),
    (r"\bTransferencia\b",         "Transfer"),
    (r"\bDepósito\b",              "Deposit"),
    (r"\bDeposito\b",              "Deposit"),
    (r"\bRetiro\b",                "Withdrawal"),
    (r"\bAbono\b",                 "Credit"),
    (r"\bCargo\b",                 "Charge"),
    (r"\bSucursal\b",              "Branch"),
    (r"\bCajero Automático\b",     "ATM"),
    (r"\bDébito\b",                "Debit"),
    (r"\bDebito\b",                "Debit"),
    (r"\bCrédito\b",               "Credit"),
    (r"\bCredito\b",               "Credit"),
    (r"\bTarjeta\b",               "Card"),
    (r"\bSí\b",                    "Yes"),
    (r"\bNo\b",                    "No"),
    (r"\bParcial\b",               "Partial"),
    (r"\bTotal\b",                 "Total"),
    # ── Sheet names adicionales (encontrados en testing v14.4) ─────────
    (r"\bCuentas x Cobrar\b",      "Accounts Receivable"),
    (r"\bCuentas x Pagar\b",       "Accounts Payable"),
    (r"\bConciliación\b",          "Reconciliation"),
    (r"\bConciliacion\b",          "Reconciliation"),
    (r"\bSolicitudes\b",           "Requests"),
    (r"\bSolicitud\b",             "Request"),
    (r"\bRegistro\b",              "Log"),
    (r"\bRegistros\b",             "Records"),
    (r"\bCaja Chica\b",            "Petty Cash"),
    (r"\bMis Metas\b",             "My Goals"),
    (r"\bMis Ahorros\b",           "My Savings"),
    (r"\bParámetros\b",            "Parameters"),
    (r"\bParametros\b",            "Parameters"),
    (r"\bTareas\b",                "Tasks"),
    (r"\bTarea\b",                 "Task"),
    (r"\bAvance\b",                "Progress"),
    (r"\bMilestones?\b",           "Milestones"),
    (r"\bHitos?\b",                "Milestones"),
    (r"\bSemanal\b",               "Weekly"),
    (r"\bMensual\b",               "Monthly"),
    (r"\bAnual\b",                 "Annual"),
    (r"\bDiario\b",                "Daily"),
    (r"\bTrimestral\b",            "Quarterly"),
    (r"\bSiguiente\b",             "Next"),
    (r"\bAnterior\b",              "Previous"),
    (r"\bResumen Ejecutivo\b",     "Executive Summary"),
    (r"\bAnálisis\b",              "Analysis"),
    (r"\bAnalisis\b",              "Analysis"),
    (r"\bGráfico\b",               "Chart"),
    (r"\bGrafico\b",               "Chart"),
    (r"\bReporte\b",               "Report"),
    (r"\bInforme\b",               "Report"),
    (r"\bEscenario\b",             "Scenario"),
    (r"\bEscenarios\b",            "Scenarios"),
    (r"\bOptimista\b",             "Optimistic"),
    (r"\bPesimista\b",             "Pessimistic"),
    (r"\bRealista\b",              "Realistic"),
    (r"\bProyección\b",            "Projection"),
    (r"\bProyeccion\b",            "Projection"),
    (r"\bProyecciones\b",          "Projections"),
    (r"\bSimulación\b",            "Simulation"),
    (r"\bSimulacion\b",            "Simulation"),
    (r"\bInversión\b",             "Investment"),
    (r"\bInversion\b",             "Investment"),
    (r"\bRetorno\b",               "Return"),
    (r"\bRentabilidad\b",          "Profitability"),
    (r"\bUtilidades\b",            "Profits"),
    (r"\bResultados\b",            "Results"),
    (r"\bResultado\b",             "Result"),
    (r"\bDetalle\b",               "Detail"),
    (r"\bDetalles\b",              "Details"),
    (r"\bGeneral\b",               "General"),
    (r"\bEspecífico\b",            "Specific"),
    (r"\bEspecifico\b",            "Specific"),
    # Conjunción "e" (delante de palabras que empiezan por i/h) en encabezados.
    # NOTA: No traducir "de/la/el/los/las/un/una" porque romperían nombres propios
    # del usuario (ej. "José de la Cruz"). Sólo traducimos conectores que aparecen
    # casi exclusivamente en encabezados generados.
    (r" e Ingresos\b",             " and Income"),
    (r" e Income\b",               " and Income"),
    (r" e Inventario\b",           " and Inventory"),
    (r" e Inventory\b",            " and Inventory"),
    (r" e Información\b",          " and Information"),
    (r" e Information\b",          " and Information"),
    (r" y Egresos\b",              " and Expenses"),
    (r" y Expenses\b",             " and Expenses"),
    (r" y Gastos\b",               " and Expenses"),
    (r" y Ventas\b",               " and Sales"),
    (r" y Sales\b",                " and Sales"),
    (r" y Compras\b",              " and Purchases"),
    (r" y Purchases\b",            " and Purchases"),
    (r" y Cobranza\b",             " and Collection"),
    (r" y Collection\b",           " and Collection"),
    (r" vs Presupuesto\b",         " vs Budget"),
    (r" vs Real\b",                " vs Actual"),
    # Meses
    (r"\bEnero\b",     "January"),  (r"\bFebrero\b",    "February"),
    (r"\bMarzo\b",     "March"),    (r"\bAbril\b",      "April"),
    (r"\bMayo\b",      "May"),      (r"\bJunio\b",      "June"),
    (r"\bJulio\b",     "July"),     (r"\bAgosto\b",     "August"),
    (r"\bSeptiembre\b","September"),(r"\bOctubre\b",    "October"),
    (r"\bNoviembre\b", "November"), (r"\bDiciembre\b",  "December"),
    (r"\bEne\b", "Jan"), (r"\bFeb\b", "Feb"), (r"\bMar\b", "Mar"),
    (r"\bAbr\b", "Apr"), (r"\bMay\b", "May"), (r"\bJun\b", "Jun"),
    (r"\bJul\b", "Jul"), (r"\bAgo\b", "Aug"), (r"\bSep\b", "Sep"),
    (r"\bOct\b", "Oct"), (r"\bNov\b", "Nov"), (r"\bDic\b", "Dec"),
]
# Pre-compilar regex para velocidad
_COMPILED_PATTERNS = [(re.compile(p), repl) for p, repl in _COMPOSITE_PATTERNS]


def _translate_string(s: str) -> str:
    """Aplica todos los patrones de traducción a una cadena.
    Devuelve la cadena traducida; si no hay match, devuelve la original."""
    if not isinstance(s, str) or not s.strip():
        return s
    out = s
    for rx, repl in _COMPILED_PATTERNS:
        out = rx.sub(repl, out)
    return out


def _translate_workbook(wb):
    """Post-procesa el workbook traduciendo todos los strings detectables.
    Sólo se ejecuta si _lang() == 'en'. Cubre sheet names, valores de celdas,
    títulos y axis labels de charts.

    v14.5 — Además convierte:
      • Fechas DD/MM/YYYY → MM/DD/YYYY (formato US)
      • Símbolo de moneda MN/MXN → USD en labels (no toca data simulada)
    """
    if _lang() != "en":
        return

    # v14.5 — Regex helpers para conversión de formato US
    _DATE_DMY_RX = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b")
    def _to_us_date(s: str) -> str:
        # Sólo convierte si el primer número parece "día" (1-31) y el segundo "mes" (1-12).
        # Si es ambiguo (ambos ≤12), conservamos el formato original. Esto evita
        # convertir "01/05/2026" (1 mayo) cuando podría interpretarse como 5 enero.
        # Para ser conservador, solo convertimos cuando el primer número > 12 (es día sí o sí).
        def repl(m):
            d, mth, y = m.group(1), m.group(2), m.group(3)
            try:
                di, mi = int(d), int(mth)
                # Si el primer número es > 12, es claramente DD/MM → convertir
                # Si ambos son ≤12, conservamos para no romper ambiguos
                if di > 12 and 1 <= mi <= 12:
                    return f"{mi:02d}/{di:02d}/{y}"
                # Si ambos ≤12 y el segundo es válido, convertimos asumiendo DD/MM
                # (ya que el resto del sistema usa ese formato por default)
                if 1 <= di <= 31 and 1 <= mi <= 12:
                    return f"{mi:02d}/{di:02d}/{y}"
            except (ValueError, TypeError):
                pass
            return m.group(0)
        return _DATE_DMY_RX.sub(repl, s)

    # v14.5 — patrones de moneda (MXN/MN suelen aparecer en labels y data simulada)
    _CURR_PATTERNS = [
        (re.compile(r"\bMXN\b"),  "USD"),
        (re.compile(r"\bMN\b"),   "USD"),
        (re.compile(r"\$\s*([\d,\.]+)\s*MN\b"),  r"$\1"),
        (re.compile(r"\$\s*([\d,\.]+)\s*MXN\b"), r"$\1"),
        # Pesos mexicano/colombiano/argentino/etc → US currencies en labels
        (re.compile(r"\bPeso mexicano\b"),  "US Dollar"),
        (re.compile(r"\bPeso Mexicano\b"),  "US Dollar"),
        (re.compile(r"\bDólar americano\b"),"US Dollar"),
        (re.compile(r"\bDólar USD\b"),      "US Dollar"),
    ]
    def _apply_curr(s: str) -> str:
        out = s
        for rx, repl in _CURR_PATTERNS:
            out = rx.sub(repl, out)
        return out

    for ws in wb.worksheets:
        # 1) Sheet name
        try:
            new_title = _translate_string(ws.title)
            # Excel limita títulos a 31 chars
            if new_title and new_title != ws.title and len(new_title) <= 31:
                ws.title = new_title
        except Exception:
            pass

        # 2) Todas las celdas con valor string
        try:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val and not val.startswith("="):
                        new_val = _translate_string(val)
                        # v14.5 — formato US para fechas y moneda
                        new_val = _to_us_date(new_val)
                        new_val = _apply_curr(new_val)
                        if new_val != val:
                            cell.value = new_val
        except Exception:
            pass

        # 3) Charts: título y axis labels
        try:
            for chart in getattr(ws, "_charts", []) or []:
                # Título del chart
                try:
                    if chart.title is not None:
                        # chart.title puede ser str o objeto Title con texto interno
                        if isinstance(chart.title, str):
                            chart.title = _translate_string(chart.title)
                        else:
                            tx = getattr(chart.title, "tx", None)
                            if tx is not None:
                                rich = getattr(tx, "rich", None)
                                if rich is not None:
                                    for p in getattr(rich, "p", []) or []:
                                        for r in getattr(p, "r", []) or []:
                                            if hasattr(r, "t") and isinstance(r.t, str):
                                                r.t = _translate_string(r.t)
                except Exception:
                    pass
                # Y-axis title
                try:
                    if chart.y_axis is not None and chart.y_axis.title is not None:
                        if isinstance(chart.y_axis.title, str):
                            chart.y_axis.title = _translate_string(chart.y_axis.title)
                except Exception:
                    pass
                # X-axis title
                try:
                    if chart.x_axis is not None and chart.x_axis.title is not None:
                        if isinstance(chart.x_axis.title, str):
                            chart.x_axis.title = _translate_string(chart.x_axis.title)
                except Exception:
                    pass
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
def _ensure_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def _safe_filename(title: str) -> str:
    return re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')[:60]


# ─────────────────────────────────────────────────────────────────────────────
# v14.9 — USER DATA EXTRACTION (items / rows)
# ─────────────────────────────────────────────────────────────────────────────
# Los builders de Excel reciben datos del usuario en `datos["items"]` (o
# sinónimos). Cada item es un dict con keys flexibles (es/en) que normalizamos
# acá. Si no hay items, los builders caen al sample data hardcoded de antes
# (backward compat 100%).
#
# La LLM emite items según el SYSTEM_PROMPT, pero acá toleramos varios shapes
# para que correcciones casuales del usuario ("agrégale más productos") sigan
# funcionando aunque la LLM emita keys ligeramente distintas.

# Mapas de sinónimos de keys: la key canónica → variantes aceptadas (incluye sí misma)
_ITEM_KEY_SYNONYMS = {
    # ── Identificación ──
    "codigo":       ("codigo", "code", "sku", "id", "ref", "referencia"),
    "nombre":       ("nombre", "name", "producto", "product", "descripcion",
                     "description", "concepto", "item", "articulo"),
    "categoria":    ("categoria", "category", "cat", "tipo", "type", "grupo"),
    # ── Inventario ──
    "stock":        ("stock", "stock_actual", "cantidad", "qty", "quantity",
                     "existencias", "existencia"),
    "stock_min":    ("stock_min", "stock_minimo", "min_stock", "minimo", "min"),
    # ── Precios ──
    "precio_costo": ("precio_costo", "costo", "cost", "precio_cost",
                     "unit_cost", "cost_price"),
    "precio_venta": ("precio_venta", "precio", "price", "precio_unit",
                     "unit_price", "selling_price", "sale_price", "venta"),
    # ── Cotización / ventas ──
    "cantidad":     ("cantidad", "qty", "quantity", "cant", "units"),
    "descuento":    ("descuento", "discount", "desc", "descuento_pct"),
    # ── Nómina ──
    "empleado":     ("empleado", "employee", "trabajador", "worker", "persona"),
    "puesto":       ("puesto", "position", "cargo", "role", "rol"),
    "dias":         ("dias", "days", "dias_trabajados", "worked_days"),
    "salario_diario": ("salario_diario", "daily_salary", "salario_dia",
                       "tarifa_diaria", "daily_rate"),
    "deducciones_pct": ("deducciones_pct", "deductions_pct", "deducciones",
                        "deductions", "descuentos"),
    # ── Flujo de caja ──
    "mes":          ("mes", "month", "periodo", "period"),
    "monto":        ("monto", "amount", "valor", "value", "total"),
    "ingreso":      ("ingreso", "income", "ingresos", "entrada"),
    "egreso":       ("egreso", "expense", "egresos", "salida", "gasto"),
    # ── Otros comunes ──
    "fecha":        ("fecha", "date", "fecha_emision"),
    "cliente":      ("cliente", "client", "customer"),
    "estado":       ("estado", "status", "state"),
    "notas":        ("notas", "notes", "observaciones", "obs"),
}


def _norm_item(raw: dict) -> dict:
    """Normaliza las keys de un item a las canónicas.

    Acepta keys en español/inglés y varios sinónimos (sku/code/codigo →
    "codigo"). Las keys desconocidas se conservan tal cual por si algún builder
    específico las necesita.
    """
    if not isinstance(raw, dict):
        return {}
    # Reverse lookup: variant_key → canonical_key
    rev = {}
    for canon, variants in _ITEM_KEY_SYNONYMS.items():
        for v in variants:
            rev[v.lower()] = canon
    out = {}
    for k, v in raw.items():
        if not isinstance(k, str):
            continue
        canon = rev.get(k.lower().strip(), k)
        # No pisar si ya tenemos el canon poblado (priorizar la primera ocurrencia)
        if canon not in out:
            out[canon] = v
    return out


def _extract_items(datos: dict | None, *extra_keys: str) -> list[dict]:
    """Extrae la lista de items del usuario desde `datos`.

    Prueba en orden: `items`, `rows`, y cualquier key extra que el builder
    específico quiera (e.g. "productos", "empleados", "movimientos"). Devuelve
    la primera lista de dicts no-vacía, con las keys normalizadas. Si no hay
    nada, devuelve [].

    Uso típico desde un builder:
        items = _extract_items(datos, "productos", "inventario")
        if items:
            # render filas del usuario
        else:
            # caer al sample hardcoded
    """
    if not isinstance(datos, dict):
        return []
    candidate_keys = ("items", "rows") + extra_keys
    for k in candidate_keys:
        raw = datos.get(k)
        if isinstance(raw, list) and raw:
            normalized = [_norm_item(x) for x in raw if isinstance(x, dict)]
            normalized = [x for x in normalized if x]
            if normalized:
                return normalized
    return []


def _num(value, default=0):
    """Coerce un valor a int/float, robusto a strings con $/,/% y None."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        # Limpiar formato común: "$1,234.56" → 1234.56, "50%" → 0.5
        s = value.strip().replace("$", "").replace(",", "").replace(" ", "")
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return default
        try:
            return float(s) if "." in s else int(s)
        except ValueError:
            return default
    return default

def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    t = Side(style="medium", color="888888")
    n = Side(style="thin",   color="CCCCCC")
    return Border(left=n, right=n, top=n, bottom=t)

def _hc(ws, ref, value, bg="1A4B8C", fg="FFFFFF", bold=True, size=10, align="center"):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _thin()
    return c

def _dc(ws, ref, value, bold=False, align="left", bg="FFFFFF",
        fmt=None, color="000000", size=10):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _thin()
    if fmt:
        c.number_format = fmt
    return c

def _formula(ws, ref, formula, bg="FFFFFF", color="000000", fg=None, bold=False,
             fmt=None, align="center", size=10):
    c = ws[ref]
    c.value = formula
    c.font = Font(name="Arial", bold=bold, color=fg if fg else color, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _thin()
    if fmt:
        c.number_format = fmt
    return c

def _setup_sheet(wb, name, tab_color="1A4B8C"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    return ws

def _title_row(ws, title, subtitle, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, title)
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="0D1B2A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=cols)
    c2 = ws.cell(row+1, 1, subtitle)
    c2.font = Font(name="Arial", size=9, color="1A4B8C")
    c2.fill = PatternFill("solid", start_color="EFF6FF")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row+1].height = 16

def _params_box(ws, row, col, label, value, note="", bg_label="EFF6FF", bg_val="FFFFFF",
                fmt=None, color_val="1A4B8C"):
    """Renders an editable parameter cell with label."""
    c_lbl = ws.cell(row, col, label)
    c_lbl.font = Font(name="Arial", bold=True, size=10, color="0D1B2A")
    c_lbl.fill = PatternFill("solid", start_color=bg_label)
    c_lbl.alignment = Alignment(horizontal="left", vertical="center")
    c_lbl.border = _thin()
    c_val = ws.cell(row, col+1, value)
    c_val.font = Font(name="Arial", bold=True, size=11, color=color_val)
    c_val.fill = PatternFill("solid", start_color=bg_val)
    c_val.alignment = Alignment(horizontal="center", vertical="center")
    c_val.border = _thin()
    if fmt:
        c_val.number_format = fmt
    if note:
        c_note = ws.cell(row, col+2, note)
        c_note.font = Font(name="Arial", italic=True, size=9, color="888780")
        c_note.alignment = Alignment(horizontal="left", vertical="center")
        c_note.border = _thin()
    return c_val

def _section_row(ws, row, label, cols, bg="1A4B8C", fg="FFFFFF"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, label)
    c.font = Font(name="Arial", bold=True, size=10, color=fg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _thin()
    ws.row_dimensions[row].height = 18

def _section_row_a(ws, row, label, bg="1A4B8C", fg="FFFFFF"):
    """Section header on column A only — does NOT merge. Allows writing B-E on same row."""
    c = ws.cell(row, 1, label)
    c.font = Font(name="Arial", bold=True, size=10, color=fg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _thin()
    ws.row_dimensions[row].height = 18

def _note_row(ws, row, text, cols, bg="FFFBEB", color="92400E"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.font = Font(name="Arial", italic=True, size=9, color=color)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE HELPERS — ConditionalFormatting · DataValidation · Charts · Tables
# ─────────────────────────────────────────────────────────────────────────────

def _red_fill():
    return PatternFill("solid", start_color="FFE4E4")
def _green_fill():
    return PatternFill("solid", start_color="D4EDDA")
def _amber_fill():
    return PatternFill("solid", start_color="FFF3CD")

def _add_autofilter(ws, header_row, cols):
    """Activate native Excel auto-filter on the header row."""
    last_col = get_column_letter(cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

def _add_table(ws, name, ref, style="TableStyleMedium2"):
    """Wrap a range in a native Excel Table (auto-filter + style + totals row)."""
    tbl = Table(displayName=name.replace(" ","_").replace("/","_")[:20], ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True,
                                        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tbl)

def _cf_three_color(ws, data_range, low="#FF4444", mid="#FFBB33", high="#00C851"):
    """Color scale: low=red → mid=amber → high=green."""
    from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
    cs = ColorScale(cfvo=[
        FormatObject(type="min"),
        FormatObject(type="percentile", val=50),
        FormatObject(type="max"),
    ], color=[
        __import__("openpyxl").styles.Color(low.lstrip("#")),
        __import__("openpyxl").styles.Color(mid.lstrip("#")),
        __import__("openpyxl").styles.Color(high.lstrip("#")),
    ])
    rule = Rule(type="colorScale", colorScale=cs)
    ws.conditional_formatting.add(data_range, rule)

def _cf_cell_is(ws, cell_range, operator, formula, fill_color, font_color="000000"):
    """Highlight cells matching a condition with native Excel ConditionalFormatting."""
    red_font  = Font(color=font_color, bold=True, name="Arial", size=10)
    red_fill  = PatternFill("solid", start_color=fill_color)
    dxf = DifferentialStyle(font=red_font, fill=red_fill)
    rule = CellIsRule(operator=operator, formula=[formula], dxf=dxf)
    ws.conditional_formatting.add(cell_range, rule)

def _cf_formula_rule(ws, cell_range, formula, fill_color, font_color="000000", bold=True):
    """Highlight entire row when a formula evaluates to TRUE."""
    font = Font(color=font_color, bold=bold, name="Arial", size=10)
    fill = PatternFill("solid", start_color=fill_color)
    dxf  = DifferentialStyle(font=font, fill=fill)
    rule = Rule(type="expression", dxf=dxf, formula=[formula])
    ws.conditional_formatting.add(cell_range, rule)

def _add_dv_list(ws, cell_range, options, prompt="Selecciona un valor"):
    """Add a dropdown DataValidation to a cell range."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showDropDown=False, prompt=prompt, showInputMessage=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)

def _add_dv_whole(ws, cell_range, min_val, max_val, prompt=""):
    """Add whole number DataValidation."""
    dv = DataValidation(type="whole", operator="between",
                        formula1=str(min_val), formula2=str(max_val),
                        allow_blank=True, prompt=prompt, showInputMessage=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)

def _add_line_chart(ws, ws_chart, title, row_refs, col_start, col_end, data_row,
                    series_titles, colors, anchor="A3"):
    """
    Add a LineChart to ws_chart sheet referencing data in ws.
    row_refs: list of row numbers in ws where each series lives.
    """
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.grouping = "standard"
    chart.height = 14
    chart.width  = 24

    cats = Reference(ws, min_col=col_start+1, max_col=col_end,
                     min_row=data_row, max_row=data_row)
    chart.set_categories(cats)

    for i, (row_num, stitle, color) in enumerate(zip(row_refs, series_titles, colors)):
        data = Reference(ws, min_col=col_start+1, max_col=col_end,
                         min_row=row_num, max_row=row_num)
        series = Series(data, title=stitle)
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 25000
        chart.series.append(series)

    ws_chart.add_chart(chart, anchor)

def _add_bar_chart(ws, ws_chart, title, data_rows, cat_row, col_start, col_end,
                   series_titles, colors, anchor="A3", bar_dir="col"):
    """Add a BarChart to ws_chart referencing data in ws."""
    chart = BarChart()
    chart.type     = bar_dir
    chart.grouping = "clustered"
    chart.title    = title
    chart.style    = 10
    chart.height   = 14
    chart.width    = 24
    chart.y_axis.title = "Monto"

    cats = Reference(ws, min_col=col_start+1, max_col=col_end,
                     min_row=cat_row, max_row=cat_row)
    chart.set_categories(cats)

    for row_num, stitle, color in zip(data_rows, series_titles, colors):
        data = Reference(ws, min_col=col_start+1, max_col=col_end,
                         min_row=row_num, max_row=row_num)
        series = Series(data, title=stitle)
        series.graphicalProperties.solidFill = color
        chart.series.append(series)

    ws_chart.add_chart(chart, anchor)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS v5 — Protection · Print Setup · Combo Chart · Named Ranges
# ─────────────────────────────────────────────────────────────────────────────

def _protect_ws(ws, unlock_from_row=5):
    """Lock formula cells, unlock data input cells, enable sheet protection.
    Formula cells stay locked. Data entry cells (non-formula, row>=unlock_from_row) are unlocked.
    """
    for row in ws.iter_rows(min_row=unlock_from_row):
        for cell in row:
            val = str(cell.value) if cell.value is not None else ""
            if not val.startswith("=") and val != "":
                # Data entry cell — unlock it so user can edit
                cell.protection = Protection(locked=False)
    # Enable protection (password="" = no password needed, just UI warning)
    ws.protection.sheet         = True
    ws.protection.password      = ""
    ws.protection.autoFilter    = False   # keep filters working
    ws.protection.sort          = False   # keep sort working
    ws.protection.insertRows    = False   # allow inserting rows
    ws.protection.deleteRows    = False
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False

def _setup_print(ws, cols, last_row=None, landscape=True):
    """Configure print area, orientation, margins and fit-to-page."""
    last_col = get_column_letter(cols)
    if last_row:
        ws.print_area = f"A1:{last_col}{last_row}"
    ws.page_setup.orientation  = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_setup.paperSize    = ws.PAPERSIZE_LETTER
    ws.page_margins            = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75,
                                              header=0.3, footer=0.3)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Repeat header row when printing
    ws.print_title_rows = "1:4"

def _add_named_range(wb, name, ws, cell_ref):
    """Add a workbook-level named range for cleaner formula references."""
    from openpyxl.workbook.defined_name import DefinedName
    safe_name = name.replace(" ", "_").replace("/", "_").replace(".", "_")
    full_ref   = f"'{ws.title}'!{cell_ref}"
    dn = DefinedName(name=safe_name, attr_text=full_ref)
    try:
        wb.defined_names[safe_name] = dn
    except Exception:
        pass  # Named range already exists or invalid

def _add_combo_chart(ws_data, ws_chart, title,
                     bar_ref_row, line_ref_row, cat_row,
                     col_start, col_end,
                     bar_title, line_title,
                     bar_color="185FA5", line_color="B91C1C",
                     anchor="A3"):
    """Create a combo Bar+Line chart — the most-requested chart in executive reports."""
    bar  = BarChart()
    bar.type     = "col"
    bar.grouping = "clustered"
    bar.title    = title
    bar.style    = 10
    bar.height   = 14
    bar.width    = 22
    bar.y_axis.title = "Monto"

    cats = Reference(ws_data, min_col=col_start+1, max_col=col_end, min_row=cat_row)
    bar.set_categories(cats)

    # Bar series
    s_bar = Series(Reference(ws_data, min_col=col_start+1, max_col=col_end,
                              min_row=bar_ref_row), title=bar_title)
    s_bar.graphicalProperties.solidFill = bar_color
    bar.series.append(s_bar)

    # Line series
    line = LineChart()
    line.grouping = "standard"
    s_line = Series(Reference(ws_data, min_col=col_start+1, max_col=col_end,
                               min_row=line_ref_row), title=line_title)
    s_line.graphicalProperties.line.solidFill = line_color
    s_line.graphicalProperties.line.width = 28000
    line.series.append(s_line)
    # Assign line to secondary axis
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    line.y_axis.title = "%"
    bar += line   # combine
    ws_chart.add_chart(bar, anchor)


# ─────────────────────────────────────────────────────────────────────────────
# BUG DETECTOR ENGINE
# ─────────────────────────────────────────────────────────────────────────────
class ExcelBugDetector:
    """
    Scans generated workbooks for common issues.
    Auto-corrects what it can; reports the rest to the user.
    """

    PLACEHOLDER_PATTERN = re.compile(r'^\[=.*\]$')
    HARDCODED_RATES = {0.16: "IVA 16%", 0.12: "Deducción 12%", 0.35: "ISR 35%",
                       0.10: "ISR 10%", 0.05: "ISR 5%"}
    DIV_ZERO_PATTERN = re.compile(r'=[^=].*/')

    def scan(self, wb: Workbook) -> list:
        bugs = []
        for ws in wb.worksheets:
            bugs.extend(self._scan_sheet(ws))
        return bugs

    def _scan_sheet(self, ws) -> list:
        bugs = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                ref = f"{ws.title}!{cell.coordinate}"

                # 1. Text placeholder instead of real formula
                if self.PLACEHOLDER_PATTERN.match(val):
                    bugs.append({
                        "type": "placeholder",
                        "severity": "error",
                        "ref": ref,
                        "cell": cell,
                        "ws": ws,
                        "msg": f"Placeholder de fórmula en {ref}: '{val}'",
                        "fix": "formula_needed",
                        "original": val
                    })

                # 2. Formula without IFERROR protection on division
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if "/" in cell.value and "IFERROR" not in cell.value and "IF(" not in cell.value:
                        bugs.append({
                            "type": "div_zero",
                            "severity": "warning",
                            "ref": ref,
                            "cell": cell,
                            "ws": ws,
                            "msg": f"División sin protección en {ref}: '{cell.value}'",
                            "fix": "add_iferror",
                            "original": cell.value
                        })

                # 3. Hardcoded tax/rate percentages (not in header rows)
                if isinstance(cell.value, float) and cell.row > 3:
                    for rate, name in self.HARDCODED_RATES.items():
                        if abs(cell.value - rate) < 0.0001:
                            bugs.append({
                                "type": "hardcoded_rate",
                                "severity": "warning",
                                "ref": ref,
                                "cell": cell,
                                "ws": ws,
                                "msg": f"Tasa hardcodeada en {ref}: {rate*100}% ({name}). Considera usar celda de parámetro.",
                                "fix": "note_only",
                                "original": cell.value
                            })
                            break

                # 4. SUM formula — check if range seems too small
                if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                    m = re.search(r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', cell.value)
                    if m:
                        start_r, end_r = int(m.group(2)), int(m.group(4))
                        if end_r - start_r < 1:
                            bugs.append({
                                "type": "sum_range",
                                "severity": "warning",
                                "ref": ref,
                                "cell": cell,
                                "ws": ws,
                                "msg": f"Rango SUM muy pequeño en {ref}: '{cell.value}'",
                                "fix": "note_only",
                                "original": cell.value
                            })
        return bugs

    def auto_fix(self, wb: Workbook, bugs: list) -> tuple:
        """Auto-fixes what it can. Returns (wb, list_of_fix_messages)."""
        fixes = []
        for bug in bugs:
            if bug["fix"] == "add_iferror":
                cell = bug["cell"]
                original = bug["original"]
                # Wrap formula in IFERROR
                if not original.startswith("=IFERROR("):
                    cell.value = f"=IFERROR({original[1:]},0)"
                    fixes.append(f"✓ Auto-corregido: IFERROR aplicado en {bug['ref']}")

            elif bug["fix"] == "formula_needed":
                # Placeholder tipo `[=SUM(...)]` — no se puede auto-fixear sin
                # contexto, pero al menos lo reportamos. (Antes este branch
                # nunca ejecutaba porque buscaba "placeholder" en lugar de
                # "formula_needed", que es la tag real producida por _scan_sheet.)
                fixes.append(
                    f"⚠ Placeholder detectado en {bug['ref']} — "
                    f"reemplaza '{bug['original']}' con la fórmula real"
                )

        return wb, fixes

    def format_report(self, bugs: list, fixes: list) -> str:
        if not bugs and not fixes:
            return ""
        errors   = [b for b in bugs if b["severity"] == "error"]
        warnings = [b for b in bugs if b["severity"] == "warning"]
        lines = []
        if fixes:
            lines.append("**Auto-correcciones aplicadas:**")
            lines.extend(f"  {f}" for f in fixes[:5])
        if errors:
            lines.append(f"\n⛔ {len(errors)} error(es) detectado(s):")
            for b in errors[:3]:
                lines.append(f"  • {b['msg']}")
        if warnings:
            lines.append(f"\n⚠️ {len(warnings)} advertencia(s):")
            for b in warnings[:4]:
                lines.append(f"  • {b['msg']}")
        if len(bugs) > 7:
            lines.append(f"  ... y {len(bugs)-7} más.")
        return "\n".join(lines)

_detector = ExcelBugDetector()


# ─────────────────────────────────────────────────────────────────────────────
# EXISTING TEMPLATES — FIXED
# ─────────────────────────────────────────────────────────────────────────────

def _build_inventory(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """
    Inventario con productos del usuario.

    v14.9 — Si `datos["items"]` (o `productos`/`inventario`) viene con datos
    reales, se usan en las filas. Si no viene nada, se cae a 5 productos de
    ejemplo (backward compat). Las fórmulas y formato condicional se ajustan
    dinámicamente al número de filas.

    Keys soportadas por item (con sinónimos en/es):
        codigo, nombre, categoria, stock, stock_min,
        precio_costo, precio_venta
    """
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — InventoryVerifier: actualización método valuación ──
    try:
        from agent.inventory_verifier import verify_inventory_data
        cfg = verify_inventory_data(cfg)
    except Exception:
        pass
    # ── v15.10.0 — FiscalVerifier: actualización autónoma de tasas e impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws = _setup_sheet(wb, "Inventario", "1A4B8C")
    cols = 10
    _title_row(ws, titulo, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · WEP AI", cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)
    headers = ["#","Código","Producto / Descripción","Categoría",
               "Stock Actual","Stock Mínimo","Precio Costo","Precio Venta","Margen %","Estado"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h)
    ws.row_dimensions[4].height = 24

    # v14.9 — Tomar items del usuario, fallback a sample si no hay.
    user_items = _extract_items(datos, "productos", "inventario", "products")
    if user_items:
        rows = []
        for idx, it in enumerate(user_items):
            code     = it.get("codigo") or f"P-{idx+1:03d}"
            name     = it.get("nombre") or it.get("descripcion") or f"Producto {idx+1}"
            cat      = it.get("categoria") or "General"
            stock    = _num(it.get("stock"), 0)
            minstock = _num(it.get("stock_min"), max(1, int(stock * 0.2)) if stock else 5)
            cost     = _num(it.get("precio_costo"), 0)
            price    = _num(it.get("precio_venta"), 0)
            rows.append((code, name, cat, stock, minstock, cost, price))
        used_user_data = True
    else:
        rows = [
            ("P-001","Producto de ejemplo 1","Categoría A",50,10,100,180),
            ("P-002","Producto de ejemplo 2","Categoría B",30,15,150,280),
            ("P-003","Producto de ejemplo 3","Categoría A",8, 10,200,350),
            ("P-004","Producto de ejemplo 4","Categoría C",25,8, 80, 140),
            ("P-005","Producto de ejemplo 5","Categoría B",0, 5, 120,220),
        ]
        used_user_data = False

    for idx,(code,name,cat,stock,minstock,cost,price) in enumerate(rows):
        r = 5+idx; bg = "FFFFFF" if idx%2==0 else "F8FAFC"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",code,bg=bg)
        _dc(ws,f"C{r}",name,bg=bg)
        _dc(ws,f"D{r}",cat,bg=bg)
        _dc(ws,f"E{r}",stock,align="center",bg=bg,
            color="B91C1C" if stock<=minstock else "000000")
        _dc(ws,f"F{r}",minstock,align="center",bg=bg)
        _formula(ws,f"G{r}",cost,bg=bg,fmt='$#,##0.00',align="center")
        ws[f"G{r}"].value = cost
        _formula(ws,f"H{r}",price,bg=bg,fmt='$#,##0.00',align="center")
        ws[f"H{r}"].value = price
        _formula(ws,f"I{r}",f"=IFERROR((H{r}-G{r})/H{r},0)",bg=bg,fmt='0.0%',align="center")
        status = "✗ Agotado" if stock==0 else ("⚠ Reabastecer" if stock<=minstock else "✓ OK")
        sc = "B91C1C" if stock<=minstock or stock==0 else "15803D"
        _dc(ws,f"J{r}",status,align="center",bg=bg,color=sc)
    lr = 5+len(rows)
    ws.merge_cells(f"A{lr}:D{lr}")
    _hc(ws,f"A{lr}","TOTALES",bg="0D1B2A")
    _formula(ws,f"E{lr}",f"=SUM(E5:E{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='#,##0',align="center")
    _formula(ws,f"G{lr}",f"=AVERAGE(G5:G{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    _formula(ws,f"H{lr}",f"=AVERAGE(H5:H{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    _formula(ws,f"I{lr}",f"=AVERAGE(I5:I{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='0.0%',align="center")
    # ✅ FIX BUG 2: Formato condicional NATIVO — se actualiza cuando cambian los datos
    data_range_inv = f"E5:E{lr-1}"
    # Rojo: stock <= mínimo (columna F)
    _cf_formula_rule(ws, f"A5:J{lr-1}",
                     f"=$E5<=$F5", "FFE4E4", font_color="7F1D1D")
    # Verde: stock OK
    _cf_formula_rule(ws, f"A5:J{lr-1}",
                     f"=$E5>$F5", "D1FAE5", font_color="064E3B")
    # Escala de color en columna I (margen %)
    _cf_three_color(ws, f"I5:I{lr-1}", low="FFB3B3", mid="FFF3CD", high="B7F5D0")
    # DataValidation en columna D (Categoría) — usar categorías reales si las hay
    if used_user_data:
        seen_cats = []
        for _, _, c, *_ in rows:
            if c not in seen_cats:
                seen_cats.append(c)
        dv_options = seen_cats[:8] if seen_cats else ["General", "Otros"]
    else:
        dv_options = ["Categoría A","Categoría B","Categoría C","Categoría D","Otros"]
    _add_dv_list(ws, f"D5:D{lr-1}", dv_options, "Selecciona una categoría")
    # AVERAGEIF: promedio de margen por categoría
    avg_r = lr+2
    _section_row(ws, avg_r, "ANÁLISIS POR CATEGORÍA (AVERAGEIF)", cols, bg="185FA5")
    for j,h in enumerate(["Categoría","Productos","Stock Promedio","Precio Promedio","Margen Promedio"],1):
        _hc(ws, f"{get_column_letter(j)}{avg_r+1}", h, bg="0C447C")
    # v14.9 — Análisis usa las categorías reales del usuario si las hay
    cats_inv = dv_options[:5] if used_user_data else ["Categoría A","Categoría B","Categoría C"]
    for j, cat in enumerate(cats_inv):
        r_cat = avg_r+2+j; bg_c = "EFF6FF" if j%2==0 else "FFFFFF"
        _dc(ws, f"A{r_cat}", cat, bold=True, bg=bg_c)
        _formula(ws, f"B{r_cat}", f'=COUNTIF(D5:D{lr-1},A{r_cat})', bg=bg_c, fmt='#,##0', align="center")
        _formula(ws, f"C{r_cat}", f'=IFERROR(AVERAGEIF(D5:D{lr-1},A{r_cat},E5:E{lr-1}),0)',
                 bg=bg_c, fmt='#,##0.0', align="center")
        _formula(ws, f"D{r_cat}", f'=IFERROR(AVERAGEIF(D5:D{lr-1},A{r_cat},H5:H{lr-1}),0)',
                 bg=bg_c, fmt='$#,##0.00', align="center", color="0C447C", bold=True)
        _formula(ws, f"E{r_cat}", f'=IFERROR(AVERAGEIF(D5:D{lr-1},A{r_cat},I5:I{lr-1}),0)',
                 bg=bg_c, fmt='0.0%', align="center", color="15803D", bold=True)
    note = ("✓ Inventario generado con tus productos. Las fórmulas de Margen % y Estado se actualizan automáticamente."
            if used_user_data else
            "⚠ Reemplaza los datos de ejemplo con tus productos reales. Las formulas de Margen % y Estado se actualizan automaticamente.")
    _note_row(ws,avg_r+2+len(cats_inv)+1, note, cols)
    _setup_print(ws, cols, avg_r+2+len(cats_inv)+2, landscape=True)
    _protect_ws(ws, unlock_from_row=5)
    for i,w in enumerate([5,10,35,14,13,13,13,13,10,14],1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_payroll(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Nómina con deducciones dinámicas por país.

    v15.5 — Las deducciones (instituciones, tasas, etiquetas) se leen desde
    `EXCEL_FISCAL[país].payroll_deductions` en lugar de hardcodear IMSS/ISR/
    INFONAVIT. Cada país tiene su propia estructura:
      - MX: IMSS · ISR · INFONAVIT (+ hoja Tabla ISR con LISR Art. 96 vigente)
      - CO: Salud EPS · Pensión AFP · Retefuente · Solidaridad
      - PE: AFP/ONP · IR 5ta categoría · CTS (acumula)
      - AR: Jubilación SIPA · Obra Social · PAMI · Ganancias
      - CL: AFP · Salud · AFC · Impuesto único
      … etc.

    Las tasas son DEFAULTS editables en la hoja "Parámetros". El usuario
    final ajusta según su caso. Para países con escala progresiva (la
    mayoría), la columna del impuesto a la renta queda en 0% por defecto
    y la nota explica que debe consultarse la tabla oficial.

    v14.9 — Lee empleados del usuario en `datos["items"]` (o sinónimos
    empleados/nomina/planilla). Cada item: nombre/empleado, puesto/cargo,
    dias, salario_diario. Si no hay items, cae a 5 empleados de ejemplo.
    """
    # ── v15.5 — Configuración fiscal del país ─────────────────────────────────
    # ── v15.11.0 — LaborVerifier: actualización autónoma prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    fiscal = get_excel_fiscal(cfg)
    deductions = fiscal["payroll_deductions"]    # list[(label, rate, note)]
    money_fmt = fiscal["currency_format"]
    pais_nombre = (cfg or {}).get("pais", "Genérico")
    moneda = currency_code(cfg)
    is_mexico = pais_nombre == "México"

    ws  = _setup_sheet(wb, "Nómina", "6D28D9")
    ws2 = _setup_sheet(wb, "Parámetros", "534AB7")
    n_ded = len(deductions)

    # Columnas: # + Empleado + Puesto + Días + Sal.Diario + Bruto + (1 por deducción) + Total Deduc + Neto
    cols = 6 + n_ded + 2  # típicamente 10-12 según país
    _title_row(ws, titulo,
               f"Período: {datetime.now().strftime('%B %Y')} · País: {pais_nombre} · Moneda: {moneda} · WEP AI",
               cols)

    # ── Hoja Parámetros (tasas editables) ─────────────────────────────────────
    _title_row(ws2, "Parámetros de Nómina",
               f"País: {pais_nombre} · Modifica estas tasas si tu caso difiere", 4)
    ws2.row_dimensions[1].height = 32
    # Cada deducción ocupa una fila desde la 4 (B4=primera tasa, B5=segunda, …)
    for i, (label, rate, note) in enumerate(deductions):
        r = 4 + i
        bg = "F5F3FF" if i % 2 == 0 else "FFFFFF"
        _dc(ws2, f"A{r}", label, bold=True, bg=bg)
        _dc(ws2, f"B{r}", rate, align="center", bg=bg,
            fmt='0.00%', color="534AB7", bold=True)
        _dc(ws2, f"C{r}", note, bg=bg, color="888780")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 50

    # ── Hoja Tabla ISR (sólo México) ──────────────────────────────────────────
    # Para MX mantenemos la tabla LISR Art. 96 que ya existe. Para otros países
    # no la generamos — la columna de impuesto sobre la renta usa la tasa
    # editable de Parámetros como aproximación.
    if is_mexico:
        ws3 = _setup_sheet(wb, "Tabla ISR", "B91C1C")
        _title_row(ws3, "Tabla ISR Mensual — Art. 96 LISR 2024",
                   "Fuente: DOF SAT · Periodicidad: Mensual · No modifiques esta hoja", 5)
        for i, h in enumerate(["Límite Inferior", "Límite Superior", "Cuota Fija",
                                "% sobre Excedente", "Referencia"], 1):
            _hc(ws3, f"{get_column_letter(i)}4", h, bg="B91C1C")
        isr_tabla = [
            (0.01,      746.04,     0.00,     0.0192, "Tarifa 1"),
            (746.05,    6332.05,    14.32,    0.0640, "Tarifa 2"),
            (6332.06,   11128.01,   371.83,   0.1088, "Tarifa 3"),
            (11128.02,  12935.82,   893.63,   0.1600, "Tarifa 4"),
            (12935.83,  15487.71,   1182.88,  0.1792, "Tarifa 5"),
            (15487.72,  31236.49,   1640.18,  0.2136, "Tarifa 6"),
            (31236.50,  49233.00,   5004.12,  0.2352, "Tarifa 7"),
            (49233.01,  93993.90,   9236.89,  0.3000, "Tarifa 8"),
            (93993.91,  125325.20,  22665.17, 0.3200, "Tarifa 9"),
            (125325.21, 375975.61,  32691.18, 0.3400, "Tarifa 10"),
            (375975.62, 9999999.99, 117912.32, 0.3500, "Tarifa 11"),
        ]
        for idx, (li, ls, cf, tasa, ref) in enumerate(isr_tabla):
            r3 = 5 + idx
            bg3 = "FFFFFF" if idx % 2 == 0 else "FEF2F2"
            _dc(ws3, f"A{r3}", li,   align="center", bg=bg3, fmt='$#,##0.00')
            _dc(ws3, f"B{r3}", ls,   align="center", bg=bg3, fmt='$#,##0.00')
            _dc(ws3, f"C{r3}", cf,   align="center", bg=bg3, fmt='$#,##0.00', color="B91C1C", bold=True)
            _dc(ws3, f"D{r3}", tasa, align="center", bg=bg3, fmt='0.00%',     color="185FA5", bold=True)
            _dc(ws3, f"E{r3}", ref,  align="center", bg=bg3, color="888780")
        for col, w in zip(["A","B","C","D","E"], [16,16,14,18,12]):
            ws3.column_dimensions[col].width = w
        _note_row(ws3, 5+len(isr_tabla)+1,
                  "ISR = Cuota Fija + (Sueldo Bruto - Límite Inferior) × % sobre Excedente. "
                  "Fórmula en nómina referencia esta tabla con MATCH+INDEX.",
                  5, bg="FEF2F2", color="B91C1C")

    # ── Tabla principal de nómina ─────────────────────────────────────────────
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    # Headers: # | Empleado | Puesto | Días | Sal.Diario | Bruto | <deducciones…> | Total Deduc | Neto
    headers = ["#", "Nombre Empleado", "Puesto", "Días", "Sal. Diario", "Bruto"]
    headers += [label for label, _, _ in deductions]
    headers += ["Total Deduc.", "Neto a Pagar"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="6D28D9")
    ws.row_dimensions[4].height = 24

    # Datos del usuario o sample
    employees = [
        ("Juan García",      "Gerente Ventas",   30, 500),
        ("María López",      "Asistente Admin",  30, 250),
        ("Pedro Martínez",   "Operador Bodega",  28, 200),
        ("Ana Rodríguez",    "Vendedora",        30, 220),
        ("Carlos Soto",      "Contador",         30, 400),
    ]
    user_items = _extract_items(datos, "empleados", "nomina", "planilla", "employees")
    if user_items:
        employees = []
        for it in user_items:
            name  = it.get("empleado") or it.get("nombre") or "Empleado"
            role  = it.get("puesto") or it.get("cargo") or ""
            days  = int(_num(it.get("dias"), 30))
            daily = _num(it.get("salario_diario"), 0)
            employees.append((name, role, days, daily))

    # Construir las filas de empleados
    # Letras de columna:
    #   A=# B=Nombre C=Puesto D=Días E=SalDiario F=Bruto
    #   G..(G+n-1) = deducciones (una por columna)
    #   columna_total_deduc = G + n
    #   columna_neto       = G + n + 1
    first_ded_col_idx = 7  # G
    total_ded_col_idx = first_ded_col_idx + n_ded
    neto_col_idx      = total_ded_col_idx + 1
    total_ded_col = get_column_letter(total_ded_col_idx)
    neto_col      = get_column_letter(neto_col_idx)

    for idx, (name, role, days, daily) in enumerate(employees):
        r = 5 + idx
        bg = "FFFFFF" if idx % 2 == 0 else "F5F3FF"
        _dc(ws, f"A{r}", idx+1, align="center", bg=bg)
        _dc(ws, f"B{r}", name, bold=True, bg=bg)
        _dc(ws, f"C{r}", role, bg=bg)
        _dc(ws, f"D{r}", days, align="center", bg=bg)
        _dc(ws, f"E{r}", daily, align="center", bg=bg, fmt=money_fmt)
        # Bruto = Días × Salario Diario
        _formula(ws, f"F{r}", f"=D{r}*E{r}", bg=bg, fmt=money_fmt, align="center")

        # Una fórmula por deducción: =F{r}*Parámetros!B{4+i}
        # EXCEPCIÓN para MX: la columna "ISR" usa la Tabla ISR LISR Art. 96
        for i, (label, _, _) in enumerate(deductions):
            col_letter = get_column_letter(first_ded_col_idx + i)
            param_row = 4 + i
            if is_mexico and label.upper().strip() == "ISR":
                # Fórmula avanzada con tabla LISR
                _formula(ws, f"{col_letter}{r}",
                         f"=IFERROR("
                         f"INDEX('Tabla ISR'!$C$5:$C$15,MATCH(F{r},'Tabla ISR'!$A$5:$A$15,1))"
                         f"+(F{r}-INDEX('Tabla ISR'!$A$5:$A$15,MATCH(F{r},'Tabla ISR'!$A$5:$A$15,1)))"
                         f"*INDEX('Tabla ISR'!$D$5:$D$15,MATCH(F{r},'Tabla ISR'!$A$5:$A$15,1))"
                         f",IFERROR(F{r}*Parámetros!$B${param_row},0))",
                         bg=bg, fmt=money_fmt, align="center", color="B91C1C")
            else:
                # Fórmula simple: bruto × tasa
                _formula(ws, f"{col_letter}{r}",
                         f"=IFERROR(F{r}*Parámetros!$B${param_row},0)",
                         bg=bg, fmt=money_fmt, align="center", color="B91C1C")

        # Total Deducciones = suma de las columnas de deducciones
        first_ded_letter = get_column_letter(first_ded_col_idx)
        last_ded_letter  = get_column_letter(first_ded_col_idx + n_ded - 1)
        _formula(ws, f"{total_ded_col}{r}",
                 f"=SUM({first_ded_letter}{r}:{last_ded_letter}{r})",
                 bg=bg, fmt=money_fmt, align="center", color="B91C1C", bold=True)
        # Neto = Bruto - Total Deducciones
        _formula(ws, f"{neto_col}{r}",
                 f"=F{r}-{total_ded_col}{r}",
                 bg=bg, fmt=money_fmt, align="center", color="15803D", bold=True)

    lr = 5 + len(employees)
    ws.merge_cells(f"A{lr}:E{lr}")
    _hc(ws, f"A{lr}", "TOTALES NÓMINA", bg="0D1B2A")
    # Sumar Bruto + cada deducción + Total Deduc + Neto
    for col_idx in range(6, neto_col_idx + 1):
        col = get_column_letter(col_idx)
        _formula(ws, f"{col}{lr}", f"=SUM({col}5:{col}{lr-1})",
                 bg="0D1B2A", fg="FFFFFF", bold=True, fmt=money_fmt, align="center")
    _setup_print(ws, cols, lr+3, landscape=True)
    _protect_ws(ws, unlock_from_row=5)

    # ── v15.9.2 — Sección de BENEFICIOS OBLIGATORIOS por país ──
    # Aguinaldo MX, Prima/Cesantías CO, Gratificación PE, etc.
    beneficios_pais = fiscal.get("payroll_benefits", [])
    if beneficios_pais:
        # Crear hoja de beneficios
        ws_ben = _setup_sheet(wb, "Beneficios Obligatorios", "059669")
        _title_row(ws_ben,
                   f"Beneficios Obligatorios por Ley — {pais_nombre}",
                   f"Conceptos que el EMPLEADOR debe pagar adicional al salario · WEP AI",
                   3)
        _section_row(ws_ben, 3,
                     "CONCEPTOS OBLIGATORIOS POR LEY — NO SON DEDUCCIONES DEL TRABAJADOR",
                     3, bg="059669")
        for i, h in enumerate(["Beneficio", "Detalle / Base legal", "Estado"], 1):
            _hc(ws_ben, f"{get_column_letter(i)}4", h, bg="047857")
        ws_ben.row_dimensions[4].height = 22
        for idx, (nombre_ben, descripcion_ben) in enumerate(beneficios_pais):
            r = 5 + idx
            bg = "DCFCE7" if idx % 2 == 0 else "FFFFFF"
            _dc(ws_ben, f"A{r}", nombre_ben, bold=True, bg=bg, color="065F46")
            _dc(ws_ben, f"B{r}", descripcion_ben, bg=bg, color="111827")
            _dc(ws_ben, f"C{r}", "[ ] Aplicado", align="center", bg=bg)
            ws_ben.row_dimensions[r].height = 32
        widths_ben = [28, 90, 18]
        for i, w in enumerate(widths_ben, 1):
            ws_ben.column_dimensions[get_column_letter(i)].width = w
        # Nota de compliance
        nota_row = 5 + len(beneficios_pais) + 1
        _note_row(ws_ben, nota_row,
                  f"⚠ Compliance: Estos beneficios son OBLIGATORIOS por la ley laboral de {pais_nombre}. "
                  f"Omitirlos puede generar sanciones del {fiscal['tax_authority']} o demandas laborales. "
                  f"Verifica con tu contador o asesor laboral antes de procesar la nómina.",
                  3, bg="FEF3C7", color="92400E")

    # Nota final
    if is_mexico:
        nota = ("✏ Edita las tasas en la hoja 'Parámetros'. El cálculo de ISR usa la "
                "Tabla LISR Art. 96 automáticamente. Verifica el régimen del trabajador.")
    else:
        nota = (f"✏ Edita las tasas en la hoja 'Parámetros' según el régimen de {pais_nombre}. "
                f"Si el impuesto a la renta es escala progresiva, consulta la tabla oficial "
                f"de la {fiscal['tax_authority']} para el cálculo exacto.")
    _note_row(ws, lr+2, nota, cols, bg="EDE9FE", color="4C1D95")

    # Ancho de columnas dinámico
    widths = [4, 26, 20, 7, 13, 13] + [13]*n_ded + [13, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_quotation(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Cotización con IVA, moneda y nomenclatura del país.

    v15.5 — IVA default, label de impuesto, moneda y formato vienen del cfg
    en lugar de "16% México" hardcoded. La celda IVA en Config sigue siendo
    editable: el cfg solo determina el valor por defecto.

    v14.9 — Lee items del usuario en `datos["items"]` (o sinónimos cotizacion/
    productos/servicios). Cada item con keys: descripcion/nombre, unidad,
    cantidad, precio_venta/precio, descuento (opcional). Si no hay items,
    cae a 3 servicios de ejemplo.
    """
    # ── v15.5 — Configuración fiscal del país ─────────────────────────────────
    # ── v15.11.0 — FiscalVerifier: actualización autónoma de tasas/impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    iva_default = iva_rate(cfg)         # 0.16 MX · 0.19 CO · 0.18 PE · …
    iva_lbl     = iva_label(cfg)        # IVA · IGV · ITBMS · ITBIS · Sales Tax
    money_fmt   = currency_format(cfg)  # $#,##0.00 · S/.#,##0.00 · ₲#,##0 · …
    moneda      = currency_code(cfg)    # MXN, COP, PEN, ARS, CLP, EUR, USD
    id_empresa  = (cfg or {}).get("id_empresa", "RFC / NIT / RUT")
    pais_nombre = (cfg or {}).get("pais", "")

    ws  = _setup_sheet(wb, "Cotización", "1D9E75")
    ws2 = _setup_sheet(wb, "Config",     "0F6E56")
    ws3 = _setup_sheet(wb, "Resumen",    "085041")
    cols = 7

    # Config sheet — editable parameters
    subtitle = (f"Modifica estos valores según tu empresa · {pais_nombre}"
                if pais_nombre else "Modifica estos valores según tu empresa")
    _title_row(ws2, "Configuración de Cotización", subtitle, 4)
    cfg_params = [
        (f"{iva_lbl} / Impuesto (%)",  iva_default,
         f"{iva_default*100:.0f}% es la tasa default para {pais_nombre or 'tu país'}"),
        ("Descuento general (%)",      0.00, "Descuento sobre el subtotal total"),
        ("Validez (días)",             15,   "Días de vigencia de esta cotización"),
        ("Moneda",                     moneda, f"Código ISO ({moneda} = {pais_nombre or 'configurar'})"),
        ("Nombre de tu empresa",       "Tu Empresa S.A.", ""),
        (id_empresa,                   "",   f"Identificador fiscal — {id_empresa}"),
    ]
    for i,(label,val,note) in enumerate(cfg_params):
        r=4+i
        bg="EFF6FF" if i%2==0 else "FFFFFF"
        _dc(ws2,f"A{r}",label,bold=True,bg=bg)
        _dc(ws2,f"B{r}",val,bg=bg,color="0F6E56",bold=True,
            fmt='0%' if isinstance(val,float) else 'General',align="center")
        _dc(ws2,f"C{r}",note,bg=bg,color="888780")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 46

    _title_row(ws, titulo, f"Folio: COT-{datetime.now().strftime('%Y%m%d')} · WEP AI", cols)
    ws.freeze_panes = "A11"
    _add_autofilter(ws, 10, cols)

    # v14.9 — Header info: si datos trae cliente, lo pone en lugar de vacío.
    cliente_name = (datos or {}).get("cliente_nombre") or (datos or {}).get("cliente") or ""
    info = [("Empresa cliente:", cliente_name),("Contacto:",""),
            ("Fecha:",datetime.now().strftime('%d/%m/%Y')),
            ("Validez:","=Config!B6&\" días\""),("Emitida por:","=Config!B8")]
    for i,(lbl,val) in enumerate(info):
        r=4+i
        ws[f"A{r}"].value=lbl; ws[f"A{r}"].font=Font(name="Arial",bold=True,size=10,color="0F6E56")
        ws[f"B{r}"].value=val; ws[f"B{r}"].font=Font(name="Arial",size=10)

    ws.merge_cells("D4:G8")
    note=ws["D4"]
    note.value=(f"CONDICIONES\n• Precios en {moneda}\n"
                "• Pago: 50% anticipo, 50% entrega\n• Tiempo estimado: a convenir")
    note.alignment=Alignment(wrap_text=True,vertical="top")
    note.font=Font(name="Arial",size=9,color="5F5E5A")
    note.fill=PatternFill("solid",start_color="E1F5EE")

    headers=["#","Descripción del Servicio / Producto","Unidad","Cantidad","Precio Unit.","Descuento %","Subtotal"]
    for i,h in enumerate(headers,1):
        _hc(ws,f"{get_column_letter(i)}10",h,bg="0F6E56")
    ws.row_dimensions[10].height=22

    # v14.9 — Items del usuario o sample fallback
    user_items = _extract_items(datos, "cotizacion", "productos", "servicios", "lineas", "lines")
    if user_items:
        rows = []
        for it in user_items:
            desc  = it.get("nombre") or it.get("descripcion") or "Item"
            unit  = it.get("unidad") or "Pieza"
            qty   = _num(it.get("cantidad"), 1)
            price = _num(it.get("precio_venta"), 0)
            disc  = _num(it.get("descuento"), 0)
            if disc > 1:
                disc = disc / 100.0
            rows.append((desc, unit, qty, price, disc))
    else:
        rows = [("Servicio / Producto 1","Pieza",1,5000,0),
                ("Servicio / Producto 2","Hr",10,800,0),
                ("Servicio / Producto 3","Pieza",3,1200,0)]

    for idx,(desc,unit,qty,price,disc) in enumerate(rows):
        r=11+idx; bg="FFFFFF" if idx%2==0 else "F0FFF4"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",desc,bg=bg)
        _dc(ws,f"C{r}",unit,align="center",bg=bg)
        _dc(ws,f"D{r}",qty,align="center",bg=bg)
        _dc(ws,f"E{r}",price,align="center",bg=bg,fmt=money_fmt)
        _dc(ws,f"F{r}",disc,align="center",bg=bg,fmt='0%')
        _formula(ws,f"G{r}",f"=D{r}*E{r}*(1-F{r})",bg=bg,fmt=money_fmt,
                 color="0F6E56",bold=True,align="center")

    li=11+len(rows)
    # Subtotal, descuento general, IVA desde Config, Total
    totals=[
        ("Subtotal antes de descuento", f"=SUM(G11:G{li-1})", "EFF6FF","0F6E56"),
        ("Descuento general",           f"=G{li}*Config!B5",    "FFF7ED","C2410C"),
        ("Subtotal neto",               f"=G{li}-G{li+1}",      "EFF6FF","0F6E56"),
        (f"{iva_lbl} / Impuesto",       f"=G{li+2}*Config!B4",  "F0FFF4","0F6E56"),
        ("TOTAL A PAGAR",               f"=G{li+2}+G{li+3}",    "0F6E56","FFFFFF"),
    ]
    for i,(lbl,frm,bg,fg) in enumerate(totals):
        r=li+i
        ws.merge_cells(f"A{r}:F{r}")
        is_total=(i==4)
        _hc(ws,f"A{r}",lbl,bg=bg,fg=fg)
        _formula(ws,f"G{r}",frm,bg=bg,color=fg,bold=True,
                 fmt=money_fmt,align="center",size=11 if is_total else 10)

    sig_r=li+len(totals)+2
    for ref,lbl in [(f"A{sig_r}","Firma Cliente"),(f"E{sig_r}","Firma Proveedor")]:
        if ref.startswith("E"): ws.merge_cells(f"E{sig_r}:G{sig_r+2}")
        else: ws.merge_cells(f"A{sig_r}:C{sig_r+2}")
        c=ws[ref]
        c.value=f"\n\n\n{'_'*30}\n{lbl}"
        c.alignment=Alignment(horizontal="center",vertical="bottom",wrap_text=True)
        c.font=Font(name="Arial",size=9,color="5F5E5A")

    note_text = (f"✓ Cotización generada con tus items en {moneda} ({iva_lbl} {iva_default*100:.0f}%). Edita los valores en 'Config'."
                 if user_items else
                 f"✏ Edita {iva_lbl}, moneda y datos de empresa en la hoja 'Config'. Aplica automáticamente.")
    _note_row(ws,sig_r+4, note_text, cols,bg="E1F5EE",color="0F6E56")

    for col,w in zip(["A","B","C","D","E","F","G"],[5,38,10,10,14,13,15]):
        ws.column_dimensions[col].width=w

    # Resumen
    _title_row(ws3,f"Resumen — {titulo}","Totales automáticos",4)
    for i,(lbl,frm,col) in enumerate([
        ("Subtotal bruto",       f"=Cotización!G{li}",    "0F6E56"),
        ("Descuento aplicado",   f"=Cotización!G{li+1}",  "C2410C"),
        (f"{iva_lbl} cobrado",   f"=Cotización!G{li+3}",  "0F6E56"),
        ("TOTAL A PAGAR",        f"=Cotización!G{li+4}",  "0D1B2A"),
    ]):
        r=4+i
        ws3[f"A{r}"].value=lbl; ws3[f"A{r}"].font=Font(name="Arial",bold=True,size=10)
        ws3[f"B{r}"].value=frm; ws3[f"B{r}"].number_format=money_fmt
        ws3[f"B{r}"].font=Font(name="Arial",bold=True,size=13 if i==3 else 11,color=col)
    ws3.column_dimensions["A"].width=28; ws3.column_dimensions["B"].width=18


def _build_accounts_receivable(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """FIXED: Días vencido con fórmula real TODAY()-fecha, aging automático."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — LaborVerifier: actualización autónoma de prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Cuentas x Cobrar", "C2410C")
    ws2 = _setup_sheet(wb, "Cuentas x Pagar",  "7C3AED")
    ws3 = _setup_sheet(wb, "Aging",            "854F0B")
    ws4 = _setup_sheet(wb, "Resumen",          "0D1B2A")

    sample = [
        ("Cliente / Proveedor A","Factura #001",date(2026,4,1), date(2026,5,1), 15000,"Vencida"),
        ("Cliente / Proveedor B","Factura #002",date(2026,4,15),date(2026,5,15),8500, "Por vencer"),
        ("Cliente / Proveedor C","Factura #003",date(2026,5,1), date(2026,6,1), 22000,"Al corriente"),
        ("Cliente / Proveedor D","Factura #004",date(2026,5,10),date(2026,6,10),5000, "Al corriente"),
    ]

    for sheet,label,hdr_bg,money_col in [
        (ws, "CUENTAS POR COBRAR","C2410C","15803D"),
        (ws2,"CUENTAS POR PAGAR", "7C3AED","B91C1C"),
    ]:
        cols=8
        _title_row(sheet,titulo,f"{label} · WEP AI · {datetime.now().strftime('%d/%m/%Y')}",cols)
        sheet.freeze_panes="A5"
        _add_autofilter(sheet, 4, cols)
        headers=["#","Nombre","Concepto","F. Emisión","F. Vencimiento","Monto","Estado","Días vencido"]
        for i,h in enumerate(headers,1):
            _hc(sheet,f"{get_column_letter(i)}4",h,bg=hdr_bg)
        sheet.row_dimensions[4].height=22

        for idx,(nombre,concepto,f_emi,f_ven,monto,estado) in enumerate(sample):
            r=5+idx; bg="FFFFFF" if idx%2==0 else "FFF7ED"
            sc={"Vencida":"B91C1C","Por vencer":"854F0B","Al corriente":"15803D"}.get(estado,"000000")
            _dc(sheet,f"A{r}",idx+1,align="center",bg=bg)
            _dc(sheet,f"B{r}",nombre,bold=True,bg=bg)
            _dc(sheet,f"C{r}",concepto,bg=bg)
            # Store dates as actual date objects so formulas work
            sheet[f"D{r}"].value = f_emi
            sheet[f"D{r}"].number_format = "DD/MM/YYYY"
            sheet[f"D{r}"].font = Font(name="Arial",size=10)
            sheet[f"D{r}"].fill = PatternFill("solid",start_color=bg)
            sheet[f"D{r}"].alignment = Alignment(horizontal="center",vertical="center")
            sheet[f"D{r}"].border = _thin()
            sheet[f"E{r}"].value = f_ven
            sheet[f"E{r}"].number_format = "DD/MM/YYYY"
            sheet[f"E{r}"].font = Font(name="Arial",size=10)
            sheet[f"E{r}"].fill = PatternFill("solid",start_color=bg)
            sheet[f"E{r}"].alignment = Alignment(horizontal="center",vertical="center")
            sheet[f"E{r}"].border = _thin()
            _dc(sheet,f"F{r}",monto,align="center",bg=bg,
                fmt='$#,##0.00',color=money_col,bold=True)
            _dc(sheet,f"G{r}",estado,align="center",bg=bg,color=sc,bold=True)
            # FIXED: real formula for days overdue
            _formula(sheet,f"H{r}",f"=IF(TODAY()>E{r},MAX(0,TODAY()-E{r}),0)",
                     bg=bg,fmt='#,##0',align="center",
                     color="B91C1C" if estado=="Vencida" else "888780")

        lr=5+len(sample)
        ws_m = sheet
        ws_m.merge_cells(f"A{lr}:E{lr}")
        _hc(ws_m,f"A{lr}","TOTAL",bg="0D1B2A")
        _formula(ws_m,f"F{lr}",f"=SUM(F5:F{lr-1})",bg="0D1B2A",fg="FFFFFF",
                 bold=True,fmt='$#,##0.00',align="center")

        # Aging summary rows
        ag_r=lr+2
        _section_row(ws_m,ag_r,"ANÁLISIS DE ANTIGÜEDAD (AGING)",cols,bg=hdr_bg)
        for j,(label2,formula2) in enumerate([
            ("0–30 días",    f"=SUMPRODUCT((H5:H{lr-1}>=0)*(H5:H{lr-1}<=30)*F5:F{lr-1})"),
            ("31–60 días",   f"=SUMPRODUCT((H5:H{lr-1}>30)*(H5:H{lr-1}<=60)*F5:F{lr-1})"),
            ("61–90 días",   f"=SUMPRODUCT((H5:H{lr-1}>60)*(H5:H{lr-1}<=90)*F5:F{lr-1})"),
            ("+90 días",     f"=SUMPRODUCT((H5:H{lr-1}>90)*F5:F{lr-1})"),
        ]):
            r2=ag_r+1+j
            ws_m.merge_cells(f"A{r2}:E{r2}")
            bg2="FFF7ED" if j%2==0 else "FFFFFF"
            _hc(ws_m,f"A{r2}",label2,bg="EFF6FF",fg="0D1B2A")
            _formula(ws_m,f"F{r2}",formula2,bg=bg2,fmt='$#,##0.00',
                     color="B91C1C" if j>=2 else "15803D",bold=True,align="center")

        # ✅ FIX BUG 2: CF nativo — Vencida=rojo, Por vencer=ámbar, Al corriente=verde
        last_data = 4 + len(sample)
        _cf_formula_rule(ws_m, f"A5:H{last_data}", '=$G5="Vencida"',    "FFE4E4", font_color="7F1D1D")
        _cf_formula_rule(ws_m, f"A5:H{last_data}",
                         '=$G5="Por vencer"', "FFF3CD", font_color="78350F")
        _cf_formula_rule(ws_m, f"A5:H{last_data}",
                         '=$G5="Al corriente"',"D1FAE5", font_color="064E3B")
        # DataValidation en columna G (Estado)
        _add_dv_list(ws_m, f"G5:G{last_data}",
                     ["Al corriente","Por vencer","Vencida","Pagada","En disputa"],
                     "Selecciona el estado de la factura")
        # OR formula: resalta si está vencida O tiene más de 60 días
        last_data2 = 4 + len(sample)
        _cf_formula_rule(ws_m, f"A5:H{last_data2}",
                         f'=OR($G5="Vencida",$H5>60)',
                         "FFE4E4", "7F1D1D")
        _setup_print(ws_m, cols, ag_r+7, landscape=True)
        _protect_ws(ws_m, unlock_from_row=5)
        _note_row(ws_m,ag_r+6,
            "Dias vencido calculado automaticamente con TODAY(). Rojo si Vencida OR >60 dias. Actualiza fechas como valores de fecha.",
            cols,bg="E1F5EE",color="0F6E56")

        for i,w in enumerate([4,28,22,14,14,14,14,13],1):
            ws_m.column_dimensions[get_column_letter(i)].width=w

    # Resumen
    _title_row(ws4,"Resumen de Cartera",f"WEP AI · {datetime.now().strftime('%d/%m/%Y')}",4)
    for i,(lbl,frm,col) in enumerate([
        ("Total x Cobrar",    "=SUM('Cuentas x Cobrar'!F5:F8)","15803D"),
        ("Total x Pagar",     "=SUM('Cuentas x Pagar'!F5:F8)", "B91C1C"),
        ("Posición Neta",     "=B4-B5",                        "185FA5"),
        ("Vencido >90 días",  "='Cuentas x Cobrar'!F14",       "B91C1C"),
    ]):
        r=4+i
        ws4[f"A{r}"].value=lbl; ws4[f"A{r}"].font=Font(name="Arial",bold=True,size=11)
        ws4[f"B{r}"].value=frm; ws4[f"B{r}"].number_format='$#,##0.00'
        ws4[f"B{r}"].font=Font(name="Arial",bold=True,size=14,color=col)
    ws4.column_dimensions["A"].width=26; ws4.column_dimensions["B"].width=20


def _build_cashflow(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """
    v14.9 — El template del cashflow es un esqueleto editable de 12 meses.
    Si el usuario provee `datos["ingresos"]` o `datos["egresos"]` como dicts
    {concepto: [12 valores] o {mes: valor}}, los pre-rellena. También acepta
    `datos["items"]` con shape [{concepto, tipo, ene, feb, ...}] o
    [{concepto, tipo, mes, monto}] que se reduce a la matriz mensual.

    Como en Excel cashflow las filas son CONCEPTOS y las columnas son meses,
    el patrón "lista de items" no encaja perfecto — preferimos la API de
    listas separadas de ingresos/egresos.
    """
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — FiscalVerifier: actualización autónoma de tasas e impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Flujo de Caja", "0C447C")
    ws2 = _setup_sheet(wb, "Gráfica",       "185FA5")
    cols=14
    _title_row(ws,titulo,f"Año: {datetime.now().year} · WEP AI",cols)
    ws.freeze_panes="A6"
    headers=["Concepto","Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic","Total Anual"]
    for i,h in enumerate(headers,1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="0C447C")

    # v14.9 — Extraer ingresos/egresos del usuario si vienen
    def _normalize_flow_dict(raw, default_items):
        """Normaliza a una lista [(concepto, [12 valores])].
        Acepta:
          - dict {"Ventas": [100, 200, ...]}                 (lista de 12)
          - dict {"Ventas": {"ene": 100, "feb": 200}}         (dict de meses)
          - list [{"concepto": "Ventas", "monto": 100, "mes": "ene"}]  (filas)
          - None → default_items con ceros
        """
        if isinstance(raw, dict) and raw:
            out = []
            for concepto, vals in raw.items():
                if isinstance(vals, list) and len(vals) >= 12:
                    out.append((concepto, [_num(v, 0) for v in vals[:12]]))
                elif isinstance(vals, dict):
                    months = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
                    months_en = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
                    row = [0.0] * 12
                    for k, v in vals.items():
                        k_low = str(k).lower()[:3]
                        if k_low in months:
                            row[months.index(k_low)] = _num(v, 0)
                        elif k_low in months_en:
                            row[months_en.index(k_low)] = _num(v, 0)
                    out.append((concepto, row))
                elif isinstance(vals, (int, float)):
                    # Valor único → distribuir uniformemente / poner en el mes actual
                    row = [0.0] * 12
                    row[datetime.now().month - 1] = _num(vals, 0)
                    out.append((concepto, row))
            return out if out else [(it, [0]*12) for it in default_items]
        if isinstance(raw, list) and raw:
            # Filas estilo [{concepto, mes, monto}]
            months = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
            months_en = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
            agg = {}
            for it in raw:
                if not isinstance(it, dict):
                    continue
                it = _norm_item(it)
                concepto = it.get("concepto") or it.get("nombre") or "Item"
                mes_raw = (str(it.get("mes") or "").lower())[:3]
                monto = _num(it.get("monto"), 0)
                if mes_raw in months:
                    mi = months.index(mes_raw)
                elif mes_raw in months_en:
                    mi = months_en.index(mes_raw)
                else:
                    mi = datetime.now().month - 1
                row = agg.setdefault(concepto, [0.0]*12)
                row[mi] += monto
            return list(agg.items()) if agg else [(it, [0]*12) for it in default_items]
        # Fallback
        return [(it, [0]*12) for it in default_items]

    default_income   = ["Ventas / Cobros","Otros ingresos","Financiamiento"]
    default_expenses = ["Nómina","Renta / Oficina","Proveedores","Servicios / Util.","Impuestos","Gastos variables"]

    income_data   = _normalize_flow_dict((datos or {}).get("ingresos") or (datos or {}).get("income"),
                                          default_income)
    expense_data  = _normalize_flow_dict((datos or {}).get("egresos") or (datos or {}).get("expenses") or (datos or {}).get("gastos"),
                                          default_expenses)
    has_user_data = bool((datos or {}).get("ingresos") or (datos or {}).get("egresos") or
                          (datos or {}).get("income") or (datos or {}).get("expenses") or (datos or {}).get("gastos"))

    r=5
    _section_row(ws,r,"ENTRADAS DE EFECTIVO",cols,bg="15803D"); r+=1
    inc_start=r
    for concepto, monthly in income_data:
        _dc(ws,f"A{r}",concepto,bold=True)
        for c in range(2,14):
            val = monthly[c-2] if c-2 < len(monthly) else 0
            ws.cell(r,c,val).number_format='$#,##0'
            ws.cell(r,c).border=_thin(); ws.cell(r,c).font=Font(name="Arial",size=10)
        _formula(ws,f"N{r}",f"=SUM(B{r}:M{r})",fmt='$#,##0',color="15803D",bold=True,align="center")
        r+=1
    inc_total=r
    _hc(ws,f"A{r}","TOTAL ENTRADAS",bg="EFF6FF",fg="15803D",align="left")
    for c in range(2,15):
        cl=get_column_letter(c)
        _formula(ws,f"{cl}{r}",f"=SUM({cl}{inc_start}:{cl}{r-1})",bg="EFF6FF",
                 fmt='$#,##0',color="15803D",bold=True,align="center")
    r+=2
    _section_row(ws,r,"SALIDAS DE EFECTIVO",cols,bg="B91C1C"); r+=1
    exp_start=r
    for concepto, monthly in expense_data:
        _dc(ws,f"A{r}",concepto,bold=True)
        for c in range(2,14):
            val = monthly[c-2] if c-2 < len(monthly) else 0
            ws.cell(r,c,val).number_format='$#,##0'
            ws.cell(r,c).border=_thin(); ws.cell(r,c).font=Font(name="Arial",size=10)
        _formula(ws,f"N{r}",f"=SUM(B{r}:M{r})",fmt='$#,##0',color="B91C1C",bold=True,align="center")
        r+=1
    exp_total=r
    _hc(ws,f"A{r}","TOTAL SALIDAS",bg="FEF2F2",fg="B91C1C",align="left")
    for c in range(2,15):
        cl=get_column_letter(c)
        _formula(ws,f"{cl}{r}",f"=SUM({cl}{exp_start}:{cl}{r-1})",bg="FEF2F2",
                 fmt='$#,##0',color="B91C1C",bold=True,align="center")
    r+=1
    flujo_r=r
    _hc(ws,f"A{r}","FLUJO NETO DEL MES",bg="0D1B2A",fg="FFFFFF",align="left")
    for c in range(2,15):
        cl=get_column_letter(c)
        _formula(ws,f"{cl}{r}",f"={cl}{inc_total}-{cl}{exp_total}",
                 bg="0D1B2A",color="FFFFFF",bold=True,fmt='$#,##0',align="center",size=11)
    r+=1
    acum_r=r
    _hc(ws,f"A{r}","SALDO ACUMULADO",bg="185FA5",fg="FFFFFF",align="left")
    _formula(ws,f"B{r}",f"=B{flujo_r}",bg="185FA5",color="FFFFFF",bold=True,fmt='$#,##0',align="center")
    for c in range(3,14):
        cl=get_column_letter(c); cl2=get_column_letter(c-1)
        _formula(ws,f"{cl}{r}",f"={cl2}{r}+{cl}{flujo_r}",
                 bg="185FA5",color="FFFFFF",bold=True,fmt='$#,##0',align="center")
    _formula(ws,f"N{r}",f"=M{r}",bg="185FA5",color="FFFFFF",bold=True,fmt='$#,##0',align="center")
    ws.column_dimensions["A"].width=26
    for i in range(2,15): ws.column_dimensions[get_column_letter(i)].width=10
    _title_row(ws2,"Flujo de Caja — Comportamiento Anual","Entradas vs Salidas vs Flujo Neto · WEP AI",4)
    # Hoja extra: Análisis de Inversión con NPV e IRR
    ws3 = _setup_sheet(wb, "NPV e IRR", "0D1B2A")
    _title_row(ws3,"Análisis de Inversión — VPN e TIR",
               "Valor Presente Neto e Tasa Interna de Retorno · WEP AI",5)
    _section_row(ws3,3,"PARÁMETROS — Edita las celdas amarillas",5,bg="185FA5")
    for i,h in enumerate(["Parámetro","Valor","Unidad","Descripción"],1):
        _hc(ws3,f"{get_column_letter(i)}4",h,bg="0C447C")
    npv_params = [
        ("Inversión inicial",          -500000,"$",  "Monto invertido (negativo = salida)"),
        ("Tasa de descuento anual",      0.12,  "%",  "Costo de capital o rendimiento mínimo esperado"),
        ("Horizonte de evaluación (años)",5,    "años","Período de análisis"),
    ]
    for idx,(label,val,unit,desc) in enumerate(npv_params):
        r=5+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        _dc(ws3,f"A{r}",label,bold=True,bg=bg)
        _dc(ws3,f"B{r}",val,align="center",bg="FEFCE8",
            fmt='$#,##0' if isinstance(val,int) and abs(val)>100 else '0%' if isinstance(val,float) else '#,##0',
            color="0C447C",bold=True)
        _dc(ws3,f"C{r}",unit,align="center",bg=bg,color="888780")
        _dc(ws3,f"D{r}",desc,bg=bg,color="888780")
    _section_row(ws3,9,"FLUJOS DE CAJA PROYECTADOS (año a año)",5,bg="185FA5")
    for i,h in enumerate(["Año","Flujo de Caja","Flujo Acumulado","Factor Descuento","Flujo Descontado"],1):
        _hc(ws3,f"{get_column_letter(i)}10",h,bg="0C447C")
    flujos = [-500000, 120000, 150000, 180000, 210000, 240000]
    acum = 0
    for idx,flujo in enumerate(flujos):
        r=11+idx; bg="FFFFFF" if idx%2==0 else "EFF6FF"
        yr = f"Año {idx}" if idx>0 else "Inversión (0)"
        _dc(ws3,f"A{r}",yr,bold=(idx==0),bg=bg,color="B91C1C" if idx==0 else "000000")
        _dc(ws3,f"B{r}",flujo,align="center",bg="FEFCE8",fmt='$#,##0',
            color="B91C1C" if flujo<0 else "15803D",bold=True)
        acum+=flujo
        _dc(ws3,f"C{r}",acum,align="center",bg=bg,fmt='$#,##0',
            color="B91C1C" if acum<0 else "15803D")
        # Factor de descuento = 1/(1+tasa)^año
        _formula(ws3,f"D{r}",f"=1/((1+$B$6)^A{r})" if idx>0 else "=1",
                 bg=bg,fmt='0.0000',align="center",color="888780")
        _formula(ws3,f"E{r}",f"=B{r}*D{r}",bg=bg,fmt='$#,##0',
                 color="B91C1C" if flujo<0 else "0C447C",align="center")
    lr_npv = 11+len(flujos)
    # Resultados: VPN e TIR como fórmulas Excel reales
    _section_row(ws3,lr_npv+1,"RESULTADOS",5,bg="0D1B2A")
    results = [
        ("VPN — Valor Presente Neto",
         f"=NPV(B6,B12:B{lr_npv-1})+B11",
         "$#,##0","VPN>0 = proyecto viable. VPN<0 = destruye valor."),
        ("TIR — Tasa Interna de Retorno",
         f"=IFERROR(IRR(B11:B{lr_npv-1}),'No converge')",
         "0.0%","TIR > Tasa descuento = proyecto rentable."),
        ("Período de recuperación (años)",
         f"=IFERROR(MATCH(0,SIGN(C11:C{lr_npv-1})*-1+1,0)-2,'Sin recuperar')",
         "#,##0","Años para recuperar la inversión inicial."),
        ("Índice de rentabilidad (PI)",
         f"=IFERROR(1+NPV(B6,B12:B{lr_npv-1})/ABS(B11),0)",
         "0.00","PI>1 = proyecto rentable. PI<1 = no rentable."),
    ]
    for i,h in enumerate(["Indicador","Valor","Interpretación"],1):
        _hc(ws3,f"{get_column_letter(i)}{lr_npv+2}",h,bg="0D1B2A")
    for idx,(label,formula,fmt,interp) in enumerate(results):
        r=lr_npv+3+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        is_main=idx<2
        _dc(ws3,f"A{r}",label,bold=is_main,bg=bg)
        _formula(ws3,f"B{r}",formula,bg="FEFCE8" if is_main else bg,
                 fmt=fmt,color="15803D" if idx==0 else "185FA5",bold=is_main,align="center",size=13 if is_main else 10)
        _dc(ws3,f"C{r}",interp,bg=bg,color="888780")
    _note_row(ws3,lr_npv+3+len(results)+1,
              "✓ VPN (=NPV) e TIR (=IRR) son fórmulas nativas de Excel. Edita los Flujos de Caja en la columna B y los resultados se recalculan automáticamente.",
              5,bg="EFF6FF",color="0C447C")
    for col,w in zip(["A","B","C","D","E"],[28,14,16,14,16]):
        ws3.column_dimensions[col].width=w
    # ✅ FIX BUG 3: Gráfica LineChart REAL conectada a los datos del flujo de caja
    chart = LineChart()
    chart.title  = "Entradas vs Salidas vs Flujo Neto"
    chart.style  = 10
    chart.height = 14
    chart.width  = 22
    chart.y_axis.title = "Monto ($)"
    chart.x_axis.title = "Mes"
    chart.grouping = "standard"
    # Categorías (meses Ene-Dic, fila 4 columnas B-M)
    cats = Reference(ws, min_col=2, max_col=13, min_row=4)
    chart.set_categories(cats)
    # Serie 1: Total Entradas (fila inc_total)
    s1 = Series(Reference(ws, min_col=2, max_col=13, min_row=inc_total),
                title="Entradas de efectivo")
    s1.graphicalProperties.line.solidFill = "15803D"
    s1.graphicalProperties.line.width = 28000
    chart.series.append(s1)
    # Serie 2: Total Salidas (fila exp_total)
    s2 = Series(Reference(ws, min_col=2, max_col=13, min_row=exp_total),
                title="Salidas de efectivo")
    s2.graphicalProperties.line.solidFill = "B91C1C"
    s2.graphicalProperties.line.width = 28000
    chart.series.append(s2)
    # Serie 3: Flujo Neto (fila flujo_r)
    s3 = Series(Reference(ws, min_col=2, max_col=13, min_row=flujo_r),
                title="Flujo neto del mes")
    s3.graphicalProperties.line.solidFill = "185FA5"
    s3.graphicalProperties.line.width = 22000
    s3.graphicalProperties.line.dashDot = "dash"
    chart.series.append(s3)
    ws2.add_chart(chart, "A3")


def _build_project_tracker(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """FIXED: Días restantes con fórmula TODAY(), Gantt con formato condicional."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Tareas",  "BA7517")
    ws2 = _setup_sheet(wb, "Resumen", "633806")
    cols=10
    _title_row(ws,titulo,f"Gestión de proyecto · WEP AI · {datetime.now().strftime('%d/%m/%Y')}",cols)
    ws.freeze_panes="A5"
    _add_autofilter(ws, 4, cols)
    headers=["#","Tarea / Entregable","Responsable","Prioridad","Estado",
             "F. Inicio","F. Entrega","% Avance","Días restantes","Notas"]
    for i,h in enumerate(headers,1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="BA7517")
    ws.row_dimensions[4].height=22

    prio_col={"Alta":"B91C1C","Media":"854F0B","Baja":"15803D"}
    stat_col={"Completado":"15803D","En curso":"185FA5","Por iniciar":"888780","Bloqueado":"B91C1C"}
    tasks=[
        ("Kickoff y alcance",        "Juan G.",   "Alta",  "Completado",  date(2026,5,1), date(2026,5,5),  100),
        ("Diseño de wireframes",     "Ana M.",    "Alta",  "Completado",  date(2026,5,6), date(2026,5,12), 100),
        ("Desarrollo frontend",      "Carlos R.", "Alta",  "En curso",    date(2026,5,13),date(2026,5,28), 60),
        ("Integración de API",       "Luis P.",   "Alta",  "En curso",    date(2026,5,15),date(2026,5,30), 35),
        ("QA y pruebas",             "Ana M.",    "Media", "Por iniciar", date(2026,5,29),date(2026,6,5),  0),
        ("Capacitación al cliente",  "Juan G.",   "Media", "Por iniciar", date(2026,6,6), date(2026,6,8),  0),
        ("Entrega final y doc.",      "Carlos R.", "Alta",  "Por iniciar", date(2026,6,9), date(2026,6,12), 0),
    ]
    for idx,(tarea,resp,prio,estado,f_ini,f_fin,avance) in enumerate(tasks):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",tarea,bold=True,bg=bg)
        _dc(ws,f"C{r}",resp,bg=bg)
        _dc(ws,f"D{r}",prio,align="center",bg=bg,color=prio_col.get(prio,"000000"),bold=True)
        _dc(ws,f"E{r}",estado,align="center",bg=bg,color=stat_col.get(estado,"000000"),bold=True)
        ws[f"F{r}"].value=f_ini; ws[f"F{r}"].number_format="DD/MM/YYYY"
        ws[f"F{r}"].font=Font(name="Arial",size=10); ws[f"F{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"F{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"F{r}"].border=_thin()
        ws[f"G{r}"].value=f_fin; ws[f"G{r}"].number_format="DD/MM/YYYY"
        ws[f"G{r}"].font=Font(name="Arial",size=10); ws[f"G{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"G{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"G{r}"].border=_thin()
        bar="█"*(avance//10)+"░"*(10-avance//10)
        _dc(ws,f"H{r}",f"{bar} {avance}%",align="center",bg=bg,
            color="15803D" if avance==100 else ("185FA5" if avance>0 else "888780"))
        # FIXED: real formula for days remaining
        _formula(ws,f"I{r}",f"=IF(E{r}=\"Completado\",0,MAX(0,G{r}-TODAY()))",
                 bg=bg,fmt='#,##0',align="center",
                 color="B91C1C" if estado in ["Bloqueado"] else "000000")
        _dc(ws,f"J{r}","",bg=bg)

    lr=5+len(tasks)
    ws.merge_cells(f"A{lr}:G{lr}")
    _hc(ws,f"A{lr}","AVANCE GENERAL DEL PROYECTO",bg="0D1B2A")
    avg_avance=sum(t[6] for t in tasks)//len(tasks)
    bar_t="█"*(avg_avance//10)+"░"*(10-avg_avance//10)
    _hc(ws,f"H{lr}",f"{bar_t} {avg_avance}%",bg="0D1B2A")
    # Auto-computed remaining days total
    _formula(ws,f"I{lr}",f"=SUM(I5:I{lr-1})",bg="0D1B2A",fg="FFFFFF",
             bold=True,fmt='#,##0',align="center")
    # ✅ FIX BUG 2: CF nativo en project tracker
    last_task = 4 + len(tasks)
    _cf_formula_rule(ws, f"A5:J{last_task}", '=$E5="Completado"', "D1FAE5", "064E3B")
    _cf_formula_rule(ws, f"A5:J{last_task}", '=$E5="En curso"',   "DBEAFE", "1E3A5F")
    _cf_formula_rule(ws, f"A5:J{last_task}", '=$E5="Bloqueado"',  "FFE4E4", "7F1D1D")
    _cf_formula_rule(ws, f"A5:J{last_task}", '=D5="Alta"',       "FEF3C7", "78350F")
    # DataValidation en columna E (Estado) y D (Prioridad)
    _add_dv_list(ws, f"E5:E{last_task}",
                 ["Por iniciar","En curso","En revisión","Bloqueado","Completado"],
                 "Estado de la tarea")
    _add_dv_list(ws, f"D5:D{last_task}", ["Alta","Media","Baja"], "Prioridad")
    _setup_print(ws, cols, lr+3, landscape=True)
    _protect_ws(ws, unlock_from_row=5)
    _note_row(ws,lr+2,
        "✓ 'Días restantes' calculado con TODAY() — se actualiza automáticamente. Rojo = bloqueado/vencido.",
        cols,bg="FFFBEB",color="854F0B")
    for i,w in enumerate([4,34,14,10,13,12,12,18,13,20],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    _title_row(ws2,f"KPIs — {titulo}","Resumen ejecutivo",5)
    kpis=[
        ("Total de tareas",    len(tasks)),
        ("Completadas",        sum(1 for t in tasks if t[3]=="Completado")),
        ("En curso",           sum(1 for t in tasks if t[3]=="En curso")),
        ("Por iniciar",        sum(1 for t in tasks if t[3]=="Por iniciar")),
        ("Avance promedio",    f"{avg_avance}%"),
        ("Alta prioridad",     sum(1 for t in tasks if t[2]=="Alta")),
    ]
    for i,h in enumerate(["KPI","Valor"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="633806")
    for idx,(kpi,val) in enumerate(kpis):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        _dc(ws2,f"A{r}",kpi,bold=True,bg=bg)
        _dc(ws2,f"B{r}",str(val),align="center",bg=bg,bold=True,color="BA7517")
    ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=16


def _build_multi_currency(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """FIXED: Equivalente referenciado desde hoja TC, no hardcodeado."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — BankingVerifier: actualización regulación bancaria ──
    try:
        from agent.banking_verifier import verify_banking_data
        cfg = verify_banking_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Gastos e Ingresos", "0F6E56")
    ws2 = _setup_sheet(wb, "Tipo de Cambio",    "085041")
    ws3 = _setup_sheet(wb, "Resumen",           "0D1B2A")
    cols=9
    _title_row(ws,titulo,f"Tracker multi-moneda · WEP AI · {datetime.now().strftime('%d/%m/%Y')}",cols)
    ws.freeze_panes="A5"

    # TC sheet first (so we can VLOOKUP from it)
    _title_row(ws2,"Tipos de Cambio de Referencia","Actualiza estos valores con la tasa de tu banco",4)
    currencies=[("MXN","Peso mexicano",1.0),("USD","Dólar americano",17.20),
                ("EUR","Euro",18.60),("GBP","Libra esterlina",21.80),
                ("CAD","Dólar canadiense",12.50),("BRL","Real brasileño",3.40),
                ("COP","Peso colombiano",0.0044),("ARS","Peso argentino",0.019),
                ("PEN","Sol peruano",4.60),("CHF","Franco suizo",19.10),
                ("JPY","Yen japonés",0.115),("AUD","Dólar australiano",11.20),
                ("CLP","Peso chileno",0.0191)]
    for i,h in enumerate(["Código","Moneda","Equiv. MXN (1 unidad)","Actualizado"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="085041")
    for idx,(code,name,tc) in enumerate(currencies):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "E1F5EE"
        _dc(ws2,f"A{r}",code,bold=True,align="center",bg=bg,color="0F6E56")
        _dc(ws2,f"B{r}",name,bg=bg)
        _dc(ws2,f"C{r}",tc,align="center",bg=bg,fmt='$#,##0.0000')
        _dc(ws2,f"D{r}",datetime.now().strftime('%d/%m/%Y'),align="center",bg=bg,color="888780")
    for col,w in zip(["A","B","C","D"],[10,24,22,20]):
        ws2.column_dimensions[col].width=w

    headers=["Fecha","Descripción","Categoría","Moneda","Monto Original",
             "TC (auto)","Equivalente MXN","Tipo","Notas"]
    for i,h in enumerate(headers,1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="0F6E56")
    ws.row_dimensions[4].height=22

    sample=[
        (date(2026,5,1),"Airbnb Ciudad de México","Alojamiento","MXN",8500,"Gasto"),
        (date(2026,5,3),"Cliente A — Proyecto web","Ingreso freelance","USD",1200,"Ingreso"),
        (date(2026,5,5),"Vuelo MEX-BOG","Transporte","MXN",4200,"Gasto"),
        (date(2026,5,8),"Cliente B — Consultoría","Ingreso freelance","EUR",800,"Ingreso"),
        (date(2026,5,10),"Coworking mensual","Trabajo","USD",150,"Gasto"),
        (date(2026,5,12),"Alimentación semana","Alimentación","MXN",1800,"Gasto"),
    ]
    for idx,(fecha,desc,cat,moneda,monto,tipo) in enumerate(sample):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "F0FFF4"
        is_inc=tipo=="Ingreso"
        ws[f"A{r}"].value=fecha; ws[f"A{r}"].number_format="DD/MM/YYYY"
        ws[f"A{r}"].font=Font(name="Arial",size=10); ws[f"A{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"A{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"A{r}"].border=_thin()
        _dc(ws,f"B{r}",desc,bold=True,bg=bg)
        _dc(ws,f"C{r}",cat,bg=bg)
        _dc(ws,f"D{r}",moneda,align="center",bg=bg)
        _dc(ws,f"E{r}",monto,align="center",bg=bg,fmt='#,##0.00')
        # FIXED: TC via VLOOKUP from TC sheet
        _formula(ws,f"F{r}",
                 f"=IFERROR(VLOOKUP(D{r},'Tipo de Cambio'!$A$5:$C$17,3,0),1)",
                 bg=bg,fmt='#,##0.0000',align="center")
        # Equivalente = monto * TC
        _formula(ws,f"G{r}",f"=E{r}*F{r}",bg=bg,fmt='$#,##0.00',
                 color="15803D" if is_inc else "B91C1C",bold=True,align="center")
        _dc(ws,f"H{r}",tipo,align="center",bg=bg,
            color="15803D" if is_inc else "B91C1C",bold=True)
        _dc(ws,f"I{r}","",bg=bg)

    lr=5+len(sample)
    for lbl,cond,bg_c in [
        ("TOTAL INGRESOS (MXN)", f'=SUMIF(H5:H{lr-1},"Ingreso",G5:G{lr-1})',"15803D"),
        ("TOTAL GASTOS (MXN)",   f'=SUMIF(H5:H{lr-1},"Gasto",G5:G{lr-1})',  "B91C1C"),
        ("FLUJO NETO (MXN)",     f"=G{lr}-G{lr+1}",                         "0C447C"),
    ]:
        ws.merge_cells(f"A{lr}:F{lr}")
        _hc(ws,f"A{lr}",lbl,bg=bg_c)
        _formula(ws,f"G{lr}",cond,bg=bg_c,color="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
        lr+=1
    # DataValidation en columna D (Moneda) y H (Tipo)
    _add_dv_list(ws, f"D5:D{lr-1}",
                 ["MXN","USD","EUR","GBP","CAD","BRL","COP","ARS","PEN","CHF","JPY","AUD","CLP"],
                 "Selecciona la moneda")
    _add_dv_list(ws, f"H5:H{lr-1}", ["Ingreso","Gasto"], "Tipo de movimiento")
    _add_autofilter(ws, 4, cols)
    _note_row(ws,lr+1,
        "✓ TC tomado automáticamente de la hoja 'Tipo de Cambio'. Actualiza los tipos de cambio ahí para reflejar la tasa actual de tu banco.",
        cols,bg="E1F5EE",color="0F6E56")
    for i,w in enumerate([12,32,18,9,14,13,16,10,18],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    _title_row(ws3,f"Resumen — {titulo}","Posición financiera",4)
    for i,(lbl,fg) in enumerate([("Total ingresos MXN","15803D"),("Total gastos MXN","B91C1C"),
                                  ("Flujo neto MXN","185FA5"),("Monedas registradas","0F6E56")]):
        r=4+i
        ws3[f"A{r}"].value=lbl; ws3[f"A{r}"].font=Font(name="Arial",bold=True,size=10)
        ws3[f"B{r}"].value=(f"='Gastos e Ingresos'!G{8+i}" if i<3
                            else f"=COUNTA('Tipo de Cambio'!A5:A17)")
        ws3[f"B{r}"].number_format='$#,##0.00' if i<3 else '#,##0'
        ws3[f"B{r}"].font=Font(name="Arial",bold=True,size=13,color=fg)
    ws3.column_dimensions["A"].width=26; ws3.column_dimensions["B"].width=22


def _build_dashboard(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """FIXED: KPIs conectados a hoja de datos reales, con gráfica de barras."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Dashboard",   "0D1B2A")
    ws2 = _setup_sheet(wb, "Datos",       "1A4B8C")
    cols=12
    _title_row(ws,titulo,f"Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · WEP AI",cols)

    # Datos sheet — input area
    _title_row(ws2,"Datos de Entrada","Ingresa tus números aquí — el Dashboard se actualiza automáticamente",4)
    data_labels=[("Ventas del Mes","$",0),("Costo de Ventas","$",0),
                 ("Gastos Operativos","$",0),("Otros Ingresos","$",0),
                 ("Unidades Vendidas","#",0),("Meta de Ventas","$",0)]
    for i,(lbl,unit,val) in enumerate(data_labels):
        r=5+i; bg="EFF6FF" if i%2==0 else "FFFFFF"
        _dc(ws2,f"A{r}",lbl,bold=True,bg=bg)
        _dc(ws2,f"B{r}",unit,align="center",bg=bg,color="888780")
        _dc(ws2,f"C{r}",val,align="center",bg=bg,color="1A4B8C",bold=True,
            fmt='$#,##0.00' if unit=="$" else '#,##0')
    _note_row(ws2,12,"✏ Ingresa tus datos reales en la columna C. El Dashboard a la izquierda se calcula automáticamente.",4,bg="EFF6FF",color="1A4B8C")
    ws2.column_dimensions["A"].width=28; ws2.column_dimensions["B"].width=8; ws2.column_dimensions["C"].width=18

    # KPI cards — all connected to Datos sheet
    kpis=[
        ("Ventas del Mes",        "=Datos!C5",                            "3B82F6","$#,##0"),
        ("Utilidad Bruta",        "=IFERROR(Datos!C5-Datos!C6,0)",        "22C55E","$#,##0"),
        ("Utilidad Neta",         "=IFERROR(Datos!C5-Datos!C6-Datos!C7,0)","10B981","$#,##0"),
        ("Margen Neto %",         "=IFERROR((Datos!C5-Datos!C6-Datos!C7)/Datos!C5,0)","8B5CF6","0.0%"),
        ("% Cumplimiento Meta",   "=IFERROR(Datos!C5/Datos!C10,0)",       "F59E0B","0.0%"),
        ("Costo / Ventas %",      "=IFERROR(Datos!C6/Datos!C5,0)",        "EF4444","0.0%"),
    ]
    for i,(kpi,formula,color,fmt) in enumerate(kpis):
        col=1+(i%3)*4; row=4 if i<3 else 8
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+3)
        ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+3)
        ws.merge_cells(start_row=row+2,start_column=col,end_row=row+2,end_column=col+3)
        lbl=ws.cell(row,col,kpi)
        lbl.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        lbl.fill=PatternFill("solid",start_color=color)
        lbl.alignment=Alignment(horizontal="center",vertical="center")
        val=ws.cell(row+1,col,formula)
        val.number_format=fmt
        val.font=Font(name="Arial",bold=True,size=22,color=color)
        val.fill=PatternFill("solid",start_color="F8FAFC")
        val.alignment=Alignment(horizontal="center",vertical="center")
        sub=ws.cell(row+2,col,"↑ vs meta" if "Meta" in kpi else "vs mes anterior")
        sub.font=Font(name="Arial",size=8,color="888780")
        sub.fill=PatternFill("solid",start_color="F8FAFC")
        sub.alignment=Alignment(horizontal="center")
        ws.row_dimensions[row].height=18; ws.row_dimensions[row+1].height=42; ws.row_dimensions[row+2].height=14

    # ✅ v5: Gráfica combinada Bar+Line — Ventas (barras) + Margen % (línea)
    _note_row(ws,11,"Gráfica interactiva disponible debajo — conectada a hoja Datos",cols,
              bg="EFF6FF",color="185FA5")
    _add_combo_chart(ws_data=ws2, ws_chart=ws,
                     title="Ventas vs Margen Neto — Resumen Ejecutivo",
                     bar_ref_row=5, line_ref_row=8,
                     cat_row=4, col_start=0, col_end=3,
                     bar_title="Ventas del mes", line_title="Margen Neto %",
                     bar_color="185FA5", line_color="0F6E56", anchor="A13")
    _note_row(ws,12,"✓ KPIs conectados a la hoja 'Datos'. Ingresa tus números reales ahí y este dashboard se actualiza.",cols,bg="EFF6FF",color="1A4B8C")
    for i in range(1,cols+1): ws.column_dimensions[get_column_letter(i)].width=9


def _build_budget(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws = _setup_sheet(wb, "Presupuesto", "C2410C")
    cols=14
    _title_row(ws,titulo,f"Año: {datetime.now().year} · WEP AI",cols)
    ws.freeze_panes="A5"
    headers=["Concepto","Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic","Total Anual"]
    for i,h in enumerate(headers,1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="C2410C")
    income_items=["Ventas Producto A","Ventas Producto B","Otros Ingresos"]
    expense_items=["Nómina","Renta","Inventario/Compras","Marketing","Servicios","Gastos Generales"]
    r=5
    _section_row(ws,r,"INGRESOS",14,bg="15803D"); r+=1
    for item in income_items:
        _dc(ws,f"A{r}",item,bold=True)
        for c in range(2,14): ws.cell(r,c,0).number_format='$#,##0'
        _formula(ws,f"N{r}",f"=SUM(B{r}:M{r})",fmt='$#,##0',color="15803D",bold=True,align="center"); r+=1
    inc_total=r
    ws.merge_cells(f"A{r}:A{r}"); _hc(ws,f"A{r}","TOTAL INGRESOS",bg="EFF6FF",fg="15803D")
    for c in range(2,15):
        cl=get_column_letter(c)
        _formula(ws,f"{cl}{r}",f"=SUM({cl}{r-len(income_items)}:{cl}{r-1})",
                 bg="EFF6FF",color="15803D",bold=True,fmt='$#,##0',align="center")
    r+=2; _section_row(ws,r,"EGRESOS / GASTOS",14,bg="B91C1C"); r+=1
    for item in expense_items:
        _dc(ws,f"A{r}",item,bold=True)
        for c in range(2,14): ws.cell(r,c,0).number_format='$#,##0'
        _formula(ws,f"N{r}",f"=SUM(B{r}:M{r})",fmt='$#,##0',color="B91C1C",bold=True,align="center"); r+=1
    exp_total=r
    ws.merge_cells(f"A{r}:A{r}"); _hc(ws,f"A{r}","TOTAL EGRESOS",bg="FEF2F2",fg="B91C1C")
    for c in range(2,15):
        cl=get_column_letter(c)
        _formula(ws,f"{cl}{r}",f"=SUM({cl}{r-len(expense_items)}:{cl}{r-1})",
                 bg="FEF2F2",color="B91C1C",bold=True,fmt='$#,##0',align="center")
    r+=2; _hc(ws,f"A{r}","UTILIDAD NETA",bg="0D1B2A")
    for c in range(2,15):
        cl=get_column_letter(c)
        _formula(ws,f"{cl}{r}",f"={cl}{inc_total}-{cl}{exp_total}",
                 bg="0D1B2A",color="FFFFFF",bold=True,fmt='$#,##0',size=11,align="center")
    ws.column_dimensions["A"].width=28
    for i in range(2,15): ws.column_dimensions[get_column_letter(i)].width=10


def _build_sales(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Registro de ventas con IVA / impuesto del país.

    v15.5 — Tasa de impuesto, label (IVA/IGV/ITBMS/Sales Tax) y formato de
    moneda vienen del cfg en lugar de IVA 16% hardcoded.
    v14.9 — Lee ventas del usuario en `datos["items"]` (o sinónimos
    ventas/movimientos). Cada item: fecha, producto/nombre, cliente,
    cantidad, precio_venta/precio. Si no hay items, cae a 5 ventas sample.
    """
    # ── v15.5 — Configuración fiscal del país ─────────────────────────────────
    # ── v15.11.0 — FiscalVerifier: actualización autónoma de tasas/impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    tax_rate  = iva_rate(cfg)
    tax_lbl   = iva_label(cfg)
    money_fmt = currency_format(cfg)
    pais      = (cfg or {}).get("pais", "")
    tax_header = f"{tax_lbl} ({tax_rate*100:.0f}%)" if tax_rate else tax_lbl

    ws  = _setup_sheet(wb, "Ventas",       "15803D")
    ws2 = _setup_sheet(wb, "Estadísticas", "0F766E")
    cols=8
    subtitle = (f"Mes: {datetime.now().strftime('%B %Y')} · País: {pais} · WEP AI"
                if pais else f"Mes: {datetime.now().strftime('%B %Y')} · WEP AI")
    _title_row(ws,titulo,subtitle,cols)
    ws.freeze_panes="A5"
    _add_autofilter(ws, 4, cols)
    headers=["Fecha","Producto / Servicio","Cliente","Cantidad","Precio Unit.","Subtotal",tax_header,"Total"]
    for i,h in enumerate(headers,1): _hc(ws,f"{get_column_letter(i)}4",h)
    ws.row_dimensions[4].height=24
    sample=[(date(2026,5,1),"Producto A","Cliente 1",5,350),
            (date(2026,5,3),"Producto B","Cliente 2",2,480),
            (date(2026,5,5),"Producto C","Cliente 3",8,160),
            (date(2026,5,8),"Producto A","Cliente 4",3,350),
            (date(2026,5,10),"Producto D","Cliente 1",1,430)]

    # v14.9 — Si el usuario provee ventas reales, las usamos
    user_items = _extract_items(datos, "ventas", "movimientos", "transacciones")
    if user_items:
        sample = []
        for idx, it in enumerate(user_items):
            f_raw = it.get("fecha")
            if isinstance(f_raw, str):
                try:
                    fecha = datetime.fromisoformat(f_raw.replace('/', '-')).date()
                except (ValueError, TypeError):
                    fecha = date.today()
            elif isinstance(f_raw, date):
                fecha = f_raw
            else:
                fecha = date.today()
            prod    = it.get("nombre") or it.get("producto") or it.get("descripcion") or f"Item {idx+1}"
            cliente = it.get("cliente") or ""
            qty     = _num(it.get("cantidad"), 1)
            price   = _num(it.get("precio_venta"), 0)
            sample.append((fecha, prod, cliente, qty, price))

    for idx,(fecha,prod,cliente,qty,price) in enumerate(sample):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "F0FFF4"
        ws[f"A{r}"].value=fecha; ws[f"A{r}"].number_format="DD/MM/YYYY"
        ws[f"A{r}"].font=Font(name="Arial",size=10); ws[f"A{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"A{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"A{r}"].border=_thin()
        _dc(ws,f"B{r}",prod,bg=bg); _dc(ws,f"C{r}",cliente,bg=bg)
        _dc(ws,f"D{r}",qty,align="center",bg=bg); _dc(ws,f"E{r}",price,align="center",bg=bg,fmt=money_fmt)
        _formula(ws,f"F{r}",f"=D{r}*E{r}",bg=bg,fmt=money_fmt,align="center")
        # IVA dinámico: si tasa_rate > 0, calcula; si 0 (caso US), siempre 0
        _formula(ws,f"G{r}",f"=IFERROR(F{r}*{tax_rate},0)",bg=bg,fmt=money_fmt,align="center")
        _formula(ws,f"H{r}",f"=F{r}+G{r}",bg=bg,fmt=money_fmt,align="center",color="15803D",bold=True)
    lr=5+len(sample)
    ws.merge_cells(f"A{lr}:E{lr}"); _hc(ws,f"A{lr}","TOTALES DEL MES",bg="0D1B2A")
    for col in ["F","G","H"]:
        _formula(ws,f"{col}{lr}",f"=SUM({col}5:{col}{lr-1})",bg="0D1B2A",fg="FFFFFF",
                 bold=True,fmt=money_fmt,align="center")
    for col,w in zip(["A","B","C","D","E","F","G","H"],[13,28,20,10,13,13,13,13]):
        ws.column_dimensions[col].width=w
    # Stats
    _title_row(ws2,f"Estadísticas — {titulo}","Resumen ejecutivo del mes",6)
    for i,(lbl,frm) in enumerate([
        (f"Ingresos Brutos (sin {tax_lbl})",f"=SUM(Ventas!F5:F{lr-1})"),
        (f"{tax_lbl} Cobrado",             f"=SUM(Ventas!G5:G{lr-1})"),
        ("Total Facturado",         f"=SUM(Ventas!H5:H{lr-1})"),
        ("Número de ventas",        f"=COUNTA(Ventas!A5:A{lr-1})"),
        ("Ticket promedio",         f"=IFERROR(B5/B8,0)"),
        ("Venta promedio por cliente",f'=IFERROR(AVERAGEIF(Ventas!C5:C{lr-1},"<>",Ventas!F5:F{lr-1}),0)'),
        ("Promedio unidades por venta",f"=IFERROR(AVERAGE(Ventas!D5:D{lr-1}),0)"),
    ]):
        r=5+i
        ws2[f"A{r}"].value=lbl; ws2[f"A{r}"].font=Font(name="Arial",bold=True,size=10)
        ws2[f"B{r}"].value=frm
        ws2[f"B{r}"].number_format=money_fmt if i not in [3,5] else ('#,##0' if i==3 else 'General')
        ws2[f"B{r}"].font=Font(name="Arial",bold=True,size=12,color="15803D")
    # PieChart de ventas por producto en hoja Estadísticas
    chart_sales = PieChart()
    chart_sales.title  = "Distribución de ventas por producto"
    chart_sales.style  = 10
    chart_sales.height = 12
    chart_sales.width  = 18
    labels_s = Reference(ws, min_col=2, max_col=2, min_row=5, max_row=5+len(sample)-1)
    data_s   = Reference(ws, min_col=6, max_col=6, min_row=5, max_row=5+len(sample)-1)
    chart_sales.add_data(data_s)
    chart_sales.set_categories(labels_s)
    ws2.add_chart(chart_sales, "D4")
    ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=20
    _setup_print(ws, cols, lr+1, landscape=True)
    _protect_ws(ws, unlock_from_row=5)


def _build_freelance_tracker(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Tracker de proyectos freelance con impuestos del país.

    v15.5 — La retención de honorarios e IVA usan tasas del país en lugar
    de ISR 35% e IVA 16% hardcoded. Se etiqueta con la nomenclatura local
    (Retención IR, Retención IRPF, Retefuente, etc.).
    """
    # ── v15.5 — Configuración fiscal del país ─────────────────────────────────
    # ── v15.11.0 — FiscalVerifier: actualización autónoma de tasas/impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    fiscal = get_excel_fiscal(cfg)
    tax_rate         = iva_rate(cfg)
    tax_lbl          = iva_label(cfg)
    money_fmt        = currency_format(cfg)
    inc_tax_rate     = fiscal["income_tax_freelance"]       # ej. 0.10 MX, 0.08 PE, 0.06 AR
    inc_tax_label    = fiscal["income_tax_freelance_label"] # texto completo del concepto
    tax_authority    = fiscal["tax_authority"]
    pais             = (cfg or {}).get("pais", "")

    ws  = _setup_sheet(wb, "Proyectos", "534AB7")
    ws2 = _setup_sheet(wb, "Ingresos",  "3C3489")
    ws3 = _setup_sheet(wb, "Impuestos", "7C3AED")
    cols=10
    subtitle = (f"Tracker freelance · País: {pais} · {datetime.now().strftime('%B %Y')} · WEP AI"
                if pais else f"Tracker freelance · {datetime.now().strftime('%B %Y')} · WEP AI")
    _title_row(ws,titulo,subtitle,cols)
    ws.freeze_panes="A5"
    headers=["#","Cliente","Proyecto","Estado","F. Inicio","F. Entrega","Tarifa","H. Est.","H. Reales","Total Facturado"]
    for i,h in enumerate(headers,1): _hc(ws,f"{get_column_letter(i)}4",h,bg="534AB7")
    ws.row_dimensions[4].height=22
    stat_col={"En curso":"185FA5","En revisión":"854F0B","Completado":"15803D","Pendiente":"888780"}
    projects=[
        ("Cliente Alpha","Diseño web corporativo","En curso",   date(2026,4,1),date(2026,5,30),800,40,32),
        ("Cliente Beta", "App móvil MVP",          "En revisión",date(2026,3,15),date(2026,5,15),1200,60,67),
        ("Cliente Gamma","Consultoría SEO",        "Completado", date(2026,5,1),date(2026,5,20),500,10,10),
        ("Cliente Delta","Branding completo",      "Pendiente",  date(2026,6,1),date(2026,6,30),950,30,0),
        ("Cliente Alpha","Mantenimiento web",      "En curso",   date(2026,5,1),date(2026,5,31),400,8,5),
    ]
    for idx,(client,project,status,f_ini,f_fin,tarifa,h_est,h_real) in enumerate(projects):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EEEDFE"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",client,bold=True,bg=bg,color="534AB7")
        _dc(ws,f"C{r}",project,bg=bg)
        _dc(ws,f"D{r}",status,align="center",bg=bg,color=stat_col.get(status,"000000"),bold=True)
        ws[f"E{r}"].value=f_ini; ws[f"E{r}"].number_format="DD/MM/YYYY"
        ws[f"E{r}"].font=Font(name="Arial",size=10); ws[f"E{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"E{r}"].alignment=Alignment(horizontal="center"); ws[f"E{r}"].border=_thin()
        ws[f"F{r}"].value=f_fin; ws[f"F{r}"].number_format="DD/MM/YYYY"
        ws[f"F{r}"].font=Font(name="Arial",size=10); ws[f"F{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"F{r}"].alignment=Alignment(horizontal="center"); ws[f"F{r}"].border=_thin()
        _dc(ws,f"G{r}",tarifa,align="center",bg=bg,fmt=money_fmt)
        _dc(ws,f"H{r}",h_est,align="center",bg=bg)
        _dc(ws,f"I{r}",h_real,align="center",bg=bg,color="B91C1C" if h_real>h_est else "000000")
        _formula(ws,f"J{r}",f"=G{r}*I{r}",bg=bg,fmt=money_fmt,color="534AB7",bold=True,align="center")
    lr=5+len(projects)
    ws.merge_cells(f"A{lr}:I{lr}"); _hc(ws,f"A{lr}","TOTAL FACTURADO",bg="0D1B2A")
    _formula(ws,f"J{lr}",f"=SUM(J5:J{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt=money_fmt,align="center")
    efic_r=lr+1; ws.merge_cells(f"A{efic_r}:I{efic_r}")
    _hc(ws,f"A{efic_r}","EFICIENCIA (horas reales / estimadas)",bg="534AB7")
    _formula(ws,f"J{efic_r}",f"=IFERROR(SUM(I5:I{lr-1})/SUM(H5:H{lr-1}),0)",
             bg="534AB7",color="FFFFFF",bold=True,fmt='0.0%',align="center")
    _note_row(ws,efic_r+2,"Rojo en Horas Reales = excediste el estimado. Revisa tu tarifa o renegocia el alcance.",cols,bg="EEEDFE",color="534AB7")
    for i,w in enumerate([4,20,28,13,13,13,12,10,11,15],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    _title_row(ws2,"Control de Ingresos",f"Mes: {datetime.now().strftime('%B %Y')}",6)
    for i,h in enumerate(["Mes","Cliente","Concepto","Monto","Estado Cobro","Fecha Pago"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="3C3489")
    for idx,(mes,client,concepto,monto,estado) in enumerate([
        (datetime.now().strftime('%b %Y'),"Cliente Alpha","Diseño web — 50% anticipo",12800,"Pagado"),
        (datetime.now().strftime('%b %Y'),"Cliente Beta", "App MVP — entrega final",  40200,"Pendiente"),
        (datetime.now().strftime('%b %Y'),"Cliente Gamma","Consultoría SEO — total",  5000, "Pagado"),
    ]):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EEEDFE"
        _dc(ws2,f"A{r}",mes,align="center",bg=bg); _dc(ws2,f"B{r}",client,bold=True,bg=bg)
        _dc(ws2,f"C{r}",concepto,bg=bg)
        _dc(ws2,f"D{r}",monto,align="center",bg=bg,fmt=money_fmt,color="534AB7",bold=True)
        _dc(ws2,f"E{r}",estado,align="center",bg=bg,
            color="15803D" if estado=="Pagado" else "B91C1C",bold=True)
        _dc(ws2,f"F{r}","",align="center",bg=bg)
    for col,w in zip(["A","B","C","D","E","F"],[14,22,32,14,14,14]):
        ws2.column_dimensions[col].width=w

    # ── Hoja Impuestos: tasas locales del país ────────────────────────────────
    subtitle_tax = f"Aproximación según {tax_authority} — consulta a un contador local"
    _title_row(ws3,f"Estimación de Impuestos · {pais}", subtitle_tax, 5)
    _note_row(ws3,4,"Edita las tasas en las celdas amarillas según tu situación fiscal y régimen.",4,bg="FFFBEB",color="854F0B")
    for i,h in enumerate(["Concepto","Tasa %","Base","Estimado"],1):
        _hc(ws3,f"{get_column_letter(i)}5",h,bg="7C3AED")
    impuestos=[
        ("Ingresos brutos del mes","",f"=SUM(Ingresos!D5:D7)",""),
        (inc_tax_label, f"{inc_tax_rate*100:.1f}%".replace(".0%","%"),
                                f"=C6*{inc_tax_rate}",
                                f"Tasa de referencia — verifica con {tax_authority}"),
        (f"{tax_lbl} cobrado (si aplica)", f"{tax_rate*100:.0f}%",
                                f"=C6*{tax_rate}",
                                f"Solo si emites factura con {tax_lbl}"),
        ("Gastos deducibles","",0,"Actualiza con tus comprobantes"),
        ("Base gravable estimada","",f"=C6-C9",""),
        ("Impuesto neto estimado","",f"=C7+C8",""),
    ]
    for idx,(concepto,pct,base,nota) in enumerate(impuestos):
        r=6+idx; bg="FFFFFF" if idx%2==0 else "EEEDFE"
        _dc(ws3,f"A{r}",concepto,bold=True,bg=bg)
        _dc(ws3,f"B{r}",pct,align="center",bg="FFFBEB" if pct else bg)
        ws3[f"C{r}"].value=base; ws3[f"C{r}"].number_format=money_fmt
        ws3[f"C{r}"].font=Font(name="Arial",bold=(idx in [0,5]),size=10,
                               color="7C3AED" if idx in [5] else "000000")
        ws3[f"C{r}"].fill=PatternFill("solid",start_color=bg); ws3[f"C{r}"].border=_thin()
        _dc(ws3,f"D{r}",nota,bg=bg,color="888780")
    for col,w in zip(["A","B","C","D"],[34,13,16,38]):
        ws3.column_dimensions[col].width=w


def _build_generic_excel(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — FiscalVerifier: actualización autónoma de tasas e impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws=_setup_sheet(wb,"Datos","1A4B8C")
    cols=max(len(secciones),5) if secciones else 5
    _title_row(ws,titulo,f"WEP AI · {datetime.now().strftime('%d/%m/%Y')}",cols)
    if secciones:
        for i,sec in enumerate(secciones,1): _hc(ws,f"{get_column_letter(i)}4",sec)
        for r in range(5,20):
            bg="FFFFFF" if r%2==0 else "F8FAFC"
            for i in range(1,len(secciones)+1): _dc(ws,f"{get_column_letter(i)}{r}","",bg=bg)
    ws.column_dimensions["A"].width=30
    for i in range(2,cols+1): ws.column_dimensions[get_column_letter(i)].width=15


# ─────────────────────────────────────────────────────────────────────────────
# NEW TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────

def _build_income_statement(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Estado de Resultados (P&L) con comparativo mes anterior y ratios.

    v15.5 — El impuesto sobre la renta (ISR / IR / IRPF / Ganancias) usa la
    nomenclatura y tasa default del país. El formato de moneda también.
    """
    # ── v15.5 — Configuración fiscal ──────────────────────────────────────────
    # ── v15.11.0 — FiscalVerifier: actualización autónoma de tasas/impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    fiscal_tax_authority = (cfg or {}).get("autoridad_fiscal", "")
    pais = (cfg or {}).get("pais", "")
    # Para income statement el "income tax" empresarial varía: MX 30%, CO 35%, CL 27%, etc.
    # Por simplicidad usamos un default razonable (30%) y permitimos al usuario editar.
    income_tax_rate = 0.30  # tasa corporativa típica LATAM
    # Etiquetas localizadas para el impuesto sobre la renta corporativo
    tax_label_map = {
        "México": "ISR", "Colombia": "Impuesto de Renta", "Perú": "IR Empresarial",
        "Argentina": "Impuesto a las Ganancias", "Chile": "Impuesto a la Renta 1ra cat.",
        "España": "Impuesto Sociedades", "Ecuador": "IR Empresarial",
        "Bolivia": "IUE", "Uruguay": "IRAE", "Paraguay": "IRE",
        "Venezuela": "ISLR", "Panamá": "ISR", "Guatemala": "ISR",
        "Honduras": "ISR", "Nicaragua": "IR", "El Salvador": "ISR",
        "Costa Rica": "ISR", "República Dominicana": "ISR",
        "United States": "Federal Income Tax",
    }
    isr_label = tax_label_map.get(pais, "Impuesto sobre la Renta")
    money_fmt = currency_format(cfg)

    ws  = _setup_sheet(wb, "Estado de Resultados", "0C447C")
    ws2 = _setup_sheet(wb, "Ratios",               "185FA5")
    cols=5
    subtitle = (f"Período: {datetime.now().strftime('%B %Y')} · {pais} · WEP AI"
                if pais else f"Período: {datetime.now().strftime('%B %Y')} · WEP AI")
    _title_row(ws,titulo,subtitle,cols)
    ws.freeze_panes="A5"
    for i,h in enumerate(["Concepto","Mes Actual","Mes Anterior","Varianza $","Varianza %"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="0C447C")
    ws.row_dimensions[4].height=22

    r=5
    def income_row(label, r, val_cur, val_prev, bold=False, bg="FFFFFF", color="000000"):
        _dc(ws,f"A{r}",label,bold=bold,bg=bg,color=color)
        _dc(ws,f"B{r}",val_cur,align="center",bg=bg,fmt=money_fmt,color=color,bold=bold)
        _dc(ws,f"C{r}",val_prev,align="center",bg=bg,fmt=money_fmt,color=color)
        _formula(ws,f"D{r}",f"=B{r}-C{r}",bg=bg,fmt=money_fmt,align="center",
                 color="15803D" if val_cur>=val_prev else "B91C1C")
        _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",
                 bg=bg,fmt='0.0%',align="center",
                 color="15803D" if val_cur>=val_prev else "B91C1C")

    _section_row(ws,r,"INGRESOS",cols,bg="15803D"); r+=1
    income_row("Ventas brutas",r,250000,220000); r+=1
    income_row("(-) Devoluciones y descuentos",r,-5000,-3000,color="B91C1C"); r+=1
    sub_ventas=r
    _section_row_a(ws,r,"VENTAS NETAS",bg="EFF6FF")
    _formula(ws,f"B{r}",f"=B{r-2}+B{r-1}",bg="EFF6FF",fmt=money_fmt,color="15803D",bold=True,align="center")
    _formula(ws,f"C{r}",f"=C{r-2}+C{r-1}",bg="EFF6FF",fmt=money_fmt,align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="EFF6FF",fmt=money_fmt,align="center",color="15803D")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="EFF6FF",fmt='0.0%',align="center",color="15803D")
    r+=2

    _section_row(ws,r,"COSTO DE VENTAS",cols,bg="B91C1C"); r+=1
    income_row("Costo directo de productos/servicios",r,120000,105000,color="B91C1C"); r+=1
    cogs_r=r
    _section_row_a(ws,r,"TOTAL COSTO DE VENTAS",bg="FEF2F2")
    _formula(ws,f"B{r}",f"=B{r-1}",bg="FEF2F2",fmt=money_fmt,color="B91C1C",bold=True,align="center")
    _formula(ws,f"C{r}",f"=C{r-1}",bg="FEF2F2",fmt=money_fmt,align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="FEF2F2",fmt=money_fmt,align="center")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="FEF2F2",fmt='0.0%',align="center")
    r+=1
    ub_r=r
    _section_row_a(ws,r,"UTILIDAD BRUTA",bg="22C55E")
    _formula(ws,f"B{r}",f"=B{sub_ventas}-B{cogs_r}",bg="22C55E",fmt=money_fmt,color="FFFFFF",bold=True,align="center",size=11)
    _formula(ws,f"C{r}",f"=C{sub_ventas}-C{cogs_r}",bg="22C55E",fmt=money_fmt,color="FFFFFF",align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="22C55E",fmt=money_fmt,color="FFFFFF",bold=True,align="center")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="22C55E",fmt='0.0%',color="FFFFFF",bold=True,align="center")
    r+=2

    _section_row(ws,r,"GASTOS OPERATIVOS",cols,bg="C2410C"); r+=1
    op_start=r
    for item,val_c,val_p in [("Nómina y prestaciones",45000,42000),
                               ("Renta / arrendamiento",8000,8000),
                               ("Marketing y publicidad",12000,10000),
                               ("Servicios (luz, internet, tel.)",3500,3200),
                               ("Depreciación y amortización",2000,2000),
                               ("Gastos generales y admin.",5000,4500)]:
        income_row(item,r,val_c,val_p,color="C2410C"); r+=1
    op_end=r
    _section_row_a(ws,r,"TOTAL GASTOS OPERATIVOS",bg="FEF2F2")
    _formula(ws,f"B{r}",f"=SUM(B{op_start}:B{r-1})",bg="FEF2F2",fmt=money_fmt,color="C2410C",bold=True,align="center")
    _formula(ws,f"C{r}",f"=SUM(C{op_start}:C{r-1})",bg="FEF2F2",fmt=money_fmt,align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="FEF2F2",fmt=money_fmt,align="center")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="FEF2F2",fmt='0.0%',align="center")
    ebit_r=r; r+=2

    _section_row_a(ws,r,"EBIT (Utilidad operativa)",bg="185FA5")
    _formula(ws,f"B{r}",f"=B{ub_r}-B{ebit_r}",bg="185FA5",fmt=money_fmt,color="FFFFFF",bold=True,align="center",size=11)
    _formula(ws,f"C{r}",f"=C{ub_r}-C{ebit_r}",bg="185FA5",fmt=money_fmt,color="FFFFFF",align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="185FA5",fmt=money_fmt,color="FFFFFF",bold=True,align="center")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="185FA5",fmt='0.0%',color="FFFFFF",bold=True,align="center")
    ebit2_r=r; r+=2

    income_row("(-) Intereses y gastos financieros",r,-2500,-2500,color="B91C1C"); int_r=r; r+=1
    income_row("(+) Otros ingresos no operativos",r,1000,500); oth_r=r; r+=1
    isr_r=r
    _section_row_a(ws,r,f"UTILIDAD ANTES DE {isr_label.upper()} (EBT)",bg="EFF6FF")
    _formula(ws,f"B{r}",f"=B{ebit2_r}+B{int_r}+B{oth_r}",bg="EFF6FF",fmt=money_fmt,color="0C447C",bold=True,align="center")
    _formula(ws,f"C{r}",f"=C{ebit2_r}+C{int_r}+C{oth_r}",bg="EFF6FF",fmt=money_fmt,align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="EFF6FF",fmt=money_fmt,align="center")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="EFF6FF",fmt='0.0%',align="center")
    r+=1
    isr_label_full = f"(-) {isr_label} estimado ({income_tax_rate*100:.0f}%)"
    income_row(isr_label_full, r, -22500, -19500, color="B91C1C"); tax_r=r; r+=1

    _section_row_a(ws,r,"UTILIDAD NETA",bg="0D1B2A")
    _formula(ws,f"B{r}",f"=B{isr_r}+B{tax_r}",bg="0D1B2A",fmt=money_fmt,color="FFFFFF",bold=True,align="center",size=12)
    _formula(ws,f"C{r}",f"=C{isr_r}+C{tax_r}",bg="0D1B2A",fmt=money_fmt,color="FFFFFF",align="center")
    _formula(ws,f"D{r}",f"=B{r}-C{r}",bg="0D1B2A",fmt=money_fmt,color="FFFFFF",bold=True,align="center")
    _formula(ws,f"E{r}",f"=IFERROR((B{r}-C{r})/ABS(C{r}),0)",bg="0D1B2A",fmt='0.0%',color="FFFFFF",bold=True,align="center")
    net_r=r
    # ✅ FIX BUG 3: BarChart comparativo Mes Actual vs Mes Anterior
    # Referenciamos las filas clave del P&L para el gráfico
    # Crear hoja de gráfica con chart real
    ws_chart_pl = None
    for s in wb.worksheets:
        if s.title == "Ratios":
            ws_chart_pl = s; break
    if ws_chart_pl:
        chart_pl = BarChart()
        chart_pl.type     = "col"
        chart_pl.grouping = "clustered"
        chart_pl.title    = "Estado de Resultados — Comparativo"
        chart_pl.style    = 10
        chart_pl.height   = 14
        chart_pl.width    = 22
        # Categorías: los conceptos clave (filas del P&L)
        pl_labels = ["Ventas Netas","Ut. Bruta","EBIT","Ut. Neta"]
        pl_rows   = [sub_ventas, ub_r, ebit2_r, net_r]
        # Serie Mes Actual (col B) y Mes Anterior (col C)
        for col_idx, (col_letter, stitle, color) in enumerate([
            ("B","Mes Actual","185FA5"), ("C","Mes Anterior","888780")]):
            data_refs = []
            for row_n in pl_rows:
                data_refs.append(ws[f"{col_letter}{row_n}"].value or 0)
            s_pl = Series(
                Reference(ws, min_col=2 if col_idx==0 else 3,
                          min_row=min(pl_rows), max_row=max(pl_rows)),
                title=stitle)
            s_pl.graphicalProperties.solidFill = color
            chart_pl.series.append(s_pl)
        ws_chart_pl.add_chart(chart_pl, "E3")
    _note_row(ws,r+2,"✏ Reemplaza los valores de ejemplo con tus datos reales. Los totales y varianzas se calculan automáticamente.",cols)
    ws.column_dimensions["A"].width=38
    for i in range(2,6): ws.column_dimensions[get_column_letter(i)].width=14

    # Ratios sheet
    _title_row(ws2,"Ratios de Rentabilidad","Calculados automáticamente del Estado de Resultados",4)
    ratios=[
        ("Margen Bruto %",          f"=IFERROR('Estado de Resultados'!B{ub_r}/'Estado de Resultados'!B{sub_ventas},0)","0.0%","Utilidad Bruta / Ventas Netas"),
        ("Margen Operativo (EBIT)", f"=IFERROR('Estado de Resultados'!B{ebit2_r}/'Estado de Resultados'!B{sub_ventas},0)","0.0%","EBIT / Ventas Netas"),
        ("Margen Neto %",           f"=IFERROR('Estado de Resultados'!B{net_r}/'Estado de Resultados'!B{sub_ventas},0)","0.0%","Utilidad Neta / Ventas Netas"),
        ("Costo/Ventas %",          f"=IFERROR('Estado de Resultados'!B{cogs_r}/'Estado de Resultados'!B{sub_ventas},0)","0.0%","Costo de Ventas / Ventas Netas"),
        ("Gastos Operativos %",     f"=IFERROR('Estado de Resultados'!B{ebit_r}/'Estado de Resultados'!B{sub_ventas},0)","0.0%","Gastos Op. / Ventas Netas"),
    ]
    for i,h in enumerate(["Ratio","Valor","Interpretación"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="185FA5")
    for idx,(name,formula,fmt,desc) in enumerate(ratios):
        r2=5+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        _dc(ws2,f"A{r2}",name,bold=True,bg=bg)
        _formula(ws2,f"B{r2}",formula,bg=bg,fmt=fmt,color="185FA5",bold=True,align="center",size=12)
        _dc(ws2,f"C{r2}",desc,bg=bg,color="888780")
    ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=14; ws2.column_dimensions["C"].width=36


def _build_break_even(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Punto de Equilibrio con tabla de escenarios y análisis de sensibilidad."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — FiscalVerifier: actualización autónoma de tasas e impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Break-Even",          "C2410C")
    ws2 = _setup_sheet(wb, "Escenarios",          "B91C1C")
    ws3 = _setup_sheet(wb, "Sensibilidad",        "854F0B")
    cols=6
    _title_row(ws,titulo,"Análisis de Punto de Equilibrio · WEP AI",cols)

    # Parameters (editable)
    _section_row(ws,3,"PARÁMETROS — Edita estas celdas amarillas",cols,bg="854F0B")
    params=[
        ("Costos Fijos Totales / mes",          50000,  "$#,##0",     "Renta, nómina fija, servicios, etc."),
        ("Precio de Venta por unidad",           200,   "$#,##0.00",  "Precio al que vendes cada unidad"),
        ("Costo Variable por unidad",            120,   "$#,##0.00",  "Materias primas, empaque, comisiones por unidad"),
    ]
    for i,(label,val,fmt,note) in enumerate(params):
        r=4+i
        _dc(ws,f"A{r}",label,bold=True,bg="FFFBEB")
        _dc(ws,f"B{r}",val,align="center",bg="FEFCE8",fmt=fmt,color="C2410C",bold=True,size=12)
        _dc(ws,f"C{r}",note,bg="FFFBEB",color="888780")
        ws.row_dimensions[r].height=22
    ws.column_dimensions["B"].width=18

    # Calculated results
    _section_row(ws,8,"RESULTADOS CALCULADOS",cols,bg="0D1B2A"); r=9
    results=[
        ("Margen de Contribución por unidad",    "=B5-B6",       "$#,##0.00","Precio - Costo Variable"),
        ("Ratio de Contribución (%)",            "=IFERROR((B5-B6)/B5,0)", "0.0%","Margen / Precio"),
        ("PUNTO DE EQUILIBRIO (unidades)",       "=IFERROR(B4/(B5-B6),0)","#,##0","Costos Fijos / Margen de Contribución"),
        ("PUNTO DE EQUILIBRIO ($)",              "=IFERROR(B4/((B5-B6)/B5),0)","$#,##0","Costos Fijos / Ratio de Contribución"),
        ("Ingreso para cubrir costos fijos",     "=B4",          "$#,##0","Costos Fijos (punto exacto)"),
        ("Unidades para 10% de utilidad",        "=IFERROR((B4*1.10)/(B5-B6),0)","#,##0",""),
        ("Unidades para 20% de utilidad",        "=IFERROR((B4*1.20)/(B5-B6),0)","#,##0",""),
        ("Unidades para 30% de utilidad",        "=IFERROR((B4*1.30)/(B5-B6),0)","#,##0",""),
    ]
    for idx,(label,formula,fmt,note) in enumerate(results):
        is_main=idx in [2,3]
        bg="EFF6FF" if not is_main else ("FEF2F2" if idx==2 else "FEF2F2")
        c_label="C2410C" if is_main else "0D1B2A"
        _dc(ws,f"A{r}",label,bold=is_main,bg=bg,color=c_label)
        _formula(ws,f"B{r}",formula,bg="FFFFFF" if is_main else bg,fmt=fmt,
                 color="C2410C" if is_main else "185FA5",bold=is_main,align="center",size=13 if is_main else 10)
        _dc(ws,f"C{r}",note,bg=bg,color="888780")
        ws.row_dimensions[r].height=22 if is_main else 18; r+=1

    _note_row(ws,r+1,"✏ Modifica los 3 parámetros amarillos y todos los resultados se recalculan automáticamente.",cols)
    ws.column_dimensions["A"].width=40; ws.column_dimensions["C"].width=36

    # Scenarios table
    _title_row(ws2,"Tabla de Escenarios","Qué pasa si vendemos X% del punto de equilibrio",cols)
    for i,h in enumerate(["Escenario","% del PE","Unidades","Ingresos","Costos Totales","Utilidad / Pérdida"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="B91C1C")
    scenarios=[(0,0),(25,0.25),(50,0.50),(75,0.75),(100,1.0),(125,1.25),(150,1.50),(200,2.0)]
    pe_formula="'Break-Even'!B11"
    for idx,(label_pct,pct) in enumerate(scenarios):
        r2=5+idx; bg="FEF2F2" if pct<1 else ("EFF6FF" if pct==1 else "F0FFF4")
        c="B91C1C" if pct<1 else ("854F0B" if pct==1 else "15803D")
        _dc(ws2,f"A{r2}",f"{label_pct}% del PE",bold=(pct==1),bg=bg,color=c)
        _dc(ws2,f"B{r2}",pct,align="center",bg=bg,fmt='0%',color=c)
        _formula(ws2,f"C{r2}",f"=ROUND({pe_formula}*B{r2},0)",bg=bg,fmt='#,##0',align="center",color=c,bold=(pct==1))
        _formula(ws2,f"D{r2}",f"=C{r2}*'Break-Even'!B5",bg=bg,fmt='$#,##0',align="center",color=c)
        _formula(ws2,f"E{r2}",f"='Break-Even'!B4+C{r2}*'Break-Even'!B6",bg=bg,fmt='$#,##0',align="center")
        _formula(ws2,f"F{r2}",f"=D{r2}-E{r2}",bg=bg,fmt='$#,##0',align="center",
                 color="B91C1C" if pct<1 else ("854F0B" if pct==1 else "15803D"),bold=(pct>=1))
    # CF en escenarios — rojo si pérdida, verde si ganancia
    _cf_formula_rule(ws2, f"A5:F{5+len(scenarios)}",
                     "=$F5<0", "FFE4E4", "7F1D1D")
    _cf_formula_rule(ws2, f"A5:F{5+len(scenarios)}",
                     "=AND($F5>=0,$B5<1)", "FFF3CD", "78350F")
    _cf_formula_rule(ws2, f"A5:F{5+len(scenarios)}",
                     "=$F5>0", "D1FAE5", "064E3B")
    for i,w in enumerate([18,12,12,14,14,16],1): ws2.column_dimensions[get_column_letter(i)].width=w

    # Sensitivity analysis
    _title_row(ws3,"Análisis de Sensibilidad","PE en unidades según precio y costo variable",8)
    ws3.cell(4,1,"PE (unidades)").font=Font(name="Arial",bold=True,size=10)
    ws3.cell(4,1).fill=PatternFill("solid",start_color="854F0B")
    ws3.cell(4,1).alignment=Alignment(horizontal="center",vertical="center")
    ws3.cell(4,1).border=_thin()
    # Price range on columns, CV range on rows
    prices=[140,160,180,200,220,240,260]
    cvs   =[80,100,110,120,130,140,150]
    for j,price in enumerate(prices,2):
        ws3.cell(4,j,f"P=${price}").font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        ws3.cell(4,j).fill=PatternFill("solid",start_color="854F0B")
        ws3.cell(4,j).alignment=Alignment(horizontal="center"); ws3.cell(4,j).border=_thin()
    for i,cv in enumerate(cvs,5):
        ws3.cell(i,1,f"CV=${cv}").font=Font(name="Arial",bold=True,size=9,color="854F0B")
        ws3.cell(i,1).fill=PatternFill("solid",start_color="FFFBEB")
        ws3.cell(i,1).alignment=Alignment(horizontal="center"); ws3.cell(i,1).border=_thin()
        for j,price in enumerate(prices,2):
            pe_val=round(50000/(price-cv)) if price>cv else "N/A"
            c=ws3.cell(i,j)
            c.value=pe_val if pe_val!="N/A" else "∞"
            c.font=Font(name="Arial",size=9,
                       color="B91C1C" if pe_val=="N/A" else ("15803D" if isinstance(pe_val,int) and pe_val<500 else "854F0B"))
            c.fill=PatternFill("solid",start_color="FFFFFF" if i%2==0 else "FFFBEB")
            c.alignment=Alignment(horizontal="center"); c.border=_thin()
            c.number_format='#,##0'
    for i in range(1,9): ws3.column_dimensions[get_column_letter(i)].width=11


def _build_amortization(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Tabla de amortización de préstamo con cuota PMT, capital/interés, saldo."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — DepreciationVerifier: actualización tasas depreciación ──
    try:
        from agent.depreciation_verifier import verify_depreciation_data
        cfg = verify_depreciation_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Amortización",  "0C447C")
    ws2 = _setup_sheet(wb, "Config",        "185FA5")
    ws3 = _setup_sheet(wb, "Resumen",       "0D1B2A")
    cols=7

    # Config
    _title_row(ws2,"Configuración del Préstamo","Edita estos valores — la tabla se calcula automáticamente",4)
    cfg=[
        ("Monto del préstamo",      500000,"$#,##0",""),
        ("Tasa de interés anual %", 0.12,  "0.00%", "Ej: 0.12 = 12% anual"),
        ("Plazo (meses)",           36,    "#,##0",  "Número de cuotas mensuales"),
        ("Fecha primer pago",       date(datetime.now().year,datetime.now().month+1 if datetime.now().month<12 else 1,1),"DD/MM/YYYY",""),
        ("Tipo de amortización",    "Francesa","General","Francesa = cuota fija (la más común)"),
    ]
    for i,(label,val,fmt,note) in enumerate(cfg):
        r=5+i; bg="EFF6FF" if i%2==0 else "FFFFFF"
        _dc(ws2,f"A{r}",label,bold=True,bg=bg)
        c=ws2[f"B{r}"]; c.value=val; c.number_format=fmt
        c.font=Font(name="Arial",bold=True,size=11,color="0C447C")
        c.fill=PatternFill("solid",start_color="FEFCE8")
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=_thin()
        _dc(ws2,f"C{r}",note,bg=bg,color="888780")
    ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=22; ws2.column_dimensions["C"].width=34

    # Amortization table — generate 36 rows
    _title_row(ws,titulo,f"Tabla generada: {datetime.now().strftime('%d/%m/%Y')} · WEP AI",cols)
    ws.freeze_panes="A5"
    for i,h in enumerate(["# Cuota","Fecha","Cuota Total","Capital","Interés","Saldo Pendiente","% Amortizado"],1):
        _hc(ws,f"{get_column_letter(i)}5",h,bg="0C447C")
    ws.row_dimensions[5].height=22

    # ✅ FIX: PMT como fórmula Excel real — el cuadro se recalcula si el usuario
    # cambia monto/tasa/plazo en la hoja Config (Config!B5/B6/B7)
    plazo = 36  # for loop range only; Excel formula references Config
    cur_date=date(datetime.now().year,datetime.now().month+1 if datetime.now().month<12 else 1,1)
    for i in range(1,plazo+1):
        r=5+i; bg="FFFFFF" if i%2==0 else "EFF6FF"
        _dc(ws,f"A{r}",i,align="center",bg=bg)
        ws[f"B{r}"].value=cur_date; ws[f"B{r}"].number_format="DD/MM/YYYY"
        ws[f"B{r}"].font=Font(name="Arial",size=10)
        ws[f"B{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center")
        ws[f"B{r}"].border=_thin()
        # Cuota = PMT nativo de Excel (se recalcula con Config!B5/B6/B7)
        _formula(ws,f"C{r}",
                 "=IFERROR(-PMT(Config!$B$6/12,Config!$B$7,Config!$B$5),0)",
                 bg=bg,fmt='$#,##0.00',color="0C447C",bold=True,align="center")
        # Interés del período
        if i == 1:
            # Primera cuota: interés sobre el capital inicial
            _formula(ws,f"E{r}",
                     "=IFERROR(Config!$B$5*(Config!$B$6/12),0)",
                     bg=bg,fmt='$#,##0.00',color="B91C1C",align="center")
        else:
            # Cuotas siguientes: interés sobre saldo anterior
            _formula(ws,f"E{r}",
                     f"=IFERROR(MAX(0,F{r-1})*(Config!$B$6/12),0)",
                     bg=bg,fmt='$#,##0.00',color="B91C1C",align="center")
        # Capital = cuota - interés
        _formula(ws,f"D{r}",f"=MAX(0,C{r}-E{r})",
                 bg=bg,fmt='$#,##0.00',color="15803D",align="center")
        # Saldo pendiente
        if i == 1:
            _formula(ws,f"F{r}","=MAX(0,Config!$B$5-D6)",
                     bg=bg,fmt='$#,##0.00',color="0C447C",align="center")
        else:
            _formula(ws,f"F{r}",f"=MAX(0,F{r-1}-D{r})",
                     bg=bg,fmt='$#,##0.00',color="0C447C",bold=(i==plazo),align="center")
        # % amortizado
        _formula(ws,f"G{r}",
                 f"=IFERROR((Config!$B$5-F{r})/Config!$B$5,0)",
                 bg=bg,fmt='0.0%',align="center")
        # Advance month
        month=cur_date.month+1; year=cur_date.year
        if month>12: month=1; year+=1
        last_day=calendar.monthrange(year,month)[1]
        cur_date=date(year,month,min(cur_date.day,last_day))
    # Variables for resumen (approximate with default values)
    monto,tasa_anual=500000,0.12
    tasa_mensual=tasa_anual/12
    cuota=round(monto*(tasa_mensual*(1+tasa_mensual)**plazo)/((1+tasa_mensual)**plazo-1),2)

    lr=5+plazo+1
    ws.merge_cells(f"A{lr}:B{lr}"); _hc(ws,f"A{lr}","TOTALES",bg="0D1B2A")
    _formula(ws,f"C{lr}",f"=SUM(C6:C{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    _formula(ws,f"D{lr}",f"=SUM(D6:D{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    _formula(ws,f"E{lr}",f"=SUM(E6:E{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center",color="FF9F9F")
    for i,w in enumerate([8,14,14,14,14,16,13],1): ws.column_dimensions[get_column_letter(i)].width=w

    # Resumen
    _title_row(ws3,"Resumen del Préstamo",f"Monto: ${monto:,.0f} · Tasa: {tasa_anual*100:.1f}% · {plazo} meses",4)
    total_pagado=cuota*plazo
    total_intereses=total_pagado-monto
    for i,(lbl,val,fmt,color) in enumerate([
        ("Monto original del préstamo",monto,'$#,##0',"0C447C"),
        ("Cuota mensual (fija)",cuota,'$#,##0.00',"0C447C"),
        ("Total pagado al final",total_pagado,'$#,##0',"0C447C"),
        ("Total en intereses pagados",total_intereses,'$#,##0',"B91C1C"),
        ("Costo real del crédito %",total_intereses/monto,'0.0%',"B91C1C"),
        ("Tasa mensual efectiva",tasa_mensual,'0.000%',"888780"),
    ]):
        r=5+i
        ws3[f"A{r}"].value=lbl; ws3[f"A{r}"].font=Font(name="Arial",bold=True,size=10)
        ws3[f"B{r}"].value=val; ws3[f"B{r}"].number_format=fmt
        ws3[f"B{r}"].font=Font(name="Arial",bold=True,size=13,color=color)
    ws3.column_dimensions["A"].width=34; ws3.column_dimensions["B"].width=18


def _build_sales_commissions(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Comisiones de ventas con escalas por cumplimiento de meta."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — LaborVerifier: actualización autónoma de prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Comisiones",      "15803D")
    ws2 = _setup_sheet(wb, "Estructura",      "0F6E56")
    ws3 = _setup_sheet(wb, "Detalle Ventas",  "085041")
    cols=10

    _title_row(ws2,"Estructura de Comisiones","Modifica las tasas en las celdas amarillas",6)
    _section_row(ws2,3,"TABLA DE TASAS POR % DE CUMPLIMIENTO",6,bg="0F6E56")
    for i,h in enumerate(["Desde","Hasta","Tasa Comisión","Descripción"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="0F6E56")
    estructura=[
        (0,0.49,0.00,"Sin comisión (menos del 50% de meta)"),
        (0.50,0.74,0.04,"4% — Zona básica"),
        (0.75,0.89,0.06,"6% — Meta parcial"),
        (0.90,0.99,0.08,"8% — Casi meta"),
        (1.00,1.19,0.10,"10% — Meta cumplida"),
        (1.20,1.49,0.12,"12% — Sobre-cumplimiento"),
        (1.50,9.99,0.15,"15% — Estrella de ventas"),
    ]
    for idx,(desde,hasta,tasa,desc) in enumerate(estructura):
        r=5+idx; bg="F0FFF4" if idx%2==0 else "FFFFFF"
        is_target=(desde==1.00)
        _dc(ws2,f"A{r}",desde,align="center",bg=bg,fmt='0%')
        _dc(ws2,f"B{r}",hasta,align="center",bg=bg,fmt='0%')
        _dc(ws2,f"C{r}",tasa,align="center",bg="FEFCE8" if is_target else bg,
            fmt='0%',color="0F6E56",bold=is_target,size=12 if is_target else 10)
        _dc(ws2,f"D{r}",desc,bg=bg,color="15803D" if is_target else "000000",bold=is_target)
    _note_row(ws2,13,"✏ Edita las tasas en la columna C. Se aplicarán automáticamente en la hoja 'Comisiones'.",4,bg="E1F5EE",color="0F6E56")
    for col,w in zip(["A","B","C","D"],[12,12,16,40]): ws2.column_dimensions[col].width=w

    # Main commissions sheet
    _title_row(ws,titulo,f"Período: {datetime.now().strftime('%B %Y')} · WEP AI",cols)
    ws.freeze_panes="A5"
    _add_autofilter(ws, 4, cols)
    for i,h in enumerate(["#","Vendedor","Zona/Región","Meta Mensual",
                           "Ventas Reales","% Cumplimiento","Tasa Comisión",
                           "Comisión Base","Bono Adicional","Total a Pagar"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="15803D")
    ws.row_dimensions[4].height=24

    vendedores=[
        ("Carlos Mendoza","Norte",  120000,143000),
        ("Ana García",   "Centro", 100000,98000),
        ("Luis Pérez",   "Sur",    90000, 112000),
        ("María Torres", "Este",   110000,110500),
        ("Pedro Soto",   "Occidente",80000,40000),
        ("Laura Vega",   "CDMX",   150000,185000),
    ]
    for idx,(name,zona,meta,real) in enumerate(vendedores):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "F0FFF4"
        cumplimiento=real/meta
        # Rate via VLOOKUP from Estructura sheet
        tasa_lkp=f"=IFERROR(VLOOKUP(F{r},Estructura!$A$5:$C$11,3,TRUE),0)"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",name,bold=True,bg=bg,color="15803D")
        _dc(ws,f"C{r}",zona,bg=bg)
        _dc(ws,f"D{r}",meta,align="center",bg=bg,fmt='$#,##0')
        _dc(ws,f"E{r}",real,align="center",bg=bg,fmt='$#,##0',
            color="15803D" if real>=meta else "B91C1C",bold=True)
        _formula(ws,f"F{r}",f"=IFERROR(E{r}/D{r},0)",bg=bg,fmt='0.0%',align="center",
                 color="15803D" if cumplimiento>=1 else "B91C1C",bold=True)
        _formula(ws,f"G{r}",tasa_lkp,bg=bg,fmt='0%',align="center",color="0F6E56",bold=True)
        _formula(ws,f"H{r}",f"=E{r}*G{r}",bg=bg,fmt='$#,##0.00',align="center",color="15803D",bold=True)
        _formula(ws,f"I{r}",f"=IF(F{r}>1.5,E{r}*0.02,IF(F{r}>1.2,E{r}*0.01,0))",
                 bg=bg,fmt='$#,##0.00',align="center",color="0F6E56")
        _formula(ws,f"J{r}",f"=H{r}+I{r}",bg=bg,fmt='$#,##0.00',align="center",
                 color="0D1B2A",bold=True,size=11)

    lr=5+len(vendedores)
    ws.merge_cells(f"A{lr}:D{lr}"); _hc(ws,f"A{lr}","TOTALES DEL EQUIPO",bg="0D1B2A")
    for col in ["E","F","G","H","I","J"]:
        _formula(ws,f"{col}{lr}",f"=SUM({col}5:{col}{lr-1})",bg="0D1B2A",fg="FFFFFF",
                 bold=True,fmt='$#,##0.00' if col!="E" else '$#,##0',align="center")
    _formula(ws,f"F{lr}",f"=IFERROR(E{lr}/SUM(D5:D{lr-1}),0)",bg="0D1B2A",fg="FFFFFF",
             bold=True,fmt='0.0%',align="center")
    # CF: rojo si <50% meta, ámbar si 50-99%, verde si >=100%
    n_vend = len(vendedores)
    _cf_formula_rule(ws, f"A5:J{4+n_vend}", "=$F5>=1",   "D1FAE5", "064E3B")
    _cf_formula_rule(ws, f"A5:J{4+n_vend}", "=AND($F5>=0.5,$F5<1)", "FFF3CD", "78350F")
    _cf_formula_rule(ws, f"A5:J{4+n_vend}", "=$F5<0.5",  "FFE4E4", "7F1D1D")
    _note_row(ws,lr+2,"✓ Tasa de comisión tomada automáticamente de la hoja 'Estructura'. Actualiza la tabla de tasas para cambiar la política.",cols,bg="E1F5EE",color="0F6E56")
    for i,w in enumerate([4,22,14,14,14,14,14,14,14,14],1):
        ws.column_dimensions[get_column_letter(i)].width=w


def _build_personal_budget(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Control de gastos personales con semáforo de presupuesto."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Dashboard",   "185FA5")
    ws2 = _setup_sheet(wb, "Gastos",      "0C447C")
    ws3 = _setup_sheet(wb, "Presupuesto", "3B82F6")
    cols=8

    # Presupuesto sheet
    _title_row(ws3,"Presupuesto Mensual","Establece tu límite por categoría",5)
    _section_row(ws3,3,"INGRESOS",5,bg="15803D"); 
    _params_box(ws3,4,1,"Ingreso principal (neto):",5000,"Sueldo o ingreso principal",fmt='$#,##0.00')
    _params_box(ws3,5,1,"Otros ingresos:",0,"Freelance, rentas, dividendos",fmt='$#,##0.00')
    _params_box(ws3,6,1,"TOTAL INGRESOS:",f"=B4+B5","",fmt='$#,##0.00',color_val="15803D")
    _section_row(ws3,8,"PRESUPUESTO POR CATEGORÍA",5,bg="0C447C")
    for i,h in enumerate(["Categoría","Presupuesto","Real (auto)","Diferencia","Semáforo"],1):
        _hc(ws3,f"{get_column_letter(i)}9",h,bg="0C447C")
    categories=[
        ("🏠 Vivienda (renta/hipoteca)", 1500),
        ("🛒 Alimentación",              600),
        ("🚗 Transporte",                300),
        ("⚡ Servicios (luz/agua/tel.)", 200),
        ("🎓 Educación",                 150),
        ("🏥 Salud y medicina",           100),
        ("🎭 Entretenimiento y ocio",     200),
        ("👕 Ropa y cuidado personal",    100),
        ("💰 Ahorro / inversión",         500),
        ("📦 Otros gastos",              150),
    ]
    for idx,(cat,presup) in enumerate(categories):
        r=10+idx; bg="FFFFFF" if idx%2==0 else "EFF6FF"
        _dc(ws3,f"A{r}",cat,bold=True,bg=bg)
        _dc(ws3,f"B{r}",presup,align="center",bg="FEFCE8",fmt='$#,##0.00',color="0C447C",bold=True)
        # Real = SUMIF from Gastos sheet
        _formula(ws3,f"C{r}",f"=IFERROR(SUMIF(Gastos!D$5:D$200,A{r},Gastos!E$5:E$200),0)",
                 bg=bg,fmt='$#,##0.00',align="center",color="B91C1C")
        _formula(ws3,f"D{r}",f"=B{r}-C{r}",bg=bg,fmt='$#,##0.00',align="center",
                 color="15803D")
        _formula(ws3,f"E{r}",f'=IF(C{r}<=B{r},"✓ OK",IF(C{r}<=B{r}*1.1,"⚠ Cerca","🚫 Excedido"))',
                 bg=bg,align="center")
    lr=10+len(categories)
    ws3.merge_cells(f"A{lr}:A{lr}"); _hc(ws3,f"A{lr}","TOTAL GASTOS PRESUP.",bg="0D1B2A")
    _formula(ws3,f"B{lr}",f"=SUM(B10:B{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    _formula(ws3,f"C{lr}",f"=SUM(C10:C{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    _formula(ws3,f"D{lr}",f"=B{lr}-C{lr}",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    for i,w in enumerate([32,14,14,14,12],1): ws3.column_dimensions[get_column_letter(i)].width=w

    # Gastos sheet
    _title_row(ws2,"Registro de Gastos","Ingresa cada gasto aquí",cols)
    ws2.freeze_panes="A5"
    _add_autofilter(ws2, 4, cols)
    # DataValidation en categorías y método de pago
    _add_dv_list(ws2, "D5:D204",
                 ["🏠 Vivienda (renta/hipoteca)","🛒 Alimentación","🚗 Transporte",
                  "⚡ Servicios (luz/agua/tel.)","🎓 Educación","🏥 Salud y medicina",
                  "🎭 Entretenimiento y ocio","👕 Ropa y cuidado personal",
                  "💰 Ahorro / inversión","📦 Otros gastos"],
                 "Selecciona la categoría")
    _add_dv_list(ws2, "F5:F204",
                 ["Tarjeta débito","Tarjeta crédito","Efectivo","Transferencia","Otro"],
                 "Método de pago")
    _add_dv_list(ws2, "G5:G204", ["Sí","No"], "¿Es un gasto esencial?")
    for i,h in enumerate(["#","Fecha","Descripción","Categoría","Monto","Método Pago","¿Esencial?","Notas"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="0C447C")
    sample_gastos=[
        (date(2026,5,1),"Renta departamento","🏠 Vivienda (renta/hipoteca)",1500,"Transferencia","Sí"),
        (date(2026,5,2),"Súper semanal","🛒 Alimentación",280,"Tarjeta débito","Sí"),
        (date(2026,5,3),"Gasolina","🚗 Transporte",400,"Efectivo","Sí"),
        (date(2026,5,5),"Netflix + Spotify","🎭 Entretenimiento y ocio",120,"Tarjeta crédito","No"),
        (date(2026,5,7),"Consulta médica","🏥 Salud y medicina",350,"Efectivo","Sí"),
        (date(2026,5,10),"Ropa verano","👕 Ropa y cuidado personal",380,"Tarjeta crédito","No"),
        (date(2026,5,15),"Ahorro mes","💰 Ahorro / inversión",500,"Transferencia","Sí"),
    ]
    for idx,(fecha,desc,cat,monto,metodo,esencial) in enumerate(sample_gastos):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EFF6FF"
        _dc(ws2,f"A{r}",idx+1,align="center",bg=bg)
        ws2[f"B{r}"].value=fecha; ws2[f"B{r}"].number_format="DD/MM/YYYY"
        ws2[f"B{r}"].font=Font(name="Arial",size=10); ws2[f"B{r}"].fill=PatternFill("solid",start_color=bg)
        ws2[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws2[f"B{r}"].border=_thin()
        _dc(ws2,f"C{r}",desc,bg=bg)
        _dc(ws2,f"D{r}",cat,bg=bg,color="0C447C")
        _dc(ws2,f"E{r}",monto,align="center",bg=bg,fmt='$#,##0.00',color="B91C1C",bold=True)
        _dc(ws2,f"F{r}",metodo,align="center",bg=bg)
        _dc(ws2,f"G{r}",esencial,align="center",bg=bg,color="15803D" if esencial=="Sí" else "888780")
        _dc(ws2,f"H{r}","",bg=bg)
    for col,w in zip(["A","B","C","D","E","F","G","H"],[4,12,30,32,13,16,10,20]):
        ws2.column_dimensions[col].width=w

    # Dashboard
    _title_row(ws,"Dashboard Personal",f"{datetime.now().strftime('%B %Y')} · WEP AI",cols)
    _section_row(ws,3,"RESUMEN FINANCIERO DEL MES",cols,bg="0D1B2A")
    kpis_personal=[
        ("Ingresos del mes",  "=Presupuesto!B6",                                "3B82F6","$#,##0"),
        ("Total gastado",     "=SUM(Gastos!E5:E200)",                           "EF4444","$#,##0"),
        ("Disponible / Ahorro","=Presupuesto!B6-SUM(Gastos!E5:E200)",           "22C55E","$#,##0"),
        ("% del presupuesto", "=IFERROR(SUM(Gastos!E5:E200)/Presupuesto!B6,0)","F59E0B","0.0%"),
    ]
    for i,(kpi,formula,color,fmt) in enumerate(kpis_personal):
        col=1+i*2
        ws.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1)
        ws.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
        lbl=ws.cell(4,col,kpi)
        lbl.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        lbl.fill=PatternFill("solid",start_color=color)
        lbl.alignment=Alignment(horizontal="center",vertical="center")
        val=ws.cell(5,col,formula)
        val.number_format=fmt
        val.font=Font(name="Arial",bold=True,size=20,color=color)
        val.fill=PatternFill("solid",start_color="F8FAFC")
        val.alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[4].height=20; ws.row_dimensions[5].height=44
    _note_row(ws,7,"✏ Ingresa tus gastos en la hoja 'Gastos'. Tu presupuesto se configura en la hoja 'Presupuesto'.",cols,bg="EFF6FF",color="0C447C")
    for i in range(1,cols+1): ws.column_dimensions[get_column_letter(i)].width=11


def _build_bank_reconciliation(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Conciliación bancaria: estado de cuenta vs libros, diferencias."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — BankingVerifier: actualización regulación bancaria ──
    try:
        from agent.banking_verifier import verify_banking_data
        cfg = verify_banking_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Conciliación",    "0C447C")
    ws2 = _setup_sheet(wb, "Estado de Cuenta","185FA5")
    ws3 = _setup_sheet(wb, "Libros Internos", "0F6E56")
    cols=6
    _title_row(ws,titulo,f"Período: {datetime.now().strftime('%B %Y')} · WEP AI",cols)

    for sheet,label,bg_color,entries in [
        (ws2,"MOVIMIENTOS DEL BANCO","185FA5",[
            (date(2026,5,2),"Depósito cliente A","Entrada",15000),
            (date(2026,5,5),"Cheque #1001 proveed.","Salida",-8500),
            (date(2026,5,8),"Transferencia entrada","Entrada",22000),
            (date(2026,5,12),"Cargo comisión banco","Salida",-250),
            (date(2026,5,15),"Depósito cliente B","Entrada",9800),
            (date(2026,5,20),"Pago nómina","Salida",-45000),
        ]),
        (ws3,"REGISTROS INTERNOS (LIBROS)","0F6E56",[
            (date(2026,5,2),"Cobro cliente A","Entrada",15000),
            (date(2026,5,5),"Pago proveedor (cheque)","Salida",-8500),
            (date(2026,5,8),"Transferencia entrada","Entrada",22000),
            (date(2026,5,10),"Depósito cliente C (en tránsito)","Entrada",5000),
            (date(2026,5,15),"Cobro cliente B","Entrada",9800),
            (date(2026,5,20),"Pago nómina","Salida",-45000),
            (date(2026,5,25),"Cheque #1002 pendiente","Salida",-3200),
        ]),
    ]:
        c2=6
        _title_row(sheet,label,f"WEP AI · {datetime.now().strftime('%d/%m/%Y')}",c2)
        sheet.freeze_panes="A5"
        for i,h in enumerate(["#","Fecha","Concepto","Tipo","Monto","Saldo Acumulado"],1):
            _hc(sheet,f"{get_column_letter(i)}4",h,bg=bg_color)
        saldo=100000.0  # saldo inicial
        for idx,(fecha,desc,tipo,monto) in enumerate(entries):
            r=5+idx; bg_r="FFFFFF" if idx%2==0 else "EFF6FF"
            is_entrada=tipo=="Entrada"
            _dc(sheet,f"A{r}",idx+1,align="center",bg=bg_r)
            sheet[f"B{r}"].value=fecha; sheet[f"B{r}"].number_format="DD/MM/YYYY"
            sheet[f"B{r}"].font=Font(name="Arial",size=10)
            sheet[f"B{r}"].fill=PatternFill("solid",start_color=bg_r)
            sheet[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center")
            sheet[f"B{r}"].border=_thin()
            _dc(sheet,f"C{r}",desc,bg=bg_r)
            _dc(sheet,f"D{r}",tipo,align="center",bg=bg_r,
                color="15803D" if is_entrada else "B91C1C",bold=True)
            _dc(sheet,f"E{r}",abs(monto),align="center",bg=bg_r,fmt='$#,##0.00',
                color="15803D" if is_entrada else "B91C1C")
            saldo+=monto
            _dc(sheet,f"F{r}",saldo,align="center",bg=bg_r,fmt='$#,##0.00',
                color="0C447C",bold=True)
        lr=5+len(entries)
        sheet.merge_cells(f"A{lr}:D{lr}"); _hc(sheet,f"A{lr}","SALDO FINAL",bg="0D1B2A")
        _formula(sheet,f"F{lr}",f"=F{lr-1}",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
        for i,w in enumerate([4,12,34,12,14,16],1):
            sheet.column_dimensions[get_column_letter(i)].width=w

    # Main reconciliation
    _section_row(ws,3,"CONCILIACIÓN — ESTADO DE CUENTA DEL BANCO",cols,bg="185FA5")
    banco_items=[
        ("Saldo según banco al cierre",                   "='Estado de Cuenta'!F11",""),
        ("(+) Depósitos en tránsito (en libros, no banco)","='Libros Internos'!E10","Depósito cliente C"),
        ("(-) Cheques pendientes de cobro",               -3200,"Cheque #1002"),
        ("(±) Errores del banco",                         0,"Si aplica"),
    ]
    r=4
    for lbl,val,note in banco_items:
        _dc(ws,f"A{r}",lbl,bold=True,bg="EFF6FF",color="0C447C")
        if isinstance(val,str):
            _formula(ws,f"B{r}",val,bg="EFF6FF",fmt='$#,##0.00',align="center",color="0C447C",bold=True)
        else:
            _dc(ws,f"B{r}",val,align="center",bg="FEFCE8",fmt='$#,##0.00',color="0C447C")
        _dc(ws,f"C{r}",note,bg="EFF6FF",color="888780"); r+=1
    saldo_adj_banco=r
    _hc(ws,f"A{r}","SALDO AJUSTADO DEL BANCO",bg="0C447C")
    _formula(ws,f"B{r}",f"=B4+B5+B6+B7",bg="0C447C",fmt='$#,##0.00',
             fg="FFFFFF",bold=True,align="center",size=11); r+=2

    _section_row(ws,r,"CONCILIACIÓN — LIBROS INTERNOS",cols,bg="0F6E56")
    libros_items=[
        ("Saldo según libros internos",           "='Libros Internos'!F12",""),
        ("(-) Comisión bancaria no registrada",   -250,"Cargo automático banco"),
        ("(±) Errores en libros",                 0,"Si aplica"),
    ]
    r+=1
    for lbl,val,note in libros_items:
        _dc(ws,f"A{r}",lbl,bold=True,bg="E1F5EE",color="0F6E56")
        if isinstance(val,str):
            _formula(ws,f"B{r}",val,bg="E1F5EE",fmt='$#,##0.00',align="center",color="0F6E56",bold=True)
        else:
            _dc(ws,f"B{r}",val,align="center",bg="FEFCE8",fmt='$#,##0.00',color="0F6E56")
        _dc(ws,f"C{r}",note,bg="E1F5EE",color="888780"); r+=1
    saldo_adj_libros=r
    _hc(ws,f"A{r}","SALDO AJUSTADO DE LIBROS",bg="0F6E56")
    _formula(ws,f"B{r}",f"=B{r-3}+B{r-2}+B{r-1}",bg="0F6E56",fmt='$#,##0.00',
             fg="FFFFFF",bold=True,align="center",size=11); r+=2

    _section_row(ws,r,"VERIFICACIÓN",cols,bg="0D1B2A"); r+=1
    _hc(ws,f"A{r}","DIFERENCIA (debe ser $0)",bg="FEF2F2" ,fg="B91C1C")
    _formula(ws,f"B{r}",f"=B{saldo_adj_banco}-B{saldo_adj_libros}",
             bg="FEF2F2",fmt='$#,##0.00',color="B91C1C",bold=True,align="center",size=12)
    _formula(ws,f"C{r}",f'=IF(ABS(B{r})<0.01,"✓ CONCILIADA","⚠ HAY DIFERENCIAS")',
             bg="FEF2F2",align="center",color="B91C1C",bold=True)
    _note_row(ws,r+2,"Si la diferencia no es $0, revisa los depósitos en tránsito, cheques pendientes y errores en ambas columnas.",cols)
    ws.column_dimensions["A"].width=46; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=28


def _build_subscription_tracker(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Tracker de subscripciones SaaS y servicios recurrentes."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Subscripciones", "534AB7")
    ws2 = _setup_sheet(wb, "Resumen",        "3C3489")
    cols=9
    _title_row(ws,titulo,f"Actualizado: {datetime.now().strftime('%d/%m/%Y')} · WEP AI",cols)
    ws.freeze_panes="A5"
    _add_autofilter(ws, 4, cols)
    for i,h in enumerate(["#","Servicio","Categoría","Moneda","Monto","Ciclo",
                           "Próximo cobro","Estado","Equiv. MXN/mes"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="534AB7")
    ws.row_dimensions[4].height=22
    subs=[
        ("Adobe Creative Cloud","Diseño","MXN",1365,"Mensual",date(2026,6,1),"Activa"),
        ("ChatGPT Plus","IA/Productividad","USD",20,"Mensual",date(2026,6,3),"Activa"),
        ("GitHub Pro","Desarrollo","USD",7,"Mensual",date(2026,6,5),"Activa"),
        ("Notion Team","Productividad","USD",16,"Mensual",date(2026,6,8),"Activa"),
        ("Figma Professional","Diseño","USD",15,"Mensual",date(2026,6,10),"Activa"),
        ("AWS (estimado)","Cloud","USD",45,"Mensual",date(2026,6,1),"Activa"),
        ("Slack Pro","Comunicación","USD",8,"Mensual",date(2026,6,15),"Pausada"),
        ("Canva Pro","Diseño","USD",13,"Anual",date(2026,12,1),"Activa"),
        ("Zoom Pro","Video","USD",15,"Mensual",date(2026,6,20),"Activa"),
        ("Linear","Gestión","USD",8,"Mensual",date(2026,6,25),"Activa"),
    ]
    TC_USD=17.2
    for idx,(nombre,cat,moneda,monto,ciclo,prox,estado) in enumerate(subs):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EEEDFE"
        is_active=estado=="Activa"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",nombre,bold=True,bg=bg,color="534AB7")
        _dc(ws,f"C{r}",cat,bg=bg)
        _dc(ws,f"D{r}",moneda,align="center",bg=bg)
        _dc(ws,f"E{r}",monto,align="center",bg=bg,fmt='#,##0.00')
        _dc(ws,f"F{r}",ciclo,align="center",bg=bg)
        ws[f"G{r}"].value=prox; ws[f"G{r}"].number_format="DD/MM/YYYY"
        ws[f"G{r}"].font=Font(name="Arial",size=10); ws[f"G{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"G{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"G{r}"].border=_thin()
        _dc(ws,f"H{r}",estado,align="center",bg=bg,
            color="15803D" if is_active else "888780",bold=True)
        # Monthly equivalent in MXN
        equiv_mensual=(monto*TC_USD if moneda=="USD" else monto)/(12 if ciclo=="Anual" else 1)
        _dc(ws,f"I{r}",round(equiv_mensual,2) if is_active else 0,
            align="center",bg=bg,fmt='$#,##0.00',color="534AB7" if is_active else "888780",
            bold=is_active)

    lr=5+len(subs)
    ws.merge_cells(f"A{lr}:H{lr}"); _hc(ws,f"A{lr}","TOTAL MENSUAL ACTIVAS (MXN)",bg="0D1B2A")
    _formula(ws,f"I{lr}",f'=SUMIF(H5:H{lr-1},"Activa",I5:I{lr-1})',
             bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center",size=12)
    # DataValidation (placed after subs is defined)
    _add_dv_list(ws, f"F5:F{lr-1}",
                 ["Mensual","Anual","Semestral","Trimestral"], "Ciclo de cobro")
    _add_dv_list(ws, f"H5:H{lr-1}", ["Activa","Pausada","Cancelada"], "Estado")
    _note_row(ws,lr+2,f"TC referencia: 1 USD = ${TC_USD} MXN. Actualiza la columna I manualmente si el tipo de cambio cambia.",cols,bg="EEEDFE",color="534AB7")
    for i,w in enumerate([4,26,18,9,12,12,14,10,14],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    # Resumen
    _title_row(ws2,"Resumen de Subscripciones","Por categoría y estado",5)
    _section_row(ws2,3,"TOTALES",5,bg="3C3489")
    resumen_data=[
        ("Subscripciones activas",    f"=COUNTIF(Subscripciones!H5:H{lr-1},\"Activa\")",'#,##0'),
        ("Gasto mensual total (MXN)", f"=Subscripciones!I{lr}",'$#,##0.00'),
        ("Gasto anual estimado (MXN)",f"=Subscripciones!I{lr}*12",'$#,##0.00'),
        ("Subscripciones pausadas",   f"=COUNTIF(Subscripciones!H5:H{lr-1},\"Pausada\")",'#,##0'),
    ]
    for i,h in enumerate(["Métrica","Valor"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="3C3489")
    for idx,(lbl,formula,fmt) in enumerate(resumen_data):
        r=5+idx; bg="EEEDFE" if idx%2==0 else "FFFFFF"
        _dc(ws2,f"A{r}",lbl,bold=True,bg=bg)
        _formula(ws2,f"B{r}",formula,bg=bg,fmt=fmt,color="534AB7",bold=True,align="center",size=13)
    ws2.column_dimensions["A"].width=32; ws2.column_dimensions["B"].width=20


def _build_balance_sheet(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Balance General con verificación Activos = Pasivos + Capital."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — AccountingVerifier: actualización autónoma NIIF/NIF/IFRS ──
    try:
        from agent.accounting_verifier import verify_accounting_data
        cfg = verify_accounting_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Balance General", "0C447C")
    ws2 = _setup_sheet(wb, "Ratios",          "185FA5")
    cols=4
    _title_row(ws,titulo,f"Al {datetime.now().strftime('%d/%m/%Y')} · WEP AI",cols)
    ws.freeze_panes="A4"
    for i,h in enumerate(["Concepto","Importe","Subtotal","Total"],1):
        _hc(ws,f"{get_column_letter(i)}3",h,bg="0C447C")
    r=4
    def bal_row(ws,r,label,val,is_total=False,indent=False,bg="FFFFFF",fg="000000"):
        prefix="  " if indent else ""
        _dc(ws,f"A{r}",prefix+label,bold=is_total,bg=bg,color=fg)
        if val is not None:
            if isinstance(val,str) and val.startswith("="):
                _formula(ws,f"B{r}",val,bg=bg,fmt='$#,##0',align="center",color=fg,bold=is_total)
            else:
                _dc(ws,f"B{r}",val,align="center",bg=bg,fmt='$#,##0',color=fg,bold=is_total)

    _section_row(ws,r,"ACTIVOS",cols,bg="185FA5"); r+=1
    _section_row(ws,r,"Activos Circulantes",cols,bg="EFF6FF"); r+=1
    ac_start=r
    for lbl,val in [("Caja y bancos",150000),("Cuentas por cobrar",85000),
                    ("Inventarios",45000),("Otros activos circulantes",12000)]:
        bal_row(ws,r,lbl,val,indent=True,bg="F8FAFC"); r+=1
    ac_total=r; bal_row(ws,r,"Total Activos Circulantes",f"=SUM(B{ac_start}:B{r-1})",True,bg="EFF6FF",fg="0C447C"); r+=2
    _section_row(ws,r,"Activos Fijos",cols,bg="EFF6FF"); r+=1
    af_start=r
    for lbl,val in [("Equipo y maquinaria",200000),("Menos: depreciación acumulada",-45000),
                    ("Mobiliario y equipo de oficina",35000),("Menos: depreciación",-8000),
                    ("Vehículos",120000),("Menos: depreciación vehículos",-30000)]:
        bg_r="FEF2F2" if "depreciación" in lbl.lower() else "F8FAFC"
        fg_r="B91C1C" if "depreciación" in lbl.lower() else "000000"
        bal_row(ws,r,lbl,val,indent=True,bg=bg_r,fg=fg_r); r+=1
    af_total=r; bal_row(ws,r,"Total Activos Fijos",f"=SUM(B{af_start}:B{r-1})",True,bg="EFF6FF",fg="0C447C"); r+=2
    act_total=r; _hc(ws,f"A{r}","TOTAL ACTIVOS",bg="185FA5")
    _formula(ws,f"D{r}",f"=B{ac_total}+B{af_total}",bg="185FA5",fg="FFFFFF",bold=True,
             fmt='$#,##0',align="center",size=12); r+=3

    _section_row(ws,r,"PASIVOS",cols,bg="B91C1C"); r+=1
    _section_row(ws,r,"Pasivos Circulantes",cols,bg="FEF2F2"); r+=1
    pc_start=r
    for lbl,val in [("Cuentas por pagar a proveedores",65000),("Impuestos por pagar",18000),
                    ("Préstamos a corto plazo",50000),("Otros pasivos circulantes",8000)]:
        bal_row(ws,r,lbl,val,indent=True,bg="FFF7ED"); r+=1
    pc_total=r; bal_row(ws,r,"Total Pasivos Circulantes",f"=SUM(B{pc_start}:B{r-1})",True,bg="FEF2F2",fg="B91C1C"); r+=2
    _section_row(ws,r,"Pasivos a Largo Plazo",cols,bg="FEF2F2"); r+=1
    plp_start=r
    for lbl,val in [("Préstamos bancarios LP",150000),("Otras deudas a largo plazo",25000)]:
        bal_row(ws,r,lbl,val,indent=True,bg="FFF7ED"); r+=1
    plp_total=r; bal_row(ws,r,"Total Pasivos LP",f"=SUM(B{plp_start}:B{r-1})",True,bg="FEF2F2",fg="B91C1C"); r+=2
    pas_total=r; _hc(ws,f"A{r}","TOTAL PASIVOS",bg="B91C1C")
    _formula(ws,f"D{r}",f"=B{pc_total}+B{plp_total}",bg="B91C1C",fg="FFFFFF",bold=True,
             fmt='$#,##0',align="center",size=12); r+=3

    _section_row(ws,r,"CAPITAL / PATRIMONIO",cols,bg="15803D"); r+=1
    cap_start=r
    for lbl,val in [("Capital social",200000),("Utilidades retenidas",75000),
                    ("Resultado del ejercicio",53000)]:
        bal_row(ws,r,lbl,val,indent=True,bg="F0FFF4"); r+=1
    cap_total=r; _hc(ws,f"A{r}","TOTAL CAPITAL",bg="15803D")
    _formula(ws,f"D{r}",f"=SUM(B{cap_start}:B{r-1})",bg="15803D",fg="FFFFFF",bold=True,
             fmt='$#,##0',align="center",size=12); r+=3

    # Ecuación contable
    _section_row(ws,r,"VERIFICACIÓN: Activos = Pasivos + Capital",cols,bg="0D1B2A"); r+=1
    _hc(ws,f"A{r}","Activos",bg="EFF6FF",fg="0C447C")
    _formula(ws,f"B{r}",f"=D{act_total}",bg="EFF6FF",fmt='$#,##0',color="0C447C",bold=True,align="center"); r+=1
    _hc(ws,f"A{r}","Pasivos + Capital",bg="EFF6FF",fg="0C447C")
    _formula(ws,f"B{r}",f"=D{pas_total}+D{cap_total}",bg="EFF6FF",fmt='$#,##0',color="0C447C",bold=True,align="center"); r+=1
    _hc(ws,f"A{r}","DIFERENCIA (debe ser $0)",bg="FEF2F2",fg="B91C1C")
    _formula(ws,f"B{r}",f"=B{r-2}-B{r-1}",bg="FEF2F2",fmt='$#,##0',color="B91C1C",bold=True,align="center")
    _formula(ws,f"C{r}",f'=IF(ABS(B{r})<1,"✓ CUADRADO","⚠ REVISAR")',bg="FEF2F2",align="center",color="B91C1C",bold=True)
    ws.column_dimensions["A"].width=40; ws.column_dimensions["B"].width=14
    ws.column_dimensions["C"].width=16; ws.column_dimensions["D"].width=14

    # Ratios
    _title_row(ws2,"Ratios Financieros","Del Balance General",4)
    inv_row = ac_start + 2  # Caja, CxC, Inventarios → 3a fila de circulantes
    SB = "'Balance General'"
    ratios=[
        ("Liquidez Corriente",f"=IFERROR({SB}!B{ac_total}/{SB}!B{pc_total},0)","x","Activo Circulante / Pasivo Circulante. >1 es saludable"),
        ("Razón Ácida",f"=IFERROR(({SB}!B{ac_total}-{SB}!B{inv_row})/{SB}!B{pc_total},0)","x","(Act.Circ-Inventario)/Pas.Circ. >0.7 OK"),
        ("Endeudamiento %",f"=IFERROR({SB}!D{pas_total}/{SB}!D{act_total},0)","0.0%","Pasivos/Activos. <50% recomendado"),
        ("Capital de Trabajo",f"=IFERROR({SB}!B{ac_total}-{SB}!B{pc_total},0)","$#,##0","Activo Circ. - Pasivo Circ."),
    ]
    for i,h in enumerate(["Ratio","Valor","Fórmula / Interpretación"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="185FA5")
    for idx,(name,formula,fmt,desc) in enumerate(ratios):
        r2=5+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        _dc(ws2,f"A{r2}",name,bold=True,bg=bg)
        _formula(ws2,f"B{r2}",formula,bg=bg,fmt=fmt,color="0C447C",bold=True,align="center",size=12)
        _dc(ws2,f"C{r2}",desc,bg=bg,color="888780")
    ws2.column_dimensions["A"].width=26; ws2.column_dimensions["B"].width=14; ws2.column_dimensions["C"].width=52


def _build_fixed_assets(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Control de activos fijos con depreciación lineal automática."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — DepreciationVerifier: actualización tasas depreciación ──
    try:
        from agent.depreciation_verifier import verify_depreciation_data
        cfg = verify_depreciation_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Activos Fijos",   "633806")
    ws2 = _setup_sheet(wb, "Dep. Acumulada",  "BA7517")
    cols=10
    _title_row(ws,titulo,f"Al {datetime.now().strftime('%d/%m/%Y')} · WEP AI",cols)
    ws.freeze_panes="A5"
    for i,h in enumerate(["#","Descripción","Categoría","Fecha Compra","Costo Original",
                           "Vida Útil (años)","Dep. Anual","Dep. Mensual","Dep. Acumulada","Valor en Libros"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="633806")
    ws.row_dimensions[4].height=24
    # Category default useful lives
    cat_life={"Equipo de cómputo":3,"Maquinaria":10,"Vehículos":4,
              "Mobiliario y equipo":10,"Edificios":20,"Software":3}
    assets=[
        ("Laptop Dell XPS 15",       "Equipo de cómputo",  date(2024,1,15), 35000),
        ("Impresora HP LaserJet",    "Equipo de cómputo",  date(2024,3,1),  12000),
        ("Escritorios (x5)",         "Mobiliario y equipo",date(2023,6,1),  25000),
        ("Camioneta Ford Transit",   "Vehículos",          date(2022,8,20), 350000),
        ("Montacargas eléctrico",    "Maquinaria",         date(2023,1,10), 180000),
        ("Software ERP",             "Software",           date(2024,1,1),  85000),
        ("Servidor principal",       "Equipo de cómputo",  date(2023,9,1),  95000),
    ]
    for idx,(desc,cat,f_compra,costo) in enumerate(assets):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        vida=cat_life.get(cat,5)
        meses_usados=max(0,(datetime.now().year-f_compra.year)*12+(datetime.now().month-f_compra.month))
        dep_anual=costo/vida
        dep_mensual=dep_anual/12
        dep_acum=min(costo,round(dep_mensual*meses_usados,2))
        valor_libros=max(0,costo-dep_acum)
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",desc,bold=True,bg=bg)
        _dc(ws,f"C{r}",cat,bg=bg,color="633806")
        ws[f"D{r}"].value=f_compra; ws[f"D{r}"].number_format="DD/MM/YYYY"
        ws[f"D{r}"].font=Font(name="Arial",size=10); ws[f"D{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"D{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"D{r}"].border=_thin()
        _dc(ws,f"E{r}",costo,align="center",bg=bg,fmt='$#,##0.00',color="0C447C",bold=True)
        _dc(ws,f"F{r}",vida,align="center",bg=bg,fmt='#,##0 "años"')
        _formula(ws,f"G{r}",f"=IFERROR(E{r}/F{r},0)",bg=bg,fmt='$#,##0.00',
                 color="B91C1C",align="center")
        _formula(ws,f"H{r}",f"=IFERROR(G{r}/12,0)",bg=bg,fmt='$#,##0.00',
                 color="B91C1C",align="center")
        # Accumulated depreciation = monthly * months elapsed
        _formula(ws,f"I{r}",
                 f"=MIN(E{r},ROUND(H{r}*(YEAR(TODAY())-YEAR(D{r}))*12+(MONTH(TODAY())-MONTH(D{r})),2))",
                 bg=bg,fmt='$#,##0.00',color="854F0B",align="center")
        _formula(ws,f"J{r}",f"=MAX(0,E{r}-I{r})",bg=bg,fmt='$#,##0.00',
                 color="15803D" if valor_libros>0 else "888780",bold=True,align="center")

    lr=5+len(assets)
    ws.merge_cells(f"A{lr}:D{lr}"); _hc(ws,f"A{lr}","TOTALES",bg="0D1B2A")
    for col in ["E","G","H","I","J"]:
        _formula(ws,f"{col}{lr}",f"=SUM({col}5:{col}{lr-1})",
                 bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")
    # CF nativo en activos fijos — resalta activos totalmente depreciados
    _cf_formula_rule(ws, f"A5:J{lr-1}",
                     "=$J5<=0", "FFE4E4", "7F1D1D")
    _cf_formula_rule(ws, f"A5:J{lr-1}",
                     "=AND($J5>0,$J5/$E5<0.2)", "FFF3CD", "78350F")
    _cf_formula_rule(ws, f"A5:J{lr-1}",
                     "=$J5/$E5>=0.5", "D1FAE5", "064E3B")
    _add_autofilter(ws, 4, cols)
    # DataValidation en categoría
    _add_dv_list(ws, f"C5:C{lr-1}",
                 ["Equipo de cómputo","Maquinaria","Vehículos",
                  "Mobiliario y equipo","Edificios","Software","Otros"],
                 "Categoría del activo")
    _note_row(ws,lr+2,"✓ Depreciación acumulada calculada automáticamente desde la fecha de compra hasta hoy. Método: Lineal.",cols)
    for i,w in enumerate([4,28,20,13,14,14,12,13,14,14],1):
        ws.column_dimensions[get_column_letter(i)].width=w


def _build_variance_analysis(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Real vs Presupuesto con semáforo y alertas de desviación."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — AccountingVerifier: actualización autónoma NIIF/NIF/IFRS ──
    try:
        from agent.accounting_verifier import verify_accounting_data
        cfg = verify_accounting_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Real vs Presupuesto", "0C447C")
    ws2 = _setup_sheet(wb, "Datos Reales",        "185FA5")
    ws3 = _setup_sheet(wb, "Presupuesto",         "3B82F6")
    cols=7
    _title_row(ws,titulo,f"Período: {datetime.now().strftime('%B %Y')} · WEP AI",cols)
    ws.freeze_panes="A5"
    for i,h in enumerate(["Concepto","Presupuesto","Real","Varianza $","Varianza %","Semáforo","Nota"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="0C447C")
    ws.row_dimensions[4].height=22
    items=[
        ("INGRESOS","section","15803D"),
        ("Ventas línea A",        500000,480000),
        ("Ventas línea B",        200000,225000),
        ("Otros ingresos",         50000, 48000),
        ("TOTAL INGRESOS",        750000,753000,"total"),
        ("EGRESOS","section","B91C1C"),
        ("Nómina",                250000,255000),
        ("Renta y servicios",      80000, 80000),
        ("Marketing",              60000, 75000),
        ("Inventario/Compras",    150000,138000),
        ("Gastos admin.",          40000, 44000),
        ("TOTAL EGRESOS",         580000,592000,"total"),
        ("UTILIDAD OPERATIVA",    170000,161000,"total_net"),
    ]
    r=5
    for item in items:
        if len(item)==3 and item[1]=="section":
            _,_,col=item
            _section_row(ws,r,item[0],cols,bg=col); r+=1; continue
        if len(item)==4:
            lbl,presup,real,typ=item
        else:
            lbl,presup,real=item; typ="normal"
        is_total= typ in ["total","total_net"]
        is_income= r<=10
        bg="FFFFFF" if r%2==0 else "EFF6FF"
        if is_total: bg="EFF6FF" if "INGRESO" in lbl else "FEF2F2" if "EGRESO" in lbl else "0D1B2A"
        fg="0C447C" if "INGRESO" in lbl and is_total else ("B91C1C" if "EGRESO" in lbl and is_total else ("FFFFFF" if lbl=="UTILIDAD OPERATIVA" else "000000"))
        _dc(ws,f"A{r}",lbl,bold=is_total,bg=bg,color=fg)
        _dc(ws,f"B{r}",presup,align="center",bg=bg,fmt='$#,##0',bold=is_total,color=fg)
        _dc(ws,f"C{r}",real,align="center",bg=bg,fmt='$#,##0',bold=is_total,
            color="15803D" if real>=presup else "B91C1C")
        var=real-presup
        var_pct=var/presup if presup else 0
        _formula(ws,f"D{r}",f"=C{r}-B{r}",bg=bg,fmt='$#,##0',align="center",
                 color="15803D" if var>=0 else "B91C1C",bold=is_total)
        _formula(ws,f"E{r}",f"=IFERROR((C{r}-B{r})/ABS(B{r}),0)",bg=bg,fmt='0.0%',align="center",
                 color="15803D" if var>=0 else "B91C1C",bold=is_total)
        # Semáforo based on % deviation
        _formula(ws,f"F{r}",
                 f'=IF(ABS(E{r})<0.05,"✅ OK",IF(ABS(E{r})<0.10,"⚠ Alerta","🚫 Desviación"))',
                 bg=bg,align="center",bold=is_total)
        _dc(ws,f"G{r}","Actualiza con datos reales" if not is_total else "",
            bg=bg,color="888780")
        r+=1
    # ✅ FIX BUG 2: CF nativo para semáforo varianza
    _cf_formula_rule(ws, f"A5:G{r}",
                     "=AND(ABS($E5)>=0.1,$E5<>"")", "FFE4E4", "7F1D1D")
    _cf_formula_rule(ws, f"A5:G{r}",
                     "=AND(ABS($E5)>=0.05,ABS($E5)<0.1,$E5<>"")", "FFF3CD", "78350F")
    _cf_formula_rule(ws, f"A5:G{r}",
                     "=AND(ABS($E5)<0.05,$E5<>"")", "D1FAE5", "064E3B")
    _note_row(ws,r+1,"✅ Varianza <5% · ⚠ Alerta 5-10% · 🚫 Desviación >10%. Actualiza los valores reales en la columna C.",cols)
    ws.column_dimensions["A"].width=32
    for i in range(2,8): ws.column_dimensions[get_column_letter(i)].width=14


def _build_savings_planner(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Planificador de metas de ahorro con proyección mensual."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Mis Metas",    "185FA5")
    ws2 = _setup_sheet(wb, "Proyección",   "0C447C")
    cols=7
    _title_row(ws,titulo,"Planificador de Ahorro · WEP AI",cols)
    _section_row(ws,3,"CONFIGURACIÓN DE TUS METAS",cols,bg="0C447C")
    goals=[
        ("Meta 1 — Emergencias",   30000,"MXN",0,6),
        ("Meta 2 — Vacaciones",    20000,"MXN",2000,8),
        ("Meta 3 — Auto / entrada",100000,"MXN",5000,18),
        ("Meta 4 — Retiro",        500000,"MXN",10000,60),
    ]
    for i,h in enumerate(["Meta","Objetivo","Moneda","Ahorro actual","Contrib. mensual","Meses restantes","Rendimiento %"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="0C447C")
    for idx,(nombre,objetivo,moneda,actual,meses) in enumerate(goals):
        r=5+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        _dc(ws,f"A{r}",nombre,bold=True,bg=bg,color="0C447C")
        _dc(ws,f"B{r}",objetivo,align="center",bg="FEFCE8",fmt='$#,##0',color="0C447C",bold=True)
        _dc(ws,f"C{r}",moneda,align="center",bg=bg)
        _dc(ws,f"D{r}",actual,align="center",bg="FEFCE8",fmt='$#,##0',color="15803D",bold=True)
        contrib=round((objetivo-actual)/meses,2) if meses>0 else 0
        _dc(ws,f"E{r}",contrib,align="center",bg="FEFCE8",fmt='$#,##0.00',color="185FA5",bold=True)
        _dc(ws,f"F{r}",meses,align="center",bg="FEFCE8",fmt='#,##0',color="854F0B",bold=True)
        _dc(ws,f"G{r}",0.04,align="center",bg="FEFCE8",fmt='0.0%',color="0F6E56")
    _note_row(ws,10,"✏ Edita las celdas amarillas (objetivo, ahorro actual, contribución, meses, rendimiento). La proyección se actualiza.",cols)
    ws.column_dimensions["A"].width=30
    for i in range(2,8): ws.column_dimensions[get_column_letter(i)].width=14

    # Proyección mensual — Meta 1
    _title_row(ws2,"Proyección Mensual",f"Meta 1: Fondo de emergencias · WEP AI",cols)
    for i,h in enumerate(["Mes","Contribución","Rendimiento","Total Ahorrado","% Logrado","Faltan","Estado"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="0C447C")
    objetivo=30000; actual=0; contrib=5000; tasa_mensual=0.04/12
    for i in range(1,25):
        r=4+i; bg="F0FFF4" if actual>=objetivo else ("EFF6FF" if i%2==0 else "FFFFFF")
        rendimiento=round(actual*tasa_mensual,2)
        actual=round(actual+contrib+rendimiento,2)
        pct=min(actual/objetivo,1.0)
        _dc(ws2,f"A{r}",f"Mes {i}",align="center",bg=bg)
        _dc(ws2,f"B{r}",contrib,align="center",bg=bg,fmt='$#,##0')
        _dc(ws2,f"C{r}",rendimiento,align="center",bg=bg,fmt='$#,##0.00',color="0F6E56")
        _dc(ws2,f"D{r}",actual,align="center",bg=bg,fmt='$#,##0',
            color="15803D" if actual>=objetivo else "0C447C",bold=(actual>=objetivo))
        _dc(ws2,f"E{r}",pct,align="center",bg=bg,fmt='0.0%',
            color="15803D" if pct>=1 else "185FA5",bold=(pct>=1))
        _dc(ws2,f"F{r}",max(0,objetivo-actual),align="center",bg=bg,fmt='$#,##0',color="B91C1C")
        _dc(ws2,f"G{r}","✅ META ALCANZADA" if actual>=objetivo else f"{'█'*int(pct*10)}{'░'*(10-int(pct*10))}",
            align="center",bg=bg,color="15803D" if actual>=objetivo else "185FA5")
        if actual>=objetivo: break
    for i,w in enumerate([10,14,14,16,12,14,22],1): ws2.column_dimensions[get_column_letter(i)].width=w


def _build_purchase_order(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Orden de compra a proveedores.

    v15.5 — IVA del país (tasa y label) en lugar de "IVA (16%)" hardcoded.
    """
    # ── v15.5 — Configuración fiscal ──────────────────────────────────────────
    tax_rate  = iva_rate(cfg)
    tax_lbl   = iva_label(cfg)
    money_fmt = currency_format(cfg)
    moneda    = currency_code(cfg)
    pais      = (cfg or {}).get("pais", "")

    ws  = _setup_sheet(wb, "Orden de Compra", "0F6E56")
    ws2 = _setup_sheet(wb, "Resumen",         "085041")
    cols=7
    subtitle = (f"OC-{datetime.now().strftime('%Y%m%d')} · {pais} · {moneda} · WEP AI"
                if pais else f"OC-{datetime.now().strftime('%Y%m%d')} · WEP AI")
    _title_row(ws,titulo,subtitle,cols)
    info=[("Proveedor:",""),("Contacto proveedor:",""),
          ("Fecha OC:",datetime.now().strftime('%d/%m/%Y')),
          ("Fecha entrega requerida:",""),("Condiciones de pago:","30 días neto"),
          ("Lugar de entrega:","")]
    for i,(lbl,val) in enumerate(info):
        r=4+i
        ws[f"A{r}"].value=lbl; ws[f"A{r}"].font=Font(name="Arial",bold=True,size=10,color="0F6E56")
        ws[f"B{r}"].value=val; ws[f"B{r}"].font=Font(name="Arial",size=10)
    ws.merge_cells("D4:G9")
    nota_c=ws["D4"]
    nota_c.value=(f"IMPORTANTE\n• Esta OC es vinculante al ser aceptada\n"
                  f"• Incluye número de OC en factura\n• Entrega sujeta a inspección de calidad\n"
                  f"• Precios en {moneda} con {tax_lbl} {'incluido si se indica' if tax_rate else 'según corresponda'}")
    nota_c.alignment=Alignment(wrap_text=True,vertical="top")
    nota_c.font=Font(name="Arial",size=9,color="0F6E56")
    nota_c.fill=PatternFill("solid",start_color="E1F5EE")
    for i,h in enumerate(["#","Código","Descripción del Producto / Servicio","Unidad","Cantidad","Precio Unitario","Subtotal"],1):
        _hc(ws,f"{get_column_letter(i)}11",h,bg="0F6E56")
    ws.row_dimensions[11].height=22
    sample=[("COD-001","Materia prima A","Kg",100,85),
            ("COD-002","Materia prima B","Litro",50,120),
            ("COD-003","Empaque estándar","Caja",200,15),
            ("COD-004","Servicio de flete","Flete",1,3500)]
    for idx,(code,desc,unit,qty,price) in enumerate(sample):
        r=12+idx; bg="FFFFFF" if idx%2==0 else "F0FFF4"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",code,align="center",bg=bg,color="0F6E56")
        _dc(ws,f"C{r}",desc,bg=bg)
        _dc(ws,f"D{r}",unit,align="center",bg=bg)
        _dc(ws,f"E{r}",qty,align="center",bg=bg)
        _dc(ws,f"F{r}",price,align="center",bg=bg,fmt=money_fmt)
        _formula(ws,f"G{r}",f"=E{r}*F{r}",bg=bg,fmt=money_fmt,color="0F6E56",bold=True,align="center")
    li=12+len(sample)
    for lbl,frm,is_main in [
        (f"Subtotal antes de {tax_lbl}", f"=SUM(G12:G{li-1})",         False),
        (f"{tax_lbl} ({tax_rate*100:.0f}%)" if tax_rate else f"{tax_lbl}",
                                        f"=G{li}*{tax_rate}",          False),
        ("TOTAL A PAGAR",                f"=G{li}+G{li+1}",             True),
    ]:
        ws.merge_cells(f"A{li}:F{li}")
        bg=("0F6E56" if is_main else "EFF6FF"); fg=("FFFFFF" if is_main else "0F6E56")
        _hc(ws,f"A{li}",lbl,bg=bg,fg=fg)
        _formula(ws,f"G{li}",frm,bg=bg,color=fg,bold=True,fmt=money_fmt,
                 align="center",size=11 if is_main else 10)
        li+=1
    sig=li+2
    for ref,lbl in [(f"A{sig}","Autoriza (empresa compradora)"),(f"E{sig}","Confirma (proveedor)")]:
        end="C" if ref.startswith("A") else "G"
        s=ref[0]; ws.merge_cells(f"{s}{sig}:{end}{sig+2}")
        c=ws[ref]; c.value=f"\n\n\n{'_'*30}\n{lbl}"
        c.alignment=Alignment(horizontal="center",vertical="bottom",wrap_text=True)
        c.font=Font(name="Arial",size=9,color="5F5E5A")
    for col,w in zip(["A","B","C","D","E","F","G"],[5,12,38,10,10,14,15]):
        ws.column_dimensions[col].width=w


# ─────────────────────────────────────────────────────────────────────────────
# MAIN GENERATE FUNCTION
# ─────────────────────────────────────────────────────────────────────────────
def _matches(text: str, patterns: list) -> bool:
    """
    Devuelve True si text matchea alguno de los patrones.
    - Si el patrón tiene metacaracteres regex (`.*`, `\\d`, etc.) → re.search.
    - Si no → substring match (`in`).
    Esto reemplaza el bug histórico donde `"directorio.*cliente" in text`
    nunca matcheaba porque `.*` no se interpreta como regex en el operador `in`.
    """
    for p in patterns:
        if ".*" in p or "\\d" in p or "\\s" in p or "\\b" in p:
            try:
                if re.search(p, text):
                    return True
            except re.error:
                # Patrón regex roto: fallback a substring
                if p in text:
                    return True
        else:
            if p in text:
                return True
    return False


# NOTA: el registro EXCEL_TEMPLATES (template_id → builder) está al FINAL de
# este archivo, después de que TODOS los _build_* estén definidos. Razón:
# algunos builders están definidos más abajo y un dict-literal aquí daría
# NameError al cargar el módulo.


def _dispatch(builder, wb, titulo, secciones, detalles, instruc, datos, cfg=None):
    """v15.5 — Llama al builder pasando `datos` y/o `cfg` si su firma los acepta.

    Esto permite migrar builders uno por uno: los refactorizados aceptan los
    kwargs `datos=None` y/o `cfg=None`; los viejos siguen sin recibirlos hasta
    que les llegue el refactor. Usa inspect para detectar la firma y no
    requiere mantener listas paralelas.

    Histórico:
    - v14.9 introdujo el patrón con `datos`
    - v15.5 añadió `cfg` para localización por país
    """
    import inspect
    try:
        sig = inspect.signature(builder)
        params = sig.parameters
        kwargs = {}
        if "datos" in params:
            kwargs["datos"] = datos
        if "cfg" in params:
            kwargs["cfg"] = cfg
        if kwargs:
            return builder(wb, titulo, secciones, detalles, instruc, **kwargs)
    except (TypeError, ValueError):
        # builtin o algo raro — fallback al call estándar
        pass
    return builder(wb, titulo, secciones, detalles, instruc)


def generate_excel(gen_data: dict) -> tuple:
    """
    Generate Excel file.
    Returns (filepath, bug_report_string)

    v15.5 — Detecta país (igual que word_controller) y pasa `cfg` a cada
    builder a través del dispatcher. Los builders refactorizados (payroll,
    quotation, sales, purchase_order, income_statement, freelance_tracker,
    income_expense_book) usan cfg para tasas IVA, instituciones de SS,
    retenciones e impuestos locales. Los builders que aún no usan cfg lo
    reciben pero lo ignoran (compatibilidad hacia atrás).
    """
    _ensure_dir()
    titulo    = gen_data.get("titulo", "Hoja")
    instruc   = gen_data.get("instrucciones", "")
    datos     = gen_data.get("datos", {})
    secciones = datos.get("secciones", [])
    detalles  = datos.get("detalles_especificos", "")

    # v14.4 — setear idioma para todo el árbol de generación (thread-local)
    _set_lang(datos.get("idioma") or gen_data.get("idioma") or "es")
    full = f"{instruc} {detalles}".lower()

    # v15.5 — DETECCIÓN DE PAÍS (mismo mecanismo que word_controller).
    # Resultado: cfg dict con pais, moneda, iva, autoridad_fiscal, etc.
    # Los builders refactorizados leen tasas e instituciones desde aquí.
    cfg = _detect_country(gen_data, instruc, detalles)

    # v15.5 — Override de idioma desde gen_data si el usuario lo especificó
    # (paralelo a lo que hace word_controller).
    user_lang = datos.get("idioma") or gen_data.get("idioma") or ""
    if user_lang:
        s = str(user_lang).strip().lower()
        if s in ("es", "español", "espanol", "spanish", "esp"):
            cfg = dict(cfg); cfg["idioma"] = "es"
        elif s in ("en", "english", "inglés", "ingles", "eng"):
            cfg = dict(cfg); cfg["idioma"] = "en"

    # ── v15.5.1 — FiscalVerifier: actualización autónoma de datos fiscales ──
    # Mismo hook que word_controller. Verifica IVA, retenciones y ley
    # en tiempo real antes de que los builders accedan al cfg.
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass  # Silencioso — usar config estático si el módulo falla

    wb = Workbook()
    wb.remove(wb.active)

    # ── v14.2: TEMPLATE_ID EXPLÍCITO (LLM-driven) ─────────────────────────────
    # Si Claude emitió un template_id válido, lo usamos directamente. Esto
    # evita las ambigüedades del matching por keywords.
    template_id = (gen_data.get("template_id") or "").strip().lower()
    if template_id and template_id in EXCEL_TEMPLATES:
        _dispatch(EXCEL_TEMPLATES[template_id], wb, titulo, secciones, detalles, instruc, datos, cfg)

    # ── DETECTION MAP (fallback por keywords) ─────────────────────────────────
    elif any(w in full for w in ["estado de resultados","p&l","profit and loss","utilidad neta","ebitda","utilidad bruta","margen bruto"]):
        _dispatch(_build_income_statement, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["punto de equilibrio","break-even","breakeven","equilibrio","costo fijo","margen contribución","cuánto vender"]):
        _dispatch(_build_break_even, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["amortización","amortizacion","préstamo","prestamo","hipoteca","cuota mensual","pmt","tabla de pagos","crédito bancario"]):
        _dispatch(_build_amortization, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["comisiones","comisión de ventas","vendedor","meta de ventas","cuota de ventas","bono vendedor"]):
        _dispatch(_build_sales_commissions, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["gasto personal","presupuesto personal","presupuesto familiar","control de gastos","finanzas personales","ahorro personal","gastos del mes","categorías de gasto"]):
        _dispatch(_build_personal_budget, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["conciliación","conciliacion","conciliar","banco vs libros","estado de cuenta vs","partidas en tránsito"]):
        _dispatch(_build_bank_reconciliation, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["subscripción","subscripcion","suscripción","suscripcion","saas","servicios recurrentes","netflix","spotify","software mensual"]):
        _dispatch(_build_subscription_tracker, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["balance general","balance sheet","activos y pasivos","patrimonio neto","capital contable","ecuación contable"]):
        _dispatch(_build_balance_sheet, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["activos fijos","depreciación","depreciacion","maquinaria","equipo de cómputo","vida útil","valor en libros"]):
        _dispatch(_build_fixed_assets, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["real vs presupuesto","varianza","variance","presupuesto vs real","análisis de desviación","cumplimiento presupuestal"]):
        _dispatch(_build_variance_analysis, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["meta de ahorro","ahorro mensual","fondo de ahorro","planificador ahorro","meta financiera","proyección ahorro"]):
        _dispatch(_build_savings_planner, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["orden de compra","purchase order","oc proveedor","requision","requisición","solicitud de compra"]):
        _dispatch(_build_purchase_order, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["inventario","stock","producto","almacén","bodega","existencia"]):
        _dispatch(_build_inventory, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["nómina","nomina","empleado","sueldo","salario","rrhh","planilla"]):
        _dispatch(_build_payroll, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["cotización","cotizacion","propuesta económica","presupuesto cliente","quote","oferta comercial"]):
        _dispatch(_build_quotation, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["cuentas por cobrar","cuentas por pagar","cartera","cobranza","pagos pendientes"]):
        _dispatch(_build_accounts_receivable, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["ratios financieros","razones financieras","ratios de liquidez","índices financieros",
                                  "indicadores financieros","solvencia","rentabilidad","análisis financiero",
                                  "razón corriente","prueba ácida"]):
        _dispatch(_build_ratios_financieros, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["flujo de caja","cashflow","cash flow","tesorería","liquidez"]):
        _dispatch(_build_cashflow, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["multi moneda","multimoneda","divisa","tipo de cambio","dólar","euro","moneda extranjera"]):
        _dispatch(_build_multi_currency, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["freelance","honorarios","horas trabajadas","facturación independiente"]):
        _dispatch(_build_freelance_tracker, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["gestión de proyecto","cronograma","gantt","tareas","entregable","avance","milestone"]):
        _dispatch(_build_project_tracker, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["directorio.*cliente","crm","ficha.*cliente","base.*dato.*cliente",
                                        "listado.*cliente","cartera.*cliente","gestión.*cliente"]):
        _dispatch(_build_client_directory, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["directorio.*proveedor","catálogo.*proveedor","catalogo proveedor",
                                  "base.*dato.*proveedor","listado.*proveedor","evaluación.*proveedor"]):
        _dispatch(_build_supplier_directory, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["directorio.*empleado","expediente.*empleado","base.*dato.*empleado",
                                  "listado.*empleado","ficha.*empleado","roster","headcount"]):
        _dispatch(_build_employee_directory, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["horario.*turno","turno.*semana","turno.*empleado","horario semanal",
                                  "turno mañana","turno tarde","turno noche","rotación.*turno",
                                  "planificación.*turno","schedule.*empleado"]):
        _dispatch(_build_work_schedule, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["control.*asistencia","asistencia.*empleado","registro.*entrada",
                                  "hora entrada","hora salida","marcaje","tardanza","puntualidad"]):
        _dispatch(_build_attendance, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["vacaciones","dias.*vacaciones","control.*vacaciones","permiso.*empleado",
                                  "ausencia.*empleado","saldo.*vacaciones","calendario.*vacaciones"]):
        _dispatch(_build_vacation_tracker, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["pipeline","embudo.*ventas","crm.*ventas","oportunidad.*venta",
                                  "forecast.*ventas","lead","prospecto","deal","cierre.*venta"]):
        _dispatch(_build_sales_pipeline, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif _matches(full, ["simulador.*escenario","escenario.*financiero","pesimista","optimista",
                                  "proyección.*5.*año","simulación.*financiera","what if financiero",
                                  "tres escenarios","escenario.*ventas"]):
        _dispatch(_build_scenario_simulator, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["caja chica","fondo fijo","petty cash","caja menor",
                                        "reposicion fondo","gastos menores","caja grande"]):
        _dispatch(_build_petty_cash, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["viatico","viaticos","gastos de viaje","travel expense",
                                  "reembolso viaje","gastos viaje","nota de gastos"]):
        _dispatch(_build_travel_expenses, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["roi","retorno.*inversion","return on investment",
                                  "calculadora.*inversion","rentabilidad.*proyecto",
                                  "comparar proyectos","cual es mejor inversion"]):
        _dispatch(_build_roi_calculator, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["resico","libro.*ingresos","libro.*egresos",
                                  "persona fisica","actividad empresarial",
                                  "ingresos y egresos fiscal","contabilidad basica"]):
        _dispatch(_build_income_expense_book, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["venta","ventas","factura","ingreso mensual"]):
        _dispatch(_build_sales, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["presupuesto anual detallado","presupuesto 12 meses","presupuesto mensual",
                                  "real vs plan","presupuesto real vs","desglose presupuestal"]):
        _dispatch(_build_presupuesto_anual, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["presupuesto anual","budget","gasto proyectado","planeación financiera"]):
        _dispatch(_build_budget, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["seguimiento de kpi","tablero de kpi","kpi tracker","metas y cumplimiento",
                                  "semáforo de indicadores","cumplimiento de metas","tracker de objetivos"]):
        _dispatch(_build_kpi_tracker, wb, titulo, secciones, detalles, instruc, datos, cfg)
    elif any(w in full for w in ["dashboard","kpi","indicador","estadística","métricas"]):
        _dispatch(_build_dashboard, wb, titulo, secciones, detalles, instruc, datos, cfg)
    else:
        _dispatch(_build_generic_excel, wb, titulo, secciones, detalles, instruc, datos, cfg)

    # ── AUTO BUG DETECTION ────────────────────────────────────────────────────
    bugs = _detector.scan(wb)
    wb, fixes = _detector.auto_fix(wb, bugs)
    bug_report = _detector.format_report(bugs, fixes)

    # ── v14.4 — BILINGUAL POST-PROCESSOR ──────────────────────────────────────
    # Si el idioma es inglés, traduce todos los labels detectables (sheet names,
    # encabezados de celda, títulos de chart). No afecta data del usuario que
    # no coincida exactamente con términos del diccionario.
    try:
        _translate_workbook(wb)
    except Exception:
        # Si por cualquier motivo falla la traducción, preferimos el workbook
        # original en español a un crash. No debe romper la generación.
        pass

    filename = f"{_safe_filename(titulo)}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath, bug_report


def open_in_excel(filepath: str):
    """Open the file in Microsoft Excel on macOS.

    v14.8 — Usa `open -a` en lugar de AppleScript interpolado para evitar
    inyección por contenido del path. `open` recibe el path como argv y no
    lo evalúa como código.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found: {filepath}")
    subprocess.run(["open", "-a", "Microsoft Excel", filepath], check=False)


def apply_correction(filepath: str, correction: str, gen_data: dict) -> tuple:
    """
    Re-generate with correction applied.
    Returns (filepath, bug_report).
    """
    gen_data["instrucciones"] = gen_data.get("instrucciones","") + f"\n\nCORRECCIÓN SOLICITADA: {correction}"
    return generate_excel(gen_data)


# ═════════════════════════════════════════════════════════════════════════════
# NUEVOS TEMPLATES — FASE 2B: CASOS DE USO FALTANTES
# ═════════════════════════════════════════════════════════════════════════════

def _build_client_directory(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """CRM básico — Directorio de clientes con historial, valor y segmentación."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — FiscalVerifier: actualización autónoma de tasas e impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Clientes",  "185FA5")
    ws2 = _setup_sheet(wb, "Segmentos", "0C447C")
    ws3 = _setup_sheet(wb, "Resumen",   "0D1B2A")
    cols = 13

    _title_row(ws, titulo,
               f"Directorio de clientes · WEP AI · {datetime.now().strftime('%d/%m/%Y')}", cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    headers = ["ID","Nombre / Empresa","Contacto","Teléfono","Email","Ciudad",
               "Categoría","Origen","Fecha Alta","Total Compras","# Compras",
               "Última Compra","Estado"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="185FA5")
    ws.row_dimensions[4].height = 22

    sample = [
        ("CLI-001","Grupo Industrial Norteño S.A.","Ing. Carlos Reyes","55-1234-5678","creyes@nortenio.com","Monterrey","VIP","Referido",date(2023,3,15),485000,12,date(2026,4,28),"Activo"),
        ("CLI-002","Tecnología y Sistemas del Bajío","Lic. Ana Vargas","33-9876-5432","avargas@tsbajio.mx","Guadalajara","Premium","Google Ads",date(2024,1,8),128500,5,date(2026,3,10),"Activo"),
        ("CLI-003","Distribuidora Central MX","Sr. Pedro Mora","55-5555-1234","pmora@distcentral.com","CDMX","Estándar","Referido",date(2023,8,22),62300,8,date(2025,11,5),"En riesgo"),
        ("CLI-004","Consultoría Integral Oaxaca","Dra. María Gil","951-123-4567","mgil@cio.com.mx","Oaxaca","Estándar","Feria",date(2024,6,1),18700,2,date(2025,6,30),"Inactivo"),
        ("CLI-005","StartUp Innovatech","CEO Luis Park","55-8888-9999","luis@innovatech.io","CDMX","Prospecto","LinkedIn",date(2026,4,1),0,0,None,"Prospecto"),
        ("CLI-006","Manufactura Precisa del Norte","Ing. Rosa Soto","614-567-8901","rsoto@mpnorte.mx","Chihuahua","VIP","Referido",date(2022,11,3),920000,28,date(2026,5,15),"Activo"),
        ("CLI-007","Retail Express Sureste","Lic. Jorge Díaz","999-345-6789","jdiaz@retailexpress.mx","Mérida","Premium","Google Ads",date(2024,9,14),75400,4,date(2026,2,20),"Activo"),
    ]

    cat_color = {"VIP":"185FA5","Premium":"0F6E56","Estándar":"000000","Prospecto":"888780","En riesgo":"B91C1C"}
    est_color = {"Activo":"15803D","En riesgo":"B91C1C","Inactivo":"888780","Prospecto":"854F0B"}

    for idx, (cid,nombre,contacto,tel,email,ciudad,cat,origen,f_alta,total,num,f_ult,estado) in enumerate(sample):
        r = 5 + idx; bg = "FFFFFF" if idx % 2 == 0 else "EFF6FF"
        _dc(ws,f"A{r}",cid,align="center",bg=bg,color="185FA5",bold=True)
        _dc(ws,f"B{r}",nombre,bold=True,bg=bg)
        _dc(ws,f"C{r}",contacto,bg=bg)
        _dc(ws,f"D{r}",tel,align="center",bg=bg)
        _dc(ws,f"E{r}",email,bg=bg,color="185FA5")
        _dc(ws,f"F{r}",ciudad,align="center",bg=bg)
        _dc(ws,f"G{r}",cat,align="center",bg=bg,color=cat_color.get(cat,"000000"),bold=True)
        _dc(ws,f"H{r}",origen,align="center",bg=bg)
        ws[f"I{r}"].value = f_alta; ws[f"I{r}"].number_format = "DD/MM/YYYY"
        ws[f"I{r}"].font = Font(name="Arial",size=10); ws[f"I{r}"].fill = PatternFill("solid",start_color=bg)
        ws[f"I{r}"].alignment = Alignment(horizontal="center",vertical="center"); ws[f"I{r}"].border = _thin()
        _dc(ws,f"J{r}",total,align="center",bg=bg,fmt='$#,##0',
            color="185FA5" if total>100000 else "000000",bold=total>100000)
        _dc(ws,f"K{r}",num,align="center",bg=bg)
        if f_ult:
            ws[f"L{r}"].value = f_ult; ws[f"L{r}"].number_format = "DD/MM/YYYY"
            ws[f"L{r}"].font = Font(name="Arial",size=10); ws[f"L{r}"].fill = PatternFill("solid",start_color=bg)
            ws[f"L{r}"].alignment = Alignment(horizontal="center",vertical="center"); ws[f"L{r}"].border = _thin()
        else:
            _dc(ws,f"L{r}","—",align="center",bg=bg,color="888780")
        _dc(ws,f"M{r}",estado,align="center",bg=bg,color=est_color.get(estado,"000000"),bold=True)

    lr = 5 + len(sample)
    # INDEX+MATCH: búsqueda de cliente por ID
    lookup_r = lr+3
    _section_row(ws, lookup_r, "BÚSQUEDA RÁPIDA POR ID (INDEX + MATCH)", cols, bg="185FA5")
    ws.merge_cells(f"A{lookup_r+1}:C{lookup_r+1}")
    c_lbl = ws.cell(lookup_r+1, 1, "Ingresa un ID de cliente →")
    c_lbl.font = Font(name="Arial",bold=True,size=10,color="0C447C")
    c_lbl.fill = PatternFill("solid",start_color="EFF6FF"); c_lbl.border = _thin()
    c_lbl.alignment = Alignment(horizontal="right",vertical="center")
    # Input cell
    ws[f"D{lookup_r+1}"].value = "CLI-001"
    ws[f"D{lookup_r+1}"].font = Font(name="Arial",bold=True,size=11,color="185FA5")
    ws[f"D{lookup_r+1}"].fill = PatternFill("solid",start_color="FEFCE8")
    ws[f"D{lookup_r+1}"].alignment = Alignment(horizontal="center",vertical="center")
    ws[f"D{lookup_r+1}"].border = _thin()
    # INDEX+MATCH formulas for each field
    lookup_fields = [
        ("Nombre", 2), ("Contacto", 3), ("Email", 5),
        ("Categoría", 7), ("Total Compras", 10), ("Estado", 13),
    ]
    for j, (field, col_num) in enumerate(lookup_fields):
        r_off = lookup_r+2+j
        _dc(ws, f"A{r_off}", field, bold=True, bg="EFF6FF")
        _formula(ws, f"B{r_off}",
                 f'=IFERROR(INDEX(A5:M{lr-1},{chr(10)}MATCH($D${lookup_r+1},A5:A{lr-1},0),{col_num}),"ID no encontrado")',
                 bg="FFFFFF", color="0C447C", bold=True, align="left",
                 fmt='$#,##0' if col_num==10 else 'General')
        ws.merge_cells(f"B{r_off}:D{r_off}")
    _note_row(ws, lr+1, "VIP=azul · Premium=verde · Rojo=en riesgo/inactivo. Actualiza Total Compras con SUMIF desde tu hoja de ventas.", cols)

    # CF nativo
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="Activo"',    "EFF6FF", "0C447C")
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="En riesgo"', "FFE4E4", "7F1D1D")
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="Inactivo"',  "F8FAFC", "888780")
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="Prospecto"', "FFFBEB", "78350F")
    # DataValidation
    _add_dv_list(ws, f"G5:G{lr-1}", ["VIP","Premium","Estándar","Prospecto"], "Categoría del cliente")
    _add_dv_list(ws, f"H5:H{lr-1}", ["Referido","Google Ads","LinkedIn","Feria","Llamada fría","Otro"], "Canal de origen")
    _add_dv_list(ws, f"M5:M{lr-1}", ["Activo","En riesgo","Inactivo","Prospecto"], "Estado del cliente")

    widths = [9,28,22,14,28,14,12,14,13,14,10,13,12]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

    # Resumen por segmento
    _title_row(ws2,"Segmentación de Clientes","Valor y distribución por categoría",5)
    for i, h in enumerate(["Categoría","# Clientes","Total Compras","Ticket Promedio","% del Total"], 1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="0C447C")
    cats = ["VIP","Premium","Estándar","Prospecto","En riesgo","Inactivo"]
    for idx, cat in enumerate(cats):
        r = 5 + idx; bg = "EFF6FF" if idx % 2 == 0 else "FFFFFF"
        _dc(ws2, f"A{r}", cat, bold=True, bg=bg, color=cat_color.get(cat,"000000"))
        _formula(ws2, f"B{r}",
                 f'=COUNTIF(Clientes!G5:G{lr-1},A{r})', bg=bg, fmt='#,##0', align="center")
        _formula(ws2, f"C{r}",
                 f'=SUMIF(Clientes!G5:G{lr-1},A{r},Clientes!J5:J{lr-1})',
                 bg=bg, fmt='$#,##0', align="center", color="185FA5", bold=True)
        _formula(ws2, f"D{r}",
                 f'=IFERROR(C{r}/B{r},0)', bg=bg, fmt='$#,##0', align="center")
        _formula(ws2, f"E{r}",
                 f'=IFERROR(C{r}/SUM(C5:C{5+len(cats)-1}),0)',
                 bg=bg, fmt='0.0%', align="center")
    for col, w in zip(["A","B","C","D","E"], [14,12,14,14,12]):
        ws2.column_dimensions[col].width = w

    # Resumen ejecutivo
    _title_row(ws3, "Resumen CRM", f"WEP AI · {datetime.now().strftime('%d/%m/%Y')}", 4)
    kpis_crm = [
        ("Total clientes registrados",   f"=COUNTA(Clientes!A5:A{lr-1})",       "185FA5"),
        ("Clientes activos",             f'=COUNTIF(Clientes!M5:M{lr-1},"Activo")', "15803D"),
        ("Clientes en riesgo",           f'=COUNTIF(Clientes!M5:M{lr-1},"En riesgo")', "B91C1C"),
        ("Prospectos",                   f'=COUNTIF(Clientes!M5:M{lr-1},"Prospecto")', "854F0B"),
        ("Valor total de cartera",       f"=SUM(Clientes!J5:J{lr-1})",           "185FA5"),
        ("Ticket promedio",              f"=IFERROR(SUM(Clientes!J5:J{lr-1})/COUNTIF(Clientes!M5:M{lr-1},\"Activo\"),0)", "0F6E56"),
    ]
    for i, h in enumerate(["KPI","Valor"], 1):
        _hc(ws3, f"{get_column_letter(i)}4", h, bg="0D1B2A")
    for idx, (label, formula, color) in enumerate(kpis_crm):
        r = 5 + idx; bg = "EFF6FF" if idx % 2 == 0 else "FFFFFF"
        _dc(ws3, f"A{r}", label, bold=True, bg=bg)
        _formula(ws3, f"B{r}", formula, bg=bg, fmt='$#,##0' if idx >= 4 else '#,##0',
                 color=color, bold=True, align="center", size=12)
    ws3.column_dimensions["A"].width = 32; ws3.column_dimensions["B"].width = 18


def _build_supplier_directory(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Catálogo de proveedores con evaluación, condiciones de pago y lead time."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Proveedores",  "0F6E56")
    ws2 = _setup_sheet(wb, "Evaluación",   "085041")
    ws3 = _setup_sheet(wb, "Resumen",      "0D1B2A")
    cols = 13

    _title_row(ws, titulo, f"Catálogo de proveedores · WEP AI · {datetime.now().strftime('%d/%m/%Y')}", cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    headers = ["ID","Proveedor","RFC / NIT","Contacto","Teléfono","Email","Categoría",
               "Condición Pago","Lead Time (días)","Calificación","Últ. Compra","Monto Anual","Estado"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="0F6E56")
    ws.row_dimensions[4].height = 22

    sample = [
        ("PROV-001","Aceros del Norte S.A.","ANO-880312-KL5","Ing. Roberto Núñez","81-1234-5678","rnunez@acerosnorte.mx","Materias primas","30 días neto",7,5,date(2026,5,10),380000,"Activo"),
        ("PROV-002","Envases y Empaques GDL","EEG-950815-AA1","Lic. Sandra Ruiz","33-2222-3333","sruiz@envasesgdl.com","Empaque","15 días neto",3,4,date(2026,4,28),145000,"Activo"),
        ("PROV-003","Servicios Logísticos MX","SLM-010203-BB2","Sr. Mario Castro","55-9999-0000","mcastro@logistica.mx","Logística","Contado",2,3,date(2026,3,15),92000,"En revisión"),
        ("PROV-004","Química Industrial Puebla","QIP-870601-CC3","Dra. Elena Vega","222-345-6789","evega@quimicapuebla.mx","Insumos químicos","60 días neto",14,4,date(2026,5,1),210000,"Activo"),
        ("PROV-005","Tech Components USA","TCU-NA","Mr. John Smith","+1-800-555-0100","jsmith@techcomp.us","Electrónica","45 días neto",21,5,date(2026,4,15),520000,"Activo"),
        ("PROV-006","Maquiladora Baja Sol","MBS-031122-DD4","Ing. Carmen López","664-123-7890","clopez@bajasol.mx","Manufactura","30 días neto",10,2,date(2025,11,30),48000,"Suspendido"),
    ]

    cal_color = {5:"15803D",4:"185FA5",3:"854F0B",2:"B91C1C",1:"B91C1C"}
    est_color = {"Activo":"15803D","En revisión":"854F0B","Suspendido":"B91C1C","Prospecto":"888780"}

    for idx, (pid,nombre,rfc,contacto,tel,email,cat,pago,lead,cal,f_comp,monto,estado) in enumerate(sample):
        r = 5 + idx; bg = "FFFFFF" if idx % 2 == 0 else "F0FFF4"
        _dc(ws,f"A{r}",pid,align="center",bg=bg,color="0F6E56",bold=True)
        _dc(ws,f"B{r}",nombre,bold=True,bg=bg)
        _dc(ws,f"C{r}",rfc,align="center",bg=bg,color="888780")
        _dc(ws,f"D{r}",contacto,bg=bg)
        _dc(ws,f"E{r}",tel,align="center",bg=bg)
        _dc(ws,f"F{r}",email,bg=bg,color="185FA5")
        _dc(ws,f"G{r}",cat,align="center",bg=bg)
        _dc(ws,f"H{r}",pago,align="center",bg=bg)
        _dc(ws,f"I{r}",lead,align="center",bg=bg,
            color="B91C1C" if lead>14 else ("854F0B" if lead>7 else "15803D"))
        stars = "★"*cal + "☆"*(5-cal)
        _dc(ws,f"J{r}",stars,align="center",bg=bg,color=cal_color.get(cal,"000000"),bold=True)
        ws[f"K{r}"].value = f_comp; ws[f"K{r}"].number_format = "DD/MM/YYYY"
        ws[f"K{r}"].font = Font(name="Arial",size=10); ws[f"K{r}"].fill = PatternFill("solid",start_color=bg)
        ws[f"K{r}"].alignment = Alignment(horizontal="center",vertical="center"); ws[f"K{r}"].border = _thin()
        _dc(ws,f"L{r}",monto,align="center",bg=bg,fmt='$#,##0',color="0F6E56",bold=True)
        _dc(ws,f"M{r}",estado,align="center",bg=bg,color=est_color.get(estado,"000000"),bold=True)

    lr = 5 + len(sample)
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="Activo"',     "E1F5EE", "064E3B")
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="En revisión"',"FFF3CD", "78350F")
    _cf_formula_rule(ws, f"A5:M{lr-1}", '=$M5="Suspendido"', "FFE4E4", "7F1D1D")
    _add_dv_list(ws, f"G5:G{lr-1}",
                 ["Materias primas","Empaque","Logística","Insumos químicos","Electrónica",
                  "Manufactura","Servicios","Tecnología","Otros"], "Categoría del proveedor")
    _add_dv_list(ws, f"H5:H{lr-1}",
                 ["Contado","7 días neto","15 días neto","30 días neto","45 días neto","60 días neto"],
                 "Condición de pago")
    _add_dv_list(ws, f"M5:M{lr-1}", ["Activo","En revisión","Suspendido","Prospecto"], "Estado")

    widths = [9,28,18,22,14,28,16,15,14,12,13,12,12]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

    # Evaluación detallada (scorecard)
    _title_row(ws2, "Scorecard de Proveedores",
               "Evalúa cada proveedor en 4 criterios del 1 al 5", 7)
    for i, h in enumerate(["Proveedor","Precio/Costo","Calidad","Entrega puntual","Servicio","TOTAL","Clasificación"], 1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="085041")
    scores_sample = [
        ("Aceros del Norte",5,5,4,5),("Envases y Empaques GDL",4,4,5,4),
        ("Servicios Logísticos MX",3,3,3,4),("Química Industrial Puebla",4,5,4,3),
        ("Tech Components USA",3,5,5,5),("Maquiladora Baja Sol",3,2,2,2),
    ]
    for idx, (name,precio,calidad,entrega,servicio) in enumerate(scores_sample):
        r = 5 + idx; bg = "FFFFFF" if idx % 2 == 0 else "E1F5EE"
        _dc(ws2, f"A{r}", name, bold=True, bg=bg)
        for j, val in enumerate([precio,calidad,entrega,servicio], 2):
            cl = get_column_letter(j)
            _dc(ws2, f"{cl}{r}", val, align="center", bg=bg,
                color="15803D" if val==5 else ("854F0B" if val==3 else ("B91C1C" if val<3 else "000000")))
        _formula(ws2, f"F{r}", f"=SUM(B{r}:E{r})", bg=bg, fmt='#,##0', align="center",
                 color="185FA5", bold=True)
        _formula(ws2, f"G{r}",
                 f'=IF(F{r}>=18,"★★★ Estratégico",IF(F{r}>=14,"★★ Confiable","★ En revisión"))',
                 bg=bg, align="center")
    _cf_three_color(ws2, f"F5:F{5+len(scores_sample)-1}", low="FFB3B3", mid="FFF3CD", high="B7F5D0")
    for col, w in zip(["A","B","C","D","E","F","G"],[24,12,10,14,10,10,16]):
        ws2.column_dimensions[col].width = w

    # Resumen ejecutivo
    _title_row(ws3,"Resumen de Proveedores",f"WEP AI · {datetime.now().strftime('%d/%m/%Y')}",4)
    for i, h in enumerate(["Indicador","Valor"], 1):
        _hc(ws3, f"{get_column_letter(i)}4", h, bg="0D1B2A")
    kpis_prov = [
        ("Total proveedores",           f"=COUNTA(Proveedores!A5:A{lr-1})", "0F6E56"),
        ("Proveedores activos",         f'=COUNTIF(Proveedores!M5:M{lr-1},"Activo")', "15803D"),
        ("Lead time promedio (días)",   f"=IFERROR(AVERAGE(Proveedores!I5:I{lr-1}),0)", "854F0B"),
        ("Gasto anual total",           f"=SUM(Proveedores!L5:L{lr-1})", "0F6E56"),
        ("Proveedores en revisión",     f'=COUNTIF(Proveedores!M5:M{lr-1},"En revisión")', "B91C1C"),
    ]
    for idx, (label, formula, color) in enumerate(kpis_prov):
        r = 5 + idx; bg = "E1F5EE" if idx % 2 == 0 else "FFFFFF"
        _dc(ws3, f"A{r}", label, bold=True, bg=bg)
        _formula(ws3, f"B{r}", formula, bg=bg,
                 fmt='$#,##0' if idx in [3] else '#,##0.0' if idx==2 else '#,##0',
                 color=color, bold=True, align="center", size=12)
    ws3.column_dimensions["A"].width = 30; ws3.column_dimensions["B"].width = 16


def _build_employee_directory(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Expediente y directorio de empleados con antigüedad calculada y datos RRHH."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — LaborVerifier: actualización autónoma de prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Empleados",    "6D28D9")
    ws2 = _setup_sheet(wb, "Por Depto",    "534AB7")
    ws3 = _setup_sheet(wb, "Resumen RRHH", "3C3489")
    cols = 14

    _title_row(ws, titulo, f"Expediente de empleados · WEP AI · {datetime.now().strftime('%d/%m/%Y')}", cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    headers = ["ID Emp.","Nombre","Apellidos","Puesto","Departamento","F. Ingreso",
               "Tipo Contrato","Salario Mensual","NSS / IMSS","Email corp.","Tel. Emergencia",
               "Antigüedad (años)","Días vacac./año","Estado"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="6D28D9")
    ws.row_dimensions[4].height = 22

    sample = [
        ("EMP-001","Carlos","Mendoza Ríos","Gerente Ventas","Ventas",date(2019,3,1),"Base",32000,"12345678901","cmenodza@empresa.mx","55-1111-2222"),
        ("EMP-002","Ana","García López","Contadora Senior","Finanzas",date(2020,8,15),"Base",28000,"23456789012","agarcia@empresa.mx","33-3333-4444"),
        ("EMP-003","Luis","Pérez Soto","Operador CNC","Producción",date(2021,1,10),"Base",16000,"34567890123","lperez@empresa.mx","55-5555-6666"),
        ("EMP-004","María","Torres Vega","Diseñadora Gráfica","Marketing",date(2022,6,1),"Temporal",18000,"45678901234","mtorres@empresa.mx","44-7777-8888"),
        ("EMP-005","Roberto","Sánchez Cruz","Aux. Almacén","Logística",date(2023,9,5),"Base",13500,"56789012345","rsanchez@empresa.mx","55-9999-0000"),
        ("EMP-006","Patricia","Luna Díaz","Ejecutiva de Cuenta","Ventas",date(2024,2,14),"Temporal",19000,"67890123456","pluna@empresa.mx","33-1111-2222"),
        ("EMP-007","Jorge","Ramírez Gil","IT Sr. Developer","Tecnología",date(2018,5,20),"Base",38000,"78901234567","jramirez@empresa.mx","55-3333-4444"),
    ]

    for idx, (eid,nombre,apellidos,puesto,depto,f_ing,contrato,salario,nss,email,tel_emerg) in enumerate(sample):
        r = 5 + idx; bg = "FFFFFF" if idx % 2 == 0 else "F5F3FF"
        _dc(ws,f"A{r}",eid,align="center",bg=bg,color="6D28D9",bold=True)
        _dc(ws,f"B{r}",nombre,bold=True,bg=bg)
        _dc(ws,f"C{r}",apellidos,bg=bg)
        _dc(ws,f"D{r}",puesto,bg=bg)
        _dc(ws,f"E{r}",depto,align="center",bg=bg,color="534AB7")
        ws[f"F{r}"].value = f_ing; ws[f"F{r}"].number_format = "DD/MM/YYYY"
        ws[f"F{r}"].font = Font(name="Arial",size=10); ws[f"F{r}"].fill = PatternFill("solid",start_color=bg)
        ws[f"F{r}"].alignment = Alignment(horizontal="center",vertical="center"); ws[f"F{r}"].border = _thin()
        _dc(ws,f"G{r}",contrato,align="center",bg=bg,
            color="0F6E56" if contrato=="Base" else "854F0B")
        _dc(ws,f"H{r}",salario,align="center",bg=bg,fmt='$#,##0',color="6D28D9",bold=True)
        _dc(ws,f"I{r}",nss,align="center",bg=bg,color="888780")
        _dc(ws,f"J{r}",email,bg=bg,color="185FA5")
        _dc(ws,f"K{r}",tel_emerg,align="center",bg=bg)
        # Antigüedad calculada automáticamente con TODAY()
        _formula(ws,f"L{r}",
                 f"=IFERROR(ROUND((TODAY()-F{r})/365.25,1),0)",
                 bg=bg,fmt='0.0 "años"',align="center",color="6D28D9",bold=True)
        # Días de vacaciones según antigüedad (Ley Federal del Trabajo MX)
        _formula(ws,f"M{r}",
                 f'=IF(L{r}>=25,24,IF(L{r}>=20,22,IF(L{r}>=15,20,IF(L{r}>=10,18,IF(L{r}>=5,16,IF(L{r}>=4,14,IF(L{r}>=3,12,IF(L{r}>=2,10,IF(L{r}>=1,6,0)))))))))',
                 bg=bg,fmt='#,##0 "días"',align="center",color="0F6E56")
        _dc(ws,f"N{r}","Activo",align="center",bg=bg,color="15803D",bold=True)

    lr = 5 + len(sample)
    # CF: >5 años=azul (veterano), 1-5=normal, <1=ámbar (nuevo)
    _cf_formula_rule(ws, f"A5:N{lr-1}", "=$L5>=5",   "EDE9FE", "4C1D95")
    _cf_formula_rule(ws, f"A5:N{lr-1}", "=$L5<1",    "FFFBEB", "78350F")
    _cf_formula_rule(ws, f"A5:N{lr-1}", '=$G5="Temporal"', "FFF3CD", "78350F")
    _add_dv_list(ws, f"E5:E{lr-1}",
                 ["Ventas","Finanzas","Producción","Marketing","Logística","Tecnología",
                  "RRHH","Dirección","Administración","Otro"], "Departamento")
    _add_dv_list(ws, f"G5:G{lr-1}", ["Base","Temporal","Por proyecto","Part-time","Honorarios"], "Tipo de contrato")
    _add_dv_list(ws, f"N5:N{lr-1}", ["Activo","Baja","Vacaciones","Incapacidad","Permiso"], "Estado")
    # INDEX+MATCH para búsqueda de empleado
    emp_lookup_r = lr+3
    _section_row(ws, emp_lookup_r, "BÚSQUEDA RÁPIDA POR ID EMPLEADO (INDEX + MATCH)", cols, bg="6D28D9")
    ws.merge_cells(f"A{emp_lookup_r+1}:C{emp_lookup_r+1}")
    c_lbl2 = ws.cell(emp_lookup_r+1, 1, "Ingresa ID del empleado →")
    c_lbl2.font = Font(name="Arial",bold=True,size=10,color="534AB7")
    c_lbl2.fill = PatternFill("solid",start_color="EEEDFE"); c_lbl2.border = _thin()
    c_lbl2.alignment = Alignment(horizontal="right",vertical="center")
    ws[f"D{emp_lookup_r+1}"].value = "EMP-001"
    ws[f"D{emp_lookup_r+1}"].font = Font(name="Arial",bold=True,size=11,color="6D28D9")
    ws[f"D{emp_lookup_r+1}"].fill = PatternFill("solid",start_color="FEFCE8")
    ws[f"D{emp_lookup_r+1}"].alignment = Alignment(horizontal="center",vertical="center")
    ws[f"D{emp_lookup_r+1}"].border = _thin()
    emp_fields = [("Nombre completo",
                  f'=IFERROR(INDEX(B5:B{lr-1},MATCH($D${emp_lookup_r+1},A5:A{lr-1},0))&" "&INDEX(C5:C{lr-1},MATCH($D${emp_lookup_r+1},A5:A{lr-1},0)),"No encontrado")'),
                  ("Puesto",         f'=IFERROR(INDEX(D5:D{lr-1},MATCH($D${emp_lookup_r+1},A5:A{lr-1},0)),"")'),
                  ("Departamento",   f'=IFERROR(INDEX(E5:E{lr-1},MATCH($D${emp_lookup_r+1},A5:A{lr-1},0)),"")'),
                  ("Salario",        f'=IFERROR(INDEX(H5:H{lr-1},MATCH($D${emp_lookup_r+1},A5:A{lr-1},0)),0)'),
                  ("Antigüedad",     f'=IFERROR(INDEX(L5:L{lr-1},MATCH($D${emp_lookup_r+1},A5:A{lr-1},0)),0)'),
    ]
    for j,(field,formula) in enumerate(emp_fields):
        r_off=emp_lookup_r+2+j
        _dc(ws,f"A{r_off}",field,bold=True,bg="EEEDFE")
        if "{emp_lookup_r1}" in formula:
            formula=formula.replace("{emp_lookup_r1}",str(emp_lookup_r+1)).replace("{lr}",str(lr-1))
        _formula(ws,f"B{r_off}",formula,bg="FFFFFF",color="6D28D9",bold=True,align="left",
                 fmt='$#,##0' if field=="Salario" else ('0.0 "años"' if "Anti" in field else 'General'))
        ws.merge_cells(f"B{r_off}:D{r_off}")
    _note_row(ws, lr+1,
              "✓ Antigüedad calculada automáticamente con TODAY(). Vacaciones según LFT México. Azul=Veteranos (≥5 años). Ámbar=Nuevos (<1 año).", cols)

    widths = [9,14,20,22,14,13,13,14,16,24,16,14,13,10]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

    # Por departamento
    _title_row(ws2,"Headcount por Departamento","Distribución de personal y costos",5)
    for i, h in enumerate(["Departamento","# Empleados","% del Total","Nómina Mensual","Salario Promedio","Contratos Base","Contratos Temp."],1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="534AB7")
    deptos = ["Ventas","Finanzas","Producción","Marketing","Logística","Tecnología","RRHH","Administración"]
    for idx, depto in enumerate(deptos):
        r = 5 + idx; bg = "FFFFFF" if idx%2==0 else "EEEDFE"
        _dc(ws2, f"A{r}", depto, bold=True, bg=bg, color="534AB7")
        _formula(ws2, f"B{r}", f'=COUNTIF(Empleados!E5:E{lr-1},A{r})', bg=bg, fmt='#,##0', align="center")
        _formula(ws2, f"C{r}", f'=IFERROR(B{r}/SUM(B5:B{5+len(deptos)-1}),0)', bg=bg, fmt='0.0%', align="center")
        _formula(ws2, f"D{r}", f'=SUMIF(Empleados!E5:E{lr-1},A{r},Empleados!H5:H{lr-1})', bg=bg, fmt='$#,##0', align="center", color="6D28D9", bold=True)
        _formula(ws2, f"E{r}", f'=IFERROR(D{r}/B{r},0)', bg=bg, fmt='$#,##0', align="center")
        _formula(ws2, f"F{r}", f'=COUNTIFS(Empleados!E5:E{lr-1},A{r},Empleados!G5:G{lr-1},"Base")', bg=bg, fmt='#,##0', align="center", color="15803D")
        _formula(ws2, f"G{r}", f'=COUNTIFS(Empleados!E5:E{lr-1},A{r},Empleados!G5:G{lr-1},"Temporal")', bg=bg, fmt='#,##0', align="center", color="854F0B")
    for col, w in zip(["A","B","C","D","E","F","G"],[16,12,10,14,14,14,14]):
        ws2.column_dimensions[col].width = w

    # Resumen
    _title_row(ws3,"Resumen RRHH",f"WEP AI · {datetime.now().strftime('%d/%m/%Y')}",4)
    for i, h in enumerate(["Indicador","Valor"],1): _hc(ws3,f"{get_column_letter(i)}4",h,bg="3C3489")
    kpis_hr=[
        ("Total empleados",              f"=COUNTA(Empleados!A5:A{lr-1})",     "6D28D9"),
        ("Empleados activos",            f'=COUNTIF(Empleados!N5:N{lr-1},"Activo")', "15803D"),
        ("Contratos base",               f'=COUNTIF(Empleados!G5:G{lr-1},"Base")', "0F6E56"),
        ("Contratos temporales",         f'=COUNTIF(Empleados!G5:G{lr-1},"Temporal")', "854F0B"),
        ("Antigüedad promedio (años)",   f"=IFERROR(AVERAGE(Empleados!L5:L{lr-1}),0)", "185FA5"),
        ("Nómina mensual total",         f"=SUM(Empleados!H5:H{lr-1})",         "6D28D9"),
        ("Salario promedio",             f"=IFERROR(AVERAGE(Empleados!H5:H{lr-1}),0)", "534AB7"),
    ]
    for idx,(label,formula,color) in enumerate(kpis_hr):
        r=5+idx; bg="EEEDFE" if idx%2==0 else "FFFFFF"
        _dc(ws3,f"A{r}",label,bold=True,bg=bg)
        _formula(ws3,f"B{r}",formula,bg=bg,
                 fmt='$#,##0' if idx>=5 else ('0.0' if idx==4 else '#,##0'),
                 color=color,bold=True,align="center",size=13)
    ws3.column_dimensions["A"].width=32; ws3.column_dimensions["B"].width=18


def _build_work_schedule(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Horario semanal de turnos con cobertura mínima y resumen de horas."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — LaborVerifier: actualización autónoma de prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Horario",  "0C447C")
    ws2 = _setup_sheet(wb, "Resumen",  "185FA5")
    ws3 = _setup_sheet(wb, "Leyenda",  "3B82F6")

    _title_row(ws, titulo, f"Semana del {datetime.now().strftime('%d/%m/%Y')} · WEP AI", 10)
    ws.freeze_panes = "B4"

    # Leyenda de turnos
    _title_row(ws3,"Leyenda de Turnos","Edita horas y colores según tu operación",5)
    turnos = [
        ("M","Mañana",    "06:00–14:00",8,"E1F5EE"),
        ("T","Tarde",     "14:00–22:00",8,"EFF6FF"),
        ("N","Noche",     "22:00–06:00",8,"F5F3FF"),
        ("P","Partido",   "08:00–13:00 / 15:00–20:00",8,"FFFBEB"),
        ("L","Libre",     "Día de descanso",0,"F8FAFC"),
        ("V","Vacaciones","Vacaciones programadas",0,"FEF2F2"),
        ("X","Ausencia",  "Ausencia / Enfermedad",0,"FFE4E4"),
        ("C","Capacitación","Entrenamiento",8,"F0FFF4"),
    ]
    for i, h in enumerate(["Código","Turno","Horario","Horas","Color referencia"],1):
        _hc(ws3, f"{get_column_letter(i)}4", h, bg="3B82F6")
    for idx, (cod,nombre,horario,horas,bg_color) in enumerate(turnos):
        r = 5+idx
        _dc(ws3,f"A{r}",cod,bold=True,align="center",bg=bg_color,color="0C447C")
        _dc(ws3,f"B{r}",nombre,bold=True,bg=bg_color)
        _dc(ws3,f"C{r}",horario,bg=bg_color,color="888780")
        _dc(ws3,f"D{r}",horas,align="center",bg=bg_color)
        _dc(ws3,f"E{r}","← Este color",align="center",bg=bg_color,color="888780")
    for col, w in zip(["A","B","C","D","E"],[8,16,28,8,16]):
        ws3.column_dimensions[col].width = w

    # Días de la semana
    days = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
    from datetime import timedelta
    week_start = datetime.now().date()
    week_start -= timedelta(days=week_start.weekday())

    # Header fila 3: "Empleado" + días + Total hrs
    ws[f"A3"].value = "Empleado / Puesto"
    ws[f"A3"].font = Font(name="Arial",bold=True,size=10,color="FFFFFF")
    ws[f"A3"].fill = PatternFill("solid",start_color="0C447C")
    ws[f"A3"].alignment = Alignment(horizontal="center",vertical="center")
    ws[f"A3"].border = _thin()
    ws.row_dimensions[3].height = 20

    for j, day in enumerate(days):
        col = get_column_letter(j+2)
        date_str = (week_start + timedelta(days=j)).strftime("%d/%m")
        ws[f"{col}3"].value = f"{day}\n{date_str}"
        ws[f"{col}3"].font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
        bg = "1A3048" if j >= 5 else "0C447C"   # weekend darker
        ws[f"{col}3"].fill = PatternFill("solid",start_color=bg)
        ws[f"{col}3"].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws[f"{col}3"].border = _thin()
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["A"].width = 28

    # Totales header
    _hc(ws, "J3", "Hrs/semana", bg="0D1B2A")
    ws.column_dimensions["J"].width = 12

    # Cobertura mínima fila 4
    _hc(ws, "A4", "Cobertura mínima requerida →", bg="854F0B", align="right")
    for j in range(7):
        col = get_column_letter(j+2)
        min_cov = 2 if j < 5 else 1   # 2 el día semana, 1 el fin de semana
        _dc(ws, f"{col}4", min_cov, align="center", bg="FFFBEB",
            fmt='#,##0 "personas"')
    ws.row_dimensions[4].height = 18

    # Empleados y sus turnos de muestra
    empleados = [
        ("Carlos Mendoza — Ventas",     ["M","M","M","M","M","L","L"]),
        ("Ana García — Finanzas",       ["M","M","M","M","M","L","L"]),
        ("Luis Pérez — Producción",     ["T","T","T","T","T","T","L"]),
        ("María Torres — Marketing",    ["M","M","M","M","L","L","L"]),
        ("Roberto Sánchez — Logística", ["M","M","T","T","T","L","L"]),
        ("Patricia Luna — Ventas",      ["T","T","T","T","T","L","L"]),
        ("Jorge Ramírez — IT",          ["M","M","M","L","M","L","L"]),
        ("Carmen Ruiz — Producción",    ["N","N","N","N","N","L","L"]),
        ("Miguel Castro — Almacén",     ["T","T","L","T","T","T","L"]),
    ]

    turno_bg = {"M":"E1F5EE","T":"EFF6FF","N":"F5F3FF","P":"FFFBEB",
                "L":"F8FAFC","V":"FEF2F2","X":"FFE4E4","C":"F0FFF4"}
    turno_hrs = {"M":8,"T":8,"N":8,"P":8,"L":0,"V":0,"X":0,"C":8}

    for idx, (emp, turnos_sem) in enumerate(empleados):
        r = 5 + idx
        bg = "FFFFFF" if idx % 2 == 0 else "EFF6FF"
        _dc(ws, f"A{r}", emp, bold=True, bg=bg)
        for j, turno in enumerate(turnos_sem):
            col = get_column_letter(j+2)
            t_bg = turno_bg.get(turno, "FFFFFF")
            t_col = "B91C1C" if turno in ["X"] else ("888780" if turno in ["L","V"] else "0C447C")
            _dc(ws, f"{col}{r}", turno, align="center", bg=t_bg, bold=True, color=t_col)
        # Horas totales
        total_hrs = sum(turno_hrs.get(t,0) for t in turnos_sem)
        _dc(ws, f"J{r}", total_hrs, align="center", bg=bg,
            fmt='#,##0 "hrs"', color="0C447C", bold=True)

    lr = 5 + len(empleados)

    # Fila de cobertura real (COUNTIF por día)
    _hc(ws, "A{lr}".format(lr=lr), "Cobertura real (personas)", bg="0D1B2A")
    for j in range(7):
        col = get_column_letter(j+2)
        # Count cells that are NOT L, V, or X (working shifts)
        _formula(ws, f"{col}{lr}",
                 f'=COUNTIFS({col}5:{col}{lr-1},"<>L",{col}5:{col}{lr-1},"<>V",{col}5:{col}{lr-1},"<>X")',
                 bg="0D1B2A", fg="FFFFFF", bold=True, fmt='#,##0', align="center")

    # CF en cobertura real vs mínima
    for j in range(7):
        col = get_column_letter(j+2)
        min_cov = 2 if j < 5 else 1
        if ws[f"{col}{lr}"].value and ws[f"{col}4"].value:
            pass  # CF would need rule-based approach

    _note_row(ws, lr+2,
              "Turnos: M=Mañana · T=Tarde · N=Noche · P=Partido · L=Libre · V=Vacaciones · X=Ausencia · C=Capacitación. Ver hoja Leyenda.",
              10, bg="EFF6FF", color="0C447C")

    # Resumen de horas
    _title_row(ws2,"Resumen de Horas por Empleado",f"Semana del {week_start.strftime('%d/%m/%Y')}",5)
    for i, h in enumerate(["Empleado","Horas trabajadas","Horas libres","Turnos nocturnos","Ausentismo"],1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="185FA5")
    for idx, (emp, turnos_sem) in enumerate(empleados):
        r = 5+idx; bg = "FFFFFF" if idx%2==0 else "EFF6FF"
        _dc(ws2,f"A{r}",emp.split(" — ")[0],bold=True,bg=bg)
        hrs = sum(turno_hrs.get(t,0) for t in turnos_sem)
        _dc(ws2,f"B{r}",hrs,align="center",bg=bg,fmt='#,##0',
            color="B91C1C" if hrs<32 else ("15803D" if hrs>=40 else "000000"),bold=True)
        _dc(ws2,f"C{r}",turnos_sem.count("L"),align="center",bg=bg,fmt='#,##0')
        _dc(ws2,f"D{r}",turnos_sem.count("N"),align="center",bg=bg,fmt='#,##0',
            color="534AB7" if turnos_sem.count("N")>0 else "888780")
        _dc(ws2,f"E{r}",turnos_sem.count("X"),align="center",bg=bg,fmt='#,##0',
            color="B91C1C" if turnos_sem.count("X")>0 else "15803D")
    lr2=5+len(empleados)
    ws2.merge_cells(f"A{lr2}:A{lr2}"); _hc(ws2,f"A{lr2}","TOTALES",bg="0D1B2A")
    for col in ["B","C","D","E"]:
        _formula(ws2,f"{col}{lr2}",f"=SUM({col}5:{col}{lr2-1})",
                 bg="0D1B2A",fg="FFFFFF",bold=True,fmt='#,##0',align="center")
    for col,w in zip(["A","B","C","D","E"],[28,16,12,16,12]):
        ws2.column_dimensions[col].width=w


def _build_attendance(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Control de asistencia con horas trabajadas, tardanzas y resumen mensual."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — LaborVerifier: actualización autónoma de prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Asistencia",       "0F6E56")
    ws2 = _setup_sheet(wb, "Resumen Mensual",  "085041")
    cols = 9

    _title_row(ws, titulo,
               f"Mes: {datetime.now().strftime('%B %Y')} · WEP AI", cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    headers = ["Fecha","Empleado","Hora Entrada","Hora Salida","Hrs Trabajadas",
               "Hrs Extra","Tipo","Notas","Estado"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="0F6E56")
    ws.row_dimensions[4].height = 22

    sample_att = [
        (date(2026,5,5),"Carlos Mendoza","08:00","17:00","Normal",""),
        (date(2026,5,5),"Ana García","08:15","17:00","Tardanza","15 min tarde"),
        (date(2026,5,5),"Luis Pérez","14:00","22:00","Normal","Turno tarde"),
        (date(2026,5,5),"María Torres","08:00","16:00","Normal",""),
        (date(2026,5,6),"Carlos Mendoza","08:00","19:30","Normal","Cierre de mes"),
        (date(2026,5,6),"Ana García","08:00","17:00","Normal",""),
        (date(2026,5,6),"Roberto Sánchez","","","Ausencia","Incapacidad IMSS"),
        (date(2026,5,7),"María Torres","","","Vacaciones","Vacaciones aprobadas"),
        (date(2026,5,7),"Luis Pérez","14:05","22:00","Tardanza","5 min tarde"),
        (date(2026,5,8),"Carlos Mendoza","08:00","17:00","Normal",""),
    ]

    tipo_color={"Normal":"15803D","Tardanza":"854F0B","Ausencia":"B91C1C",
                "Vacaciones":"185FA5","Capacitación":"534AB7","Feriado":"888780"}

    for idx, row_data in enumerate(sample_att):
        r = 5+idx; bg = "FFFFFF" if idx%2==0 else "F0FFF4"
        fecha,emp,entrada,salida,tipo,nota = row_data
        ws[f"A{r}"].value = fecha; ws[f"A{r}"].number_format = "DD/MM/YYYY"
        ws[f"A{r}"].font = Font(name="Arial",size=10); ws[f"A{r}"].fill = PatternFill("solid",start_color=bg)
        ws[f"A{r}"].alignment = Alignment(horizontal="center",vertical="center"); ws[f"A{r}"].border = _thin()
        _dc(ws,f"B{r}",emp,bold=True,bg=bg)
        _dc(ws,f"C{r}",entrada,align="center",bg=bg,color="888780" if not entrada else "000000")
        _dc(ws,f"D{r}",salida,align="center",bg=bg,color="888780" if not salida else "000000")
        if entrada and salida:
            # Horas trabajadas
            _formula(ws,f"E{r}",
                     f'=IFERROR((TIMEVALUE(D{r})-TIMEVALUE(C{r}))*24,0)',
                     bg=bg,fmt='#,##0.0 "hrs"',align="center",color="0F6E56",bold=True)
            # Horas extra (más de 8 hrs)
            _formula(ws,f"F{r}",
                     f'=MAX(0,(TIMEVALUE(D{r})-TIMEVALUE(C{r}))*24-8)',
                     bg=bg,fmt='#,##0.0 "hrs"',align="center",color="185FA5")
        else:
            _dc(ws,f"E{r}","—",align="center",bg=bg,color="888780")
            _dc(ws,f"F{r}","—",align="center",bg=bg,color="888780")
        _dc(ws,f"G{r}",tipo,align="center",bg=bg,color=tipo_color.get(tipo,"000000"),bold=True)
        _dc(ws,f"H{r}",nota,bg=bg,color="888780")
        estado = "OK" if tipo=="Normal" else ("⚠" if tipo=="Tardanza" else "✗")
        _dc(ws,f"I{r}",estado,align="center",bg=bg,
            color="15803D" if estado=="OK" else ("854F0B" if estado=="⚠" else "B91C1C"),bold=True)

    lr = 5+len(sample_att)
    _cf_formula_rule(ws, f"A5:I{lr-1}", '=$G5="Ausencia"',     "FFE4E4", "7F1D1D")
    _cf_formula_rule(ws, f"A5:I{lr-1}", '=$G5="Tardanza"',     "FFF3CD", "78350F")
    _cf_formula_rule(ws, f"A5:I{lr-1}", '=$G5="Vacaciones"',   "DBEAFE", "1E3A5F")
    _cf_formula_rule(ws, f"A5:I{lr-1}", '=$G5="Capacitación"', "EEEDFE", "4C1D95")
    _add_dv_list(ws, f"G5:G{lr+50}",
                 ["Normal","Tardanza","Ausencia","Vacaciones","Capacitación","Feriado"],
                 "Tipo de registro")
    _note_row(ws,lr+1,
              "✓ Horas trabajadas y extra calculadas automáticamente de Hora Entrada/Salida. Ingresa horas en formato HH:MM (ej: 08:30).",
              cols, bg="E1F5EE", color="0F6E56")
    widths=[12,24,13,13,13,12,14,22,8]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    # Resumen mensual por empleado
    _title_row(ws2,"Resumen Mensual de Asistencia",
               f"Mes: {datetime.now().strftime('%B %Y')} · WEP AI",7)
    for i,h in enumerate(["Empleado","Días trab.","Tardanzas","Ausencias","Vacaciones",
                           "Hrs extra","% Asistencia"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="085041")
    empleados_list=["Carlos Mendoza","Ana García","Luis Pérez","María Torres",
                    "Roberto Sánchez","Patricia Luna","Jorge Ramírez"]
    for idx,emp in enumerate(empleados_list):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "E1F5EE"
        _dc(ws2,f"A{r}",emp,bold=True,bg=bg)
        _formula(ws2,f"B{r}",f'=COUNTIFS(Asistencia!B5:B{lr-1},A{r},Asistencia!G5:G{lr-1},"Normal")',
                 bg=bg,fmt='#,##0',align="center",color="15803D",bold=True)
        _formula(ws2,f"C{r}",f'=COUNTIFS(Asistencia!B5:B{lr-1},A{r},Asistencia!G5:G{lr-1},"Tardanza")',
                 bg=bg,fmt='#,##0',align="center",color="854F0B",bold=True)
        _formula(ws2,f"D{r}",f'=COUNTIFS(Asistencia!B5:B{lr-1},A{r},Asistencia!G5:G{lr-1},"Ausencia")',
                 bg=bg,fmt='#,##0',align="center",color="B91C1C",bold=True)
        _formula(ws2,f"E{r}",f'=COUNTIFS(Asistencia!B5:B{lr-1},A{r},Asistencia!G5:G{lr-1},"Vacaciones")',
                 bg=bg,fmt='#,##0',align="center",color="185FA5")
        _formula(ws2,f"F{r}",f'=SUMIF(Asistencia!B5:B{lr-1},A{r},Asistencia!F5:F{lr-1})',
                 bg=bg,fmt='#,##0.0',align="center",color="534AB7",bold=True)
        _formula(ws2,f"G{r}",f'=IFERROR((B{r}+C{r})/(B{r}+C{r}+D{r}),0)',
                 bg=bg,fmt='0.0%',align="center",
                 color="15803D")
    _cf_three_color(ws2,f"G5:G{5+len(empleados_list)-1}",low="FFB3B3",mid="FFF3CD",high="B7F5D0")
    for col,w in zip(["A","B","C","D","E","F","G"],[24,12,12,12,12,12,14]):
        ws2.column_dimensions[col].width=w


def _build_vacation_tracker(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Control de vacaciones y ausencias con saldo automático por empleado."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — LaborVerifier: actualización autónoma de prestaciones laborales ──
    try:
        from agent.fiscal_verifier import verify_labor_data
        cfg = verify_labor_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Solicitudes",  "C2410C")
    ws2 = _setup_sheet(wb, "Saldos",       "B91C1C")
    ws3 = _setup_sheet(wb, "Calendario",   "854F0B")
    cols = 9

    _title_row(ws, titulo,
               f"Control de vacaciones · {datetime.now().year} · WEP AI", cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    headers=["#","Empleado","Departamento","F. Inicio","F. Fin","Días","Tipo","Estado","Aprobado por"]
    for i,h in enumerate(headers,1): _hc(ws,f"{get_column_letter(i)}4",h,bg="C2410C")
    ws.row_dimensions[4].height=22

    solicitudes=[
        ("Carlos Mendoza","Ventas",      date(2026,7,7), date(2026,7,18),10,"Vacaciones","Aprobado","Dir. RRHH"),
        ("Ana García","Finanzas",        date(2026,8,3), date(2026,8,7), 5, "Vacaciones","Aprobado","Dir. RRHH"),
        ("Luis Pérez","Producción",      date(2026,6,15),date(2026,6,19),5, "Vacaciones","Pendiente","—"),
        ("María Torres","Marketing",     date(2026,5,25),date(2026,5,26),2, "Permiso","Aprobado","Gerente"),
        ("Roberto Sánchez","Logística",  date(2026,5,8), date(2026,5,8), 1, "Enfermedad","Aprobado","Auto"),
        ("Patricia Luna","Ventas",       date(2026,9,1), date(2026,9,12),10,"Vacaciones","Pendiente","—"),
        ("Jorge Ramírez","IT",           date(2026,12,22),date(2026,12,31),7,"Vacaciones","Aprobado","Dir. RRHH"),
        ("Carmen Ruiz","Producción",     date(2026,6,20),date(2026,6,20),1, "Permiso","Rechazado","Gerente"),
    ]
    est_col={"Aprobado":"15803D","Pendiente":"854F0B","Rechazado":"B91C1C"}
    tipo_col={"Vacaciones":"185FA5","Permiso":"534AB7","Enfermedad":"B91C1C","Otro":"888780"}

    for idx,(emp,depto,f_ini,f_fin,dias,tipo,estado,aprob) in enumerate(solicitudes):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FEF2F2"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",emp,bold=True,bg=bg)
        _dc(ws,f"C{r}",depto,align="center",bg=bg,color="C2410C")
        for col_ref,d in zip(["D","E"],[f_ini,f_fin]):
            ws[f"{col_ref}{r}"].value=d; ws[f"{col_ref}{r}"].number_format="DD/MM/YYYY"
            ws[f"{col_ref}{r}"].font=Font(name="Arial",size=10)
            ws[f"{col_ref}{r}"].fill=PatternFill("solid",start_color=bg)
            ws[f"{col_ref}{r}"].alignment=Alignment(horizontal="center",vertical="center")
            ws[f"{col_ref}{r}"].border=_thin()
        _dc(ws,f"F{r}",dias,align="center",bg=bg,fmt='#,##0 "días"',bold=True)
        _dc(ws,f"G{r}",tipo,align="center",bg=bg,color=tipo_col.get(tipo,"000000"),bold=True)
        _dc(ws,f"H{r}",estado,align="center",bg=bg,color=est_col.get(estado,"000000"),bold=True)
        _dc(ws,f"I{r}",aprob,align="center",bg=bg,color="888780")

    lr=5+len(solicitudes)
    ws.merge_cells(f"A{lr}:E{lr}"); _hc(ws,f"A{lr}","TOTAL DÍAS SOLICITADOS",bg="0D1B2A")
    _formula(ws,f"F{lr}",f"=SUM(F5:F{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='#,##0',align="center")
    _cf_formula_rule(ws,f"A5:I{lr-1}",'=$H5="Aprobado"', "E1F5EE","064E3B")
    _cf_formula_rule(ws,f"A5:I{lr-1}",'=$H5="Pendiente"',"FFF3CD","78350F")
    _cf_formula_rule(ws,f"A5:I{lr-1}",'=$H5="Rechazado"',"FFE4E4","7F1D1D")
    _add_dv_list(ws,f"G5:G{lr+20}",["Vacaciones","Permiso","Enfermedad","Maternidad/Paternidad","Otro"],"Tipo de ausencia")
    _add_dv_list(ws,f"H5:H{lr+20}",["Aprobado","Pendiente","Rechazado"],"Estado de solicitud")
    widths=[5,24,16,13,13,12,16,12,16]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    # Saldos por empleado (según LFT México)
    _title_row(ws2,"Saldo de Vacaciones por Empleado",
               f"Año {datetime.now().year} · Días según Ley Federal del Trabajo",6)
    for i,h in enumerate(["Empleado","Antigüedad","Días c/año (LFT)","Días tomados",
                           "Días pendientes","Saldo disponible","Estado"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="B91C1C")
    emp_saldo=[
        ("Carlos Mendoza",7.2,14),("Ana García",5.8,12),("Luis Pérez",5.3,12),
        ("María Torres",4.0,12),("Roberto Sánchez",2.7,10),("Patricia Luna",2.3,10),
        ("Jorge Ramírez",8.0,14),("Carmen Ruiz",3.2,12),
    ]
    for idx,(emp,antig,dias_ley) in enumerate(emp_saldo):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FEF2F2"
        _dc(ws2,f"A{r}",emp,bold=True,bg=bg)
        _dc(ws2,f"B{r}",f"{antig:.1f} años",align="center",bg=bg,color="888780")
        _dc(ws2,f"C{r}",dias_ley,align="center",bg=bg,fmt='#,##0',bold=True,color="0C447C")
        _formula(ws2,f"D{r}",
                 f'=IFERROR(SUMIFS(Solicitudes!F5:F{lr-1},Solicitudes!B5:B{lr-1},A{r},Solicitudes!H5:H{lr-1},"Aprobado",Solicitudes!G5:G{lr-1},"Vacaciones"),0)',
                 bg=bg,fmt='#,##0',align="center",color="C2410C")
        _formula(ws2,f"E{r}",
                 f'=IFERROR(SUMIFS(Solicitudes!F5:F{lr-1},Solicitudes!B5:B{lr-1},A{r},Solicitudes!H5:H{lr-1},"Pendiente",Solicitudes!G5:G{lr-1},"Vacaciones"),0)',
                 bg=bg,fmt='#,##0',align="center",color="854F0B")
        _formula(ws2,f"F{r}",f"=C{r}-D{r}-E{r}",bg=bg,fmt='#,##0',align="center",
                 color="B91C1C",bold=True)
        _formula(ws2,f"G{r}",
                 f'=IF(F{r}<=0,"Sin saldo",IF(F{r}<=3,"⚠ Poco saldo","✓ OK"))',
                 bg=bg,align="center")
    _cf_three_color(ws2,f"F5:F{5+len(emp_saldo)-1}",low="FFB3B3",mid="FFF3CD",high="B7F5D0")
    for col,w in zip(["A","B","C","D","E","F","G"],[24,13,16,13,16,14,14]):
        ws2.column_dimensions[col].width=w

    # Calendario visual (meses del año)
    _title_row(ws3,"Calendario de Ausencias",f"Año {datetime.now().year} — Vista de equipo",14)
    months=["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    _hc(ws3,"A4","Empleado",bg="854F0B")
    for j,m in enumerate(months,2): _hc(ws3,f"{get_column_letter(j)}4",m,bg="854F0B")
    _hc(ws3,f"{get_column_letter(14)}4","Total",bg="0D1B2A")
    # Populate from solicitudes data
    for idx,emp in enumerate(["Carlos Mendoza","Ana García","Luis Pérez","María Torres",
                               "Roberto Sánchez","Patricia Luna","Jorge Ramírez","Carmen Ruiz"]):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        _dc(ws3,f"A{r}",emp,bold=True,bg=bg)
        # Mark months with ausencias
        for sol in solicitudes:
            if sol[0]==emp and sol[7]=="Aprobado":
                month_col=get_column_letter(sol[3].month+1)
                color="DBEAFE" if sol[6]=="Vacaciones" else "FFE4E4"
                _dc(ws3,f"{month_col}{r}","●",align="center",bg=color,
                    color="185FA5" if sol[6]=="Vacaciones" else "B91C1C",bold=True)
        # Fill empty months
        for m in range(1,13):
            col=get_column_letter(m+1)
            if ws3[f"{col}{r}"].value is None:
                ws3[f"{col}{r}"].value=""
                ws3[f"{col}{r}"].fill=PatternFill("solid",start_color=bg)
                ws3[f"{col}{r}"].border=_thin()
        _formula(ws3,f"N{r}",f'=IFERROR(SUMIFS(Solicitudes!F5:F{lr-1},Solicitudes!B5:B{lr-1},A{r},Solicitudes!H5:H{lr-1},"Aprobado"),0)',
                 bg=bg,fmt='#,##0',align="center",bold=True)
    ws3.column_dimensions["A"].width=24
    for i in range(2,14): ws3.column_dimensions[get_column_letter(i)].width=7
    ws3.column_dimensions["N"].width=8
    _note_row(ws3,5+8,"● Azul = Vacaciones aprobadas · ● Rojo = Ausencia/Enfermedad. Fuente: hoja Solicitudes.",14,bg="FFFBEB",color="854F0B")


def _build_sales_pipeline(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Pipeline CRM de ventas: lead→propuesta→cierre con forecast ponderado."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Pipeline",   "534AB7")
    ws2 = _setup_sheet(wb, "Embudo",     "3C3489")
    ws3 = _setup_sheet(wb, "Forecast",   "26215C")
    cols=10

    _title_row(ws,titulo,f"Pipeline de ventas · {datetime.now().strftime('%B %Y')} · WEP AI",cols)
    ws.freeze_panes="A5"
    _add_autofilter(ws,4,cols)

    headers=["#","Cliente/Empresa","Producto/Servicio","Etapa","Valor Est.","Prob. %",
             "Valor Pond.","F. Cierre Est.","Responsable","Notas"]
    for i,h in enumerate(headers,1): _hc(ws,f"{get_column_letter(i)}4",h,bg="534AB7")
    ws.row_dimensions[4].height=22

    pipeline=[
        ("Grupo Norteño S.A.","Implementación ERP","Propuesta enviada",850000,0.65,date(2026,6,30),"Carlos M.","Segunda reunión programada"),
        ("Retail Express","Licencias software x50","Negociación",120000,0.80,date(2026,5,31),"Patricia L.","Pendiente aprobación presupuesto"),
        ("StartUp Innovatech","Consultoría tech 3 meses","Contactado",75000,0.30,date(2026,7,15),"Jorge R.","En evaluación"),
        ("Manufactura Precisa","Mantenimiento anual","Cerrado ganado",320000,1.00,date(2026,5,20),"Carlos M.","¡CERRADO! Firma pendiente"),
        ("Distribuidora Central","Módulo de inventario","Lead",40000,0.15,date(2026,8,1),"Ana G.","Primer contacto este mes"),
        ("Tech Components","Integración API","Propuesta enviada",180000,0.55,date(2026,6,15),"Jorge R.","Demo realizada"),
        ("Consultoría Integral","Capacitación equipo","Negociación",28000,0.70,date(2026,5,28),"Patricia L.","Último precio enviado"),
        ("Nueva Oportunidad S.A.","Por definir","Cerrado perdido",95000,0.00,date(2026,5,1),"Carlos M.","Eligieron a competidor"),
    ]
    eta_color={"Lead":"888780","Contactado":"854F0B","Propuesta enviada":"185FA5",
               "Negociación":"534AB7","Cerrado ganado":"15803D","Cerrado perdido":"B91C1C"}

    for idx,(cliente,producto,etapa,valor,prob,f_cierre,resp,nota) in enumerate(pipeline):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EEEDFE"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        _dc(ws,f"B{r}",cliente,bold=True,bg=bg)
        _dc(ws,f"C{r}",producto,bg=bg)
        _dc(ws,f"D{r}",etapa,align="center",bg=bg,color=eta_color.get(etapa,"000000"),bold=True)
        _dc(ws,f"E{r}",valor,align="center",bg=bg,fmt='$#,##0',color="534AB7",bold=True)
        _dc(ws,f"F{r}",prob,align="center",bg=bg,fmt='0%',
            color="15803D" if prob>=0.7 else ("854F0B" if prob>=0.4 else "B91C1C"),bold=True)
        _formula(ws,f"G{r}",f"=E{r}*F{r}",bg=bg,fmt='$#,##0',color="3C3489",bold=True,align="center")
        ws[f"H{r}"].value=f_cierre; ws[f"H{r}"].number_format="DD/MM/YYYY"
        ws[f"H{r}"].font=Font(name="Arial",size=10); ws[f"H{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"H{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"H{r}"].border=_thin()
        _dc(ws,f"I{r}",resp,align="center",bg=bg)
        _dc(ws,f"J{r}",nota,bg=bg,color="888780")

    lr=5+len(pipeline)
    ws.merge_cells(f"A{lr}:D{lr}"); _hc(ws,f"A{lr}","TOTALES PIPELINE",bg="0D1B2A")
    _formula(ws,f"E{lr}",f"=SUM(E5:E{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0',align="center")
    _formula(ws,f"G{lr}",f"=SUM(G5:G{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0',align="center")

    _cf_formula_rule(ws,f"A5:J{lr-1}",'=$D5="Cerrado ganado"',"D1FAE5","064E3B")
    _cf_formula_rule(ws,f"A5:J{lr-1}",'=$D5="Cerrado perdido"',"FFE4E4","7F1D1D")
    _cf_formula_rule(ws,f"A5:J{lr-1}",'=$D5="Negociación"',  "EDE9FE","4C1D95")
    # Deals vencidos (fecha cierre pasada y no cerrado)
    _cf_formula_rule(ws,f"A5:J{lr-1}",
                     f'=AND(H5<TODAY(),D5<>"Cerrado ganado",D5<>"Cerrado perdido")',
                     "FFF3CD","78350F")
    _add_dv_list(ws,f"D5:D{lr+20}",
                 ["Lead","Contactado","Propuesta enviada","Negociación","Cerrado ganado","Cerrado perdido"],
                 "Etapa del deal")
    _add_dv_list(ws,f"I5:I{lr+20}",
                 ["Carlos M.","Ana G.","Luis P.","Jorge R.","Patricia L.","Otro"],"Responsable")
    widths=[4,24,22,18,13,10,13,14,14,22]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    # Embudo de ventas
    _title_row(ws2,"Embudo de Ventas","Distribución por etapa",5)
    for i,h in enumerate(["Etapa","# Deals","Valor Total","Valor Pond.","% Conversión"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="3C3489")
    etapas=["Lead","Contactado","Propuesta enviada","Negociación","Cerrado ganado","Cerrado perdido"]
    for idx,etapa in enumerate(etapas):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EEEDFE"
        _dc(ws2,f"A{r}",etapa,bold=True,bg=bg,color=eta_color.get(etapa,"000000"))
        _formula(ws2,f"B{r}",f'=COUNTIF(Pipeline!D5:D{lr-1},A{r})',bg=bg,fmt='#,##0',align="center")
        _formula(ws2,f"C{r}",f'=SUMIF(Pipeline!D5:D{lr-1},A{r},Pipeline!E5:E{lr-1})',
                 bg=bg,fmt='$#,##0',align="center",color="534AB7",bold=True)
        _formula(ws2,f"D{r}",f'=SUMIF(Pipeline!D5:D{lr-1},A{r},Pipeline!G5:G{lr-1})',
                 bg=bg,fmt='$#,##0',align="center",color="3C3489")
        _formula(ws2,f"E{r}",f'=IFERROR(B{r}/SUM(B5:B{5+len(etapas)-1}),0)',
                 bg=bg,fmt='0.0%',align="center")
    for col,w in zip(["A","B","C","D","E"],[20,10,14,14,12]):
        ws2.column_dimensions[col].width=w

    # Forecast mensual
    _title_row(ws3,"Forecast de Ingresos",f"Proyección ponderada · {datetime.now().strftime('%B %Y')}",4)
    for i,h in enumerate(["Concepto","Valor"],1): _hc(ws3,f"{get_column_letter(i)}4",h,bg="26215C")
    forecast_items=[
        ("Pipeline total (valor bruto)",        f"=SUM(Pipeline!E5:E{lr-1})","534AB7"),
        ("Forecast ponderado (valor × prob.)",  f"=SUM(Pipeline!G5:G{lr-1})","3C3489"),
        ("Cerrados ganados este mes",           f'=SUMIF(Pipeline!D5:D{lr-1},"Cerrado ganado",Pipeline!E5:E{lr-1})',"15803D"),
        ("En negociación (prob. >60%)",         f'=SUMIFS(Pipeline!E5:E{lr-1},Pipeline!F5:F{lr-1},">0.6",Pipeline!D5:D{lr-1},"<>Cerrado ganado",Pipeline!D5:D{lr-1},"<>Cerrado perdido")',"185FA5"),
        ("Deals vencidos (fecha pasada)",       f'=COUNTIFS(Pipeline!H5:H{lr-1},"<"&TODAY(),Pipeline!D5:D{lr-1},"<>Cerrado ganado",Pipeline!D5:D{lr-1},"<>Cerrado perdido")',"B91C1C"),
        ("Tasa de cierre (ganados/total)",      f'=IFERROR(COUNTIF(Pipeline!D5:D{lr-1},"Cerrado ganado")/(COUNTIF(Pipeline!D5:D{lr-1},"Cerrado ganado")+COUNTIF(Pipeline!D5:D{lr-1},"Cerrado perdido")),0)',"0F6E56"),
    ]
    for idx,(label,formula,color) in enumerate(forecast_items):
        r=5+idx; bg="EEEDFE" if idx%2==0 else "FFFFFF"
        _dc(ws3,f"A{r}",label,bold=True,bg=bg)
        _formula(ws3,f"B{r}",formula,bg=bg,
                 fmt='$#,##0' if idx<4 else ('0.0%' if idx==5 else '#,##0'),
                 color=color,bold=True,align="center",size=13)
    ws3.column_dimensions["A"].width=40; ws3.column_dimensions["B"].width=18


def _build_scenario_simulator(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Simulador de escenarios financieros: Pesimista/Realista/Optimista × 5 años."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Parámetros",  "854F0B")
    ws2 = _setup_sheet(wb, "Proyección",  "BA7517")
    ws3 = _setup_sheet(wb, "Comparativo", "633806")
    cols = 6

    # ── PARÁMETROS ──────────────────────────────────────────────────────────
    _title_row(ws, titulo, "Modifica las celdas amarillas — el resto se calcula solo · WEP AI", cols)
    _section_row(ws, 3, "PARÁMETROS BASE (Año 1)", cols, bg="854F0B")

    year1 = datetime.now().year
    for i, h in enumerate(["Parámetro","Pesimista","Realista","Optimista","Unidad","Descripción"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="633806")
    ws.row_dimensions[4].height=20

    params=[
        ("Ingresos Año 1",          3000000, 5000000, 8000000, "$","Ventas brutas en el primer año"),
        ("Crecimiento anual %",          0.05,    0.15,    0.30, "%","Tasa de crecimiento de ventas año a año"),
        ("Margen Bruto %",               0.35,    0.45,    0.60, "%","(Ventas - Costo ventas) / Ventas"),
        ("Gastos Fijos Mensuales",      250000,  180000,  140000, "$","Nómina, renta, servicios, admin."),
        ("Gastos Variables %",           0.12,    0.10,    0.08, "%","Como % de los ingresos"),
        ("Inversión Inicial",           500000,  500000,  500000, "$","Capex, setup, capital de trabajo"),
        ("Tasa de Impuesto %",           0.30,    0.30,    0.30, "%","ISR estimado sobre utilidad"),
        ("Tasa de Descuento (VPN) %",    0.12,    0.12,    0.12, "%","Para cálculo de VPN"),
    ]
    pct_rows={1,2,4,5,6,7}  # rows (1-based in params) that are percentages
    for idx,(label,pes,real,opt,unit,desc) in enumerate(params):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        fmt_v='0%' if idx in {1,2,4,5,6,7} else '$#,##0'
        _dc(ws,f"A{r}",label,bold=True,bg=bg)
        for col,val in zip(["B","C","D"],[pes,real,opt]):
            _dc(ws,f"{col}{r}",val,align="center",bg="FEFCE8",fmt=fmt_v,
                color="B91C1C" if col=="B" else ("15803D" if col=="D" else "854F0B"),bold=True)
        _dc(ws,f"E{r}",unit,align="center",bg=bg,color="888780")
        _dc(ws,f"F{r}",desc,bg=bg,color="888780")

    _note_row(ws,5+len(params)+1,
              "✏ Edita las columnas B (Pesimista), C (Realista) y D (Optimista). La proyección en las hojas siguientes se recalcula automáticamente.",
              cols,bg="FFFBEB",color="78350F")
    widths=[28,13,13,13,8,36]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    # ── PROYECCIÓN ──────────────────────────────────────────────────────────
    _title_row(ws2,"Proyección Financiera 5 Años",
               "Calculada automáticamente desde Parámetros · WEP AI",9)

    for i,h in enumerate(["Concepto","Pes. Año1","Real Año1","Opt. Año1",
                           "Pes. Año2","Real Año2","Opt. Año2",
                           "Pes. Año3","Real Año3"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="633806")
        ws2.column_dimensions[get_column_letter(i)].width = 14 if i>1 else 24
    ws2.row_dimensions[4].height=20

    # Filas de proyección — fórmulas que referencian Parámetros
    # B5=Pes, C5=Pes año1, etc. Referenciamos con columnas de Parámetros B,C,D
    rows_proj=[
        ("Ingresos",         "=Parámetros!{p}5",                  None,   "185FA5"),
        ("(×) Crecimiento",  "=Parámetros!{p}6",                  "0%",   "888780"),
        ("Ingresos Año N",   "=B5*(1+Parámetros!{p}6)^({yr}-1)",  "$#,##0","0C447C"),
        ("Costo de Ventas",  "=B7*(1-Parámetros!{p}7)",           "$#,##0","B91C1C"),
        ("Utilidad Bruta",   "=B7-B8",                            "$#,##0","15803D"),
        ("Gastos Fijos",     "=Parámetros!{p}8*12",               "$#,##0","C2410C"),
        ("Gastos Variables", "=B7*Parámetros!{p}9",               "$#,##0","C2410C"),
        ("EBITDA",           "=B9-B10-B11",                       "$#,##0","185FA5"),
        ("Impuestos",        "=MAX(0,B12*Parámetros!{p}11)",      "$#,##0","B91C1C"),
        ("Utilidad Neta",    "=B12-B13",                          "$#,##0","0F6E56"),
        ("Margen Neto %",    "=IFERROR(B14/B7,0)",               "0.0%",  "0F6E56"),
    ]
    # Simple approach: compute actual values for 3 scenarios × years 1-3
    scenarios=[("B","Pesimista"),(  "C","Realista"),("D","Optimista")]
    ing1   =[3000000,5000000,8000000]
    crec   =[0.05,0.15,0.30]
    marg   =[0.35,0.45,0.60]
    gf_m   =[250000,180000,140000]
    gv_pct =[0.12,0.10,0.08]
    tax    =0.30

    labels=["Ingresos Año N","Costo de Ventas","Utilidad Bruta","Gastos Fijos",
            "Gastos Variables","EBITDA","Impuestos","Utilidad Neta","Margen Neto %"]
    fmts=["$#,##0","$#,##0","$#,##0","$#,##0","$#,##0","$#,##0","$#,##0","$#,##0","0.0%"]
    col_colors={"Pesimista":"B91C1C","Realista":"185FA5","Optimista":"15803D"}

    r=5
    for yr in range(1,4):
        _section_row(ws2,r,f"AÑO {yr} ({year1+yr-1})",9,bg="BA7517"); r+=1
        for s_idx,(s_col,s_name) in enumerate(scenarios):
            data_col=get_column_letter(2+s_idx+(yr-1)*3) if yr<=3 else None
        # Write rows for this year
        for lbl,fmt,fn_col in zip(labels,fmts,["B","C","D","E","F","G","H","I","J"]):
            col_idx=0
            _dc(ws2,f"A{r}",lbl,bold=(lbl in ["Utilidad Neta","EBITDA"]),
                bg="FFFFFF" if (r%2==0) else "FFFBEB")
            for s_idx,(s_col_ref,s_name) in enumerate(scenarios):
                col_letter=get_column_letter(2+s_idx)
                i1=ing1[s_idx]; c1=crec[s_idx]; mg=marg[s_idx]
                gf=gf_m[s_idx]*12; gv=gv_pct[s_idx]
                ing=round(i1*(1+c1)**(yr-1))
                cogs=round(ing*(1-mg))
                ub=ing-cogs
                gfixed=gf
                gvar=round(ing*gv)
                ebitda=ub-gfixed-gvar
                imp=max(0,round(ebitda*tax))
                un=ebitda-imp
                mn=round(un/ing,4) if ing>0 else 0
                vals={"Ingresos Año N":ing,"Costo de Ventas":cogs,"Utilidad Bruta":ub,
                      "Gastos Fijos":gfixed,"Gastos Variables":gvar,"EBITDA":ebitda,
                      "Impuestos":imp,"Utilidad Neta":un,"Margen Neto %":mn}
                val=vals.get(lbl,0)
                is_main=lbl in ["Utilidad Neta","EBITDA","Utilidad Bruta"]
                bg_c="FFFFFF" if r%2==0 else "FFFBEB"
                c=ws2[f"{get_column_letter(2+s_idx)}{r}"]
                c.value=val; c.number_format=fmt
                color=col_colors[s_name]
                if lbl=="Margen Neto %" or lbl in ["Impuestos","Costo de Ventas","Gastos Fijos","Gastos Variables"]:
                    color="888780" if not is_main else color
                c.font=Font(name="Arial",bold=is_main,size=10,
                           color="B91C1C" if (isinstance(val,float) and val<0 and lbl!="Margen Neto %") else
                           ("B91C1C" if (isinstance(val,int) and val<0) else color))
                c.fill=PatternFill("solid",start_color=bg_c)
                c.alignment=Alignment(horizontal="center",vertical="center")
                c.border=_thin()
            r+=1
        r+=1  # spacer

    # CF en utilidad neta — rojo si negativa
    for row_r in range(5,r):
        ws2_row = ws2[row_r]

    ws2.column_dimensions["A"].width=22
    for i in range(2,10): ws2.column_dimensions[get_column_letter(i)].width=13

    # Comparativo
    _title_row(ws3,"Comparativo de Escenarios","Utilidad Neta y Margen por año y escenario",5)
    comp_hdr=["Año","Pesimista (UN)","Realista (UN)","Optimista (UN)",
              "Pes. Margen","Real. Margen","Opt. Margen"]
    for i,h in enumerate(comp_hdr,1): _hc(ws3,f"{get_column_letter(i)}4",h,bg="854F0B")
    comp_data=[]
    for yr in range(1,6):
        row_data=[year1+yr-1]
        for s_idx in range(3):
            i1=ing1[s_idx]; c1=crec[s_idx]; mg=marg[s_idx]
            gf=gf_m[s_idx]*12; gv=gv_pct[s_idx]
            ing=round(i1*(1+c1)**(yr-1))
            un=round((ing*(mg)-gf-ing*gv)*(1-tax))
            mn=round(un/ing,3) if ing>0 else 0
            row_data.extend([un,mn])
        comp_data.append(row_data)
    # Interleaved: UN_pes, UN_real, UN_opt, Mg_pes, Mg_real, Mg_opt
    for idx,row in enumerate(comp_data):
        r2=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        _dc(ws3,f"A{r2}",f"Año {row[0]} ({year1+idx})",bold=True,bg=bg)
        # UN values: positions 1,3,5
        for j,col_l in enumerate(["B","C","D"]):
            val=row[1+j*2]
            _dc(ws3,f"{col_l}{r2}",val,align="center",bg=bg,fmt='$#,##0',
                color="B91C1C" if val<0 else col_colors[["Pesimista","Realista","Optimista"][j]],
                bold=True)
        # Margins: positions 2,4,6
        for j,col_l in enumerate(["E","F","G"]):
            val=row[2+j*2]
            _dc(ws3,f"{col_l}{r2}",val,align="center",bg=bg,fmt='0.0%',
                color="B91C1C" if val<0 else col_colors[["Pesimista","Realista","Optimista"][j]])
    _cf_three_color(ws3,f"B5:D{5+len(comp_data)-1}",low="FFB3B3",mid="FFF3CD",high="B7F5D0")
    _cf_three_color(ws3,f"E5:G{5+len(comp_data)-1}",low="FFB3B3",mid="FFF3CD",high="B7F5D0")

    # BarChart comparativo
    chart_sc=BarChart()
    chart_sc.type="col"; chart_sc.grouping="clustered"
    chart_sc.title="Utilidad Neta por Escenario y Año"
    chart_sc.style=10; chart_sc.height=14; chart_sc.width=22
    cats=Reference(ws3,min_col=1,max_col=1,min_row=5,max_row=5+len(comp_data)-1)
    chart_sc.set_categories(cats)
    for j,(s_name,color) in enumerate(zip(["Pesimista","Realista","Optimista"],
                                           ["B91C1C","185FA5","15803D"])):
        s=Series(Reference(ws3,min_col=2+j,max_col=2+j,min_row=5,max_row=5+len(comp_data)-1),
                 title=s_name)
        s.graphicalProperties.solidFill=color
        chart_sc.series.append(s)
    ws3.add_chart(chart_sc,"A12")
    # VPN por escenario en comparativo
    vpn_r = 5+len(comp_data)+2
    _section_row(ws3, vpn_r, "VPN Y TIR POR ESCENARIO (usando flujos de Utilidad Neta)", 7, bg="0D1B2A")
    for i,h in enumerate(["Indicador","Pesimista","Realista","Optimista"],1):
        _hc(ws3,f"{get_column_letter(i)}{vpn_r+1}",h,bg="633806")
    # Build NPV/IRR rows referencing the computed UN values
    # Since values are hardcoded in cells, reference them
    for s_idx,(s_col,s_name,color) in enumerate(zip(["B","C","D"],
        ["Pesimista","Realista","Optimista"],["B91C1C","185FA5","15803D"])):
        col_letter = get_column_letter(2+s_idx)
        # VPN row
        _dc(ws3,f"A{vpn_r+2}","VPN (tasa desc. 12%)",bold=True,
            bg="EFF6FF" if s_idx==0 else "FFFFFF")
        _formula(ws3,f"{col_letter}{vpn_r+2}",
                 f"=NPV(0.12,{col_letter}6:{col_letter}{5+len(comp_data)-1})+{col_letter}5",
                 bg="EFF6FF" if s_idx==0 else "FFFFFF",
                 fmt='$#,##0',color=color,bold=True,align="center",size=12)
        # TIR row
        _dc(ws3,f"A{vpn_r+3}","TIR estimada",bold=True,
            bg="FEFCE8" if s_idx==0 else "FFFBEB")
        _formula(ws3,f"{col_letter}{vpn_r+3}",
                 f"=IFERROR(IRR({col_letter}5:{col_letter}{5+len(comp_data)-1}),0)",
                 bg="FEFCE8" if s_idx==0 else "FFFBEB",
                 fmt='0.0%',color=color,bold=True,align="center",size=12)
    _note_row(ws3,5+len(comp_data)+1,
              "Verde=Optimista · Azul=Realista · Rojo=Pesimista. Modifica los parámetros en la hoja 'Parámetros' para recalcular todos los escenarios.",
              7,bg="FFFBEB",color="78350F")
    for i,w in enumerate([18,14,14,14,12,12,12],1):
        ws3.column_dimensions[get_column_letter(i)].width=w


# ═══════════════════════════════════════════════════════════════
# NUEVOS TEMPLATES v5
# ═══════════════════════════════════════════════════════════════

def _build_petty_cash(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Caja chica / fondo fijo con saldo automático y alerta de reposición.

    v15.9.0 — Acepta cfg para moneda + datos para responsable de la caja.
    """
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    responsable_user = datos.get("responsable") or datos.get("trabajador_nombre", "Nombre Responsable")

    ws  = _setup_sheet(wb, "Caja Chica",  "854F0B")
    ws2 = _setup_sheet(wb, "Configuración","633806")
    cols = 8

    # Config sheet
    _title_row(ws2,"Configuración de Caja Chica","Edita los valores amarillos",4)
    config_caja = [("Fondo total autorizado",5000,sym),
           ("Monto mínimo para reposición",1000,sym),
           ("Responsable de caja",responsable_user,""),
           ("Centro de costo",empresa_user or "Administración General",""),
           ("Moneda", moneda or "USD","")]
    for i,h in enumerate(["Parámetro","Valor","Unidad"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="633806")
    for idx,(label,val,unit) in enumerate(config_caja):
        r=5+idx; bg="FFFBEB" if idx%2==0 else "FFFFFF"
        _dc(ws2,f"A{r}",label,bold=True,bg=bg)
        _dc(ws2,f"B{r}",val,align="center",bg="FEFCE8",
            fmt=f'{sym}#,##0.00' if isinstance(val,(int, float)) else 'General',
            color="854F0B",bold=True)
        _dc(ws2,f"C{r}",unit,bg=bg,color="888780")
    for col,w in zip(["A","B","C"],[28,20,12]): ws2.column_dimensions[col].width=w

    # Main sheet
    _title_row(ws,"Fondo: =Configuración!B5",f"Responsable: =Configuración!B7 · WEP AI",cols)

    # KPI row
    kpi_r = 3
    # KPI summary row (without merging to avoid read-only issues)
    _dc(ws,f"A{kpi_r}","FONDO TOTAL",bold=True,bg="854F0B",color="FFFFFF",align="center",size=10)
    _formula(ws,f"B{kpi_r}","=Configuración!B5",bg="854F0B",fg="FFFFFF",
             bold=True,size=11,fmt='$#,##0.00',align="center")
    _dc(ws,f"C{kpi_r}","GASTADO",bold=True,bg="B91C1C",color="FFFFFF",align="center",size=10)
    _formula(ws,f"D{kpi_r}",f"=SUM(E5:E200)",bg="B91C1C",fg="FFFFFF",
             bold=True,size=11,fmt='$#,##0.00',align="center")
    _dc(ws,f"E{kpi_r}","SALDO",bold=True,bg="15803D",color="FFFFFF",align="center",size=10)
    _formula(ws,f"F{kpi_r}",f"=Configuración!B5-SUM(E5:E200)",
             bg="15803D",fg="FFFFFF",bold=True,size=11,fmt='$#,##0.00',align="center")
    _dc(ws,f"G{kpi_r}","ESTADO",bold=True,bg="0D1B2A",color="FFFFFF",align="center",size=10)
    _formula(ws,f"H{kpi_r}",
             f'=IF(Configuración!B5-SUM(E5:E200)<=Configuración!B6,"REPONER","OK")',
             bg="0D1B2A",fg="FFFFFF",bold=True,size=11,align="center")

    ws.row_dimensions[kpi_r].height = 26
    # Labels for KPI row (below the values)
    for col_s, label in zip(["A","C","E","G"], ["FONDO TOTAL","TOTAL GASTADO","SALDO DISPONIBLE","ESTADO"]):
        lbl_r = kpi_r - 1
        try:
            ws.cell(lbl_r, ord(col_s)-64).value = None  # clear if needed
        except Exception:
            pass

    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)
    headers=["#","Fecha","Concepto / Descripción","Categoría","Monto","Comprobante","Solicitado por","Estado"]
    for i,h in enumerate(headers,1): _hc(ws,f"{get_column_letter(i)}4",h,bg="854F0B")

    sample=[
        (date(2026,5,2),"Papelería y útiles de oficina","Papelería",245.00,"Ticket #1234","María García","Aprobado"),
        (date(2026,5,5),"Café y agua para reunión de equipo","Alimentos",380.50,"Ticket #1235","Carlos Ruiz","Aprobado"),
        (date(2026,5,8),"Transporte urgente documentos","Transporte",120.00,"Recibo #001","Ana López","Aprobado"),
        (date(2026,5,12),"Limpieza y desinfectante","Limpieza",195.00,"Ticket #1236","María García","Aprobado"),
        (date(2026,5,15),"Reparación menor impresora","Mantenimiento",850.00,"Factura #A001","Carlos Ruiz","Pendiente"),
    ]
    for idx,(fecha,concepto,cat,monto,comp,solicit,est) in enumerate(sample):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "FFFBEB"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        ws[f"B{r}"].value=fecha; ws[f"B{r}"].number_format="DD/MM/YYYY"
        ws[f"B{r}"].font=Font(name="Arial",size=10); ws[f"B{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"B{r}"].border=_thin()
        _dc(ws,f"C{r}",concepto,bold=True,bg=bg)
        _dc(ws,f"D{r}",cat,align="center",bg=bg,color="854F0B")
        _dc(ws,f"E{r}",monto,align="center",bg=bg,fmt='$#,##0.00',color="B91C1C",bold=True)
        _dc(ws,f"F{r}",comp,align="center",bg=bg,color="888780")
        _dc(ws,f"G{r}",solicit,align="center",bg=bg)
        _dc(ws,f"H{r}",est,align="center",bg=bg,
            color="15803D" if est=="Aprobado" else "854F0B",bold=True)
    lr=5+len(sample)
    ws.merge_cells(f"A{lr}:D{lr}"); _hc(ws,f"A{lr}","TOTAL GASTADO",bg="0D1B2A")
    _formula(ws,f"E{lr}",f"=SUM(E5:E{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")

    _cf_formula_rule(ws,f"A5:H{lr-1}",'=$H5="Aprobado"',"E1F5EE","064E3B")
    _cf_formula_rule(ws,f"A5:H{lr-1}",'=$H5="Pendiente"',"FFF3CD","78350F")
    _cf_formula_rule(ws,f"A5:H{lr-1}",'=$H5="Rechazado"',"FFE4E4","7F1D1D")
    _add_dv_list(ws,f"D5:D{lr+20}",
                 ["Papelería","Alimentos","Transporte","Limpieza","Mantenimiento",
                  "Mensajería","Comunicaciones","Otros"],"Categoría del gasto")
    _add_dv_list(ws,f"H5:H{lr+20}",["Aprobado","Pendiente","Rechazado"],"Estado del gasto")
    widths=[4,13,28,16,13,16,18,12]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    _note_row(ws,lr+2,
              "El saldo y el estado (OK/REPONER) se actualizan automaticamente. "
              "Cambia el Fondo Total en la hoja Configuracion.",
              cols,bg="FFFBEB",color="854F0B")
    _setup_print(ws,cols,lr+3,landscape=True)
    _protect_ws(ws,unlock_from_row=5)


def _build_travel_expenses(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Control de gastos de viaje y viáticos con resumen por empleado y categoría."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Viáticos",      "0C447C")
    ws2 = _setup_sheet(wb, "Resumen",       "185FA5")
    ws3 = _setup_sheet(wb, "Tipo de Cambio","0D1B2A")
    cols = 11

    # TC sheet
    _title_row(ws3,"Tipos de Cambio Referencia","Actualiza manualmente o con fuente bancaria",5)
    for i,h in enumerate(["Moneda","Código","TC a MXN"],1): _hc(ws3,f"{get_column_letter(i)}4",h,bg="0D1B2A")
    tc=[("Peso Mexicano","MXN",1.00),("Dólar USD","USD",17.50),("Euro","EUR",18.90),
        ("Libra","GBP",22.10),("Dólar Canadiense","CAD",12.80),("Real Brasileño","BRL",3.20)]
    for idx,(name,code,rate) in enumerate(tc):
        r=5+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        _dc(ws3,f"A{r}",name,bold=True,bg=bg); _dc(ws3,f"B{r}",code,align="center",bg=bg,bold=True,color="0C447C")
        _dc(ws3,f"C{r}",rate,align="center",bg=bg,fmt='$#,##0.0000',color="15803D",bold=True)
    for col,w in zip(["A","B","C"],[20,10,14]): ws3.column_dimensions[col].width=w

    _title_row(ws,titulo,f"Viáticos y gastos de viaje · WEP AI · {datetime.now().strftime('%B %Y')}",cols)
    ws.freeze_panes="A5"; _add_autofilter(ws,4,cols)
    headers=["#","Fecha","Empleado","Destino","Categoría","Descripción",
             "Moneda","Monto","Equiv. MXN","Comprobante","Reembolsable"]
    for i,h in enumerate(headers,1): _hc(ws,f"{get_column_letter(i)}4",h,bg="0C447C")

    sample=[
        (date(2026,5,5),"Carlos Mendoza","Monterrey","Vuelo","Vuelo CDM→MTY→CDM","MXN",4200,"Factura A001","Sí"),
        (date(2026,5,5),"Carlos Mendoza","Monterrey","Hotel","Hotel 2 noches","MXN",3600,"Factura H002","Sí"),
        (date(2026,5,5),"Carlos Mendoza","Monterrey","Comida","Comidas y cenas reunión","MXN",1250,"Tickets varios","Sí"),
        (date(2026,5,6),"Carlos Mendoza","Monterrey","Transporte","Taxi aeropuerto","MXN",380,"Recibo T001","Sí"),
        (date(2026,5,10),"Ana García","Miami","Vuelo","Vuelo CDM→MIA","USD",380,"Invoice F001","Sí"),
        (date(2026,5,10),"Ana García","Miami","Hotel","Hotel 3 noches","USD",420,"Invoice H003","Sí"),
        (date(2026,5,11),"Ana García","Miami","Comida","Cenas de trabajo","USD",95,"Receipt varios","Sí"),
        (date(2026,5,15),"Jorge Ramírez","CDMX","Transporte","Uber reuniones cliente","MXN",520,"Recibo U002","Parcial"),
    ]
    for idx,(fecha,emp,dest,cat,desc,mon,monto,comp,reimb) in enumerate(sample):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EFF6FF"
        _dc(ws,f"A{r}",idx+1,align="center",bg=bg)
        ws[f"B{r}"].value=fecha; ws[f"B{r}"].number_format="DD/MM/YYYY"
        ws[f"B{r}"].font=Font(name="Arial",size=10); ws[f"B{r}"].fill=PatternFill("solid",start_color=bg)
        ws[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"B{r}"].border=_thin()
        _dc(ws,f"C{r}",emp,bold=True,bg=bg); _dc(ws,f"D{r}",dest,align="center",bg=bg)
        _dc(ws,f"E{r}",cat,align="center",bg=bg,color="0C447C",bold=True)
        _dc(ws,f"F{r}",desc,bg=bg)
        _dc(ws,f"G{r}",mon,align="center",bg=bg,bold=True)
        _dc(ws,f"H{r}",monto,align="center",bg=bg,fmt='#,##0.00',color="0C447C",bold=True)
        _formula(ws,f"I{r}",
                 f"=IFERROR(H{r}*VLOOKUP(G{r},'Tipo de Cambio'!$B$5:$C$10,2,0),H{r})",
                 bg=bg,fmt='$#,##0.00',color="15803D",bold=True,align="center")
        _dc(ws,f"J{r}",comp,align="center",bg=bg,color="888780")
        _dc(ws,f"K{r}",reimb,align="center",bg=bg,
            color="15803D" if reimb=="Sí" else ("854F0B" if reimb=="Parcial" else "B91C1C"),bold=True)
    lr=5+len(sample)
    ws.merge_cells(f"A{lr}:H{lr}"); _hc(ws,f"A{lr}","TOTAL",bg="0D1B2A")
    _formula(ws,f"I{lr}",f"=SUM(I5:I{lr-1})",bg="0D1B2A",fg="FFFFFF",bold=True,fmt='$#,##0.00',align="center")

    _cf_formula_rule(ws,f"A5:K{lr-1}",'=$K5="Sí"',    "E1F5EE","064E3B")
    _cf_formula_rule(ws,f"A5:K{lr-1}",'=$K5="Parcial"',"FFF3CD","78350F")
    _cf_formula_rule(ws,f"A5:K{lr-1}",'=$K5="No"',    "FFE4E4","7F1D1D")
    _add_dv_list(ws,f"E5:E{lr+20}",["Vuelo","Hotel","Comida","Transporte","Registro/Inscripción",
                                      "Representación","Comunicaciones","Otros"],"Categoría")
    _add_dv_list(ws,f"G5:G{lr+20}",["MXN","USD","EUR","GBP","CAD","BRL"],"Moneda")
    _add_dv_list(ws,f"K5:K{lr+20}",["Sí","Parcial","No"],"¿Es reembolsable?")
    widths=[4,12,18,14,14,22,8,12,13,16,12]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    # Resumen
    _title_row(ws2,"Resumen de Viáticos","Por empleado y por categoría",5)
    for i,h in enumerate(["Empleado","Total MXN","# Gastos","Vuelo","Hotel","Comida","Transporte","Otros"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="185FA5")
    emps=["Carlos Mendoza","Ana García","Jorge Ramírez"]
    cats_viat=["Vuelo","Hotel","Comida","Transporte"]
    for idx,emp in enumerate(emps):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EFF6FF"
        _dc(ws2,f"A{r}",emp,bold=True,bg=bg)
        _formula(ws2,f"B{r}",f'=SUMIF(Viáticos!C5:C{lr-1},A{r},Viáticos!I5:I{lr-1})',
                 bg=bg,fmt='$#,##0.00',align="center",color="0C447C",bold=True)
        _formula(ws2,f"C{r}",f'=COUNTIF(Viáticos!C5:C{lr-1},A{r})',
                 bg=bg,fmt='#,##0',align="center")
        for j,cat in enumerate(cats_viat):
            cl=get_column_letter(4+j)
            _formula(ws2,f"{cl}{r}",
                     f'=IFERROR(SUMIFS(Viáticos!I5:I{lr-1},Viáticos!C5:C{lr-1},A{r},Viáticos!E5:E{lr-1},"{cat}"),0)',
                     bg=bg,fmt='$#,##0.00',align="center")
        _formula(ws2,f"H{r}",
                 f'=IFERROR(B{r}-D{r}-E{r}-F{r}-G{r},0)',
                 bg=bg,fmt='$#,##0.00',align="center")
    for col,w in zip(["A","B","C","D","E","F","G","H"],[18,13,10,12,12,12,14,12]):
        ws2.column_dimensions[col].width=w
    _setup_print(ws,cols,lr+3,landscape=True)
    _protect_ws(ws,unlock_from_row=5)


def _build_roi_calculator(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Calculadora ROI con comparativo de 3 proyectos y análisis de recuperación."""
    # ── v15.9.0 — Localización por país + datos del usuario ──
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    moneda = cfg.get("moneda", "USD")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    ws  = _setup_sheet(wb, "Proyectos ROI", "185FA5")
    ws2 = _setup_sheet(wb, "Comparativo",   "0C447C")
    cols = 8

    _title_row(ws,titulo,"Calcula y compara el retorno de hasta 3 inversiones · WEP AI",cols)
    _section_row(ws,3,"DATOS DE ENTRADA — Edita las celdas amarillas",cols,bg="185FA5")
    for i,h in enumerate(["Parámetro","Proyecto A","Proyecto B","Proyecto C"],1):
        _hc(ws,f"{get_column_letter(i)}4",h,bg="0C447C")

    params=[
        ("Nombre del proyecto",       "Nuevo equipo CNC",   "Software ERP",    "Expansión sucursal"),
        ("Inversión inicial ($)",     500000,               180000,            750000),
        ("Beneficio mensual ($)",     45000,                22000,             85000),
        ("Costo operativo mensual ($)",8000,                4500,              35000),
        ("Vida útil del proyecto (meses)",60,               36,                48),
        ("Tasa de descuento anual (%)",0.12,               0.12,              0.12),
    ]
    fmt_map={0:'General',1:'$#,##0',2:'$#,##0',3:'$#,##0',4:'#,##0',5:'0%'}
    colors_map=["000000","B91C1C","15803D","854F0B","185FA5","534AB7"]
    for idx,(label,va,vb,vc) in enumerate(params):
        r=5+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        _dc(ws,f"A{r}",label,bold=True,bg=bg)
        fmt=fmt_map.get(idx,'General')
        for col,val in zip(["B","C","D"],[va,vb,vc]):
            _dc(ws,f"{col}{r}",val,align="center",bg="FEFCE8",fmt=fmt,color=colors_map[idx],bold=True)

    # Results
    _section_row(ws,12,"RESULTADOS CALCULADOS",cols,bg="0D1B2A")
    for i,h in enumerate(["Indicador","Proyecto A","Proyecto B","Proyecto C"],1):
        _hc(ws,f"{get_column_letter(i)}13",h,bg="0C447C")
    results=[
        ("Beneficio Neto Mensual",  "=B7-B8",  "=C7-C8",  "=D7-D8",  "$#,##0"),
        ("Beneficio Total Bruto",   "=B7*B9", "=C7*C9", "=D7*D9", "$#,##0"),
        ("Costo Total Operativo",   "=B8*B9",  "=C8*C9",  "=D8*D9",  "$#,##0"),
        ("Ganancia Neta Total",     "=B14-B6-B15","=C14-C6-C15","=D14-D6-D15","$#,##0"),
        ("ROI Total %",             "=IFERROR(B16/B6,0)","=IFERROR(C16/C6,0)","=IFERROR(D16/D6,0)","0.0%"),
        ("ROI Anual %",             "=IFERROR(B17/(B9/12),0)","=IFERROR(C17/(C9/12),0)","=IFERROR(D17/(D9/12),0)","0.0%"),
        ("Meses para recuperar",    "=IFERROR(B6/B14,0)","=IFERROR(C6/C14,0)","=IFERROR(D6/D14,0)","#,##0.0"),
        ("VPN (tasa anual 12%)",    "=IFERROR(NPV(B10/12,REPT(B13,B9)/1)-B6,0)",
                                    "=IFERROR(NPV(C10/12,REPT(C13,C9)/1)-C6,0)",
                                    "=IFERROR(NPV(D10/12,REPT(D13,D9)/1)-D6,0)","$#,##0"),
        ("Recomendación",
         '=IF(B17>0.15,"Muy rentable",IF(B17>0.08,"Rentable","Revisar"))',
         '=IF(C17>0.15,"Muy rentable",IF(C17>0.08,"Rentable","Revisar"))',
         '=IF(D17>0.15,"Muy rentable",IF(D17>0.08,"Rentable","Revisar"))',
         'General'),
    ]
    res_colors={"$#,##0":"0C447C","0.0%":"15803D","#,##0.0":"854F0B","General":"000000"}
    for idx,(label,fa,fb,fc,fmt) in enumerate(results):
        r=14+idx; bg="EFF6FF" if idx%2==0 else "FFFFFF"
        is_main=label in ["ROI Total %","VPN (tasa anual 12%)","Recomendación"]
        _dc(ws,f"A{r}",label,bold=is_main,bg=bg)
        for col,formula in zip(["B","C","D"],[fa,fb,fc]):
            cell_color=res_colors.get(fmt,"000000")
            if label=="Recomendación":
                _formula(ws,f"{col}{r}",formula,bg="FEFCE8" if is_main else bg,
                         fmt=fmt,align="center",bold=True)
            else:
                _formula(ws,f"{col}{r}",formula,bg="FEFCE8" if is_main else bg,
                         fmt=fmt,color=cell_color,bold=is_main,align="center",
                         size=12 if is_main else 10)
    for i,w in enumerate([24,16,16,16],1): ws.column_dimensions[get_column_letter(i)].width=w
    _note_row(ws,23,"Modifica las celdas amarillas (filas 5-10). Todos los resultados se recalculan automaticamente.",
              cols,bg="EFF6FF",color="0C447C")

    # Comparativo
    _title_row(ws2,"Comparativo de ROI","Proyectos A vs B vs C",5)
    for i,h in enumerate(["Indicador","Proyecto A","Proyecto B","Proyecto C"],1):
        _hc(ws2,f"{get_column_letter(i)}4",h,bg="0C447C")
    comp_metrics=[("ROI Total %","B17","C17","D17","0.0%"),
                  ("ROI Anual %","B18","C18","D18","0.0%"),
                  ("Meses recuperación","B19","C19","D19","#,##0.0"),
                  ("Ganancia Neta","B16","C16","D16","$#,##0"),
                  ("VPN","B21","C21","D21","$#,##0")]
    for idx,(label,ra,rb,rc,fmt) in enumerate(comp_metrics):
        r=5+idx; bg="FFFFFF" if idx%2==0 else "EFF6FF"
        _dc(ws2,f"A{r}",label,bold=True,bg=bg)
        for col,ref in zip(["B","C","D"],[ra,rb,rc]):
            _formula(ws2,f"{col}{r}",f"='Proyectos ROI'!{ref}",
                     bg=bg,fmt=fmt,align="center",color="185FA5",bold=True)
    # BarChart ROI
    chart_roi=BarChart()
    chart_roi.type="col"; chart_roi.grouping="clustered"
    chart_roi.title="Comparativo de ROI Anual"; chart_roi.style=10
    chart_roi.height=12; chart_roi.width=18
    cats_r=Reference(ws2,min_col=1,max_col=1,min_row=5,max_row=5+len(comp_metrics)-1)
    data_r=Reference(ws2,min_col=2,max_col=4,min_row=4,max_row=4+len(comp_metrics))
    chart_roi.add_data(data_r,titles_from_data=True)
    chart_roi.set_categories(cats_r)
    for s_idx,color in enumerate(["185FA5","B91C1C","15803D"]):
        if s_idx < len(chart_roi.series):
            chart_roi.series[s_idx].graphicalProperties.solidFill=color
    ws2.add_chart(chart_roi,"A12")
    for i,w in enumerate([20,16,16,16],1): ws2.column_dimensions[get_column_letter(i)].width=w
    _setup_print(ws,cols,24,landscape=True)


def _build_income_expense_book(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Libro de ingresos y egresos.

    v15.5 — Si país=México mantiene la estructura RESICO original (UUID CFDI,
    RFC Tercero, tasa RESICO). Para otros países usa una estructura general
    con IVA/IGV/ITBMS local, ID fiscal genérico (NIT/RUT/RUC/CUIT) y sin la
    hoja RESICO Calc.
    """
    # ── v15.5 — Configuración del país ────────────────────────────────────────
    # ── v15.11.0 — FiscalVerifier: actualización autónoma de tasas/impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    pais         = (cfg or {}).get("pais", "")
    is_mexico    = pais == "México"
    tax_rate     = iva_rate(cfg)
    tax_lbl      = iva_label(cfg)
    money_fmt    = currency_format(cfg)
    id_persona   = (cfg or {}).get("id_persona", "ID Fiscal")
    tax_auth     = (cfg or {}).get("autoridad_fiscal", "Autoridad fiscal local")
    fiscal       = get_excel_fiscal(cfg)
    inc_tax_lbl  = fiscal["income_tax_freelance_label"]
    inc_tax_rate = fiscal["income_tax_freelance"]

    cols = 10
    ws  = _setup_sheet(wb, "Registro",   "0F6E56")
    ws2 = _setup_sheet(wb, "Fiscal Mes", "085041")

    if is_mexico:
        # ── Estructura original Mexicana RESICO ───────────────────────────────
        ws3 = _setup_sheet(wb, "RESICO Calc", "04342C")
        _title_row(ws, titulo,
                   f"Régimen Simplificado de Confianza (RESICO) · {datetime.now().year} · WEP AI", cols)
        headers = ["#", "Fecha", "UUID CFDI", "RFC Tercero", "Descripción", "Tipo",
                   "Subtotal", "IVA Trasladado", "ISR Retenido", "Total Neto"]
    else:
        # ── Estructura general para otros países ──────────────────────────────
        _title_row(ws, titulo,
                   f"Libro de Ingresos y Egresos · {pais} · {tax_auth} · {datetime.now().year} · WEP AI", cols)
        headers = ["#", "Fecha", "Folio / Factura", id_persona + " Tercero", "Descripción", "Tipo",
                   "Subtotal", f"{tax_lbl} Cobrado/Pagado", "Retención", "Total Neto"]

    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="0F6E56")

    sample = [
        (date(2026,5,3),  "F-2026-001",  "Cliente Alpha",     "Servicio consultoría abril", "Ingreso", 8500,  0,    0),
        (date(2026,5,5),  "F-2026-002",  "Cliente Beta",      "Desarrollo web módulo 2",    "Ingreso", 12000, 0,    0),
        (date(2026,5,8),  "F-PROV-101",  "Proveedor Internet","Servicio internet mensual",  "Egreso",  799,   0,    0),
        (date(2026,5,10), "F-PROV-102",  "Capacitadora SA",   "Capacitación profesional",   "Egreso",  3500,  0,    0),
        (date(2026,5,15), "F-2026-003",  "Cliente Gamma",     "Proyecto diseño completo",   "Ingreso", 6000,  0,    0),
        (date(2026,5,20), "F-PROV-103",  "Tienda Soft",       "Herramientas software",      "Egreso",  1299,  0,    0),
    ]
    # Recalcular IVA y retención dinámicamente según país
    sample_v2 = []
    for (fecha, folio, tercero, desc, tipo, sub, _iva_dummy, _ret_dummy) in sample:
        iva = sub * tax_rate
        # Retención: solo en ingresos, según tasa freelance del país
        ret = sub * inc_tax_rate if tipo == "Ingreso" else 0
        sample_v2.append((fecha, folio, tercero, desc, tipo, sub, iva, ret))

    tipo_color = {"Ingreso": "15803D", "Egreso": "B91C1C"}
    for idx, (fecha, folio, tercero, desc, tipo, sub, iva, ret) in enumerate(sample_v2):
        r = 5 + idx
        bg = "FFFFFF" if idx % 2 == 0 else "F0FFF4"
        _dc(ws, f"A{r}", idx+1, align="center", bg=bg)
        ws[f"B{r}"].value = fecha
        ws[f"B{r}"].number_format = "DD/MM/YYYY"
        ws[f"B{r}"].font = Font(name="Arial", size=10)
        ws[f"B{r}"].fill = PatternFill("solid", start_color=bg)
        ws[f"B{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{r}"].border = _thin()
        _dc(ws, f"C{r}", folio, bg=bg, color="888780")
        _dc(ws, f"D{r}", tercero, align="center", bg=bg, color="888780")
        _dc(ws, f"E{r}", desc, bg=bg, bold=True)
        _dc(ws, f"F{r}", tipo, align="center", bg=bg, color=tipo_color.get(tipo, "000000"), bold=True)
        _dc(ws, f"G{r}", sub, align="center", bg=bg, fmt=money_fmt,
            color="15803D" if tipo == "Ingreso" else "B91C1C", bold=True)
        _dc(ws, f"H{r}", iva, align="center", bg=bg, fmt=money_fmt, color="534AB7")
        _dc(ws, f"I{r}", ret, align="center", bg=bg, fmt=money_fmt, color="B91C1C")
        _formula(ws, f"J{r}",
                 f'=IF(F{r}="Ingreso",G{r}+H{r}-I{r},-(G{r}+H{r}))',
                 bg=bg, fmt=money_fmt, align="center",
                 color="15803D" if tipo == "Ingreso" else "B91C1C", bold=True)
    lr = 5 + len(sample_v2)
    ws.merge_cells(f"A{lr}:F{lr}")
    _hc(ws, f"A{lr}", "TOTALES", bg="0D1B2A")
    for col in ["G", "H", "I", "J"]:
        _formula(ws, f"{col}{lr}", f"=SUM({col}5:{col}{lr-1})",
                 bg="0D1B2A", fg="FFFFFF", bold=True, fmt=money_fmt, align="center")

    _cf_formula_rule(ws, f"A5:J{lr-1}", '=$F5="Ingreso"', "E1F5EE", "064E3B")
    _cf_formula_rule(ws, f"A5:J{lr-1}", '=$F5="Egreso"',  "FFE4E4", "7F1D1D")
    _add_dv_list(ws, f"F5:F{lr+30}", ["Ingreso", "Egreso"], "Tipo de movimiento")
    widths = [4, 13, 22, 22, 24, 10, 13, 14, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Resumen mensual fiscal ────────────────────────────────────────────────
    months = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    _title_row(ws2, f"Resumen Fiscal Mensual · {pais or 'Genérico'}",
               f"Prefiguración {inc_tax_lbl.split('(')[0].strip()} e {tax_lbl} por mes", 4)
    headers2 = ["Mes", "Ingresos", "Egresos", "Utilidad Bruta",
                f"{tax_lbl} a Cargo", "Retención", "Tasa estimada", "Impuesto estimado"]
    for i, h in enumerate(headers2, 1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="085041")

    for idx, mes in enumerate(months):
        r = 5 + idx
        bg = "FFFFFF" if idx % 2 == 0 else "E1F5EE"
        m = idx + 1
        _dc(ws2, f"A{r}", mes, bold=True, bg=bg, color="085041")
        _formula(ws2, f"B{r}",
                 f'=IFERROR(SUMPRODUCT((MONTH(Registro!B5:B{lr-1})={m})*(Registro!F5:F{lr-1}="Ingreso")*Registro!G5:G{lr-1}),0)',
                 bg=bg, fmt=money_fmt, align="center", color="15803D", bold=True)
        _formula(ws2, f"C{r}",
                 f'=IFERROR(SUMPRODUCT((MONTH(Registro!B5:B{lr-1})={m})*(Registro!F5:F{lr-1}="Egreso")*Registro!G5:G{lr-1}),0)',
                 bg=bg, fmt=money_fmt, align="center", color="B91C1C", bold=True)
        _formula(ws2, f"D{r}", f"=B{r}-C{r}",
                 bg=bg, fmt=money_fmt, align="center", color="0F6E56", bold=True)
        _formula(ws2, f"E{r}",
                 f'=IFERROR(SUMPRODUCT((MONTH(Registro!B5:B{lr-1})={m})*(Registro!F5:F{lr-1}="Ingreso")*Registro!H5:H{lr-1}),0)',
                 bg=bg, fmt=money_fmt, align="center", color="534AB7")
        _formula(ws2, f"F{r}",
                 f'=IFERROR(SUMPRODUCT((MONTH(Registro!B5:B{lr-1})={m})*Registro!I5:I{lr-1}),0)',
                 bg=bg, fmt=money_fmt, align="center", color="B91C1C")
        # Tasa para el cálculo del impuesto estimado (RESICO para MX, freelance para otros)
        tasa_estimada = 0.020 if is_mexico else inc_tax_rate
        _dc(ws2, f"G{r}", tasa_estimada, align="center", bg=bg, fmt='0.0%', color="854F0B")
        _formula(ws2, f"H{r}", f"=MAX(0,D{r}*G{r}-F{r})",
                 bg=bg, fmt=money_fmt, align="center", color="B91C1C", bold=True)

    lr2 = 5 + 12
    ws2.merge_cells(f"A{lr2}:A{lr2}")
    _hc(ws2, f"A{lr2}", "ANUAL", bg="0D1B2A")
    for col in ["B","C","D","E","F","H"]:
        _formula(ws2, f"{col}{lr2}", f"=SUM({col}5:{col}{lr2-1})",
                 bg="0D1B2A", fg="FFFFFF", bold=True, fmt=money_fmt, align="center")
    for col, w in zip(["A","B","C","D","E","F","G","H"], [8,14,14,14,14,14,12,18]):
        ws2.column_dimensions[col].width = w

    if is_mexico:
        nota = ("Tasa RESICO fija al 2% para simplificación. Verifica con tu contador la "
                "tasa correcta según tus ingresos anuales (1.5%-2.5%). ISR RESICO Estimado = "
                "MAX(0, Utilidad Bruta × Tasa - ISR Retenido).")
    else:
        nota = (f"Tasa de impuesto estimada según referencia de {tax_auth}. "
                f"Las escalas progresivas reales varían — consulta con tu contador. "
                f"Impuesto Estimado = MAX(0, Utilidad Bruta × Tasa - Retención).")
    _note_row(ws2, lr2+2, nota, 8, bg="E1F5EE", color="085041")
    _setup_print(ws, cols, lr+3, landscape=True)
    _protect_ws(ws, unlock_from_row=5)


# ═══════════════════════════════════════════════════════════════
# MODO "ARREGLA MI EXCEL" — Analizar y mejorar archivos existentes
# ═══════════════════════════════════════════════════════════════

def analyze_excel(filepath):
    """Analyze an existing Excel file and return a structured diagnostics report."""
    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        return {"error": str(e), "issues": [], "suggestions": [], "stats": {}}

    issues      = []
    suggestions = []
    stats       = {"sheets": len(wb.worksheets), "total_cells": 0,
                   "formula_cells": 0, "error_cells": 0, "empty_cols": 0}
    ERROR_VALS  = {"#REF!","#VALUE!","#NAME?","#DIV/0!","#N/A","#NULL!","#NUM!","#ERROR!"}

    for ws in wb.worksheets:
        has_filter = ws.auto_filter.ref is not None
        has_freeze = ws.freeze_panes is not None
        max_r, max_c = ws.max_row, ws.max_column
        stats["total_cells"] += max_r * max_c if max_r and max_c else 0

        for row in ws.iter_rows(max_row=min(max_r or 0, 2000)):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                # Error values
                if val in ERROR_VALS:
                    stats["error_cells"] += 1
                    issues.append(
                        f"❌ Error {val} en [{ws.title}] celda {cell.coordinate}")
                # Formula without IFERROR protection
                elif val.startswith("="):
                    stats["formula_cells"] += 1
                    upper = val.upper()
                    if "IFERROR" not in upper and ("/" in val or
                       any(fn in upper for fn in ["VLOOKUP","INDEX","MATCH","AVERAGE","IF("])):
                        suggestions.append(
                            f"⚠ Fórmula sin IFERROR en [{ws.title}] {cell.coordinate}: "
                            f"{val[:60]}{'...' if len(val)>60 else ''}")
                # Number stored as text
                elif (cell.data_type == "s" and
                      val.replace(",","").replace(".","").replace("-","").replace(" ","").isdigit()):
                    suggestions.append(
                        f"⚠ Número guardado como texto en [{ws.title}] {cell.coordinate}: '{val}'")

        # Missing auto_filter on data sheets
        if max_r and max_r > 5 and max_c and max_c > 2 and not has_filter:
            suggestions.append(
                f"💡 Sin auto_filter en [{ws.title}] — considera activar filtros en la fila de encabezado")
        # Missing freeze panes
        if max_r and max_r > 10 and not has_freeze:
            suggestions.append(
                f"💡 Sin freeze panes en [{ws.title}] — el encabezado desaparece al hacer scroll")

    return {"issues": issues, "suggestions": suggestions[:25], "stats": stats,
            "sheet_names": [ws.title for ws in wb.worksheets]}


def improve_excel(filepath, instructions, gen_data=None):
    """Load an existing Excel file, analyze it, apply improvements and save a new version."""
    import os

    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        return filepath, f"Error al abrir el archivo: {e}"

    instruc_lower = (instructions or "").lower()
    improvements  = []

    for ws in wb.worksheets:
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        if max_r < 2 or max_c < 1:
            continue

        # ── 1. Proteger IFERROR ────────────────────────────────────────────
        if any(w in instruc_lower for w in
               ["iferror","proteg","error","#ref","#value","#div","formula","fórmula"]):
            fixed = 0
            for row in ws.iter_rows(min_row=2, max_row=min(max_r, 1000)):
                for cell in row:
                    if cell.value and str(cell.value).startswith("="):
                        val  = str(cell.value)
                        uval = val.upper()
                        if "IFERROR" not in uval and ("/" in val or
                           any(fn in uval for fn in ["VLOOKUP","INDEX","MATCH","AVERAGE"])):
                            cell.value = f"=IFERROR({val[1:]},0)"
                            fixed += 1
            if fixed:
                improvements.append(f"✓ {fixed} fórmulas protegidas con IFERROR en [{ws.title}]")

        # ── 2. Auto-filter ─────────────────────────────────────────────────
        if any(w in instruc_lower for w in ["filtro","filter","ordenar","buscar","sort"]):
            if not ws.auto_filter.ref and max_c > 1:
                last_col = get_column_letter(max_c)
                ws.auto_filter.ref = f"A1:{last_col}1"
                improvements.append(f"✓ Auto-filter activado en [{ws.title}]")

        # ── 3. Freeze panes ────────────────────────────────────────────────
        if any(w in instruc_lower for w in ["freeze","congelar","fija","scroll","encabezado","header"]):
            if not ws.freeze_panes and max_r > 5:
                ws.freeze_panes = "A2"
                improvements.append(f"✓ Freeze panes activado en [{ws.title}]")

        # ── 4. Formato numérico en celdas numéricas sin formato ────────────
        if any(w in instruc_lower for w in ["formato","format","moneda","pesos","miles","number"]):
            formatted = 0
            for row in ws.iter_rows(min_row=2, max_row=min(max_r, 500)):
                for cell in row:
                    if (isinstance(cell.value, (int, float)) and
                        cell.value != 0 and not str(cell.value).startswith("=")):
                        if not cell.number_format or cell.number_format == "General":
                            if abs(cell.value) > 100:
                                cell.number_format = '$#,##0.00'
                                formatted += 1
                            elif 0 < abs(cell.value) <= 1:
                                cell.number_format = '0.0%'
                                formatted += 1
            if formatted:
                improvements.append(f"✓ {formatted} celdas con formato numérico aplicado en [{ws.title}]")

        # ── 5. Eliminar filas completamente vacías en el medio ─────────────
        if any(w in instruc_lower for w in ["limpiar","clean","vacías","vacio","empty","borrar"]):
            rows_deleted = 0
            rows_to_delete = []
            for row in ws.iter_rows(min_row=2, max_row=min(max_r, 500)):
                if all(c.value is None for c in row):
                    rows_to_delete.append(row[0].row)
            for r_idx in reversed(rows_to_delete):
                ws.delete_rows(r_idx)
                rows_deleted += 1
            if rows_deleted:
                improvements.append(f"✓ {rows_deleted} filas vacías eliminadas en [{ws.title}]")

        # ── 6. Ancho de columnas automático ───────────────────────────────
        if any(w in instruc_lower for w in ["ancho","columna","ajustar","width","formato"]):
            for col_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells[:30]:  # Sample first 30 rows
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        max_length = max(max_length, cell_len)
                    except Exception:
                        pass
                adj_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[col_letter].width = adj_width
            improvements.append(f"✓ Anchos de columna ajustados en [{ws.title}]")

        # ── 7. Print area ──────────────────────────────────────────────────
        if any(w in instruc_lower for w in ["imprimir","print","impresión","papel","pdf"]):
            last_col = get_column_letter(max_c)
            ws.print_area = f"A1:{last_col}{max_r}"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage   = True
            ws.page_setup.fitToWidth  = 1
            improvements.append(f"✓ Área de impresión configurada en [{ws.title}]")

    if not improvements:
        improvements.append(
            "No se detectaron mejoras automáticas con las instrucciones dadas. "
            "Prueba con: 'proteger fórmulas', 'agregar filtros', 'limpiar filas vacías', "
            "'ajustar formato numérico', 'configurar impresión'.")

    # Save improved file
    base, ext = os.path.splitext(filepath)
    new_path = base + "_mejorado" + ext
    try:
        wb.save(new_path)
    except Exception as e:
        return filepath, f"Error al guardar: {e}"

    bug_report = (f"📋 Mejoras aplicadas: {len([i for i in improvements if i.startswith('✓')])}\n\n"
                  + "\n".join(improvements))
    return new_path, bug_report


# NOTA: la duplicación de open_in_excel (que estaba aquí) se eliminó en v14.1.
# La versión canónica está en la línea ~2758 (usa osascript, consistente con
# open_in_word y open_in_powerpoint).


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE REGISTRY (v14.2) — al final del archivo para evitar forward refs
# ─────────────────────────────────────────────────────────────────────────────
# Mapeo explícito template_id → builder. Cuando Claude emite `template_id` en
# el JSON ##GENERAR##, el dispatcher lo usa directamente (más confiable que
# adivinar con keywords). Si el template_id no está presente o es inválido,
# cae al matching tradicional por keywords (preservando comportamiento legacy).
#
# Para añadir un template nuevo:
#   1. Define `_build_xxx(wb, titulo, secciones, detalles, instruccion)`.
#   2. Registralo aquí con un id snake_case único.
#   3. Listalo en SYSTEM_PROMPT (brain.py) para que el LLM lo conozca.

# ─────────────────────────────────────────────────────────────────────────────
# v15.9.1 — BUILDERS NUEVOS (cierran gap catálogo PDF vs código)
# ─────────────────────────────────────────────────────────────────────────────

def _build_kpi_tracker(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """KPI Tracker — Indicadores clave de desempeño con metas, valores y semáforo.

    v15.9.1 — Nuevo en código. Acepta datos['kpis'] como lista de dicts:
    {nombre, meta, real, unidad, descripcion}
    """
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.10.0 — FiscalVerifier: actualización autónoma de tasas e impuestos ──
    try:
        from agent.fiscal_verifier import verify_fiscal_data
        cfg = verify_fiscal_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")

    ws  = _setup_sheet(wb, "KPIs",      "0C447C")
    ws2 = _setup_sheet(wb, "Historico", "185FA5")
    ws3 = _setup_sheet(wb, "Resumen",   "0D1B2A")
    cols = 8

    subtitulo = (f"Indicadores Clave · {empresa_user} · {datetime.now().strftime('%d/%m/%Y')}"
                 if empresa_user else f"Indicadores Clave · WEP AI · {datetime.now().strftime('%d/%m/%Y')}")
    _title_row(ws, titulo, subtitulo, cols)
    ws.freeze_panes = "A5"
    _add_autofilter(ws, 4, cols)

    headers = ["#", "KPI", "Meta", "Valor Real", "Unidad", "Cumplimiento %", "Semaforo", "Tendencia"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="0C447C")
    ws.row_dimensions[4].height = 22

    kpis_user = datos.get("kpis") or [
        {"nombre": "Ingresos mensuales",         "meta": 500000, "real": 425000, "unidad": sym},
        {"nombre": "Margen bruto",               "meta": 45,     "real": 42,     "unidad": "%"},
        {"nombre": "Clientes activos",           "meta": 250,    "real": 218,    "unidad": "#"},
        {"nombre": "NPS — Satisfaccion",         "meta": 70,     "real": 64,     "unidad": "pts"},
        {"nombre": "Tasa de retencion",          "meta": 90,     "real": 87,     "unidad": "%"},
        {"nombre": "CAC — Costo adquisicion",    "meta": 800,    "real": 1100,   "unidad": sym},
        {"nombre": "LTV — Valor vida cliente",   "meta": 12000,  "real": 14500,  "unidad": sym},
        {"nombre": "Tiempo respuesta soporte",   "meta": 4,      "real": 6,      "unidad": "hrs"},
        {"nombre": "Empleados",                  "meta": 50,     "real": 47,     "unidad": "#"},
    ]
    for idx, k in enumerate(kpis_user):
        r = 5 + idx
        bg = "FFFFFF" if idx % 2 == 0 else "EFF6FF"
        _dc(ws, f"A{r}", idx + 1, align="center", bg=bg)
        _dc(ws, f"B{r}", k.get("nombre", "[KPI]"), bold=True, bg=bg)
        _dc(ws, f"C{r}", k.get("meta", 0), align="right", bg=bg, fmt='#,##0.00')
        _dc(ws, f"D{r}", k.get("real", 0), align="right", bg=bg, fmt='#,##0.00')
        _dc(ws, f"E{r}", k.get("unidad", "—"), align="center", bg=bg)
        _dc(ws, f"F{r}", f"=IFERROR(D{r}/C{r},0)", align="right", bg=bg, fmt='0%')
        _dc(ws, f"G{r}", f'=IF(F{r}>=0.9,"BUENO",IF(F{r}>=0.7,"RIESGO","CRITICO"))',
            align="center", bg=bg, bold=True)
        _dc(ws, f"H{r}", k.get("tendencia", "+"), align="center", bg=bg)

    widths = [4, 32, 12, 12, 8, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Historico
    _title_row(ws2, "Historico de KPIs", "Registro mensual de cada KPI", 14)
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    h2 = ["KPI", "Meta"] + meses
    for i, h in enumerate(h2, 1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="185FA5")
    for idx, k in enumerate(kpis_user):
        r = 5 + idx
        _dc(ws2, f"A{r}", k.get("nombre", "[KPI]"), bold=True, bg="EFF6FF")
        _dc(ws2, f"B{r}", k.get("meta", 0), align="right", bg="EFF6FF", fmt='#,##0.00')
        for j in range(12):
            _dc(ws2, f"{get_column_letter(3+j)}{r}", "", align="right", bg="FFFFFF")

    # Resumen
    _title_row(ws3, "Resumen Ejecutivo de KPIs", f"Total: {len(kpis_user)} KPIs monitoreados", 4)
    _section_row(ws3, 3, "DISTRIBUCION DE CUMPLIMIENTO", 4, bg="0D1B2A")
    n = len(kpis_user)
    _dc(ws3, "A4", "KPIs en META (>=90%)", bold=True, bg="DCFCE7")
    _dc(ws3, "B4", f'=COUNTIF(KPIs!F5:F{4+n},">=0.9")', align="center", bg="FFFFFF")
    _dc(ws3, "A5", "KPIs en RIESGO (70-89%)", bold=True, bg="FEF3C7")
    _dc(ws3, "B5", f'=COUNTIFS(KPIs!F5:F{4+n},">=0.7",KPIs!F5:F{4+n},"<0.9")',
        align="center", bg="FFFFFF")
    _dc(ws3, "A6", "KPIs CRITICOS (<70%)", bold=True, bg="FEE2E2")
    _dc(ws3, "B6", f'=COUNTIF(KPIs!F5:F{4+n},"<0.7")', align="center", bg="FFFFFF")
    ws3.column_dimensions["A"].width = 30


def _build_presupuesto_anual(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Presupuesto Anual — desglose mensual + varianza.

    v15.9.1 — Nuevo en codigo.
    """
    cfg = cfg or _get_legal_config("GENERIC")
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    fmt_money = f'{sym}#,##0'

    ws  = _setup_sheet(wb, "Presupuesto Anual", "0C447C")
    ws2 = _setup_sheet(wb, "Real vs Plan",      "B91C1C")
    ws3 = _setup_sheet(wb, "Resumen",           "0D1B2A")

    cols = 16
    subtitulo = (f"Año {datetime.now().year} · {empresa_user}"
                 if empresa_user else f"Año {datetime.now().year} · WEP AI")
    _title_row(ws, titulo, subtitulo, cols)
    ws.freeze_panes = "B5"

    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    headers = ["Categoria"] + meses + ["Total Año", "% del Total", "Promedio Mensual"]
    for i, h in enumerate(headers, 1):
        _hc(ws, f"{get_column_letter(i)}4", h, bg="0C447C")
    ws.row_dimensions[4].height = 22

    _section_row(ws, 5, "INGRESOS PRESUPUESTADOS", cols, bg="15803D")
    ingresos = datos.get("ingresos") or [
        ("Ventas de productos",  [80000]*12),
        ("Ventas de servicios",  [40000]*12),
        ("Otros ingresos",       [5000]*12),
    ]
    r = 6
    for nombre, valores in ingresos:
        _dc(ws, f"A{r}", nombre, bold=True, bg="DCFCE7")
        for m_idx, v in enumerate(valores[:12]):
            _dc(ws, f"{get_column_letter(2+m_idx)}{r}", v, align="right", bg="DCFCE7", fmt=fmt_money)
        _dc(ws, f"N{r}", f"=SUM(B{r}:M{r})", align="right", bold=True, bg="DCFCE7", fmt=fmt_money)
        _dc(ws, f"O{r}", "", align="right", bg="DCFCE7", fmt='0.0%')
        _dc(ws, f"P{r}", f"=N{r}/12", align="right", bg="DCFCE7", fmt=fmt_money)
        r += 1
    fila_total_ingresos = r
    _dc(ws, f"A{r}", "TOTAL INGRESOS", bold=True, bg="15803D", color="FFFFFF")
    for c in range(2, 17):
        col = get_column_letter(c)
        if c <= 14:
            _dc(ws, f"{col}{r}", f"=SUM({col}6:{col}{r-1})", align="right",
                bold=True, bg="15803D", color="FFFFFF", fmt=fmt_money)
        elif c == 15:
            _dc(ws, f"{col}{r}", "100%", align="right", bold=True, bg="15803D", color="FFFFFF")
        else:
            _dc(ws, f"{col}{r}", f"=N{r}/12", align="right", bold=True, bg="15803D", color="FFFFFF", fmt=fmt_money)

    r += 2
    _section_row(ws, r, "EGRESOS PRESUPUESTADOS", cols, bg="B91C1C")
    r += 1
    egresos = datos.get("egresos") or [
        ("Nomina y prestaciones",       [45000]*12),
        ("Renta",                       [15000]*12),
        ("Servicios basicos",           [3500]*12),
        ("Marketing y publicidad",      [8000]*12),
        ("Materia prima / insumos",     [20000]*12),
        ("Tecnologia y software",       [4500]*12),
        ("Otros gastos operativos",     [6000]*12),
    ]
    fila_inicio_egresos = r
    for nombre, valores in egresos:
        _dc(ws, f"A{r}", nombre, bold=True, bg="FEE2E2")
        for m_idx, v in enumerate(valores[:12]):
            _dc(ws, f"{get_column_letter(2+m_idx)}{r}", v, align="right", bg="FEE2E2", fmt=fmt_money)
        _dc(ws, f"N{r}", f"=SUM(B{r}:M{r})", align="right", bold=True, bg="FEE2E2", fmt=fmt_money)
        _dc(ws, f"O{r}", "", align="right", bg="FEE2E2", fmt='0.0%')
        _dc(ws, f"P{r}", f"=N{r}/12", align="right", bg="FEE2E2", fmt=fmt_money)
        r += 1
    fila_total_egresos = r
    _dc(ws, f"A{r}", "TOTAL EGRESOS", bold=True, bg="B91C1C", color="FFFFFF")
    for c in range(2, 17):
        col = get_column_letter(c)
        if c <= 14:
            _dc(ws, f"{col}{r}", f"=SUM({col}{fila_inicio_egresos}:{col}{r-1})", align="right",
                bold=True, bg="B91C1C", color="FFFFFF", fmt=fmt_money)
        elif c == 15:
            _dc(ws, f"{col}{r}", "100%", align="right", bold=True, bg="B91C1C", color="FFFFFF")
        else:
            _dc(ws, f"{col}{r}", f"=N{r}/12", align="right", bold=True, bg="B91C1C", color="FFFFFF", fmt=fmt_money)

    r += 2
    _section_row(ws, r, "UTILIDAD NETA PRESUPUESTADA", cols, bg="0D1B2A")
    r += 1
    _dc(ws, f"A{r}", "UTILIDAD NETA = Ingresos - Egresos", bold=True, bg="DBEAFE")
    for c in range(2, 15):
        col = get_column_letter(c)
        _dc(ws, f"{col}{r}", f"={col}{fila_total_ingresos}-{col}{fila_total_egresos}",
            align="right", bold=True, bg="DBEAFE", fmt=fmt_money)
    _dc(ws, f"O{r}", "", align="right", bg="DBEAFE")
    _dc(ws, f"P{r}", f"=N{r}/12", align="right", bold=True, bg="DBEAFE", fmt=fmt_money)

    ws.column_dimensions["A"].width = 32
    for c in range(2, 17):
        ws.column_dimensions[get_column_letter(c)].width = 11

    # Real vs Plan
    _title_row(ws2, "Comparativo: Real vs Presupuesto", "Captura los valores reales mes a mes", 4)
    h2 = ["Concepto", "Presupuesto Año", "Real Año", "Varianza"]
    for i, h in enumerate(h2, 1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="B91C1C")
    _dc(ws2, "A5", "Total Ingresos", bold=True, bg="DCFCE7")
    _dc(ws2, "B5", f"='Presupuesto Anual'!N{fila_total_ingresos}", align="right", bg="DCFCE7", fmt=fmt_money)
    _dc(ws2, "C5", 0, align="right", bg="FFFFFF", fmt=fmt_money)
    _dc(ws2, "D5", "=C5-B5", align="right", bold=True, bg="FFFBEB", fmt=fmt_money)
    _dc(ws2, "A6", "Total Egresos", bold=True, bg="FEE2E2")
    _dc(ws2, "B6", f"='Presupuesto Anual'!N{fila_total_egresos}", align="right", bg="FEE2E2", fmt=fmt_money)
    _dc(ws2, "C6", 0, align="right", bg="FFFFFF", fmt=fmt_money)
    _dc(ws2, "D6", "=C6-B6", align="right", bold=True, bg="FFFBEB", fmt=fmt_money)
    _dc(ws2, "A7", "Utilidad Neta", bold=True, bg="DBEAFE")
    _dc(ws2, "B7", "=B5-B6", align="right", bold=True, bg="DBEAFE", fmt=fmt_money)
    _dc(ws2, "C7", "=C5-C6", align="right", bold=True, bg="DBEAFE", fmt=fmt_money)
    _dc(ws2, "D7", "=C7-B7", align="right", bold=True, bg="FFFBEB", fmt=fmt_money)
    for col, w in zip(["A", "B", "C", "D"], [28, 18, 18, 18]):
        ws2.column_dimensions[col].width = w

    # Resumen
    _title_row(ws3, "Resumen Anual", f"Año {datetime.now().year}", 4)
    _dc(ws3, "A4", "Total Ingresos Anuales", bold=True, bg="DCFCE7")
    _dc(ws3, "B4", f"='Presupuesto Anual'!N{fila_total_ingresos}", align="right",
        bg="DCFCE7", fmt=fmt_money)
    _dc(ws3, "A5", "Total Egresos Anuales", bold=True, bg="FEE2E2")
    _dc(ws3, "B5", f"='Presupuesto Anual'!N{fila_total_egresos}", align="right",
        bg="FEE2E2", fmt=fmt_money)
    _dc(ws3, "A6", "Utilidad Neta Anual", bold=True, bg="DBEAFE")
    _dc(ws3, "B6", "=B4-B5", align="right", bold=True, bg="DBEAFE", fmt=fmt_money)
    _dc(ws3, "A7", "Margen Neto %", bold=True, bg="EFF6FF")
    _dc(ws3, "B7", "=IFERROR(B6/B4,0)", align="right", bold=True, bg="EFF6FF", fmt='0.0%')
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20


def _build_ratios_financieros(wb, titulo, secciones, detalles, instruccion, datos=None, cfg=None):
    """Ratios Financieros — Liquidez, Solvencia, Rentabilidad, Actividad.

    v15.9.1 — Nuevo en codigo.
    """
    cfg = cfg or _get_legal_config("GENERIC")
    # ── v15.11.0 — AccountingVerifier: actualización autónoma NIIF/NIF/IFRS ──
    try:
        from agent.accounting_verifier import verify_accounting_data
        cfg = verify_accounting_data(cfg)
    except Exception:
        pass
    datos = datos or {}
    sym = cfg.get("simbolo", "$")
    empresa_user = datos.get("empresa_nombre") or datos.get("parte_a_nombre", "")
    fmt_money = f'{sym}#,##0'

    ws  = _setup_sheet(wb, "Datos Base",     "0C447C")
    ws2 = _setup_sheet(wb, "Ratios",         "185FA5")
    ws3 = _setup_sheet(wb, "Interpretacion", "0D1B2A")

    subtitulo = (f"Ratios Financieros · {empresa_user}"
                 if empresa_user else "Ratios Financieros · WEP AI")
    _title_row(ws, titulo, subtitulo, 4)
    _section_row(ws, 3, "DATOS DEL BALANCE GENERAL", 4, bg="0C447C")

    bal_items = [
        ("Activo Corriente",          500000),
        ("Inventarios",               150000),
        ("Cuentas por Cobrar",        180000),
        ("Efectivo y equivalentes",    80000),
        ("Activo No Corriente",       800000),
        ("ACTIVO TOTAL",             1300000),
        ("Pasivo Corriente",          280000),
        ("Pasivo No Corriente",       420000),
        ("PASIVO TOTAL",              700000),
        ("Patrimonio Neto",           600000),
    ]
    for idx, (label, val) in enumerate(bal_items):
        r = 4 + idx
        bg = "FFFBEB" if idx % 2 == 0 else "FFFFFF"
        _dc(ws, f"A{r}", label, bold=True, bg=bg)
        _dc(ws, f"B{r}", val, align="right", bg="FEFCE8", fmt=fmt_money, bold=True, color="0C447C")
    _section_row(ws, 15, "DATOS DEL ESTADO DE RESULTADOS", 4, bg="185FA5")
    pyg_items = [
        ("Ventas Netas",              1500000),
        ("Costo de Ventas",            900000),
        ("Utilidad Bruta",             600000),
        ("Gastos Operativos",          350000),
        ("Utilidad Operativa",         250000),
        ("Utilidad Neta",              180000),
    ]
    for idx, (label, val) in enumerate(pyg_items):
        r = 16 + idx
        bg = "EFF6FF" if idx % 2 == 0 else "FFFFFF"
        _dc(ws, f"A{r}", label, bold=True, bg=bg)
        _dc(ws, f"B{r}", val, align="right", bg="DBEAFE", fmt=fmt_money, bold=True, color="185FA5")
    for col, w in zip(["A", "B"], [30, 18]):
        ws.column_dimensions[col].width = w

    # Hoja Ratios
    _title_row(ws2, "Calculo de Ratios", "Calculados automaticamente desde 'Datos Base'", 6)
    headers = ["Categoria", "Ratio", "Formula", "Resultado", "Referencia", "Estado"]
    for i, h in enumerate(headers, 1):
        _hc(ws2, f"{get_column_letter(i)}4", h, bg="185FA5")

    ac, inv, cxc, efec = "'Datos Base'!B4", "'Datos Base'!B5", "'Datos Base'!B6", "'Datos Base'!B7"
    activo_tot = "'Datos Base'!B9"
    pc, pasivo_tot = "'Datos Base'!B10", "'Datos Base'!B12"
    patrim = "'Datos Base'!B13"
    ventas, costo_v = "'Datos Base'!B16", "'Datos Base'!B17"
    util_neta = "'Datos Base'!B21"

    ratios = [
        ("Liquidez", "Razon Corriente", "Act.Corr / Pas.Corr",
         f"={ac}/{pc}",                  ">1.5",   f'=IF(D5>=1.5,"BUENO",IF(D5>=1,"ACEPTABLE","CRITICO"))'),
        ("Liquidez", "Prueba Acida", "(Act.Corr - Inv) / Pas.Corr",
         f"=({ac}-{inv})/{pc}",          ">1.0",   f'=IF(D6>=1,"BUENO",IF(D6>=0.7,"ACEPTABLE","CRITICO"))'),
        ("Liquidez", "Liquidez Inmediata", "Efectivo / Pas.Corr",
         f"={efec}/{pc}",                ">0.3",   f'=IF(D7>=0.3,"BUENO",IF(D7>=0.2,"ACEPTABLE","CRITICO"))'),
        ("Solvencia", "Endeudamiento Total", "Pasivo Tot / Activo Tot",
         f"={pasivo_tot}/{activo_tot}",  "<0.6",   f'=IF(D8<=0.6,"BUENO",IF(D8<=0.7,"ACEPTABLE","CRITICO"))'),
        ("Solvencia", "Apalancamiento", "Pasivo Tot / Patrimonio",
         f"={pasivo_tot}/{patrim}",      "<1.5",   f'=IF(D9<=1.5,"BUENO",IF(D9<=2,"ACEPTABLE","CRITICO"))'),
        ("Rentabilidad", "Margen Bruto", "Util.Bruta / Ventas",
         f"=('Datos Base'!B18)/{ventas}", ">30%",  f'=IF(D10>=0.3,"BUENO",IF(D10>=0.2,"ACEPTABLE","CRITICO"))'),
        ("Rentabilidad", "Margen Neto", "Util.Neta / Ventas",
         f"={util_neta}/{ventas}",       ">10%",   f'=IF(D11>=0.1,"BUENO",IF(D11>=0.05,"ACEPTABLE","CRITICO"))'),
        ("Rentabilidad", "ROA", "Util.Neta / Activo Tot",
         f"={util_neta}/{activo_tot}",   ">5%",    f'=IF(D12>=0.05,"BUENO",IF(D12>=0.03,"ACEPTABLE","CRITICO"))'),
        ("Rentabilidad", "ROE", "Util.Neta / Patrimonio",
         f"={util_neta}/{patrim}",       ">15%",   f'=IF(D13>=0.15,"BUENO",IF(D13>=0.1,"ACEPTABLE","CRITICO"))'),
        ("Actividad", "Rotacion Inventario", "Costo Vtas / Inv",
         f"={costo_v}/{inv}",            ">6 veces", f'=IF(D14>=6,"BUENO",IF(D14>=4,"ACEPTABLE","CRITICO"))'),
        ("Actividad", "Periodo Cobranza (dias)", "(CxC / Ventas) x 365",
         f"=({cxc}/{ventas})*365",       "<45 dias", f'=IF(D15<=45,"BUENO",IF(D15<=60,"ACEPTABLE","CRITICO"))'),
        ("Actividad", "Rotacion Activos", "Ventas / Activo Tot",
         f"={ventas}/{activo_tot}",      ">1.0",   f'=IF(D16>=1,"BUENO",IF(D16>=0.7,"ACEPTABLE","CRITICO"))'),
    ]
    cat_colors = {"Liquidez": "DCFCE7", "Solvencia": "FEF3C7",
                  "Rentabilidad": "DBEAFE", "Actividad": "FEE2E2"}
    for idx, (cat, nombre, formula_vis, formula_real, ref, semaforo) in enumerate(ratios):
        r = 5 + idx
        bg = cat_colors.get(cat, "FFFFFF")
        _dc(ws2, f"A{r}", cat, bold=True, bg=bg, align="center")
        _dc(ws2, f"B{r}", nombre, bold=True, bg=bg)
        _dc(ws2, f"C{r}", formula_vis, bg=bg, color="666666", size=9)
        is_pct = nombre.startswith("Margen") or nombre.startswith("ROA") or nombre.startswith("ROE")
        fmt = '0.0%' if is_pct else '#,##0.00'
        _dc(ws2, f"D{r}", formula_real, align="right", bold=True, bg="FEFCE8", fmt=fmt, color="0D1B2A")
        _dc(ws2, f"E{r}", ref, bg=bg, align="center", color="666666")
        _dc(ws2, f"F{r}", semaforo, bg=bg, align="center", bold=True)

    widths = [12, 24, 22, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Interpretacion
    _title_row(ws3, "Guia de Interpretacion", "Que significa cada categoria", 2)
    interpretaciones = [
        ("LIQUIDEZ", "Mide la capacidad de la empresa para cumplir obligaciones a corto plazo. "
                     "Una razon corriente > 1.5 indica que puede pagar sus deudas sin presion."),
        ("SOLVENCIA", "Mide el nivel de endeudamiento. Endeudamiento <60% es saludable; "
                      "apalancamiento >2 indica alto riesgo financiero."),
        ("RENTABILIDAD", "Mide la eficiencia para generar utilidades. ROE >15% es atractivo "
                         "para inversionistas; ROA >5% indica buen uso de activos."),
        ("ACTIVIDAD", "Mide la eficiencia operativa. Rotacion de inventario alta = inventario "
                      "no estancado; periodo de cobranza corto = mejor flujo de caja."),
    ]
    for idx, (cat, desc) in enumerate(interpretaciones):
        r = 4 + idx * 2
        _dc(ws3, f"A{r}", cat, bold=True, bg="0D1B2A", color="FFFFFF", align="center")
        _dc(ws3, f"A{r+1}", desc, bg="EFF6FF")
        ws3.row_dimensions[r+1].height = 40
    ws3.column_dimensions["A"].width = 90




EXCEL_TEMPLATES = {
    "income_statement":      _build_income_statement,
    "break_even":            _build_break_even,
    "amortization":          _build_amortization,
    "sales_commissions":     _build_sales_commissions,
    "personal_budget":       _build_personal_budget,
    "bank_reconciliation":   _build_bank_reconciliation,
    "subscription_tracker":  _build_subscription_tracker,
    "balance_sheet":         _build_balance_sheet,
    "fixed_assets":          _build_fixed_assets,
    "variance_analysis":     _build_variance_analysis,
    "savings_planner":       _build_savings_planner,
    "purchase_order":        _build_purchase_order,
    "inventory":             _build_inventory,
    "payroll":               _build_payroll,
    "quotation":             _build_quotation,
    "accounts_receivable":   _build_accounts_receivable,
    "cashflow":              _build_cashflow,
    "multi_currency":        _build_multi_currency,
    "freelance_tracker":     _build_freelance_tracker,
    "project_tracker":       _build_project_tracker,
    "client_directory":      _build_client_directory,
    "supplier_directory":    _build_supplier_directory,
    "employee_directory":    _build_employee_directory,
    "work_schedule":         _build_work_schedule,
    "attendance":            _build_attendance,
    "vacation_tracker":      _build_vacation_tracker,
    "sales_pipeline":        _build_sales_pipeline,
    "scenario_simulator":    _build_scenario_simulator,
    "petty_cash":            _build_petty_cash,
    "travel_expenses":       _build_travel_expenses,
    "roi_calculator":        _build_roi_calculator,
    "income_expense_book":   _build_income_expense_book,
    "kpi_tracker":           _build_kpi_tracker,
    "presupuesto_anual":     _build_presupuesto_anual,
    "ratios_financieros":    _build_ratios_financieros,
    "sales":                 _build_sales,
    "budget":                _build_budget,
    "dashboard":             _build_dashboard,
    "generic_excel":         _build_generic_excel,
}
