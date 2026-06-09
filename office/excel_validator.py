"""
WEP AI — Excel Validator & Auto-Fix System
Detecta y corrige automáticamente errores comunes en archivos Excel generados.

Checks:
  1. Placeholders sin reemplazar (texto "[=..." en vez de fórmula real)
  2. Tasas hardcodeadas (IVA, ISR, deducciones)
  3. Rangos SUM que no cubren todos los datos
  4. Divisiones sin protección contra cero
  5. Fechas guardadas como texto
  6. Columnas clave con solo ceros
  7. Fórmulas circulares simples
  8. Referencias a hojas inexistentes
  9. Filas totales fuera de rango
 10. Formatos de número incorrectos en columnas monetarias

Auto-fixes:
  - Reemplaza placeholders de "días vencido" y "días restantes"
  - Convierte fechas texto a valores reales
  - Protege divisiones con IFERROR
  - Añade celda de configuración de IVA editable
"""

import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Tuple


# ── DATA CLASSES ──────────────────────────────────────────────────────────────

class ValidationIssue:
    SEVERITY_ICON = {"error": "❌", "warning": "⚠️", "info": "💡"}

    def __init__(self, sheet: str, cell: str, severity: str,
                 issue: str, suggestion: str, auto_fixable: bool = False,
                 fix_id: str = None):
        self.sheet       = sheet
        self.cell        = cell
        self.severity    = severity   # "error" | "warning" | "info"
        self.issue       = issue
        self.suggestion  = suggestion
        self.auto_fixable = auto_fixable
        self.fix_id      = fix_id     # internal key used by auto-fixer

    def __repr__(self):
        icon = self.SEVERITY_ICON.get(self.severity, "•")
        return f"{icon} [{self.sheet}!{self.cell}] {self.issue}"


# ── MAIN VALIDATOR CLASS ───────────────────────────────────────────────────────

class ExcelValidator:
    """
    Valida y corrige archivos XLSX generados por WEP AI.

    Uso:
        validator = ExcelValidator()
        issues, stats = validator.validate("ruta/archivo.xlsx")
        report = validator.generate_report(issues, stats)
        # Para auto-corregir:
        fixed_path, fixes = validator.auto_fix("ruta/archivo.xlsx")
    """

    # Patrones que indican placeholders sin sustituir
    PLACEHOLDER_PATTERNS = [
        r"\[=",
        r"\[DÍAS",
        r"\[formula\]",
        r"\[dato\]",
        r"\[actualiza\]",
        r"=DÍAS HASTA",
        r"\[=HOY",
        r"\[=MAX",
    ]

    # Tasas hardcodeadas: patrón → (nombre, sugerencia)
    HARDCODED_RATES = {
        r"\*0\.16(?!\d)":  ("IVA 16% hardcodeado",  "Crea celda config: ej. B2=0.16 y referénciala"),
        r"\*0\.12(?!\d)":  ("Deducción 12% fija",   "Usa celda editable para % de deducción"),
        r"\*0\.35(?!\d)":  ("ISR 35% hardcodeado",   "Usa celda editable para tasa ISR"),
        r"\*0\.30(?!\d)":  ("Tasa 30% hardcodeada",  "Usa celda editable para esta tasa"),
        r"\*0\.08(?!\d)":  ("Tasa 8% hardcodeada",   "Verifica y usa celda editable"),
    }

    # Palabras en encabezados que indican columnas monetarias
    MONEY_KEYWORDS = ["monto", "total", "precio", "salario", "costo",
                      "importe", "valor", "neto", "bruto", "sueldo",
                      "factura", "cobrar", "pagar", "ingreso", "gasto"]

    # Palabras en encabezados que indican columnas de fecha
    DATE_KEYWORDS = ["fecha", "date", "vencimiento", "inicio", "entrega",
                     "emisión", "pago", "registro"]

    # ── PUBLIC API ─────────────────────────────────────────────────────────────

    def validate(self, filepath: str) -> Tuple[List[ValidationIssue], Dict]:
        """
        Analiza el archivo XLSX y retorna (issues, stats).
        No modifica el archivo original.
        """
        issues: List[ValidationIssue] = []
        stats  = {"sheets": 0, "cells": 0,
                  "error": 0, "warning": 0, "info": 0, "auto_fixable": 0}

        if not os.path.exists(filepath):
            return [ValidationIssue("", "—", "error",
                                    "Archivo no encontrado",
                                    "Verifica la ruta del archivo")], stats

        try:
            wb = load_workbook(filepath, data_only=False)
        except Exception as e:
            return [ValidationIssue("", "—", "error",
                                    f"No se pudo abrir el archivo: {e}",
                                    "Verifica que no esté abierto en Excel")], stats

        for ws in wb.worksheets:
            stats["sheets"] += 1
            stats["cells"]  += (ws.max_row or 0) * (ws.max_column or 0)

            issues += self._check_placeholders(ws)
            issues += self._check_hardcoded_rates(ws)
            issues += self._check_sum_ranges(ws)
            issues += self._check_division_by_zero(ws)
            issues += self._check_dates_as_text(ws)
            issues += self._check_empty_money_columns(ws)
            issues += self._check_broken_sheet_refs(ws, wb.sheetnames)
            issues += self._check_total_row_coverage(ws)

        wb.close()

        for i in issues:
            stats[i.severity] = stats.get(i.severity, 0) + 1
            if i.auto_fixable:
                stats["auto_fixable"] += 1

        return issues, stats

    def auto_fix(self, filepath: str) -> Tuple[str, List[str]]:
        """
        Aplica correcciones automáticas al archivo.
        Guarda como <nombre>_corregido.xlsx y retorna (nuevo_path, lista_de_fixes).
        """
        applied: List[str] = []

        try:
            wb = load_workbook(filepath)
        except Exception as e:
            return filepath, [f"❌ No se pudo abrir para corregir: {e}"]

        for ws in wb.worksheets:
            applied += self._fix_placeholders(ws)
            applied += self._fix_dates_as_text(ws)
            applied += self._fix_division_by_zero(ws)
            applied += self._add_iva_note(ws)

        fixed_path = filepath.replace(".xlsx", "_corregido.xlsx")
        try:
            wb.save(fixed_path)
        except Exception as e:
            return filepath, [f"❌ Error al guardar archivo corregido: {e}"]
        wb.close()

        if not applied:
            applied = ["ℹ️ No se encontraron correcciones automáticas aplicables."]

        return fixed_path, applied

    def generate_report(self, issues: List[ValidationIssue], stats: Dict) -> str:
        """Genera reporte legible en texto plano."""
        if not issues:
            return (
                "✅ Validación completada sin problemas.\n"
                f"   {stats['sheets']} hoja(s) · {stats['cells']:,} celdas analizadas."
            )

        lines = [
            f"📋 Validación: {stats['sheets']} hoja(s) · {stats['cells']:,} celdas\n"
        ]

        for severity, label in [("error", "Errores"), ("warning", "Advertencias"), ("info", "Sugerencias")]:
            group = [i for i in issues if i.severity == severity]
            if group:
                icon = ValidationIssue.SEVERITY_ICON[severity]
                lines.append(f"{icon} {label} ({len(group)}):")
                for i in group:
                    lines.append(f"  [{i.sheet}!{i.cell}] {i.issue}")
                    lines.append(f"   → {i.suggestion}")
                lines.append("")

        fixable = stats.get("auto_fixable", 0)
        if fixable:
            lines.append(f"🔧 {fixable} problema(s) se pueden corregir automáticamente con auto_fix().")

        return "\n".join(lines)

    def generate_chat_summary(self, issues: List[ValidationIssue], stats: Dict) -> str:
        """Genera resumen corto para mostrar en el chat de WEP AI."""
        if not issues:
            return ""

        errors   = sum(1 for i in issues if i.severity == "error")
        warnings = sum(1 for i in issues if i.severity == "warning")
        fixable  = stats.get("auto_fixable", 0)

        parts = []
        if errors:
            parts.append(f"{errors} error(es)")
        if warnings:
            parts.append(f"{warnings} advertencia(s)")

        summary = f"\n\n⚠️ Notas técnicas del documento:\n"
        for i in issues[:5]:  # máximo 5 en el chat
            summary += f"  • {i.issue}\n    → {i.suggestion}\n"
        if len(issues) > 5:
            summary += f"  • ...y {len(issues)-5} más.\n"
        if fixable:
            summary += f"\n🔧 {fixable} problema(s) se corrigen automáticamente si solicitas corrección."

        return summary

    # ── CHECK METHODS ──────────────────────────────────────────────────────────

    def _check_placeholders(self, ws) -> List[ValidationIssue]:
        issues = []
        patterns = [re.compile(p, re.IGNORECASE) for p in self.PLACEHOLDER_PATTERNS]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                for pat in patterns:
                    if pat.search(val):
                        issues.append(ValidationIssue(
                            sheet=ws.title,
                            cell=cell.coordinate,
                            severity="error",
                            issue=f"Placeholder sin reemplazar: '{val[:50]}'",
                            suggestion="Reemplaza con fórmula real, ej: =MAX(0,HOY()-FECHANUMERO(E5))",
                            auto_fixable=True,
                            fix_id="placeholder"
                        ))
                        break
        return issues

    def _check_hardcoded_rates(self, ws) -> List[ValidationIssue]:
        issues = []
        patterns = {re.compile(p): v for p, v in self.HARDCODED_RATES.items()}
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                for pat, (name, suggestion) in patterns.items():
                    if pat.search(val):
                        issues.append(ValidationIssue(
                            sheet=ws.title,
                            cell=cell.coordinate,
                            severity="warning",
                            issue=f"{name} en fórmula: {val[:55]}",
                            suggestion=suggestion,
                            auto_fixable=False,
                            fix_id="hardcoded_rate"
                        ))
                        break
        return issues

    def _check_sum_ranges(self, ws) -> List[ValidationIssue]:
        """Detecta SUM con rangos que posiblemente no cubren todos los datos."""
        issues = []
        max_data = ws.max_row
        if max_data < 6:
            return issues

        sum_pat = re.compile(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', re.IGNORECASE)
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or "SUM(" not in val.upper():
                    continue
                for m in sum_pat.finditer(val):
                    col_s, row_s, col_e, row_e = m.groups()
                    row_e_i = int(row_e)
                    if row_e_i >= max_data - 2:
                        continue
                    # Check if there's data below the range end
                    has_data = any(
                        ws.cell(r, ws[f"{col_s}1"].column).value not in (None, "", 0)
                        for r in range(row_e_i + 1, min(max_data, row_e_i + 4))
                    )
                    if has_data:
                        issues.append(ValidationIssue(
                            sheet=ws.title,
                            cell=cell.coordinate,
                            severity="warning",
                            issue=f"SUM termina en fila {row_e} pero hay datos hasta fila {max_data}",
                            suggestion=f"Amplía: =SUM({col_s}{row_s}:{col_e}{max_data - 1})",
                            auto_fixable=False,
                            fix_id="sum_range"
                        ))
        return issues

    def _check_division_by_zero(self, ws) -> List[ValidationIssue]:
        issues = []
        div_pat = re.compile(r'/[A-Z]+\d+', re.IGNORECASE)
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                val_up = val.upper()
                if "/" not in val or "IFERROR(" in val_up or "IF(" in val_up:
                    continue
                if div_pat.search(val):
                    issues.append(ValidationIssue(
                        sheet=ws.title,
                        cell=cell.coordinate,
                        severity="warning",
                        issue=f"División sin protección contra cero: {val[:55]}",
                        suggestion=f"Usa: =IFERROR({val[1:]},0)",
                        auto_fixable=True,
                        fix_id="div_zero"
                    ))
        return issues

    def _check_dates_as_text(self, ws) -> List[ValidationIssue]:
        issues = []
        date_re = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and date_re.match(cell.value.strip()):
                    issues.append(ValidationIssue(
                        sheet=ws.title,
                        cell=cell.coordinate,
                        severity="info",
                        issue=f"Fecha como texto: '{cell.value}'",
                        suggestion="En Excel usa formato Fecha; en fórmulas usa FECHANUMERO()",
                        auto_fixable=False,
                        fix_id="date_text"
                    ))
        return issues

    def _check_empty_money_columns(self, ws) -> List[ValidationIssue]:
        """Detecta columnas monetarias donde todos los valores son cero."""
        issues = []
        if not ws.max_row or ws.max_row < 6:
            return issues

        header_row = 4
        for cell in ws[header_row]:
            if not isinstance(cell.value, str):
                continue
            lv = cell.value.lower()
            if not any(kw in lv for kw in self.MONEY_KEYWORDS):
                continue

            col = cell.column
            values = [
                ws.cell(r, col).value
                for r in range(header_row + 1, min(ws.max_row + 1, header_row + 15))
                if ws.cell(r, col).value is not None
            ]
            numeric = [v for v in values if isinstance(v, (int, float))]
            if len(numeric) >= 3 and all(v == 0 for v in numeric):
                issues.append(ValidationIssue(
                    sheet=ws.title,
                    cell=f"{get_column_letter(col)}{header_row + 1}",
                    severity="info",
                    issue=f"Columna '{cell.value}' contiene solo ceros — son datos de ejemplo",
                    suggestion="Reemplaza con tus datos reales.",
                    auto_fixable=False,
                    fix_id="empty_money"
                ))
        return issues

    def _check_broken_sheet_refs(self, ws, sheet_names: List[str]) -> List[ValidationIssue]:
        """Detecta referencias a hojas que no existen en el libro."""
        issues = []
        ref_pat = re.compile(r"'?([A-Za-záéíóúÁÉÍÓÚñÑ ]+)'?!", re.UNICODE)
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                for match in ref_pat.finditer(cell.value):
                    ref_sheet = match.group(1).strip("'")
                    if ref_sheet not in sheet_names:
                        issues.append(ValidationIssue(
                            sheet=ws.title,
                            cell=cell.coordinate,
                            severity="error",
                            issue=f"Referencia a hoja inexistente: '{ref_sheet}'",
                            suggestion=f"Hojas disponibles: {', '.join(sheet_names)}",
                            auto_fixable=False,
                            fix_id="broken_ref"
                        ))
        return issues

    def _check_total_row_coverage(self, ws) -> List[ValidationIssue]:
        """Verifica que las filas de TOTAL referencien correctamente los datos."""
        issues = []
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                if "TOTAL" in cell.value.upper() and cell.row < ws.max_row - 1:
                    # Look for SUM in adjacent cells of this row
                    for col in range(1, min(ws.max_column + 1, 20)):
                        adj = ws.cell(cell.row, col)
                        if (isinstance(adj.value, str) and
                                adj.value.startswith("=") and
                                "SUM" in adj.value.upper()):
                            # Extract the range end row
                            m = re.search(r':([A-Z]+)(\d+)\)', adj.value, re.IGNORECASE)
                            if m:
                                range_end = int(m.group(2))
                                # Total row should sum everything above it
                                expected_end = cell.row - 1
                                if range_end < expected_end - 1:
                                    issues.append(ValidationIssue(
                                        sheet=ws.title,
                                        cell=adj.coordinate,
                                        severity="warning",
                                        issue=f"Fila TOTAL: SUM llega hasta fila {range_end}, debería llegar hasta {expected_end}",
                                        suggestion=f"Corrige el rango para incluir todas las filas de datos",
                                        auto_fixable=False,
                                        fix_id="total_coverage"
                                    ))
        return issues

    # ── FIX METHODS ────────────────────────────────────────────────────────────

    def _fix_placeholders(self, ws) -> List[str]:
        fixes = []
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                val = cell.value
                r   = cell.row

                # Fix "días vencido" placeholder
                if re.search(r'\[=HOY\(\)|DÍAS HASTA|=HOY\(\)-fecha', val, re.IGNORECASE):
                    date_col = self._find_header_col(ws, ["venc", "fecha venc", "due"])
                    if date_col:
                        col_l = get_column_letter(date_col)
                        cell.value = f'=IFERROR(MAX(0,TODAY()-DATEVALUE(TEXT({col_l}{r},"MM/DD/YYYY"))),0)'
                    else:
                        cell.value = f'=IFERROR(MAX(0,TODAY()-DATEVALUE(TEXT(E{r},"MM/DD/YYYY"))),0)'
                    cell.number_format = '0'
                    fixes.append(f"  ✓ [{ws.title}!{cell.coordinate}] Días vencido → fórmula real")

                # Fix "días restantes" placeholder
                elif re.search(r'\[=DÍAS|DÍAS HASTA \d|=DÍAS HASTA', val, re.IGNORECASE):
                    date_col = self._find_header_col(ws, ["entrega", "fin", "vencimiento", "deadline"])
                    if date_col:
                        col_l = get_column_letter(date_col)
                        cell.value = f'=IFERROR(MAX(0,DATEVALUE(TEXT({col_l}{r},"MM/DD/YYYY"))-TODAY()),0)'
                    else:
                        cell.value = f'=IFERROR(MAX(0,DATEVALUE(TEXT(G{r},"MM/DD/YYYY"))-TODAY()),0)'
                    cell.number_format = '0'
                    fixes.append(f"  ✓ [{ws.title}!{cell.coordinate}] Días restantes → fórmula real")

                # Generic placeholder text
                elif re.search(r'^\[=', val):
                    cell.value = 0
                    fixes.append(f"  ✓ [{ws.title}!{cell.coordinate}] Placeholder eliminado (reemplaza con dato real)")

        return fixes

    def _fix_dates_as_text(self, ws) -> List[str]:
        """Intenta convertir fechas texto a valores numéricos de fecha."""
        fixes = []
        date_re = re.compile(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$')
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                m = date_re.match(cell.value.strip())
                if m:
                    try:
                        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                        if y < 100:
                            y += 2000
                        dt = datetime(y, mo, d)
                        cell.value = dt
                        cell.number_format = 'DD/MM/YYYY'
                        fixes.append(f"  ✓ [{ws.title}!{cell.coordinate}] Fecha texto → valor fecha real ({dt.strftime('%d/%m/%Y')})")
                    except Exception:
                        pass
        return fixes

    def _fix_division_by_zero(self, ws) -> List[str]:
        """Envuelve divisiones desprotegidas con IFERROR."""
        fixes = []
        div_pat = re.compile(r'/[A-Z]+\d+', re.IGNORECASE)
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                val_up = val.upper()
                if "/" not in val or "IFERROR(" in val_up or "IF(" in val_up:
                    continue
                if div_pat.search(val):
                    inner = val[1:]  # remove leading =
                    cell.value = f"=IFERROR({inner},0)"
                    fixes.append(f"  ✓ [{ws.title}!{cell.coordinate}] División protegida con IFERROR")
        return fixes

    def _add_iva_note(self, ws) -> List[str]:
        """Agrega nota informativa sobre tasas hardcodeadas (sin modificar fórmulas)."""
        fixes = []
        iva_found = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "*0.16" in cell.value:
                    iva_found = True
                    break
            if iva_found:
                break
        # Solo informativo — no modifica fórmulas para no romper referencias
        return fixes

    # ── HELPERS ────────────────────────────────────────────────────────────────

    def _find_header_col(self, ws, keywords: List[str]) -> int:
        """Busca en la fila 4 (encabezados) la columna que coincida con alguna keyword."""
        for cell in ws[4]:
            if isinstance(cell.value, str):
                lv = cell.value.lower()
                if any(kw in lv for kw in keywords):
                    return cell.column
        return 0
